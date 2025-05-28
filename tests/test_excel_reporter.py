"""
Test Suite for Excel Report Generator

This module contains comprehensive tests for the Excel Reporter functionality.

Author: Laser Trim AI System
Date: 2024
Version: 1.0.0
"""

import unittest
import os
import tempfile
import shutil
from datetime import datetime
import pandas as pd
import openpyxl
from pathlib import Path

# Import the module to test
from excel_reporter import ExcelReporter


class TestExcelReporter(unittest.TestCase):
    """Test cases for the Excel Reporter."""

    def setUp(self):
        """Set up test fixtures."""
        self.test_dir = tempfile.mkdtemp()
        self.reporter = ExcelReporter()

        # Create sample test data
        self.sample_results = {
            "file_results": [
                {
                    "filename": "8340_SN001.xlsx",
                    "model": "8340",
                    "serial": "SN001",
                    "system": "A",
                    "tracks": {
                        "TRK1": {
                            "status": "Pass",
                            "sigma_gradient": 0.0025,
                            "sigma_threshold": 0.004,
                            "sigma_pass": True,
                            "linearity_pass": True,
                            "failure_probability": 0.05,
                            "risk_category": "Low",
                            "resistance_change_percent": 2.5,
                            "trim_improvement_percent": 15.3,
                            "unit_length": 150.2,
                            "optimal_offset": 0.001
                        },
                        "TRK2": {
                            "status": "Fail",
                            "sigma_gradient": 0.0045,
                            "sigma_threshold": 0.004,
                            "sigma_pass": False,
                            "linearity_pass": True,
                            "failure_probability": 0.75,
                            "risk_category": "High",
                            "resistance_change_percent": 5.2,
                            "trim_improvement_percent": 8.7,
                            "unit_length": 149.8,
                            "optimal_offset": 0.002
                        }
                    }
                },
                {
                    "filename": "8555_SN002.xlsx",
                    "model": "8555",
                    "serial": "SN002",
                    "system": "B",
                    "status": "Pass",
                    "sigma_gradient": 0.0018,
                    "sigma_threshold": 0.003,
                    "sigma_pass": True,
                    "linearity_pass": True,
                    "failure_probability": 0.03,
                    "risk_category": "Low",
                    "resistance_change_percent": 1.8,
                    "trim_improvement_percent": 22.5,
                    "unit_length": 180.5,
                    "optimal_offset": 0.0005
                }
            ],
            "ml_predictions": {
                "next_batch_pass_rate": 0.87,
                "maintenance_due_in_days": 12,
                "quality_trend": "stable"
            }
        }

    def tearDown(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.test_dir)

    def test_report_generation(self):
        """Test basic report generation."""
        output_path = os.path.join(self.test_dir, "test_report.xlsx")

        result_path = self.reporter.generate_report(
            self.sample_results,
            output_path,
            include_ai_insights=False
        )

        # Verify file was created
        self.assertTrue(os.path.exists(result_path))
        self.assertEqual(result_path, output_path)

        # Verify it's a valid Excel file
        wb = openpyxl.load_workbook(result_path)
        self.assertIsNotNone(wb)

        # Check expected sheets exist
        expected_sheets = [
            'Executive Summary',
            'Detailed Analysis',
            'Statistical Summary',
            'Trend Analysis',
            'Quality Metrics',
            'Recommendations',
            'Raw Data'
        ]

        for sheet_name in expected_sheets:
            self.assertIn(sheet_name, wb.sheetnames)

        wb.close()

    def test_executive_summary_content(self):
        """Test executive summary sheet content."""
        output_path = os.path.join(self.test_dir, "test_report.xlsx")
        self.reporter.generate_report(self.sample_results, output_path, False)

        # Load workbook and check executive summary
        wb = openpyxl.load_workbook(output_path)
        ws = wb['Executive Summary']

        # Check title
        self.assertEqual(ws['A1'].value, 'LASER TRIM ANALYSIS REPORT')

        # Check that report date is present
        report_date_cell = ws['B3'].value
        self.assertIsNotNone(report_date_cell)

        # Check total files analyzed
        total_files = ws['B4'].value
        self.assertEqual(total_files, 2)  # We have 2 files in sample data

        wb.close()

    def test_detailed_analysis_content(self):
        """Test detailed analysis sheet content."""
        output_path = os.path.join(self.test_dir, "test_report.xlsx")
        self.reporter.generate_report(self.sample_results, output_path, False)

        # Load workbook
        wb = openpyxl.load_workbook(output_path)
        ws = wb['Detailed Analysis']

        # Check headers
        headers = [cell.value for cell in ws[3]]  # Row 3 has headers
        expected_headers = [
            'File Name', 'Model', 'Serial', 'Track', 'Status',
            'Sigma Gradient', 'Sigma Threshold', 'Sigma Pass',
            'Linearity Pass', 'Failure Probability', 'Risk Category',
            'Resistance Change (%)', 'Trim Improvement (%)',
            'Unit Length', 'Optimal Offset'
        ]

        for header in expected_headers:
            self.assertIn(header, headers)

        # Check data rows - should have 3 rows (2 tracks from first file + 1 from second)
        data_rows = 0
        for row in ws.iter_rows(min_row=4):  # Start after headers
            if row[0].value:  # Check if filename is present
                data_rows += 1

        self.assertEqual(data_rows, 3)

        wb.close()

    def test_statistical_summary(self):
        """Test statistical summary calculations."""
        output_path = os.path.join(self.test_dir, "test_report.xlsx")
        self.reporter.generate_report(self.sample_results, output_path, False)

        wb = openpyxl.load_workbook(output_path)
        ws = wb['Statistical Summary']

        # Find sigma gradient statistics
        for row in ws.iter_rows(min_row=4):
            if row[0].value == 'Sigma Gradient':
                # Check that statistics are calculated
                mean_val = row[1].value
                std_val = row[2].value
                min_val = row[3].value
                max_val = row[4].value

                self.assertIsNotNone(mean_val)
                self.assertIsNotNone(std_val)
                self.assertIsNotNone(min_val)
                self.assertIsNotNone(max_val)

                # Verify min <= mean <= max
                self.assertLessEqual(min_val, mean_val)
                self.assertLessEqual(mean_val, max_val)
                break

        wb.close()

    def test_quality_metrics(self):
        """Test quality metrics calculations."""
        output_path = os.path.join(self.test_dir, "test_report.xlsx")
        self.reporter.generate_report(self.sample_results, output_path, False)

        wb = openpyxl.load_workbook(output_path)
        ws = wb['Quality Metrics']

        # Check that KPIs are present
        kpi_found = False
        for row in ws.iter_rows():
            if row[0].value == 'First Pass Yield':
                kpi_found = True
                fpy_value = row[1].value
                self.assertIsNotNone(fpy_value)
                self.assertGreaterEqual(fpy_value, 0)
                self.assertLessEqual(fpy_value, 1)
                break

        self.assertTrue(kpi_found, "First Pass Yield KPI not found")

        wb.close()

    def test_recommendations_generation(self):
        """Test recommendations generation."""
        output_path = os.path.join(self.test_dir, "test_report.xlsx")
        self.reporter.generate_report(self.sample_results, output_path, False)

        wb = openpyxl.load_workbook(output_path)
        ws = wb['Recommendations']

        # Check headers
        headers = [cell.value for cell in ws[3]]
        expected_headers = ['Priority', 'Category', 'Recommendation', 'Expected Impact']

        for header in expected_headers:
            self.assertIn(header, headers)

        # Check that at least one recommendation exists
        has_recommendations = False
        for row in ws.iter_rows(min_row=4):
            if row[0].value:  # Priority column
                has_recommendations = True
                break

        self.assertTrue(has_recommendations, "No recommendations found")

        wb.close()

    def test_raw_data_export(self):
        """Test raw data export functionality."""
        output_path = os.path.join(self.test_dir, "test_report.xlsx")
        self.reporter.generate_report(self.sample_results, output_path, False)

        wb = openpyxl.load_workbook(output_path)
        ws = wb['Raw Data']

        # Check that data is present
        has_data = False
        for row in ws.iter_rows(min_row=2):  # Skip header
            if row[0].value:
                has_data = True
                break

        self.assertTrue(has_data, "No data in raw data sheet")

        wb.close()

    def test_overall_metrics_calculation(self):
        """Test overall metrics calculation."""
        metrics = self.reporter._calculate_overall_metrics(self.sample_results)

        # Check calculated metrics
        self.assertIn('pass_rate', metrics)
        self.assertIn('avg_sigma_gradient', metrics)
        self.assertIn('avg_failure_probability', metrics)
        self.assertIn('high_risk_count', metrics)
        self.assertIn('attention_required', metrics)

        # Verify pass rate calculation
        # We have 2 passes out of 3 tracks
        expected_pass_rate = 2 / 3
        self.assertAlmostEqual(metrics['pass_rate'], expected_pass_rate, places=2)

        # Verify high risk count
        self.assertEqual(metrics['high_risk_count'], 1)  # TRK2 is high risk

    def test_model_statistics_calculation(self):
        """Test model statistics calculation."""
        model_stats = self.reporter._calculate_model_statistics(self.sample_results)

        # Check that both models are present
        self.assertIn('8340', model_stats)
        self.assertIn('8555', model_stats)

        # Check 8340 statistics
        stats_8340 = model_stats['8340']
        self.assertEqual(stats_8340['count'], 2)  # 2 tracks
        self.assertEqual(stats_8340['pass_rate'], 0.5)  # 1 pass, 1 fail

        # Check 8555 statistics
        stats_8555 = model_stats['8555']
        self.assertEqual(stats_8555['count'], 1)
        self.assertEqual(stats_8555['pass_rate'], 1.0)  # All pass

    def test_empty_results_handling(self):
        """Test handling of empty results."""
        empty_results = {"file_results": []}
        output_path = os.path.join(self.test_dir, "empty_report.xlsx")

        # Should not raise exception
        result_path = self.reporter.generate_report(
            empty_results,
            output_path,
            include_ai_insights=False
        )

        self.assertTrue(os.path.exists(result_path))

    def test_missing_data_handling(self):
        """Test handling of missing data fields."""
        incomplete_results = {
            "file_results": [
                {
                    "filename": "test.xlsx",
                    "model": "TEST",
                    "serial": "001",
                    "status": "Pass"
                    # Missing most fields
                }
            ]
        }

        output_path = os.path.join(self.test_dir, "incomplete_report.xlsx")

        # Should handle gracefully
        result_path = self.reporter.generate_report(
            incomplete_results,
            output_path,
            include_ai_insights=False
        )

        self.assertTrue(os.path.exists(result_path))

    def test_format_creation(self):
        """Test format creation for Excel styling."""
        from xlsxwriter import Workbook

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook(tmp.name)
            formats = self.reporter._create_formats(wb)

            # Check all expected formats are created
            expected_formats = [
                'title', 'header', 'subheader', 'data', 'number',
                'percent', 'pass', 'fail', 'warning', 'insight'
            ]

            for fmt_name in expected_formats:
                self.assertIn(fmt_name, formats)
                self.assertIsNotNone(formats[fmt_name])

            wb.close()
            os.unlink(tmp.name)

    def test_kpi_calculation(self):
        """Test KPI calculations."""
        kpis = self.reporter._calculate_kpis(self.sample_results)

        # Check all KPIs are present
        expected_kpis = [
            'cpk_sigma', 'ppk_overall', 'first_pass_yield', 'rty',
            'dpm', 'mtbf', 'copq', 'oee', 'in_control_percent',
            'special_causes', 'stability_index', 'improvement_opportunity'
        ]

        for kpi in expected_kpis:
            self.assertIn(kpi, kpis)

        # Verify first pass yield
        # 2 passes out of 3 tracks
        self.assertAlmostEqual(kpis['first_pass_yield'], 2 / 3, places=2)

        # Verify DPM calculation
        expected_dpm = (1 - 2 / 3) * 1_000_000
        self.assertAlmostEqual(kpis['dpm'], expected_dpm, delta=1)

    def test_recommendations_logic(self):
        """Test recommendation generation logic."""
        recommendations = self.reporter._generate_recommendations(self.sample_results)

        # Should have recommendations
        self.assertGreater(len(recommendations), 0)

        # Check recommendation structure
        for rec in recommendations:
            self.assertIn('priority', rec)
            self.assertIn('category', rec)
            self.assertIn('recommendation', rec)
            self.assertIn('impact', rec)

            # Check priority is valid
            self.assertIn(rec['priority'], ['Critical', 'High', 'Medium', 'Low'])

        # Check that recommendations are sorted by priority
        priority_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}
        for i in range(1, len(recommendations)):
            prev_priority = priority_order.get(recommendations[i - 1]['priority'], 4)
            curr_priority = priority_order.get(recommendations[i]['priority'], 4)
            self.assertLessEqual(prev_priority, curr_priority)

    def test_dataframe_conversion(self):
        """Test conversion to pandas DataFrame."""
        df = self.reporter._results_to_dataframe(self.sample_results)

        # Check DataFrame structure
        self.assertEqual(len(df), 3)  # 3 total tracks/rows

        # Check columns exist
        expected_columns = [
            'filename', 'model', 'serial', 'track_id',
            'sigma_gradient', 'sigma_pass'
        ]

        for col in expected_columns:
            self.assertIn(col, df.columns)

        # Check data integrity
        models = df['model'].unique()
        self.assertIn('8340', models)
        self.assertIn('8555', models)


class TestExcelReporterIntegration(unittest.TestCase):
    """Integration tests for Excel Reporter with other components."""

    def setUp(self):
        """Set up integration test fixtures."""
        self.test_dir = tempfile.mkdtemp()
        self.reporter = ExcelReporter()

    def tearDown(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.test_dir)

    def test_large_dataset_handling(self):
        """Test handling of large datasets."""
        # Generate large dataset
        large_results = {"file_results": []}

        for i in range(100):
            large_results["file_results"].append({
                "filename": f"test_{i}.xlsx",
                "model": f"MODEL_{i % 5}",
                "serial": f"SN{i:04d}",
                "status": "Pass" if i % 3 != 0 else "Fail",
                "sigma_gradient": 0.001 + (i % 10) * 0.0005,
                "sigma_threshold": 0.004,
                "sigma_pass": i % 3 != 0,
                "failure_probability": (i % 10) / 10,
                "risk_category": "Low" if i % 10 < 3 else "Medium" if i % 10 < 7 else "High"
            })

        output_path = os.path.join(self.test_dir, "large_report.xlsx")

        # Should handle without issues
        result_path = self.reporter.generate_report(
            large_results,
            output_path,
            include_ai_insights=False
        )

        self.assertTrue(os.path.exists(result_path))

        # Verify file size is reasonable
        file_size = os.path.getsize(result_path)
        self.assertGreater(file_size, 50000)  # Should be at least 50KB
        self.assertLess(file_size, 10000000)  # But not more than 10MB

    def test_excel_file_validity(self):
        """Test that generated Excel files are valid and readable."""
        output_path = os.path.join(self.test_dir, "valid_test.xlsx")
        self.reporter.generate_report(
            self.sample_results,
            output_path,
            include_ai_insights=False
        )

        # Try to read with pandas
        try:
            df = pd.read_excel(output_path, sheet_name='Raw Data')
            self.assertIsNotNone(df)
            self.assertGreater(len(df), 0)
        except Exception as e:
            self.fail(f"Failed to read Excel file with pandas: {str(e)}")


def run_tests():
    """Run all tests."""
    # Create test suite
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    # Add test cases
    suite.addTests(loader.loadTestsFromTestCase(TestExcelReporter))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelReporterIntegration))

    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    return result.wasSuccessful()


if __name__ == "__main__":
    # Run tests
    success = run_tests()

    if success:
        print("\n✅ All tests passed!")

        # Generate a sample report for demonstration
        print("\nGenerating sample report for demonstration...")
        from excel_reporter import generate_sample_report

        generate_sample_report()
        print("Sample report generated: laser_trim_analysis_report.xlsx")
    else:
        print("\n❌ Some tests failed. Please check the output above.")
        exit(1)