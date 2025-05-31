"""
ML Integration Example

This script shows how to integrate the MLPredictor into your existing
laser trim analyzer for real-time predictions and alerts.
"""

import os
import sys
from datetime import datetime

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from laser_trim_analyzer.processor.processor_module import DataDrivenLaserProcessor
from ml_predictor import MLPredictor, integrate_ml_predictor
from ml_model_trainer import MLModelTrainer


def setup_ml_models():
    """Set up ML models if they don't exist."""
    db_path = os.path.join(os.path.expanduser("~"), "LaserTrimResults", "analysis_history.db")
    model_dir = os.path.join(os.path.expanduser("~"), "LaserTrimResults", "ml_models")

    # Check if models exist
    models_exist = all(
        os.path.exists(os.path.join(model_dir, f"{model}.pkl"))
        for model in ['failure_predictor', 'anomaly_detector']
    )

    if not models_exist:
        print("ML models not found. Training models...")
        if os.path.exists(db_path):
            trainer = MLModelTrainer(db_path, model_dir)
            trainer.train_all_models(days_back=365)
        else:
            print("Database not found. Cannot train models.")
            return False

    return True


def process_with_ml_predictions():
    """Example of processing files with ML predictions."""

    # Set up paths
    input_folder = input("Enter input folder path: ")
    output_folder = os.path.join(os.path.expanduser("~"), "LaserTrimResults", "ML_Analysis")
    model_dir = os.path.join(os.path.expanduser("~"), "LaserTrimResults", "ml_models")

    # Ensure ML models are available
    if not setup_ml_models():
        print("Cannot proceed without ML models.")
        return

    # Create output directory
    os.makedirs(output_folder, exist_ok=True)
    output_dir = os.path.join(output_folder, datetime.now().strftime("%Y%m%d_%H%M%S"))

    # Initialize processor
    print("\nInitializing processor with ML capabilities...")
    processor = DataDrivenLaserProcessor(
        output_dir=output_dir,
        generate_unit_plots=True,  # Can still generate plots
        enable_database=True
    )

    # Integrate ML predictor
    integrate_ml_predictor(processor, model_dir)

    # Process folder
    print(f"\nProcessing folder: {input_folder}")
    print("ML predictions will be generated in real-time for each file.\n")

    # Custom callback to show ML alerts
    def ml_progress_callback(current, total, filename):
        print(f"\n[{current}/{total}] Processing: {filename}")

    # Process files
    results = processor.process_folder(
        input_folder,
        callback=ml_progress_callback
    )

    # Analyze ML predictions
    print("\n" + "=" * 60)
    print("ML ANALYSIS SUMMARY")
    print("=" * 60)

    total_files = len(results)
    anomalies_detected = 0
    high_risk_units = 0
    threshold_adjustments = 0

    # Collect all warnings and recommendations
    all_warnings = []
    all_recommendations = []

    for result in results:
        if 'ml_predictions' in result:
            predictions = result['ml_predictions']

            # Count anomalies
            if predictions.get('is_anomaly', False):
                anomalies_detected += 1

            # Count high risk
            if predictions.get('failure_probability', 0) > 0.7:
                high_risk_units += 1

            # Count threshold adjustments
            if predictions.get('suggested_threshold'):
                current = result.get('Sigma Threshold', 0)
                suggested = predictions['suggested_threshold']
                if current > 0 and abs(suggested - current) / current > 0.1:
                    threshold_adjustments += 1

        # Collect warnings
        if 'ml_warnings' in result:
            for warning in result['ml_warnings']:
                all_warnings.append({
                    'file': result['File'],
                    'warning': warning
                })

        # Collect recommendations
        if 'ml_recommendations' in result:
            for rec in result['ml_recommendations']:
                all_recommendations.append({
                    'file': result['File'],
                    'recommendation': rec
                })

    # Print summary
    print(f"\nTotal files processed: {total_files}")
    print(f"Anomalies detected: {anomalies_detected} ({anomalies_detected / total_files * 100:.1f}%)")
    print(f"High risk units: {high_risk_units} ({high_risk_units / total_files * 100:.1f}%)")
    print(f"Threshold adjustments needed: {threshold_adjustments}")

    # Print top warnings
    if all_warnings:
        print(f"\n⚠️  TOP ML WARNINGS ({len(all_warnings)} total):")
        # Group by warning type
        warning_types = {}
        for w in all_warnings:
            wtype = w['warning']['type']
            if wtype not in warning_types:
                warning_types[wtype] = []
            warning_types[wtype].append(w)

        for wtype, warnings in warning_types.items():
            print(f"\n  {wtype}: {len(warnings)} occurrences")
            # Show first 3 examples
            for w in warnings[:3]:
                print(f"    - {w['file']}: {w['warning']['message']}")

    # Print recommendations
    if all_recommendations:
        print(f"\n💡 ML RECOMMENDATIONS ({len(all_recommendations)} total):")
        # Show unique recommendations
        unique_recs = {}
        for r in all_recommendations:
            rec_text = r['recommendation']
            if rec_text not in unique_recs:
                unique_recs[rec_text] = []
            unique_recs[rec_text].append(r['file'])

        for rec_text, files in unique_recs.items():
            print(f"\n  {rec_text}")
            print(f"    Applies to {len(files)} files: {', '.join(files[:5])}")
            if len(files) > 5:
                print(f"    ... and {len(files) - 5} more")

    # Get ML performance stats
    if hasattr(processor, 'ml_predictor'):
        stats = processor.ml_predictor.get_performance_stats()
        print(f"\n📊 ML PERFORMANCE:")
        print(f"  Average prediction time: {stats['average_prediction_time_ms']:.1f} ms")
        print(f"  Cache hit rate: {stats['cache_hit_rate'] * 100:.1f}%")
        print(f"  Models loaded: {', '.join(stats['models_loaded'])}")

    # Generate enhanced Excel report with ML insights
    generate_ml_enhanced_report(results, output_dir)

    print(f"\n✅ Analysis complete! Results saved to: {output_dir}")


def generate_ml_enhanced_report(results, output_dir):
    """Generate an Excel report with ML predictions included."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    # Flatten results including ML predictions
    report_data = []

    for result in results:
        row = {
            'File': result.get('File'),
            'Model': result.get('Model'),
            'Serial': result.get('Serial'),
            'Overall Status': result.get('Overall Status'),
            'Sigma Gradient': result.get('Sigma Gradient'),
            'Sigma Threshold': result.get('Sigma Threshold'),
            'Sigma Pass': result.get('Sigma Pass'),
            'ML Quality Score': result.get('ml_quality_score', 'N/A')
        }

        # Add ML predictions
        if 'ml_predictions' in result:
            predictions = result['ml_predictions']
            row['ML Failure Probability'] = predictions.get('failure_probability', 'N/A')
            row['ML Anomaly Detected'] = predictions.get('is_anomaly', 'N/A')
            row['ML Suggested Threshold'] = predictions.get('suggested_threshold', 'N/A')

        # Add first warning if any
        if 'ml_warnings' in result and result['ml_warnings']:
            row['ML Warning'] = result['ml_warnings'][0]['message']
        else:
            row['ML Warning'] = ''

        # Add first recommendation if any
        if 'ml_recommendations' in result and result['ml_recommendations']:
            row['ML Recommendation'] = result['ml_recommendations'][0]
        else:
            row['ML Recommendation'] = ''

        report_data.append(row)

    # Create DataFrame and save to Excel
    df = pd.DataFrame(report_data)
    excel_path = os.path.join(output_dir, 'ML_Enhanced_Report.xlsx')

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='ML Analysis', index=False)

        # Format the Excel file
        workbook = writer.book
        worksheet = writer.sheets['ML Analysis']

        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Conditional formatting for ML Quality Score
        score_col = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'ML Quality Score':
                score_col = idx
                break

        if score_col:
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=score_col)
                try:
                    score = float(cell.value)
                    if score >= 80:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif score >= 60:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                except:
                    pass

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    print(f"\nML-enhanced Excel report saved to: {excel_path}")


def demonstrate_real_time_alerts():
    """Demonstrate real-time ML alerts during processing."""
    print("\n" + "=" * 60)
    print("REAL-TIME ML ALERTS DEMONSTRATION")
    print("=" * 60)

    # Sample data that would trigger alerts
    test_cases = [
        {
            'name': 'Normal Unit',
            'data': {
                'File': 'test_normal.xlsx',
                'Model': '8340',
                'Serial': 'N001',
                'Sigma Gradient': 0.015,
                'Sigma Threshold': 0.04,
                'Failure Probability': 0.1,
                'Sigma Pass': True
            }
        },
        {
            'name': 'High Risk Unit',
            'data': {
                'File': 'test_high_risk.xlsx',
                'Model': '8340',
                'Serial': 'H001',
                'Sigma Gradient': 0.038,
                'Sigma Threshold': 0.04,
                'Failure Probability': 0.85,
                'Sigma Pass': True
            }
        },
        {
            'name': 'Anomaly Unit',
            'data': {
                'File': 'test_anomaly.xlsx',
                'Model': '8555',
                'Serial': 'A001',
                'Sigma Gradient': 0.001,  # Unusually low
                'Sigma Threshold': 0.0025,
                'Unit Length': 500,  # Unusually high
                'Failure Probability': 0.3,
                'Sigma Pass': True
            }
        }
    ]

    # Initialize predictor (mock)
    model_dir = os.path.join(os.path.expanduser("~"), "LaserTrimResults", "ml_models")

    # Mock predictor for demonstration
    print("\nSimulating real-time ML predictions during file processing...\n")

    for test_case in test_cases:
        print(f"\n{'=' * 40}")
        print(f"Processing: {test_case['name']} - {test_case['data']['File']}")
        print(f"{'=' * 40}")

        # Simulate processing
        print(f"Sigma Gradient: {test_case['data']['Sigma Gradient']:.4f}")
        print(f"Sigma Threshold: {test_case['data']['Sigma Threshold']:.4f}")

        # Simulate ML predictions
        if test_case['name'] == 'High Risk Unit':
            print("\n🚨 ML ALERT: HIGH FAILURE RISK DETECTED!")
            print("   Failure Probability: 85%")
            print("   Recommendation: Immediate inspection required")
            print("   Action: Unit flagged for quality review")

        elif test_case['name'] == 'Anomaly Unit':
            print("\n⚠️  ML ALERT: ANOMALY DETECTED!")
            print("   Anomaly Score: -0.245 (outlier)")
            print("   Unusual Features: Very low sigma gradient, high unit length")
            print("   Recommendation: Verify measurement accuracy")

        elif test_case['name'] == 'Normal Unit':
            print("\n✅ ML Assessment: Unit within normal parameters")
            print("   Quality Score: 92/100")
            print("   No alerts or warnings")

        print("\nContinuing to next file...")


if __name__ == "__main__":
    print("ML INTEGRATION FOR LASER TRIM ANALYZER")
    print("=====================================\n")
    print("This demonstrates real-time ML predictions during analysis.\n")

    print("Options:")
    print("1. Process folder with ML predictions")
    print("2. Demonstrate real-time alerts")
    print("3. Train/update ML models")
    print("4. Exit")

    choice = input("\nSelect option (1-4): ")

    if choice == '1':
        process_with_ml_predictions()
    elif choice == '2':
        demonstrate_real_time_alerts()
    elif choice == '3':
        db_path = os.path.join(os.path.expanduser("~"), "LaserTrimResults", "analysis_history.db")
        model_dir = os.path.join(os.path.expanduser("~"), "LaserTrimResults", "ml_models")

        if os.path.exists(db_path):
            trainer = MLModelTrainer(db_path, model_dir)
            trainer.train_all_models()
        else:
            print("Database not found. Please run some analyses first.")
    elif choice == '4':
        print("Exiting...")
    else:
        print("Invalid option")