"""
Database Manager for Laser Trim Analyzer v3.

Simplified from v2's 2,900+ line manager to ~600 lines.
Focuses on essential operations with clean session management.

Operations:
- Save/retrieve analysis results
- Incremental processing tracking
- Historical data queries
- Model statistics
- QA alerts management
"""

import json
import logging
import re
import threading
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any, Union, Iterator, Tuple
from contextlib import contextmanager

from sqlalchemy import create_engine, exists, func, and_, or_, desc, text, case, select
from sqlalchemy.orm import sessionmaker, Session, joinedload, subqueryload
from sqlalchemy.pool import StaticPool
from sqlalchemy.exc import IntegrityError, OperationalError

from laser_trim_analyzer.database.models import (
    Base,
    AnalysisResult as DBAnalysisResult,
    TrackResult as DBTrackResult,
    QAAlert as DBQAAlert,
    ProcessedFile as DBProcessedFile,
    ModelSpec,
    SystemType as DBSystemType,
    StatusType as DBStatusType,
    RiskCategory as DBRiskCategory,
    AlertType as DBAlertType,
    utc_now,
)
from laser_trim_analyzer.core.models import (
    AnalysisResult,
    TrackData,
    AnalysisStatus,
    SystemType,
    RiskCategory,
)
from laser_trim_analyzer.config import get_config
from laser_trim_analyzer.utils.hashing import calculate_file_hash

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Unit identity helpers — used at write time and by the migration backfill.
# A "unit" is one physical part: same model + same shop number + same day.
# Sections (P, R, etc.) and same-day retrims all collapse to the same unit_id.
# See docs/superpowers/specs/2026-05-08-unit-level-yield-design.md.
# ---------------------------------------------------------------------------

def extract_shop_number(serial: Optional[str]) -> Optional[str]:
    """Return the leading-digit prefix of a serial, or None if absent.

    Examples: "3P" -> "3", "0009" -> "0009", "1A" -> "1", "TEST" -> None.
    Whitespace is stripped first. None / empty string return None.
    The re.ASCII flag restricts \\d to [0-9] so Unicode decimal digits
    (e.g. Arabic-Indic '٣') do not silently match.
    """
    if not serial:
        return None
    match = re.match(r"^\d+", serial.strip(), re.ASCII)
    return match.group(0) if match else None


def compute_unit_id(
    model: Optional[str],
    serial: Optional[str],
    file_date: Optional[datetime],
) -> Optional[str]:
    """Compute the canonical unit identifier for a trim record.

    Returns None if any input is missing or the serial has no shop number —
    such rows are excluded from unit-level yield queries by design.

    Format: "<model>/<shop_number>/<YYYY-MM-DD>".
    """
    if not model or file_date is None:
        return None
    shop = extract_shop_number(serial)
    if shop is None:
        return None
    return f"{model}/{shop}/{file_date.strftime('%Y-%m-%d')}"


def _model_sort_key(model: str) -> tuple:
    """
    Sort key for numerical model sorting.
    Extracts leading digits from model number for numerical comparison.
    E.g., '8340-1' -> (8340, '8340-1'), '8397-2' -> (8397, '8397-2')
    """
    base = model.split('-')[0] if model else ''
    num = int(''.join(c for c in base if c.isdigit()) or '0')
    return (num, model)


def _status_matches(status: Any, *targets: DBStatusType) -> bool:
    """Check if status matches any of the target StatusType values.

    Handles both enum values and string values (from legacy data).
    SQLAlchemy Enum columns store the enum NAME (e.g., 'PASS') but may
    have legacy data with the VALUE (e.g., 'Pass').

    Args:
        status: The status value from database (enum or string)
        *targets: One or more StatusType enum values to match against

    Returns:
        True if status matches any target
    """
    if status is None:
        return False

    # Get both names (PASS) and values (Pass) we're looking for
    target_names = {t.name for t in targets}
    target_values = {t.value for t in targets}

    # Handle both enum and string
    if isinstance(status, DBStatusType):
        return status in targets
    elif isinstance(status, str):
        # Check both name and value for backwards compatibility
        return status in target_names or status in target_values
    else:
        return False


class DatabaseError(Exception):
    """Base exception for database operations."""
    pass


class DatabaseManager:
    """
    Simplified database manager for v3.

    Features:
    - Single-file SQLite database (self-contained)
    - Context manager for safe transactions
    - Incremental processing support
    - Memory-efficient queries
    """

    def __init__(self, database_path: Optional[Path] = None):
        """
        Initialize the database manager.

        Args:
            database_path: Path to SQLite database. If None, uses config default.
        """
        config = get_config()

        if database_path is None:
            database_path = config.database.path

        # Ensure parent directory exists
        database_path = Path(database_path)
        database_path.parent.mkdir(parents=True, exist_ok=True)

        self.database_path = database_path
        self.database_url = f"sqlite:///{database_path}"

        logger.info(f"Using database: {database_path}")

        # Create engine with SQLite-appropriate settings
        # timeout=30 gives SQLite time to wait for locks instead of failing immediately
        self._engine = create_engine(
            self.database_url,
            echo=False,
            poolclass=StaticPool,  # Good for SQLite
            connect_args={
                "check_same_thread": False,
                "timeout": 30,  # Wait up to 30s for locks
            },
        )

        # Enable WAL mode for better concurrency (allows readers during writes)
        # Enable foreign key enforcement (SQLite disables it by default!)
        with self._engine.connect() as conn:
            conn.execute(text("PRAGMA journal_mode=WAL"))
            conn.execute(text("PRAGMA busy_timeout=30000"))  # 30 second timeout
            conn.execute(text("PRAGMA foreign_keys=ON"))
            conn.commit()

        # Foreign keys must be enabled on EVERY connection (SQLite per-connection).
        # With StaticPool there's only one connection, but add an event listener
        # as defense-in-depth in case the pool strategy changes later.
        from sqlalchemy import event
        @event.listens_for(self._engine, "connect")
        def _set_sqlite_pragma(dbapi_conn, connection_record):
            cursor = dbapi_conn.cursor()
            cursor.execute("PRAGMA foreign_keys=ON")
            cursor.close()

        # Create session factory
        self._SessionFactory = sessionmaker(bind=self._engine)

        # Thread-local session storage
        self._thread_local = threading.local()

        # Re-entrant DB lock. With StaticPool every thread shares one
        # underlying SQLite connection, so all session() / cursor.execute
        # operations must be serialised — not just writes — to avoid
        # SQLITE_MISUSE from concurrent reads. Re-entrant so existing
        # nested `with _write_lock: with session():` blocks still work.
        self._write_lock = threading.RLock()

        # Initialize database
        self._init_database()

    def _init_database(self) -> None:
        """Create all tables if they don't exist and run migrations."""
        try:
            Base.metadata.create_all(self._engine, checkfirst=True)
            self._run_migrations()
            logger.info("Database initialized successfully")
        except Exception as e:
            logger.error(f"Failed to initialize database: {e}")
            raise DatabaseError(f"Database initialization failed: {e}")

    @staticmethod
    def _reparse_filename(filename: str) -> tuple:
        """Re-parse a filename to extract model and serial using improved logic.

        Standalone version of ExcelParser._parse_filename() for use in migrations
        without importing the parser (avoids circular dependencies).
        """
        import re
        from pathlib import Path

        def _is_valid_suffix(suffix):
            if suffix.lower().startswith('shop'):
                return False
            if re.match(r'^[A-Za-z0-9]{1,3}$', suffix):
                return True
            if suffix.lower() == 'outer':
                return True
            return False

        name = Path(filename).stem
        parts = re.split(r'[_\s]+', name)

        # Concatenated model+sn+serial (e.g., "7928sn1040", "8340-1-sn201")
        if parts:
            concat_match = re.match(
                r'^(\d{4,}[A-Za-z]?(?:-\d+[A-Za-z]?)?)-?[sS][nN](\d+[a-zA-Z]?)$',
                parts[0]
            )
            if concat_match:
                return concat_match.group(1), concat_match.group(2)

        if len(parts) >= 2:
            model = "Unknown"
            serial = "Unknown"

            for part in parts:
                model_match = re.match(r'^(\d{4,}[A-Za-z]?)((?:-[A-Za-z0-9]+)*)$', part)
                if model_match:
                    base = model_match.group(1)
                    suffixes_str = model_match.group(2)

                    if not suffixes_str:
                        model = part
                        break

                    suffixes = suffixes_str.split('-')[1:]

                    if suffixes[-1].lower() == 'sn':
                        model = base + '-'.join([''] + suffixes[:-1]) if len(suffixes) > 1 else base
                        break

                    valid_parts = [base]
                    for s in suffixes:
                        if _is_valid_suffix(s):
                            valid_parts.append(s)
                        else:
                            model = '-'.join(valid_parts)
                            serial = s
                            break
                    else:
                        model = part
                    break

            if serial == "Unknown" and model != "Unknown":
                skip_keywords = {
                    'test', 'data', 'deg', 'ta', 'tb', 'trimmed', 'correct',
                    'scrap', 'cut', 'wiper', 'path', 'am', 'pm',
                    'fail', 'pass', 'final', 'template', 'primary',
                    'customer', 'report', 'master', 'noise', 'copy', 'of',
                    'pre', 'outer', 'redunat', 'redundant',
                }
                for i, part in enumerate(parts):
                    if part == model:
                        continue
                    if '-' in part and part.startswith(model + '-'):
                        continue
                    if part.lower() in skip_keywords:
                        continue
                    if i + 1 < len(parts) and parts[i + 1].lower() == 'deg':
                        continue
                    if re.match(r'^\d{1,4}-\d{1,2}-\d{1,4}$', part):
                        continue
                    if re.match(r'^\d{1,2}-\d{2}(-\d{2})?$', part):
                        continue
                    if re.match(r'^\d{8,}', part):
                        continue
                    if re.match(r'^[A-Z]{1,3}\d+', part, re.IGNORECASE):
                        serial = part
                        break
                    elif re.match(r'^\d+$', part):
                        serial = part
                        break

            return model, serial

        # Single-part filename: only return as model if it looks like a valid model number
        if re.match(r'^\d{4,}[A-Za-z]?$', name):
            return name, "Unknown"
        return "Unknown", "Unknown"

    def set_baseline_requalification(self, model: str, effective_date, note: str = "") -> None:
        """Record a per-model baseline requalification (AS9100 audit trail:
        who decided is the operator of this single-user tool; when and why
        are stored). The LATEST effective_date wins for training."""
        from sqlalchemy import text as _text
        with self.session() as s:
            s.execute(_text(
                "INSERT INTO baseline_requalifications (model, effective_date, note, set_at) "
                "VALUES (:m, :d, :n, :t)"),
                {"m": model, "d": str(effective_date), "n": note or "",
                 "t": datetime.now().isoformat(sep=" ", timespec="seconds")})
            s.commit()

    def get_baseline_requalification(self, model: str):
        """Latest requalification for the model -> (effective_date str, note,
        set_at str) or None. Training uses effective_date as the sample floor."""
        from sqlalchemy import text as _text
        with self.session() as s:
            row = s.execute(_text(
                "SELECT effective_date, note, set_at FROM baseline_requalifications "
                "WHERE model = :m ORDER BY set_at DESC, id DESC LIMIT 1"),
                {"m": model}).fetchone()
        return (row[0], row[1], row[2]) if row else None

    def _run_migrations(self) -> None:
        """Run database migrations for schema updates."""
        # Baseline requalification audit table (2026-07-13: per-model manual
        # baseline reset on design change — AS9100 traceability).
        try:
            with self.session() as _s:
                _s.execute(text(
                    "CREATE TABLE IF NOT EXISTS baseline_requalifications ("
                    "id INTEGER PRIMARY KEY AUTOINCREMENT, model TEXT NOT NULL, "
                    "effective_date TEXT NOT NULL, note TEXT, set_at TEXT NOT NULL)"))
                _s.execute(text(
                    "CREATE INDEX IF NOT EXISTS idx_baseline_requal_model "
                    "ON baseline_requalifications(model)"))
                _s.commit()
        except Exception:
            logger.exception("baseline_requalifications migration failed")
        needs_rematch = False

        with self.session() as session:
            # Migration: Add is_anomaly and anomaly_reason columns to track_results
            try:
                # Check if columns exist by attempting a query
                session.execute(text("SELECT is_anomaly FROM track_results LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                # Columns don't exist, add them
                logger.info("Running migration: Adding is_anomaly and anomaly_reason columns")
                try:
                    session.execute(text("ALTER TABLE track_results ADD COLUMN is_anomaly BOOLEAN DEFAULT 0"))
                    session.execute(text("ALTER TABLE track_results ADD COLUMN anomaly_reason TEXT"))
                    session.commit()
                    logger.info("Migration completed: Added anomaly detection columns")
                except Exception as e:
                    logger.warning(f"Migration warning (may already exist): {e}")

            # Migration: Add drift_baseline_cutoff_date column to model_ml_state
            try:
                session.execute(text("SELECT drift_baseline_cutoff_date FROM model_ml_state LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info("Running migration: Adding drift_baseline_cutoff_date column")
                try:
                    session.execute(text("ALTER TABLE model_ml_state ADD COLUMN drift_baseline_cutoff_date DATETIME"))
                    session.commit()
                    logger.info("Migration completed: Added drift_baseline_cutoff_date column")
                except Exception as e:
                    logger.warning(f"Migration warning (may already exist): {e}")

            # Migration: Add peak_cusum column to model_ml_state
            try:
                session.execute(text("SELECT peak_cusum FROM model_ml_state LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info("Running migration: Adding peak_cusum column")
                try:
                    session.execute(text("ALTER TABLE model_ml_state ADD COLUMN peak_cusum FLOAT DEFAULT 0"))
                    session.commit()
                    logger.info("Migration completed: Added peak_cusum column")
                except Exception as e:
                    logger.warning(f"Migration warning (may already exist): {e}")

            # Migration: Normalize status values from 'Pass' to 'PASS' format
            # SQLAlchemy stores enum NAME (PASS), not value (Pass)
            # This fixes data corrupted by bulk SQL that used .value instead of .name
            try:
                # Check if there are any title-case values that need fixing
                result = session.execute(text(
                    "SELECT COUNT(*) FROM analysis_results WHERE overall_status IN ('Pass', 'Fail', 'Warning', 'Error')"
                )).scalar()
                if result and result > 0:
                    logger.info(f"Running migration: Normalizing {result} status values to uppercase")
                    # Fix analysis_results
                    session.execute(text("UPDATE analysis_results SET overall_status = 'PASS' WHERE overall_status = 'Pass'"))
                    session.execute(text("UPDATE analysis_results SET overall_status = 'FAIL' WHERE overall_status = 'Fail'"))
                    session.execute(text("UPDATE analysis_results SET overall_status = 'WARNING' WHERE overall_status = 'Warning'"))
                    session.execute(text("UPDATE analysis_results SET overall_status = 'ERROR' WHERE overall_status = 'Error'"))
                    # Fix track_results
                    session.execute(text("UPDATE track_results SET status = 'PASS' WHERE status = 'Pass'"))
                    session.execute(text("UPDATE track_results SET status = 'FAIL' WHERE status = 'Fail'"))
                    session.execute(text("UPDATE track_results SET status = 'WARNING' WHERE status = 'Warning'"))
                    session.execute(text("UPDATE track_results SET status = 'ERROR' WHERE status = 'Error'"))
                    # Fix final_test_results / smoothness_results too (the original
                    # migration missed these tables). Idempotent if already uppercase.
                    for _tbl in ("final_test_results", "smoothness_results"):
                        for _old, _new in (("Pass", "PASS"), ("Fail", "FAIL"),
                                           ("Warning", "WARNING"), ("Error", "ERROR"),
                                           ("Untrimmed", "UNTRIMMED")):
                            try:
                                session.execute(text(
                                    f"UPDATE {_tbl} SET overall_status = '{_new}' "
                                    f"WHERE overall_status = '{_old}'"))
                            except Exception:
                                pass  # table/column may not exist on older schemas
                    session.commit()
                    logger.info("Migration completed: Status values normalized")
            except Exception as e:
                logger.warning(f"Status normalization warning: {e}")

            # Migration: Clean up "-shop" model name parsing artifacts
            # Files like "8444-shop0_date.xlsx" were incorrectly parsed with
            # model="8444-shop0" instead of model="8444", serial="shop0"
            try:
                shop_count = session.execute(text(
                    "SELECT COUNT(*) FROM analysis_results WHERE LOWER(model) LIKE '%-shop%'"
                )).scalar()
                if shop_count and shop_count > 0:
                    logger.info(f"Running migration: Cleaning up {shop_count} shop model name records")

                    # Fix analysis_results: split "8444-shop0" into model="8444", serial="shop0"
                    shop_records = session.execute(text(
                        "SELECT id, model FROM analysis_results WHERE LOWER(model) LIKE '%-shop%'"
                    )).fetchall()

                    for row in shop_records:
                        old_model = row[1]
                        # Find the "-shop" split point (case-insensitive)
                        lower = old_model.lower()
                        shop_idx = lower.find('-shop')
                        if shop_idx > 0:
                            base_model = old_model[:shop_idx]
                            new_serial = old_model[shop_idx + 1:]  # "shop0", "shop101", etc.
                            session.execute(text(
                                "UPDATE analysis_results SET model = :model, serial = :serial WHERE id = :id"
                            ), {"model": base_model, "serial": new_serial, "id": row[0]})

                    # Delete model_ml_state entries for fake shop model names
                    ml_deleted = session.execute(text(
                        "DELETE FROM model_ml_state WHERE LOWER(model) LIKE '%-shop%'"
                    )).rowcount
                    logger.info(f"Deleted {ml_deleted} fake ML state entries for shop models")

                    session.commit()
                    needs_rematch = True
                    logger.info(f"Migration completed: Cleaned up {shop_count} shop model name records")
            except Exception as e:
                logger.warning(f"Shop model cleanup warning: {e}")

            # Migration: Re-parse "Unknown" model records with improved parser logic
            # Handles: multi-hyphen models (7280-1-CT), -sn serial indicators,
            # concatenated sn patterns, "final NNN" serials, etc.
            try:
                unknown_records = session.execute(text(
                    "SELECT id, filename FROM analysis_results WHERE model = 'Unknown'"
                )).fetchall()
                if unknown_records:
                    logger.info(f"Running migration: Re-parsing {len(unknown_records)} Unknown model records")
                    fixed = 0
                    for row in unknown_records:
                        rec_id, filename = row[0], row[1]
                        model, serial = self._reparse_filename(filename)
                        if model != "Unknown":
                            session.execute(text(
                                "UPDATE analysis_results SET model = :model, serial = :serial WHERE id = :id"
                            ), {"model": model, "serial": serial, "id": rec_id})
                            fixed += 1
                    if fixed > 0:
                        session.commit()
                        needs_rematch = True
                        logger.info(f"Migration completed: Fixed {fixed} of {len(unknown_records)} Unknown model records")
                    else:
                        logger.info("Migration: No Unknown records could be re-parsed")
            except Exception as e:
                logger.warning(f"Unknown model re-parse warning: {e}")

            # Migration: Ensure all performance indexes exist
            # create_all() only creates indexes for NEW tables. Existing databases
            # may be missing indexes that were added to models.py later.
            # CREATE INDEX IF NOT EXISTS is idempotent — safe to run every startup.
            try:
                index_statements = [
                    # analysis_results indexes
                    "CREATE INDEX IF NOT EXISTS idx_filename_date ON analysis_results(filename, file_date)",
                    "CREATE INDEX IF NOT EXISTS idx_file_date ON analysis_results(file_date)",
                    "CREATE INDEX IF NOT EXISTS idx_model_serial ON analysis_results(model, serial)",
                    "CREATE INDEX IF NOT EXISTS idx_model_serial_date ON analysis_results(model, serial, file_date)",
                    "CREATE INDEX IF NOT EXISTS idx_timestamp ON analysis_results(timestamp)",
                    "CREATE INDEX IF NOT EXISTS idx_status ON analysis_results(overall_status)",
                    "CREATE INDEX IF NOT EXISTS idx_system ON analysis_results(system)",
                    "CREATE INDEX IF NOT EXISTS idx_status_timestamp ON analysis_results(overall_status, timestamp)",
                    "CREATE INDEX IF NOT EXISTS idx_model_status ON analysis_results(model, overall_status)",
                    # track_results indexes
                    "CREATE INDEX IF NOT EXISTS idx_track_analysis ON track_results(analysis_id, track_id)",
                    "CREATE INDEX IF NOT EXISTS idx_track_sigma_gradient ON track_results(sigma_gradient)",
                    "CREATE INDEX IF NOT EXISTS idx_track_sigma_pass ON track_results(sigma_pass)",
                    "CREATE INDEX IF NOT EXISTS idx_track_linearity_pass ON track_results(linearity_pass)",
                    "CREATE INDEX IF NOT EXISTS idx_track_risk_category ON track_results(risk_category)",
                    "CREATE INDEX IF NOT EXISTS idx_track_failure_probability ON track_results(failure_probability)",
                    "CREATE INDEX IF NOT EXISTS idx_track_status ON track_results(status)",
                    "CREATE INDEX IF NOT EXISTS idx_track_analysis_prob ON track_results(analysis_id, failure_probability)",
                    # final_test_results indexes
                    "CREATE INDEX IF NOT EXISTS idx_ft_filename_date ON final_test_results(filename, file_date)",
                    "CREATE INDEX IF NOT EXISTS idx_ft_model_serial ON final_test_results(model, serial)",
                    "CREATE INDEX IF NOT EXISTS idx_ft_model_serial_date ON final_test_results(model, serial, file_date)",
                    "CREATE INDEX IF NOT EXISTS idx_ft_timestamp ON final_test_results(timestamp)",
                    "CREATE INDEX IF NOT EXISTS idx_ft_status ON final_test_results(overall_status)",
                    "CREATE INDEX IF NOT EXISTS idx_ft_linked_trim ON final_test_results(linked_trim_id)",
                    "CREATE INDEX IF NOT EXISTS idx_ft_test_date ON final_test_results(test_date)",
                    # Standalone file_date index - Compare/Final Test page does
                    # ORDER BY file_date DESC LIMIT 500 with no leading filter,
                    # which the composite indexes don't satisfy.
                    "CREATE INDEX IF NOT EXISTS idx_ft_file_date ON final_test_results(file_date)",
                    # Placed last so the bulk loop on pre-Spec-1 DBs re-confirms all
                    # existing indexes before failing on this one new-column entry.
                    # The Spec 1 column migration immediately after (untrimmed_sigma_gradient
                    # block) creates both the column and this index on first upgrade.
                    "CREATE INDEX IF NOT EXISTS idx_track_untrimmed_sigma_gradient ON track_results(untrimmed_sigma_gradient)",
                ]
                created = 0
                for stmt in index_statements:
                    session.execute(text(stmt))
                    created += 1
                session.commit()
                logger.info(f"Index migration: ensured {created} indexes exist")
            except Exception as e:
                logger.warning(f"Index migration warning: {e}")

            # Migration: Add failure margin columns to track_results
            try:
                session.execute(text("SELECT max_violation FROM track_results LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info("Running migration: Adding failure margin columns")
                try:
                    session.execute(text("ALTER TABLE track_results ADD COLUMN max_violation FLOAT"))
                    session.execute(text("ALTER TABLE track_results ADD COLUMN avg_violation FLOAT"))
                    session.execute(text("ALTER TABLE track_results ADD COLUMN margin_to_spec FLOAT"))
                    session.commit()
                    logger.info("Migration completed: Added failure margin columns")
                except Exception as e:
                    logger.warning(f"Failure margin migration warning (may already exist): {e}")

            # Migration: Add measured_electrical_angle column to track_results
            try:
                session.execute(text("SELECT measured_electrical_angle FROM track_results LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info("Running migration: Adding measured_electrical_angle column")
                try:
                    session.execute(text("ALTER TABLE track_results ADD COLUMN measured_electrical_angle FLOAT"))
                    session.commit()
                    logger.info("Migration completed: Added measured_electrical_angle column")
                except Exception as e:
                    logger.warning(f"measured_electrical_angle migration warning (may already exist): {e}")

            # Migration: Add max deviation columns to track_results
            try:
                session.execute(text("SELECT max_deviation FROM track_results LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info("Running migration: Adding max deviation columns")
                try:
                    session.execute(text("ALTER TABLE track_results ADD COLUMN max_deviation FLOAT"))
                    session.execute(text("ALTER TABLE track_results ADD COLUMN max_deviation_position FLOAT"))
                    session.execute(text("ALTER TABLE track_results ADD COLUMN deviation_uniformity FLOAT"))
                    session.commit()
                    logger.info("Migration completed: Added max deviation columns")
                except Exception as e:
                    logger.warning(f"Max deviation migration warning (may already exist): {e}")

            # Migration: Add untrimmed_sigma_gradient column to track_results.
            # Spec 1 (2026-05-30): upstream element-quality signal independent
            # of post-trim sigma_gradient.  Backfilled by natural reprocess flow.
            try:
                session.execute(
                    text("SELECT untrimmed_sigma_gradient FROM track_results LIMIT 1")
                )
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info(
                    "Running migration: Adding untrimmed_sigma_gradient column"
                )
                try:
                    session.execute(text(
                        "ALTER TABLE track_results "
                        "ADD COLUMN untrimmed_sigma_gradient FLOAT"
                    ))
                    session.execute(text(
                        "CREATE INDEX IF NOT EXISTS "
                        "idx_track_untrimmed_sigma_gradient "
                        "ON track_results (untrimmed_sigma_gradient)"
                    ))
                    session.commit()
                    logger.info(
                        "Migration completed: Added untrimmed_sigma_gradient"
                    )
                except Exception as e:
                    logger.warning(
                        f"Migration warning (may already exist): {e}"
                    )

            # Migration: Add untrimmed_error_max column to track_results.
            # Spec 2 (2026-06-02): worst-case linearity error across untrimmed
            # data points; complements untrimmed_sigma_gradient as an element-
            # quality signal.  Backfilled by natural reprocess flow.
            try:
                session.execute(
                    text("SELECT untrimmed_error_max FROM track_results LIMIT 1")
                )
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info(
                    "Running migration: Adding untrimmed_error_max column"
                )
                try:
                    session.execute(text(
                        "ALTER TABLE track_results "
                        "ADD COLUMN untrimmed_error_max FLOAT"
                    ))
                    session.execute(text(
                        "CREATE INDEX IF NOT EXISTS "
                        "idx_track_untrimmed_error_max "
                        "ON track_results (untrimmed_error_max)"
                    ))
                    session.commit()
                    logger.info(
                        "Migration completed: Added untrimmed_error_max"
                    )
                except Exception as e:
                    logger.warning(
                        f"Migration warning (may already exist): {e}"
                    )

            # Migration: Add composite_trim_risk_score column to track_results.
            try:
                session.execute(
                    text("SELECT composite_trim_risk_score FROM track_results LIMIT 1")
                )
            except OperationalError:
                session.rollback()
                logger.info("Running migration: Adding composite_trim_risk_score column")
                try:
                    session.execute(text(
                        "ALTER TABLE track_results "
                        "ADD COLUMN composite_trim_risk_score FLOAT"
                    ))
                    session.execute(text(
                        "CREATE INDEX IF NOT EXISTS idx_track_composite_trim_risk_score "
                        "ON track_results (composite_trim_risk_score)"
                    ))
                    session.commit()
                    logger.info("Migration completed: Added composite_trim_risk_score")
                except Exception as e:
                    logger.warning(f"Migration warning (may already exist): {e}")

            # Migration: Create model_metric_state table for Spec 2.
            # This is a CREATE TABLE rather than ALTER TABLE because the
            # table is entirely new in V6.  Use Base.metadata.create_all
            # with checkfirst=True for idempotency.
            try:
                from laser_trim_analyzer.database.models import ModelMetricState
                ModelMetricState.__table__.create(bind=self._engine, checkfirst=True)
                session.commit()
            except Exception as e:
                session.rollback()
                logger.warning(
                    f"Migration warning for model_metric_state (may already exist): {e}"
                )

            # Migration: Retag LTS3 (System C) rows (2026-07-06). The third
            # trim system writes files format-identical to an existing system,
            # so anything processed before path-based detection landed was
            # stored as A or B. Identity marker = an 'LTS3*' DIRECTORY in the
            # path (a separator must follow, so filenames starting with LTS3
            # don't match). Idempotent: already-C rows aren't selected.
            try:
                total_retagged = 0
                for table in ("analysis_results", "final_test_results"):
                    res = session.execute(text(
                        f"UPDATE {table} SET system = 'C' "
                        f"WHERE system != 'C' AND ("
                        f"file_path LIKE '%/LTS3%/%' OR file_path LIKE '%\\LTS3%\\%'"
                        f")"
                    ))
                    total_retagged += res.rowcount or 0
                session.commit()
                if total_retagged:
                    logger.info(f"Migration: retagged {total_retagged} LTS3 rows as System C")
            except Exception as e:
                session.rollback()
                logger.warning(f"LTS3 retag migration warning: {e}")

            # Migration: Add last_row_id watermark + recent_window to
            # model_metric_state (2026-07-06). last_row_id lets
            # advance_drift_state consume same-day samples the date-only
            # filter skipped forever; recent_window makes the step-change
            # check live across restarts.
            for col, ddl in (("last_row_id", "INTEGER"), ("recent_window", "TEXT")):
                try:
                    session.execute(text(
                        f"SELECT {col} FROM model_metric_state LIMIT 1"))
                except OperationalError:
                    session.rollback()
                    logger.info(f"Running migration: Adding model_metric_state.{col}")
                    try:
                        session.execute(text(
                            f"ALTER TABLE model_metric_state ADD COLUMN {col} {ddl}"
                        ))
                        session.commit()
                        logger.info(f"Migration completed: Added {col}")
                    except Exception as e:
                        logger.warning(f"{col} migration warning (may already exist): {e}")

            # Migration: Add data_quality columns to analysis_results
            try:
                session.execute(text("SELECT data_quality FROM analysis_results LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                logger.info("Running migration: Adding data_quality columns")
                try:
                    session.execute(text(
                        "ALTER TABLE analysis_results ADD COLUMN data_quality VARCHAR(20) DEFAULT 'good'"
                    ))
                    session.execute(text(
                        "ALTER TABLE analysis_results ADD COLUMN data_quality_issues TEXT"
                    ))
                    session.commit()
                    logger.info("Migration completed: Added data_quality columns")
                except Exception as e:
                    logger.warning(f"Data quality migration warning (may already exist): {e}")

            # Migration: Add Phase 2 spec-aware optimization columns to track_results.
            # Each column gets its own try/commit so a duplicate-column error on
            # one ALTER does not roll back columns added earlier in the same
            # session — that was the prior bug (single rollback at end of loop
            # discarded successful ALTERs in the SQLAlchemy unit of work).
            phase2_columns = {
                "optimal_slope": "FLOAT DEFAULT 0.0",
                "station_compensation": "FLOAT",
                "linearity_type": "VARCHAR(30)",
                "raw_linearity_error": "FLOAT",
                "optimized_linearity_error": "FLOAT",
                "raw_fail_points": "INTEGER",
            }
            for col_name, col_type in phase2_columns.items():
                try:
                    session.execute(text(
                        f"ALTER TABLE track_results ADD COLUMN {col_name} {col_type}"
                    ))
                    session.commit()
                except Exception as e:
                    session.rollback()
                    if "duplicate column" not in str(e).lower() and "already exists" not in str(e).lower():
                        logger.warning(f"Migration error adding {col_name}: {e}")
            logger.info("Phase 2 migration: ensured spec-aware columns exist")

            # Migration: Add match_method column to final_test_results
            try:
                session.execute(text("SELECT match_method FROM final_test_results LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                try:
                    session.execute(text("ALTER TABLE final_test_results ADD COLUMN match_method VARCHAR(30)"))
                    session.commit()
                    logger.info("Migration: Added match_method column to final_test_results")
                except Exception:
                    pass

            # Migration: Add trim_pass_count column to track_results.
            # Counts how many laser-trim passes the equipment ran per track,
            # surfaced from the file's "Trim N" / "TRK<n> M" sheet layout.
            # Used as a quality indicator (1 = clean, 2+ = retrim needed).
            try:
                session.execute(text("SELECT trim_pass_count FROM track_results LIMIT 1"))
            except OperationalError:
                session.rollback()
                try:
                    session.execute(text("ALTER TABLE track_results ADD COLUMN trim_pass_count INTEGER"))
                    session.commit()
                    logger.info("Migration: Added trim_pass_count column to track_results")
                except Exception:
                    pass

            # Migration: Add unit_id column to analysis_results.
            # Canonical unit identifier "<model>/<shop>/<date>", used by the
            # unit-level yield feature (Trends chart + Excel "Yield by Unit"
            # sheet). Nullable so the app remains functional during/after
            # partial backfill; the backfill itself happens in a separate
            # migration step below so it can be retried independently.
            try:
                session.execute(text("SELECT unit_id FROM analysis_results LIMIT 1"))
            except OperationalError:
                session.rollback()
                try:
                    session.execute(text(
                        "ALTER TABLE analysis_results ADD COLUMN unit_id VARCHAR(80)"
                    ))
                    session.commit()
                    logger.info("Migration: Added unit_id column to analysis_results")
                except Exception as e:
                    session.rollback()
                    logger.warning(f"Migration error adding unit_id: {e}")
            # Ensure the index exists on both new and migrated DBs. Project
            # convention uses idx_* naming (not SQLAlchemy's auto ix_*),
            # so we manage it explicitly here rather than via index=True
            # on the column declaration.
            try:
                session.execute(text(
                    "CREATE INDEX IF NOT EXISTS idx_analysis_unit_id "
                    "ON analysis_results(unit_id)"
                ))
                session.commit()
            except Exception as e:
                session.rollback()
                logger.warning(f"Migration error creating unit_id index: {e}")

            # Migration: Backfill unit_id for existing rows. Idempotent —
            # only updates rows where unit_id IS NULL, so re-running picks up
            # where it left off (e.g. after a crash or interrupted startup).
            # Done in Python because the shop-number extraction regex lives
            # there; iterates in batches of 1000 to keep memory bounded.
            self._backfill_unit_ids(session)

            # Migration: Add aliases column to model_specs.
            # Stores pipe-separated alternate model numbers so a single spec
            # row covers cases like 1621501 and 2001621501 being the same part.
            try:
                session.execute(text("SELECT aliases FROM model_specs LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                try:
                    session.execute(text("ALTER TABLE model_specs ADD COLUMN aliases TEXT"))
                    session.commit()
                    logger.info("Migration: Added aliases column to model_specs")
                except Exception as e:
                    if "duplicate column" not in str(e).lower():
                        logger.warning(f"aliases migration warning: {e}")
                    session.rollback()

            # Migration: Add exclude_points column to model_specs
            try:
                session.execute(text("SELECT exclude_points FROM model_specs LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                try:
                    session.execute(text("ALTER TABLE model_specs ADD COLUMN exclude_points TEXT"))
                    session.commit()
                    logger.info("Migration: Added exclude_points column to model_specs")
                except Exception as e:
                    if "duplicate column" not in str(e).lower():
                        logger.warning(f"exclude_points migration warning: {e}")
                    session.rollback()

            # Migration: Add exclude_points_ft column to model_specs
            # FT files have different data point counts than trim files,
            # so they need separate exclude ranges.
            try:
                session.execute(text("SELECT exclude_points_ft FROM model_specs LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                try:
                    session.execute(text("ALTER TABLE model_specs ADD COLUMN exclude_points_ft TEXT"))
                    session.commit()
                    logger.info("Migration: Added exclude_points_ft column to model_specs")
                except Exception as e:
                    if "duplicate column" not in str(e).lower():
                        logger.warning(f"exclude_points_ft migration warning: {e}")
                    session.rollback()

            # Migration: Add open_closed column to model_specs and backfill from
            # circuit_type. The original importer wrote the Excel "Open/Closed"
            # column into circuit_type, which is misleading — Open vs Closed
            # refers to whether the resistive element is visible, not the
            # electrical circuit type. Keep circuit_type for backward compat
            # but add a correctly-named column and sync values across.
            try:
                session.execute(text("SELECT open_closed FROM model_specs LIMIT 1"))
            except OperationalError:
                session.rollback()  # Clear error state from failed probe
                try:
                    session.execute(text(
                        "ALTER TABLE model_specs ADD COLUMN open_closed VARCHAR(10)"
                    ))
                    # Backfill from circuit_type for existing rows
                    session.execute(text(
                        "UPDATE model_specs SET open_closed = circuit_type "
                        "WHERE open_closed IS NULL AND circuit_type IS NOT NULL"
                    ))
                    session.commit()
                    logger.info(
                        "Migration: Added open_closed column to model_specs "
                        "and backfilled from circuit_type"
                    )
                except Exception as e:
                    if "duplicate column" not in str(e).lower():
                        logger.warning(f"open_closed migration warning: {e}")
                    session.rollback()

            # Migration: Add spec-aware columns to final_test_tracks so the FT
            # analyzer's optimal_slope/offset/linearity_type are persisted (not
            # just held in memory for the current Process Files screen).
            ft_phase2_columns = {
                "optimal_offset": "FLOAT",
                "optimal_slope": "FLOAT DEFAULT 0.0",
                "linearity_type": "VARCHAR(30)",
            }
            for col_name, col_type in ft_phase2_columns.items():
                try:
                    session.execute(text(
                        f"ALTER TABLE final_test_tracks ADD COLUMN {col_name} {col_type}"
                    ))
                except Exception as e:
                    if "duplicate column" not in str(e).lower() and "already exists" not in str(e).lower():
                        logger.warning(f"FT migration warning adding {col_name}: {e}")
                    session.rollback()
            try:
                session.commit()
                logger.info("FT phase2 migration: ensured spec-aware columns exist on final_test_tracks")
            except Exception:
                pass

            # Migration: Add consecutive_recovered column to model_ml_state
            try:
                session.execute(text("ALTER TABLE model_ml_state ADD COLUMN consecutive_recovered INTEGER DEFAULT 0"))
                session.commit()
            except Exception as e:
                if "duplicate column" not in str(e).lower() and "already exists" not in str(e).lower():
                    logger.warning(f"consecutive_recovered migration warning: {e}")
                session.rollback()

            # Migration: Add electrical_angle_tol_type to model_specs so the
            # angle-parser qualifier ('symmetric', 'min', 'max', 'range',
            # 'bilateral') is preserved. The slope-correction rule depends on
            # this to know whether a tolerance is one-sided or two-sided.
            try:
                session.execute(text(
                    "ALTER TABLE model_specs ADD COLUMN electrical_angle_tol_type VARCHAR(12)"
                ))
                session.commit()
                logger.info("Migration: Added electrical_angle_tol_type column to model_specs")
            except Exception as e:
                if "duplicate column" not in str(e).lower() and "already exists" not in str(e).lower():
                    logger.warning(f"electrical_angle_tol_type migration warning: {e}")
                session.rollback()

            # Migration: Add theory_data and test_volts columns for slope optimization
            try:
                session.execute(text("ALTER TABLE track_results ADD COLUMN theory_data TEXT"))
                session.commit()
                logger.info("Migration: Added theory_data column to track_results")
            except Exception as e:
                if "duplicate column" not in str(e).lower() and "already exists" not in str(e).lower():
                    logger.warning(f"theory_data migration warning: {e}")
                session.rollback()

            try:
                session.execute(text("ALTER TABLE track_results ADD COLUMN test_volts FLOAT"))
                session.commit()
                logger.info("Migration: Added test_volts column to track_results")
            except Exception as e:
                if "duplicate column" not in str(e).lower() and "already exists" not in str(e).lower():
                    logger.warning(f"test_volts migration warning: {e}")
                session.rollback()

            # Migration: Add theory_data to final_test_tracks for slope optimization
            try:
                session.execute(text("ALTER TABLE final_test_tracks ADD COLUMN theory_data TEXT"))
                session.commit()
                logger.info("Migration: Added theory_data column to final_test_tracks")
            except Exception as e:
                if "duplicate column" not in str(e).lower() and "already exists" not in str(e).lower():
                    logger.warning(f"ft theory_data migration: {e}")
                session.rollback()

            # Migration: Add file_size / file_modified_date to the Final Test
            # and Smoothness result tables (2026-08-29). ProcessedFile rows had
            # them; these two didn't, so the incremental scan's stat fast-path
            # could never apply to an FT/smoothness path — every known file was
            # re-HASHED (full read over the share) on EVERY scan, and the heal
            # pass only touched processed_files so it never got better. See
            # Processor._is_processed / _load_processed_hashes.
            stat_columns = {
                "file_size": "INTEGER",
                "file_modified_date": "DATETIME",
            }
            for tbl in ("final_test_results", "smoothness_results"):
                for col_name, col_type in stat_columns.items():
                    try:
                        session.execute(text(
                            f"ALTER TABLE {tbl} ADD COLUMN {col_name} {col_type}"
                        ))
                        session.commit()
                        logger.info(f"Migration: Added {col_name} column to {tbl}")
                    except Exception as e:
                        if ("duplicate column" not in str(e).lower()
                                and "already exists" not in str(e).lower()):
                            logger.warning(f"{tbl}.{col_name} migration warning: {e}")
                        session.rollback()

            # Migration: Relax NOT NULL on sigma_gradient / sigma_threshold / sigma_pass
            # so UNTRIMMED tracks (test-sweep-only files with no laser-trim runs) can
            # be saved with sigma metrics absent. SQLite can't ALTER COLUMN nullability
            # directly, so this rebuilds track_results via the rename-table pattern.
            try:
                info = session.execute(text("PRAGMA table_info(track_results)")).fetchall()
                # PRAGMA table_info columns: cid, name, type, notnull, dflt_value, pk
                sigma_grad_notnull = next(
                    (row[3] for row in info if row[1] == 'sigma_gradient'), 0
                )
                if sigma_grad_notnull == 1:
                    logger.info(
                        "Running migration: relaxing NOT NULL on sigma_gradient / "
                        "sigma_threshold / sigma_pass in track_results"
                    )
                    create_sql = session.execute(text(
                        "SELECT sql FROM sqlite_master WHERE type='table' "
                        "AND name='track_results'"
                    )).scalar()
                    indexes = session.execute(text(
                        "SELECT name, sql FROM sqlite_master WHERE type='index' "
                        "AND tbl_name='track_results' AND sql IS NOT NULL"
                    )).fetchall()

                    new_sql = create_sql
                    for _col in ('sigma_gradient', 'sigma_threshold', 'sigma_pass'):
                        # Strip the column-level NOT NULL declaration. Preserves
                        # the type/length so the column data is untouched.
                        new_sql = re.sub(
                            rf'(\b{_col}\b\s+\w+(?:\(\d+\))?)\s+NOT\s+NULL',
                            r'\1',
                            new_sql,
                            count=1,
                            flags=re.IGNORECASE,
                        )
                    new_sql = new_sql.replace(
                        'CREATE TABLE track_results',
                        'CREATE TABLE track_results_new',
                        1,
                    )

                    # NOTE: No PRAGMA foreign_keys toggling here. SQLite ignores
                    # PRAGMA foreign_keys when issued inside an open transaction,
                    # so the typical "OFF / rebuild / ON" recipe is a no-op in
                    # this session-scoped path. The rebuild is FK-safe today
                    # because no table references track_results.id — if that
                    # changes, the recipe needs to move outside this transaction
                    # via a dedicated raw connection.
                    session.execute(text(new_sql))
                    session.execute(text(
                        "INSERT INTO track_results_new SELECT * FROM track_results"
                    ))
                    session.execute(text("DROP TABLE track_results"))
                    session.execute(text(
                        "ALTER TABLE track_results_new RENAME TO track_results"
                    ))
                    for _idx_name, _idx_sql in indexes:
                        # Re-raise on failure so the outer try/except triggers a
                        # rollback. Silently losing an index would degrade query
                        # performance without any signal to the operator.
                        session.execute(text(_idx_sql))
                    session.commit()
                    logger.info(
                        "Migration completed: sigma_gradient / sigma_threshold / "
                        "sigma_pass are now nullable"
                    )
            except Exception as e:
                session.rollback()
                logger.warning(f"sigma-nullable migration warning: {e}")

        # After session closes, re-run FT matching if model names were corrected
        if needs_rematch:
            try:
                logger.info("Re-matching Final Test records after model name cleanup...")
                stats = self.rematch_final_tests()
                logger.info(f"Post-cleanup FT rematch: {stats}")
            except Exception as e:
                logger.warning(f"FT rematch after cleanup failed: {e}")

    def _backfill_unit_ids(self, session) -> None:
        """Populate analysis_results.unit_id for rows that don't have one yet.

        Idempotent: only operates on NULL unit_id rows. Logs progress and a
        post-run sanity check (count of NULL vs non-NULL).
        """
        from laser_trim_analyzer.database.models import AnalysisResult as DBAR

        # Count rows that need backfill
        to_backfill = (
            session.query(func.count(DBAR.id))
            .filter(DBAR.unit_id.is_(None))
            .scalar()
        ) or 0
        if to_backfill == 0:
            logger.debug("unit_id backfill: nothing to do")
            return

        logger.info(f"unit_id backfill: starting on {to_backfill} rows")
        batch_size = 1000
        updated = 0
        skipped = 0  # rows that have nothing to backfill (junk serial / missing date)

        # Process in batches so we don't load 80k rows into memory at once.
        # The empty-string sentinel pattern below ensures unparseable rows
        # don't keep appearing in subsequent batches (we then convert the
        # sentinels back to NULL at the end).
        while True:
            rows = (
                session.query(DBAR.id, DBAR.model, DBAR.serial, DBAR.file_date)
                .filter(DBAR.unit_id.is_(None))
                .limit(batch_size)
                .all()
            )
            if not rows:
                break

            for row_id, model, serial, file_date in rows:
                uid = compute_unit_id(model, serial, file_date)
                if uid is None:
                    # Junk serial / missing date — mark with empty-string sentinel
                    # so the next batch query doesn't pick it up again.
                    session.execute(
                        text("UPDATE analysis_results SET unit_id = '' WHERE id = :i"),
                        {"i": row_id},
                    )
                    skipped += 1
                else:
                    session.execute(
                        text("UPDATE analysis_results SET unit_id = :u WHERE id = :i"),
                        {"u": uid, "i": row_id},
                    )
                    updated += 1
            session.commit()
            logger.info(
                f"unit_id backfill: progress {updated + skipped}/{to_backfill}"
            )

        # Restore NULL for unparseable rows.
        session.execute(text(
            "UPDATE analysis_results SET unit_id = NULL WHERE unit_id = ''"
        ))
        session.commit()

        # Post-flight sanity check
        non_null = (
            session.query(func.count(DBAR.id))
            .filter(DBAR.unit_id.isnot(None))
            .scalar()
        ) or 0
        null_now = (
            session.query(func.count(DBAR.id))
            .filter(DBAR.unit_id.is_(None))
            .scalar()
        ) or 0
        distinct_units = (
            session.query(func.count(func.distinct(DBAR.unit_id)))
            .filter(DBAR.unit_id.isnot(None))
            .scalar()
        ) or 0
        logger.info(
            f"unit_id backfill complete: "
            f"{updated} populated, {skipped} junk-serial/no-date, "
            f"{non_null} non-NULL total, {null_now} NULL total, "
            f"{distinct_units} distinct units"
        )

        # Spot-check: log three sample unit_ids so an operator can verify
        # the format on customer hardware where the DB isn't accessible.
        samples = (
            session.query(DBAR.unit_id)
            .filter(DBAR.unit_id.isnot(None))
            .limit(3)
            .all()
        )
        sample_strs = [r[0] for r in samples]
        logger.info(f"unit_id backfill sample unit_ids: {sample_strs}")

    @contextmanager
    def session(self) -> Iterator[Session]:
        """
        Provide a transactional session context.

        Usage:
            with db_manager.session() as session:
                session.add(record)
                # Auto-commits on success, rolls back on exception

        Acquires the DB lock for the lifetime of the session. With StaticPool
        every thread shares one underlying SQLite connection, and concurrent
        cursor.execute calls from different threads trip SQLITE_MISUSE
        ("bad parameter or other API misuse"). Serialising at the session
        boundary eliminates that race; reentrant so nested session() blocks
        and the explicit _write_lock acquisitions still work.
        """
        with self._write_lock:
            session = self._SessionFactory()
            try:
                yield session
                session.commit()
            except Exception as e:
                session.rollback()
                logger.error(f"Session error: {e}")
                raise
            finally:
                session.close()

    # =========================================================================
    # Analysis Results
    # =========================================================================

    def save_analysis(self, analysis: AnalysisResult) -> int:
        """
        Save a single analysis result to the database.

        Args:
            analysis: AnalysisResult from processing

        Returns:
            Database ID of the saved analysis, or -1 for final test files
        """
        # Skip Final Test files - they're already saved in processor via save_final_test
        if getattr(analysis, 'file_type', 'trim') == 'final_test':
            logger.debug(f"Skipping save_analysis for Final Test: {analysis.metadata.filename}")
            return getattr(analysis, 'final_test_id', -1) or -1

        # Skip Smoothness files - they're already saved in processor via save_smoothness_result
        if getattr(analysis, 'file_type', 'trim') == 'smoothness':
            logger.debug(f"Skipping save_analysis for Smoothness: {analysis.metadata.filename}")
            return getattr(analysis, 'smoothness_id', -1) or -1

        # session() acquires _write_lock internally (RLock-reentrant), so
        # wrapping with another `with self._write_lock` here is redundant and
        # only adds confusion to the lock graph. The session context is
        # sufficient for SQLite serialization.
        with self.session() as session:
            # Check for existing record by the DB's UNIQUE-constraint key
            # (filename, file_date, model, serial).  Pre-fix this filtered on
            # (filename, file_path), but the UNIQUE constraint is keyed on
            # the metadata tuple -- the same file on a different path string
            # (UNC vs mapped drive, folder reorg, network share migration)
            # missed this lookup and then raised IntegrityError at INSERT
            # time.  Aligning the lookup with the constraint turns these
            # re-presentations into idempotent UPDATEs.
            # See save_final_test for the parallel lesson.
            existing = session.query(DBAnalysisResult).filter(
                DBAnalysisResult.filename == analysis.metadata.filename,
                DBAnalysisResult.file_date == analysis.metadata.file_date,
                DBAnalysisResult.model == analysis.metadata.model,
                DBAnalysisResult.serial == analysis.metadata.serial,
            ).first()

            if existing:
                logger.debug(f"Updating existing analysis: {analysis.metadata.filename}")
                return self._update_existing_analysis(session, analysis)

            # No existing record, create new one
            db_analysis = self._map_analysis_to_db(analysis)
            session.add(db_analysis)
            session.flush()  # Get ID before commit

            # Record as processed file
            # ERROR results are marked success=False so they get retried
            is_success = analysis.overall_status != AnalysisStatus.ERROR
            self._record_processed_file(
                session,
                analysis.metadata.file_path,
                db_analysis.id,
                success=is_success,
            )

            logger.debug(f"Saved new analysis: {analysis.metadata.filename} (ID: {db_analysis.id})")
            return db_analysis.id

    def save_batch(self, analyses: List[AnalysisResult]) -> List[int]:
        """
        Save multiple analysis results efficiently.

        Args:
            analyses: List of AnalysisResult objects

        Returns:
            List of database IDs
        """
        saved_ids = []

        with self._write_lock:
            for analysis in analyses:
                try:
                    with self.session() as session:
                        # Skip Final Test files - they're already saved in processor
                        if getattr(analysis, 'file_type', 'trim') == 'final_test':
                            saved_ids.append(getattr(analysis, 'final_test_id', -1) or -1)
                            continue

                        # Check for existing record by the DB UNIQUE key
                        # (filename, file_date, model, serial).  Must match
                        # save_analysis; see comment there for rationale.
                        existing = session.query(DBAnalysisResult).filter(
                            DBAnalysisResult.filename == analysis.metadata.filename,
                            DBAnalysisResult.file_date == analysis.metadata.file_date,
                            DBAnalysisResult.model == analysis.metadata.model,
                            DBAnalysisResult.serial == analysis.metadata.serial,
                        ).first()

                        if existing:
                            # Update existing record
                            updated_id = self._update_existing_analysis(session, analysis)
                            saved_ids.append(updated_id)
                        else:
                            # Create new record
                            db_analysis = self._map_analysis_to_db(analysis)
                            session.add(db_analysis)
                            session.flush()

                            is_success = analysis.overall_status != AnalysisStatus.ERROR
                            self._record_processed_file(
                                session,
                                analysis.metadata.file_path,
                                db_analysis.id,
                                success=is_success,
                            )

                            saved_ids.append(db_analysis.id)
                except Exception as e:
                    logger.error(f"Failed to save analysis {getattr(analysis.metadata, 'filename', '?')}: {e}")

        logger.info(f"Saved batch of {len(saved_ids)} analyses")
        return saved_ids

    def get_analysis(self, analysis_id: int) -> Optional[AnalysisResult]:
        """
        Retrieve a single analysis by ID.

        Args:
            analysis_id: Database ID

        Returns:
            AnalysisResult or None if not found
        """
        with self.session() as session:
            db_analysis = session.get(DBAnalysisResult, analysis_id)

            if db_analysis is None:
                return None

            return self._map_db_to_analysis(db_analysis)

    def get_historical_data(
        self,
        model: Optional[str] = None,
        days_back: int = 30,
        limit: int = 1000,
        light_load: bool = False,
    ) -> List[AnalysisResult]:
        """
        Get historical analysis data with optional filtering.

        Filters and sorts by trim date (file_date), not processing date (timestamp).
        This shows files based on when they were trimmed, not when processed.

        Args:
            model: Filter by model number (optional)
            days_back: How many days back to query (based on trim date)
            limit: Maximum number of results
            light_load: When True, skip per-track JSON-blob columns
                (position_data, error_data, upper/lower limits, theory_data,
                test_volts, untrimmed_positions/errors). Use this for the
                Excel summary export which only needs scalar metrics —
                deferring the blobs cuts reconstruction time roughly 5–10×
                on a multi-GB database.

        Returns:
            List of AnalysisResult objects sorted by trim date (newest first)
        """
        with self.session() as session:
            tracks_loader = joinedload(DBAnalysisResult.tracks)
            if light_load:
                # SafeJSON columns are the bulk of track_results size and
                # decode work. The summary export never reads them, so
                # defer() them — Pydantic TrackData declares them Optional,
                # so None is a legal value end-to-end.
                # Load.defer() takes one column per call in SQLAlchemy 2.x —
                # chain it for each blob column we want to skip.
                for _col in (
                    DBTrackResult.position_data,
                    DBTrackResult.error_data,
                    DBTrackResult.upper_limits,
                    DBTrackResult.lower_limits,
                    DBTrackResult.theory_data,
                    DBTrackResult.test_volts,
                    DBTrackResult.untrimmed_positions,
                    DBTrackResult.untrimmed_errors,
                ):
                    tracks_loader = tracks_loader.defer(_col)
            # Use joinedload to fetch tracks in single query (avoids N+1)
            query = session.query(DBAnalysisResult).options(tracks_loader)

            # Filter by trim date (file_date), not processing date (timestamp)
            cutoff_date = datetime.now() - timedelta(days=days_back)
            query = query.filter(DBAnalysisResult.file_date >= cutoff_date)

            if model:
                query = query.filter(DBAnalysisResult.model == model)

            # Order by trim date (file_date), newest first
            query = query.order_by(desc(DBAnalysisResult.file_date)).limit(limit)

            results = []
            for db_analysis in query.all():
                mapped = self._map_db_to_analysis(db_analysis)
                if mapped is not None:  # Filter out failed mappings
                    results.append(mapped)

            return results

    # =========================================================================
    # Incremental Processing
    # =========================================================================

    def is_file_processed(self, file_path: Union[str, Path]) -> bool:
        """
        Check if a file has already been successfully processed.

        Uses SHA-256 hash for accurate duplicate detection even if file
        was moved or renamed. Only returns True for successful processing -
        errors will be retried. Checks both trim files and Final Test files.

        Args:
            file_path: Path to the file

        Returns:
            True if file was already successfully processed
        """
        file_path = Path(file_path)

        if not file_path.exists():
            return False

        file_hash = calculate_file_hash(file_path)

        from laser_trim_analyzer.database.models import FinalTestResult as DBFinalTestResult

        with self.session() as session:
            # Check trim files (ProcessedFile) - only successful ones
            exists = (
                session.query(DBProcessedFile)
                .filter(
                    DBProcessedFile.file_hash == file_hash,
                    DBProcessedFile.success == True
                )
                .first()
            ) is not None

            if exists:
                return True

            # Also check Final Test files
            exists = (
                session.query(DBFinalTestResult)
                .filter(DBFinalTestResult.file_hash == file_hash)
                .first()
            ) is not None

            if exists:
                return True

            # Also check Output Smoothness files
            try:
                from laser_trim_analyzer.database.models import SmoothnessResult as DBSmoothnessResult
                exists = (
                    session.query(DBSmoothnessResult)
                    .filter(DBSmoothnessResult.file_hash == file_hash)
                    .first()
                ) is not None
                if exists:
                    return True
            except Exception:
                pass  # Table may not exist yet

            return False

    def get_unprocessed_files(self, file_paths: List[Path]) -> List[Path]:
        """
        Filter a list of files to only those not yet processed.

        Efficient batch operation for incremental processing.

        Args:
            file_paths: List of file paths to check

        Returns:
            List of paths that have not been processed
        """
        if not file_paths:
            return []

        # Calculate hashes for all files
        file_hashes = {}
        for path in file_paths:
            path = Path(path)
            if path.exists():
                file_hashes[calculate_file_hash(path)] = path

        if not file_hashes:
            return []

        # Query for existing hashes across all file type tables
        hash_list = list(file_hashes.keys())
        with self.session() as session:
            existing_hashes = set(
                row.file_hash for row in
                session.query(DBProcessedFile.file_hash)
                .filter(DBProcessedFile.file_hash.in_(hash_list))
                .filter(DBProcessedFile.success == True)
                .all()
            )

            # Also check Final Test hashes
            from laser_trim_analyzer.database.models import FinalTestResult as DBFinalTestResult
            existing_hashes.update(
                row.file_hash for row in
                session.query(DBFinalTestResult.file_hash)
                .filter(DBFinalTestResult.file_hash.in_(hash_list))
                .all()
                if row.file_hash
            )

            # Also check Smoothness hashes
            try:
                from laser_trim_analyzer.database.models import SmoothnessResult as DBSmoothnessResult
                existing_hashes.update(
                    row.file_hash for row in
                    session.query(DBSmoothnessResult.file_hash)
                    .filter(DBSmoothnessResult.file_hash.in_(hash_list))
                    .all()
                    if row.file_hash
                )
            except Exception:
                pass  # Table may not exist yet

        # Return files whose hash is not in database
        return [
            path for hash_val, path in file_hashes.items()
            if hash_val not in existing_hashes
        ]

    def _record_processed_file(
        self,
        session: Session,
        file_path: Path,
        analysis_id: int,
        success: bool = True,
    ) -> None:
        """Record a file as processed.

        Args:
            session: Active database session
            file_path: Path to the processed file
            analysis_id: ID of the saved AnalysisResult
            success: False for ERROR results — allows retry on next run
        """
        file_path = Path(file_path)

        if not file_path.exists():
            return

        file_hash = calculate_file_hash(file_path)

        # Check if already recorded before inserting (avoids IntegrityError
        # which would rollback the entire transaction including parent analysis)
        existing = session.execute(
            select(DBProcessedFile.id).where(DBProcessedFile.file_hash == file_hash)
        ).scalar_one_or_none()
        if existing is None:
            processed_file = DBProcessedFile(
                filename=file_path.name,
                file_path=str(file_path),
                file_hash=file_hash,
                file_size=file_path.stat().st_size,
                file_modified_date=datetime.fromtimestamp(file_path.stat().st_mtime),
                analysis_id=analysis_id,
                success=success,
            )
            session.add(processed_file)
            session.flush()
        else:
            # Relink an existing hash row (e.g. a prior ERROR, or a file that was
            # skipped/misclassified earlier) to this analysis so a transient
            # failure isn't permanent. The old behavior silently no-op'd, leaving
            # the row stuck at its prior (e.g. success=False) state.
            session.query(DBProcessedFile).filter(
                DBProcessedFile.file_hash == file_hash
            ).update({
                DBProcessedFile.analysis_id: analysis_id,
                DBProcessedFile.success: success,
                DBProcessedFile.file_path: str(file_path),
                DBProcessedFile.filename: file_path.name,
            })
            session.flush()

    # =========================================================================
    # QA Alerts
    # =========================================================================

    def create_alert(
        self,
        analysis_id: int,
        alert_type: str,
        severity: str,
        message: str,
        track_id: Optional[str] = None,
        metric_value: Optional[float] = None,
        threshold_value: Optional[float] = None,
    ) -> int:
        """
        Create a QA alert.

        Args:
            analysis_id: Related analysis ID
            alert_type: Type of alert (from AlertType enum)
            severity: Severity level (Critical, High, Medium, Low)
            message: Alert message
            track_id: Related track (optional)
            metric_value: Value that triggered alert (optional)
            threshold_value: Threshold that was exceeded (optional)

        Returns:
            Database ID of created alert
        """
        with self.session() as session:
            alert = DBQAAlert(
                analysis_id=analysis_id,
                alert_type=DBAlertType[alert_type] if isinstance(alert_type, str) else alert_type,
                severity=severity,
                message=message,
                track_id=track_id,
                metric_value=metric_value,
                threshold_value=threshold_value,
            )
            session.add(alert)
            session.flush()

            logger.info(f"Created alert: {alert_type} - {message}")
            return alert.id

    def get_unresolved_alerts(self, limit: int = 50) -> List[Dict[str, Any]]:
        """
        Get unresolved QA alerts.

        Args:
            limit: Maximum number to return

        Returns:
            List of alert dictionaries
        """
        with self.session() as session:
            alerts = (
                session.query(DBQAAlert)
                .filter(DBQAAlert.resolved == False)
                .order_by(
                    # Critical first, then by date
                    desc(DBQAAlert.severity == "Critical"),
                    desc(DBQAAlert.created_date)
                )
                .limit(limit)
                .all()
            )

            return [
                {
                    "id": a.id,
                    "analysis_id": a.analysis_id,
                    "alert_type": a.alert_type.value if a.alert_type else None,
                    "severity": a.severity,
                    "message": a.message,
                    "created_date": a.created_date,
                    "acknowledged": a.acknowledged,
                }
                for a in alerts
            ]

    def acknowledge_alert(self, alert_id: int, acknowledged_by: str) -> bool:
        """
        Acknowledge an alert.

        Args:
            alert_id: Alert ID
            acknowledged_by: User who acknowledged

        Returns:
            True if successful
        """
        with self.session() as session:
            alert = session.get(DBQAAlert, alert_id)
            if alert:
                alert.acknowledged = True
                alert.acknowledged_by = acknowledged_by
                alert.acknowledged_date = utc_now()
                return True
            return False

    def resolve_alert(
        self,
        alert_id: int,
        resolved_by: str,
        resolution_notes: Optional[str] = None
    ) -> bool:
        """
        Resolve an alert.

        Args:
            alert_id: Alert ID
            resolved_by: User who resolved
            resolution_notes: Notes about resolution

        Returns:
            True if successful
        """
        with self.session() as session:
            alert = session.get(DBQAAlert, alert_id)
            if alert:
                if not alert.acknowledged:
                    # Auto-acknowledge when resolving
                    alert.acknowledged = True
                    alert.acknowledged_by = resolved_by
                    alert.acknowledged_date = utc_now()

                alert.resolved = True
                alert.resolved_by = resolved_by
                alert.resolved_date = utc_now()
                alert.resolution_notes = resolution_notes
                return True
            return False

    # =========================================================================
    # Dashboard Queries
    # =========================================================================

    def get_dashboard_stats(self, days_back: int = 7,
                            element_type: Optional[str] = None,
                            product_class: Optional[str] = None) -> Dict[str, Any]:
        """
        Get statistics for dashboard display.

        Filters by trim date (file_date), not processing date.
        Optionally filters by element type and/or product class via model_specs join.

        Args:
            days_back: Number of days to include (based on trim date)
            element_type: Filter to models with this element type
            product_class: Filter to models with this product class

        Returns:
            Dictionary with dashboard statistics
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Build model filter list from model_specs if filters are active
            filter_models = None
            if element_type or product_class:
                q = session.query(ModelSpec.model)
                if element_type:
                    q = q.filter(ModelSpec.element_type == element_type)
                if product_class:
                    q = q.filter(ModelSpec.product_class == product_class)
                filter_models = [r[0] for r in q.all()]
                if not filter_models:
                    # No models match — return empty stats
                    return {
                        "total_analyses": 0, "total_files": 0,
                        "passed": 0, "failed": 0, "pass_rate": 0.0,
                        "sigma_pass_rate": 0.0, "linearity_pass_rate": 0.0,
                        "total_tracks": 0, "unresolved_alerts": 0,
                        "high_risk_count": 0, "period_days": days_back,
                        "today_count": 0, "week_count": 0,
                        "daily_trend": [], "linearity_daily_trend": [],
                    }

            def _base_filter(query):
                """Apply date and optional model filters."""
                query = query.filter(DBAnalysisResult.file_date >= cutoff_date)
                if filter_models is not None:
                    query = query.filter(DBAnalysisResult.model.in_(filter_models))
                return query

            # Count analyses - filter by trim date. UNTRIMMED records belong
            # in raw totals (the file was processed) but NOT in the pass-rate
            # denominator since pass/fail isn't defined without a trim.
            total_analyses = (
                _base_filter(session.query(func.count(DBAnalysisResult.id)))
                .scalar()
            ) or 0
            trimmed_total = (
                _base_filter(session.query(func.count(DBAnalysisResult.id)))
                .filter(DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name)
                .scalar()
            ) or 0

            # Count by status - filter by trim date
            status_q = session.query(
                DBAnalysisResult.overall_status,
                func.count(DBAnalysisResult.id)
            ).filter(DBAnalysisResult.file_date >= cutoff_date)
            if filter_models is not None:
                status_q = status_q.filter(DBAnalysisResult.model.in_(filter_models))
            status_counts = status_q.group_by(DBAnalysisResult.overall_status).all()

            passed = 0
            failed = 0
            for status, count in status_counts:
                if _status_matches(status, DBStatusType.PASS):
                    passed = count
                elif _status_matches(status, DBStatusType.FAIL, DBStatusType.ERROR):
                    failed = count

            # Count unresolved alerts
            unresolved_alerts = (
                session.query(func.count(DBQAAlert.id))
                .filter(DBQAAlert.resolved == False)
                .scalar()
            ) or 0

            # Get high-risk count - filter by trim date
            high_risk_q = (
                session.query(func.count(DBTrackResult.id))
                .join(DBAnalysisResult)
                .filter(
                    DBAnalysisResult.file_date >= cutoff_date,
                    DBTrackResult.risk_category == DBRiskCategory.HIGH
                )
            )
            if filter_models is not None:
                high_risk_q = high_risk_q.filter(DBAnalysisResult.model.in_(filter_models))
            high_risk = high_risk_q.scalar() or 0

            pass_rate = (passed / trimmed_total * 100) if trimmed_total > 0 else 0.0

            # Get track-level sigma and linearity pass rates - filter by trim date.
            # Exclude UNTRIMMED tracks (sigma_pass/linearity_pass are NULL) so
            # they don't sit in the denominator and deflate the pass rate.
            track_stats_q = (
                session.query(
                    func.count(DBTrackResult.id).label('total_tracks'),
                    func.sum(case((DBTrackResult.sigma_pass == True, 1), else_=0)).label('sigma_passed'),
                    func.sum(case((DBTrackResult.linearity_pass == True, 1), else_=0)).label('linearity_passed'),
                )
                .join(DBAnalysisResult)
                .filter(
                    DBAnalysisResult.file_date >= cutoff_date,
                    DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                )
            )
            if filter_models is not None:
                track_stats_q = track_stats_q.filter(DBAnalysisResult.model.in_(filter_models))
            track_stats = track_stats_q.first()

            total_tracks = track_stats.total_tracks or 0
            sigma_passed = track_stats.sigma_passed or 0
            linearity_passed = track_stats.linearity_passed or 0

            sigma_pass_rate = (sigma_passed / total_tracks * 100) if total_tracks > 0 else 0.0
            linearity_pass_rate = (linearity_passed / total_tracks * 100) if total_tracks > 0 else 0.0

            # Total files (all time)
            total_files = (
                session.query(func.count(DBAnalysisResult.id))
                .scalar()
            ) or 0

            # Today's count (by trim date)
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            today_q = session.query(func.count(DBAnalysisResult.id)).filter(DBAnalysisResult.file_date >= today)
            if filter_models is not None:
                today_q = today_q.filter(DBAnalysisResult.model.in_(filter_models))
            today_count = today_q.scalar() or 0

            # This week's count (by trim date)
            week_start = today - timedelta(days=today.weekday())
            week_q = session.query(func.count(DBAnalysisResult.id)).filter(DBAnalysisResult.file_date >= week_start)
            if filter_models is not None:
                week_q = week_q.filter(DBAnalysisResult.model.in_(filter_models))
            week_count = week_q.scalar() or 0

            # Daily trend for the past N days (by trim date) - optimized single query
            trend_start = today - timedelta(days=days_back - 1)

            # Single query with GROUP BY date to get all days at once
            daily_q = (
                session.query(
                    func.date(DBAnalysisResult.file_date).label('day'),
                    func.count(DBAnalysisResult.id).label('total'),
                    func.sum(
                        case(
                            (DBAnalysisResult.overall_status == DBStatusType.PASS, 1),
                            else_=0
                        )
                    ).label('passed')
                )
                .filter(
                    DBAnalysisResult.file_date >= trend_start,
                    # Exclude UNTRIMMED so the per-day pass_rate uses the gradeable
                    # denominator and matches the headline (was ~20pts off).
                    DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name,
                )
            )
            if filter_models is not None:
                daily_q = daily_q.filter(DBAnalysisResult.model.in_(filter_models))
            daily_data = (
                daily_q
                .group_by(func.date(DBAnalysisResult.file_date))
                .all()
            )

            # Convert to dict for quick lookup
            daily_dict = {str(row.day): {'total': row.total, 'passed': row.passed or 0} for row in daily_data}

            # Build daily trend list (fill in zeros for missing days)
            daily_trend = []
            for i in range(days_back):
                day = trend_start + timedelta(days=i)
                day_str = day.strftime("%Y-%m-%d")
                day_data = daily_dict.get(day_str, {'total': 0, 'passed': 0})
                day_total = day_data['total']
                day_passed = day_data['passed']
                day_pass_rate = (day_passed / day_total * 100) if day_total > 0 else 0.0
                daily_trend.append({
                    "date": day.strftime("%m/%d"),
                    "total": day_total,
                    "passed": day_passed,
                    "pass_rate": day_pass_rate,
                })

            # Linearity daily trend (track-level, independent of sigma).
            # Exclude UNTRIMMED — their linearity_pass is NULL.
            lin_daily_q = (
                session.query(
                    func.date(DBAnalysisResult.file_date).label('day'),
                    func.count(DBTrackResult.id).label('total'),
                    func.sum(case((DBTrackResult.linearity_pass == True, 1), else_=0)).label('passed')
                )
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(
                    DBAnalysisResult.file_date >= trend_start,
                    DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                )
            )
            if filter_models is not None:
                lin_daily_q = lin_daily_q.filter(DBAnalysisResult.model.in_(filter_models))
            linearity_daily_data = (
                lin_daily_q
                .group_by(func.date(DBAnalysisResult.file_date))
                .all()
            )

            lin_daily_dict = {str(row.day): {'total': row.total, 'passed': row.passed or 0}
                              for row in linearity_daily_data}

            linearity_daily_trend = []
            for i in range(days_back):
                day = trend_start + timedelta(days=i)
                day_str = day.strftime("%Y-%m-%d")
                day_data = lin_daily_dict.get(day_str, {'total': 0, 'passed': 0})
                day_total = day_data['total']
                day_passed = day_data['passed']
                day_pass_rate = (day_passed / day_total * 100) if day_total > 0 else 0.0
                linearity_daily_trend.append({
                    "date": day.strftime("%m/%d"),
                    "total": day_total,
                    "passed": day_passed,
                    "pass_rate": day_pass_rate,
                })

            return {
                "total_analyses": total_analyses,
                "total_files": total_files,
                "passed": passed,
                "failed": failed,
                "pass_rate": pass_rate,
                "sigma_pass_rate": sigma_pass_rate,
                "linearity_pass_rate": linearity_pass_rate,
                "total_tracks": total_tracks,
                "unresolved_alerts": unresolved_alerts,
                "high_risk_count": high_risk,
                "period_days": days_back,
                "today_count": today_count,
                "week_count": week_count,
                "daily_trend": daily_trend,
                "linearity_daily_trend": linearity_daily_trend,
            }

    def get_pass_rate_by_category(self, category: str = "element_type",
                                  days_back: int = 30) -> List[Dict[str, Any]]:
        """
        Get pass rate grouped by element_type or product_class.

        Args:
            category: "element_type" or "product_class"
            days_back: Number of days to look back

        Returns:
            List of dicts with keys: category, total, passed, pass_rate
        """
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            spec_col = ModelSpec.element_type if category == "element_type" else ModelSpec.product_class

            results = session.query(
                spec_col.label("category"),
                func.count(DBAnalysisResult.id).label("total"),
                func.sum(
                    case(
                        (DBAnalysisResult.overall_status == DBStatusType.PASS, 1),
                        else_=0
                    )
                ).label("passed")
            ).join(
                ModelSpec,
                DBAnalysisResult.model == ModelSpec.model
            ).filter(
                DBAnalysisResult.file_date >= cutoff,
                spec_col.isnot(None)
            ).group_by(spec_col).all()

            return [
                {
                    "category": r.category,
                    "total": r.total,
                    "passed": r.passed or 0,
                    "pass_rate": ((r.passed or 0) / r.total * 100) if r.total > 0 else 0
                }
                for r in results
            ]

    def get_last_batch_stats(self) -> Dict[str, Any]:
        """
        Get statistics for the most recent processing batch.

        A "batch" is defined as files processed within a 1-hour window.
        Uses the timestamp field (when file was added to DB) to identify batches.

        Returns:
            Dictionary with last batch statistics
        """
        with self.session() as session:
            # Find the most recent processing timestamp
            latest_timestamp = (
                session.query(func.max(DBAnalysisResult.timestamp))
                .scalar()
            )

            if not latest_timestamp:
                return {
                    "has_batch": False,
                    "files_processed": 0,
                    "passed": 0,
                    "warnings": 0,
                    "failed": 0,
                    "pass_rate": 0.0,
                    "batch_start": None,
                    "batch_end": None,
                }

            # Define batch as files processed within 1 hour of the latest
            batch_start = latest_timestamp - timedelta(hours=1)

            # Get files in this batch
            batch_files = (
                session.query(DBAnalysisResult)
                .filter(DBAnalysisResult.timestamp >= batch_start)
                .all()
            )

            if not batch_files:
                return {
                    "has_batch": False,
                    "files_processed": 0,
                    "passed": 0,
                    "warnings": 0,
                    "failed": 0,
                    "pass_rate": 0.0,
                    "batch_start": None,
                    "batch_end": None,
                }

            # Count by status
            passed = sum(1 for f in batch_files if f.overall_status == DBStatusType.PASS)
            warnings = sum(1 for f in batch_files if f.overall_status == DBStatusType.WARNING)
            failed = sum(1 for f in batch_files if f.overall_status in (DBStatusType.FAIL, DBStatusType.ERROR))
            total = len(batch_files)

            # Find actual batch time range
            actual_start = min(f.timestamp for f in batch_files)
            actual_end = max(f.timestamp for f in batch_files)

            return {
                "has_batch": True,
                "files_processed": total,
                "passed": passed,
                "warnings": warnings,
                "failed": failed,
                "pass_rate": (passed / total * 100) if total > 0 else 0.0,
                "batch_start": actual_start,
                "batch_end": actual_end,
            }

    def get_overall_stats(self) -> Dict[str, Any]:
        """
        Get overall database statistics across all time.

        Returns:
            Dictionary with overall statistics including date range and model breakdown
        """
        with self.session() as session:
            # Total counts
            total_files = (
                session.query(func.count(DBAnalysisResult.id))
                .scalar()
            ) or 0

            if total_files == 0:
                return {
                    "total_files": 0,
                    "passed": 0,
                    "warnings": 0,
                    "failed": 0,
                    "pass_rate": 0.0,
                    "oldest_date": None,
                    "newest_date": None,
                    "unique_models": 0,
                }

            # Count by status
            status_counts = (
                session.query(
                    DBAnalysisResult.overall_status,
                    func.count(DBAnalysisResult.id)
                )
                .group_by(DBAnalysisResult.overall_status)
                .all()
            )

            passed = 0
            warnings = 0
            failed = 0
            for status, count in status_counts:
                if _status_matches(status, DBStatusType.PASS):
                    passed = count
                elif _status_matches(status, DBStatusType.WARNING):
                    warnings = count
                elif _status_matches(status, DBStatusType.FAIL, DBStatusType.ERROR):
                    failed += count

            # Date range (using file_date - when files were trimmed)
            date_range = (
                session.query(
                    func.min(DBAnalysisResult.file_date),
                    func.max(DBAnalysisResult.file_date)
                )
                .first()
            )
            oldest_date, newest_date = date_range if date_range else (None, None)

            # Unique models count
            unique_models = (
                session.query(func.count(func.distinct(DBAnalysisResult.model)))
                .scalar()
            ) or 0

            # Track-level pass rates (sigma and linearity). Exclude UNTRIMMED
            # tracks -- their sigma_pass/linearity_pass are NULL, so they belong
            # in neither numerator nor denominator (would deflate the rate).
            total_tracks = (
                session.query(func.count(DBTrackResult.id))
                .filter(DBTrackResult.status != DBStatusType.UNTRIMMED.name)
                .scalar()
            ) or 0

            sigma_passed = (
                session.query(func.count(DBTrackResult.id))
                .filter(DBTrackResult.sigma_pass == True)
                .scalar()
            ) or 0

            linearity_passed = (
                session.query(func.count(DBTrackResult.id))
                .filter(DBTrackResult.linearity_pass == True)
                .scalar()
            ) or 0

            # Yield denominator = trimmed (gradeable) files. UNTRIMMED test-sweeps
            # have no trim result, so excluding them keeps pass_rate honest and
            # consistent with the dashboard headline.
            trimmed_total = (
                session.query(func.count(DBAnalysisResult.id))
                .filter(DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name)
                .scalar()
            ) or 0

            return {
                "total_files": total_files,
                "trimmed_total": trimmed_total,
                "passed": passed,
                "warnings": warnings,
                "failed": failed,
                "pass_rate": (passed / trimmed_total * 100) if trimmed_total > 0 else 0.0,
                "oldest_date": oldest_date,
                "newest_date": newest_date,
                "unique_models": unique_models,
                "total_tracks": total_tracks,
                "sigma_pass_rate": (sigma_passed / total_tracks * 100) if total_tracks > 0 else 0.0,
                "linearity_pass_rate": (linearity_passed / total_tracks * 100) if total_tracks > 0 else 0.0,
            }

    def get_system_comparison(self, days_back: int = 90) -> Dict[str, Any]:
        """Get System A vs System B comparison statistics."""
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            system_data = (
                session.query(
                    DBAnalysisResult.system,
                    func.count(func.distinct(DBAnalysisResult.id)).label('total_files'),
                    func.count(DBTrackResult.id).label('total_tracks'),
                    func.sum(case(
                        (DBTrackResult.linearity_pass == True, 1), else_=0
                    )).label('lin_passed'),
                    func.sum(case(
                        (DBTrackResult.sigma_pass == True, 1), else_=0
                    )).label('sigma_passed'),
                )
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(
                    DBAnalysisResult.file_date >= cutoff_date,
                    DBAnalysisResult.system.isnot(None),
                    # Exclude UNTRIMMED tracks from the rate denominator (NULL pass flags).
                    DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                )
                .group_by(DBAnalysisResult.system)
                .all()
            )

            result = {"system_a": None, "system_b": None, "system_c": None}
            for row in system_data:
                # Direct comparison — _status_matches is for DBStatusType only
                sys_val = row.system.value if isinstance(row.system, DBSystemType) else str(row.system)
                if sys_val == DBSystemType.A.value:
                    key = "system_a"
                elif sys_val == DBSystemType.B.value:
                    key = "system_b"
                elif sys_val == DBSystemType.C.value:
                    key = "system_c"
                else:
                    continue
                total_tracks = row.total_tracks or 0
                result[key] = {
                    "total_files": row.total_files or 0,
                    "total_tracks": total_tracks,
                    "linearity_pass_rate": (row.lin_passed or 0) / total_tracks * 100 if total_tracks > 0 else 0,
                    "sigma_pass_rate": (row.sigma_passed or 0) / total_tracks * 100 if total_tracks > 0 else 0,
                }
            return result

    def get_ft_dashboard_stats(self, days_back: int = 90) -> Dict[str, Any]:
        """Get Final Test statistics for dashboard display."""
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            FinalTestTrack as DBFinalTestTrack,
        )

        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            total_ft = (
                session.query(func.count(DBFinalTestResult.id))
                .filter(DBFinalTestResult.file_date >= cutoff_date)
                .scalar()
            ) or 0

            if total_ft == 0:
                return {"total": 0, "pass_rate": 0, "linearity_pass_rate": 0,
                        "linked_count": 0, "link_rate": 0}

            ft_passed = (
                session.query(func.count(DBFinalTestResult.id))
                .filter(
                    DBFinalTestResult.file_date >= cutoff_date,
                    DBFinalTestResult.overall_status == DBStatusType.PASS,
                )
                .scalar()
            ) or 0

            linked_count = (
                session.query(func.count(DBFinalTestResult.id))
                .filter(
                    DBFinalTestResult.file_date >= cutoff_date,
                    DBFinalTestResult.linked_trim_id.isnot(None),
                )
                .scalar()
            ) or 0

            # Track-level linearity pass rate
            ft_track_stats = (
                session.query(
                    func.count(DBFinalTestTrack.id).label('total'),
                    func.sum(case(
                        (DBFinalTestTrack.linearity_pass == True, 1), else_=0
                    )).label('passed'),
                )
                .join(DBFinalTestResult)
                .filter(DBFinalTestResult.file_date >= cutoff_date)
                .first()
            )

            ft_total_tracks = (ft_track_stats.total or 0) if ft_track_stats else 0
            ft_lin_passed = (ft_track_stats.passed or 0) if ft_track_stats else 0

            return {
                "total": total_ft,
                "pass_rate": (ft_passed / total_ft * 100) if total_ft > 0 else 0,
                "linearity_pass_rate": (ft_lin_passed / ft_total_tracks * 100) if ft_total_tracks > 0 else 0,
                "linked_count": linked_count,
                "link_rate": (linked_count / total_ft * 100) if total_ft > 0 else 0,
            }

    def get_escape_overkill_analysis(self, days_back: int = 90, min_confidence: float = 0.70) -> Dict[str, Any]:
        """Analyze escapes and overkills by comparing trim vs FT linearity results.

        Escape = trim passed linearity but FT failed (bad unit shipped)
        Overkill = trim failed linearity but FT passed (unnecessarily rejected)
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
        )

        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # File-level comparison: FT linearity_pass vs trim ALL-tracks-pass
            # For each linked FT record, check if trim linearity passed (all tracks)
            linked_data = (
                session.query(
                    DBFinalTestResult.model,
                    DBFinalTestResult.linearity_pass.label('ft_lin_pass'),
                    # Trim passed linearity if minimum track pass is 1 (all passed)
                    func.min(case(
                        (DBTrackResult.linearity_pass == True, 1), else_=0
                    )).label('trim_all_pass'),
                )
                .join(DBAnalysisResult, DBFinalTestResult.linked_trim_id == DBAnalysisResult.id)
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(
                    DBFinalTestResult.file_date >= cutoff_date,
                    DBFinalTestResult.linked_trim_id.isnot(None),
                    DBFinalTestResult.linearity_pass.isnot(None),
                    DBFinalTestResult.match_confidence >= min_confidence,
                    # Exclude UNTRIMMED tracks: their NULL linearity_pass would force
                    # trim_all_pass=0, fabricating overkills and masking escapes.
                    DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                )
                .group_by(DBFinalTestResult.id, DBFinalTestResult.model, DBFinalTestResult.linearity_pass)
                .all()
            )

            if not linked_data:
                return {
                    "total_linked": 0, "escapes": 0, "overkills": 0,
                    "agreements": 0, "escape_rate": 0, "overkill_rate": 0,
                    "agreement_rate": 0, "worst_escape_models": [],
                }

            total = len(linked_data)
            escapes = 0
            overkills = 0
            true_positives = 0  # Both pass
            true_negatives = 0  # Both fail
            model_escapes = {}

            for row in linked_data:
                trim_pass = bool(row.trim_all_pass)
                ft_pass = bool(row.ft_lin_pass)

                if trim_pass and not ft_pass:
                    escapes += 1
                    model_escapes[row.model] = model_escapes.get(row.model, 0) + 1
                elif not trim_pass and ft_pass:
                    overkills += 1
                elif trim_pass and ft_pass:
                    true_positives += 1
                else:
                    true_negatives += 1

            agreements = true_positives + true_negatives
            worst_models = sorted(model_escapes.items(), key=lambda x: x[1], reverse=True)[:5]

            return {
                "total_linked": total,
                "escapes": escapes,
                "overkills": overkills,
                "agreements": agreements,
                "true_positives": true_positives,
                "true_negatives": true_negatives,
                "escape_rate": (escapes / total * 100) if total > 0 else 0,
                "overkill_rate": (overkills / total * 100) if total > 0 else 0,
                "agreement_rate": (agreements / total * 100) if total > 0 else 0,
                "worst_escape_models": [{"model": m, "count": c} for m, c in worst_models],
            }

    def get_model_trim_ft_agreement(self, model: str, days_back: Optional[int] = None,
                                    min_confidence: float = 0.70,
                                    cutoff_date: Optional[datetime] = None) -> Dict[str, Any]:
        """Per-model trim-vs-final-test view (Job 2 / 'test pass vs trim pass'): trim and FT
        linearity pass rates, escapes (passed trim but failed FT — a bad unit shipped),
        overkills (failed trim but passed FT — unnecessarily rejected), agreement, and the
        trim-pass-count distribution ('how many trim passes'). Escape/overkill serials listed
        for drill-down. Unlinked records are excluded from escape/overkill (need both stations).
        """
        from laser_trim_analyzer.database.models import FinalTestResult as DBFinalTestResult
        out: Dict[str, Any] = {
            "model": model, "trim_total": 0, "trim_pass": 0, "trim_pass_rate": None,
            "ft_total": 0, "ft_pass": 0, "ft_pass_rate": None,
            "linked": 0, "escapes": 0, "overkills": 0, "agreements": 0, "agreement_rate": None,
            "escape_units": [], "overkill_units": [],
            "trim_pass_count_avg": None, "trim_pass_count_dist": {},
        }
        with self.session() as session:
            # cutoff_date (data-anchored, from the caller) wins over the
            # wall-clock days_back — batch-loaded data can lag production.
            cutoff = cutoff_date if cutoff_date is not None else (
                (datetime.now() - timedelta(days=days_back)) if days_back else None)

            # Trim linearity pass rate (unit = analysis; pass = ALL tracks pass).
            tq = (session.query(
                    DBAnalysisResult.id,
                    func.min(case((DBTrackResult.linearity_pass == True, 1), else_=0)).label("all_pass"))
                  .join(DBTrackResult, DBTrackResult.analysis_id == DBAnalysisResult.id)
                  .filter(DBAnalysisResult.model == model,
                          DBTrackResult.status != DBStatusType.UNTRIMMED.name))
            if cutoff is not None:
                tq = tq.filter(DBAnalysisResult.file_date >= cutoff)
            trim_rows = tq.group_by(DBAnalysisResult.id).all()
            out["trim_total"] = len(trim_rows)
            out["trim_pass"] = sum(1 for r in trim_rows if r.all_pass)
            if out["trim_total"]:
                out["trim_pass_rate"] = out["trim_pass"] / out["trim_total"] * 100.0

            # Final-test linearity pass rate.
            fq = session.query(DBFinalTestResult.linearity_pass).filter(
                DBFinalTestResult.model == model, DBFinalTestResult.linearity_pass.isnot(None))
            if cutoff is not None:
                fq = fq.filter(DBFinalTestResult.file_date >= cutoff)
            ft_rows = fq.all()
            out["ft_total"] = len(ft_rows)
            out["ft_pass"] = sum(1 for (p,) in ft_rows if p)
            if out["ft_total"]:
                out["ft_pass_rate"] = out["ft_pass"] / out["ft_total"] * 100.0

            # Escapes / overkills on LINKED units (both stations present, confident match).
            lq = (session.query(
                    DBFinalTestResult.serial,
                    DBFinalTestResult.linearity_pass.label("ft_pass"),
                    func.min(case((DBTrackResult.linearity_pass == True, 1), else_=0)).label("trim_pass"))
                  .join(DBAnalysisResult, DBFinalTestResult.linked_trim_id == DBAnalysisResult.id)
                  .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                  .filter(DBFinalTestResult.model == model,
                          DBFinalTestResult.linked_trim_id.isnot(None),
                          DBFinalTestResult.linearity_pass.isnot(None),
                          DBFinalTestResult.match_confidence >= min_confidence,
                          DBTrackResult.status != DBStatusType.UNTRIMMED.name))
            if cutoff is not None:
                lq = lq.filter(DBFinalTestResult.file_date >= cutoff)
            linked = lq.group_by(DBFinalTestResult.id, DBFinalTestResult.serial,
                                 DBFinalTestResult.linearity_pass).all()
            out["linked"] = len(linked)
            agree = 0
            for r in linked:
                tp, fp = bool(r.trim_pass), bool(r.ft_pass)
                if tp and not fp:
                    out["escapes"] += 1; out["escape_units"].append(r.serial)
                elif not tp and fp:
                    out["overkills"] += 1; out["overkill_units"].append(r.serial)
                else:
                    agree += 1
            out["agreements"] = agree
            if out["linked"]:
                out["agreement_rate"] = agree / out["linked"] * 100.0

            # Trim-pass-count distribution ("how many trim passes").
            cq = (session.query(DBTrackResult.trim_pass_count, func.count())
                  .join(DBAnalysisResult, DBTrackResult.analysis_id == DBAnalysisResult.id)
                  .filter(DBAnalysisResult.model == model, DBTrackResult.trim_pass_count.isnot(None)))
            if cutoff is not None:
                cq = cq.filter(DBAnalysisResult.file_date >= cutoff)
            dist: Dict[int, int] = {}
            tot_c = 0
            sum_c = 0
            for cnt, n in cq.group_by(DBTrackResult.trim_pass_count).all():
                k = int(cnt)
                dist[k] = n
                tot_c += n
                sum_c += k * n
            out["trim_pass_count_dist"] = dist
            if tot_c:
                out["trim_pass_count_avg"] = sum_c / tot_c
        return out

    # Metrics surfaced by the Model-page History tab (Job 2b). Order = dropdown order.
    MEASUREMENT_HISTORY_METRICS = (
        "measured_electrical_angle", "untrimmed_resistance", "trimmed_resistance",
        "resistance_change_percent", "unit_length",
    )

    def get_model_measurement_history(self, model: str,
                                      days_back: Optional[int] = None,
                                      cutoff_date: Optional[datetime] = None) -> Dict[str, Any]:
        """Per-model measured-value history (Job 2b — 'pull all model data, measured angle,
        resistance values, historical linearity pass rates'). Returns each metric's
        (date, value) series over the record, summary stats, and linearity pass-rate by
        month. Trimmed tracks only (UNTRIMMED excluded so values reflect finished elements).
        """
        import statistics as _stats
        from collections import OrderedDict

        metrics = self.MEASUREMENT_HISTORY_METRICS
        out: Dict[str, Any] = {
            "model": model, "n": 0,
            "series": {m: [] for m in metrics}, "stats": {}, "passrate_periods": [],
        }
        with self.session() as session:
            # cutoff_date (data-anchored, from the caller) wins over the
            # wall-clock days_back — batch-loaded data can lag production.
            cutoff = cutoff_date if cutoff_date is not None else (
                (datetime.now() - timedelta(days=days_back)) if days_back else None)
            cols = [getattr(DBTrackResult, m) for m in metrics]
            q = (session.query(DBAnalysisResult.file_date, DBTrackResult.linearity_pass, *cols)
                 .join(DBTrackResult, DBTrackResult.analysis_id == DBAnalysisResult.id)
                 .filter(DBAnalysisResult.model == model,
                         DBTrackResult.status != DBStatusType.UNTRIMMED.name))
            if cutoff is not None:
                q = q.filter(DBAnalysisResult.file_date >= cutoff)
            rows = q.order_by(DBAnalysisResult.file_date).all()
            out["n"] = len(rows)

            per = OrderedDict()  # (year, month) -> [pass, total]
            for r in rows:
                fd, lp = r[0], r[1]
                for i, m in enumerate(metrics):
                    v = r[2 + i]
                    if fd is not None and v is not None:
                        out["series"][m].append((fd, float(v)))
                if fd is not None and lp is not None:
                    key = (fd.year, fd.month)
                    p, t = per.get(key, (0, 0))
                    per[key] = (p + (1 if lp else 0), t + 1)

            for m in metrics:
                vals = [v for _, v in out["series"][m]]
                if vals:
                    out["stats"][m] = {
                        "n": len(vals), "mean": _stats.fmean(vals),
                        "std": _stats.pstdev(vals) if len(vals) > 1 else 0.0,
                        "min": min(vals), "max": max(vals), "last": vals[-1]}
            out["passrate_periods"] = [
                (f"{y:04d}-{mo:02d}", p, t, (p / t * 100.0) if t else None)
                for (y, mo), (p, t) in per.items()]
        return out

    def get_heatmap_data(
        self, days_back: int = 90, period: str = 'week', min_samples: int = 10
    ) -> Dict[str, Any]:
        """Get model x time period pass rate matrix for heat map visualization.

        Returns:
            {models: [str], periods: [str], values: [[float]]}
            values[i][j] = pass rate for model i in period j (NaN if no data)
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Group by model and week/month period
            if period == 'month':
                period_expr = func.strftime('%Y-%m', DBAnalysisResult.file_date)
            else:
                # Week: use Monday of the week
                period_expr = func.strftime('%Y-W%W', DBAnalysisResult.file_date)

            rows = (
                session.query(
                    DBAnalysisResult.model,
                    period_expr.label('period'),
                    func.count(DBTrackResult.id).label('total'),
                    func.sum(case(
                        (DBTrackResult.linearity_pass == True, 1), else_=0
                    )).label('passed'),
                )
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(DBAnalysisResult.file_date >= cutoff_date)
                .group_by(DBAnalysisResult.model, period_expr)
                .having(func.count(DBTrackResult.id) >= min_samples)
                .all()
            )

            if not rows:
                return {"models": [], "periods": [], "values": []}

            # Build sets of unique models and periods
            model_set = set()
            period_set = set()
            data_map = {}  # (model, period) -> pass_rate

            for row in rows:
                model_set.add(row.model)
                period_set.add(row.period)
                rate = (row.passed / row.total * 100) if row.total > 0 else float('nan')
                data_map[(row.model, row.period)] = rate

            # Sort models by worst average pass rate (worst first)
            model_avgs = {}
            for model in model_set:
                rates = [data_map[(model, p)] for p in period_set if (model, p) in data_map]
                model_avgs[model] = sum(rates) / len(rates) if rates else 100
            models = sorted(model_set, key=lambda m: model_avgs[m])[:15]  # Top 15 worst
            periods = sorted(period_set)

            # Build matrix
            values = []
            for model in models:
                row = []
                for p in periods:
                    val = data_map.get((model, p), float('nan'))
                    row.append(val)
                values.append(row)

            return {"models": models, "periods": periods, "values": values}

    def get_escape_scatter_data(
        self, days_back: int = 90, max_points: int = 500,
        min_confidence: float = 0.70
    ) -> Dict[str, Any]:
        """Get paired trim/FT linearity errors for scatter plot.

        Returns:
            {trim_errors: [float], ft_errors: [float], spec_limit: float}
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            FinalTestTrack as DBFinalTestTrack,
        )

        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Join trim tracks with FT tracks via linked_trim_id
            rows = (
                session.query(
                    DBTrackResult.final_linearity_error_shifted.label('trim_lin_error'),
                    DBFinalTestTrack.linearity_error.label('ft_lin_error'),
                )
                .join(DBAnalysisResult, DBTrackResult.analysis_id == DBAnalysisResult.id)
                .join(DBFinalTestResult, DBFinalTestResult.linked_trim_id == DBAnalysisResult.id)
                .join(DBFinalTestTrack, DBFinalTestTrack.final_test_id == DBFinalTestResult.id)
                .filter(
                    DBFinalTestResult.file_date >= cutoff_date,
                    DBFinalTestResult.match_confidence >= min_confidence,
                    DBTrackResult.final_linearity_error_shifted.isnot(None),
                    DBFinalTestTrack.linearity_error.isnot(None),
                    # Match track IDs: exact match (A=A, B=B) or FT single-track maps to trim track A
                    or_(
                        DBTrackResult.track_id == DBFinalTestTrack.track_id,
                        and_(
                            DBFinalTestTrack.track_id == "default",
                            DBTrackResult.track_id.in_(["TRK1", "TRK2", "default"]),
                        ),
                    ),
                )
                .limit(max_points)
                .all()
            )

            if not rows:
                return {"trim_errors": [], "ft_errors": [], "spec_limit": 0.5}

            trim_errors = [float(r.trim_lin_error) for r in rows]
            ft_errors = [float(r.ft_lin_error) for r in rows]

            # Estimate spec limit from data (use 75th percentile as proxy)
            all_errors = sorted(trim_errors + ft_errors)
            if all_errors:
                idx = int(len(all_errors) * 0.75)
                spec_limit = float(all_errors[min(idx, len(all_errors) - 1)])
            else:
                spec_limit = 0.5

            return {
                "trim_errors": trim_errors,
                "ft_errors": ft_errors,
                "spec_limit": spec_limit,
            }

    def get_alerts(self, limit: int = 10, include_resolved: bool = False) -> List[Dict[str, Any]]:
        """
        Get recent alerts.

        Args:
            limit: Maximum number of alerts to return
            include_resolved: Whether to include resolved alerts

        Returns:
            List of alert dictionaries
        """
        with self.session() as session:
            # Join with AnalysisResult to get model info
            query = (
                session.query(DBQAAlert, DBAnalysisResult.model)
                .outerjoin(DBAnalysisResult, DBQAAlert.analysis_id == DBAnalysisResult.id)
            )

            if not include_resolved:
                query = query.filter(DBQAAlert.resolved == False)

            results = (
                query
                .order_by(DBQAAlert.created_date.desc())
                .limit(limit)
                .all()
            )

            return [
                {
                    "id": alert.id,
                    "analysis_id": alert.analysis_id,
                    "alert_type": alert.alert_type.value if alert.alert_type else "INFO",
                    "severity": alert.severity,
                    "message": alert.message,
                    "model": model or "Unknown",
                    "created_at": alert.created_date.strftime("%Y-%m-%d %H:%M") if alert.created_date else "",
                    "acknowledged": alert.acknowledged,
                    "resolved": alert.resolved,
                }
                for alert, model in results
            ]

    def get_model_stats(self, limit: int = 10) -> List[Dict[str, Any]]:
        """
        Get statistics by model number.

        Args:
            limit: Maximum number of models to return

        Returns:
            List of model statistics dictionaries
        """
        with self.session() as session:
            # Single query with conditional aggregation including track-level stats.
            # Outerjoin filters UNTRIMMED tracks at the JOIN level so they don't
            # inflate total_tracks (which is the denominator for pass rates).
            # The outerjoin still includes parent files that ONLY have UNTRIMMED
            # tracks — those rows just show 0 / 0 = 0% which is correct (no
            # trim outcome to grade).
            model_stats = (
                session.query(
                    DBAnalysisResult.model,
                    func.count(func.distinct(DBAnalysisResult.id)).label('count'),
                    func.count(func.distinct(case(
                        (DBAnalysisResult.overall_status == DBStatusType.PASS, DBAnalysisResult.id),
                        else_=None
                    ))).label('passed'),
                    func.count(DBTrackResult.id).label('total_tracks'),
                    func.sum(case((DBTrackResult.sigma_pass == True, 1), else_=0)).label('sigma_passed'),
                    func.sum(case((DBTrackResult.linearity_pass == True, 1), else_=0)).label('linearity_passed'),
                    # Gradeable file count (excludes UNTRIMMED) -- the honest
                    # denominator for the file-level pass_rate / failed count.
                    func.count(func.distinct(case(
                        (DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name,
                         DBAnalysisResult.id),
                        else_=None,
                    ))).label('trimmed_count'),
                )
                .outerjoin(
                    DBTrackResult,
                    and_(
                        DBAnalysisResult.id == DBTrackResult.analysis_id,
                        DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                    ),
                )
                .filter(DBAnalysisResult.model.isnot(None))
                .group_by(DBAnalysisResult.model)
                .order_by(func.count(func.distinct(DBAnalysisResult.id)).desc())
                .limit(limit)
                .all()
            )

            result = []
            for model, count, passed, total_tracks, sigma_passed, linearity_passed, trimmed_count in model_stats:
                passed = passed or 0
                total_tracks = total_tracks or 0
                sigma_passed = sigma_passed or 0
                linearity_passed = linearity_passed or 0
                trimmed_count = trimmed_count or 0

                # File-level yield over the gradeable (trimmed) population so
                # UNTRIMMED test-sweeps are not counted as failures.
                pass_rate = (passed / trimmed_count * 100) if trimmed_count > 0 else 0.0
                sigma_pass_rate = (sigma_passed / total_tracks * 100) if total_tracks > 0 else 0.0
                linearity_pass_rate = (linearity_passed / total_tracks * 100) if total_tracks > 0 else 0.0

                result.append({
                    "model": model,
                    "count": count,
                    "trimmed_count": trimmed_count,
                    "passed": passed,
                    "failed": trimmed_count - passed,
                    "pass_rate": pass_rate,
                    "sigma_pass_rate": sigma_pass_rate,
                    "linearity_pass_rate": linearity_pass_rate,
                })

            return result

    def get_trend_data(
        self,
        model: Optional[str] = None,
        days_back: int = 30,
        limit: int = 1000
    ) -> List[Dict[str, Any]]:
        """
        Get trend data for analysis.

        Filters by trim date (file_date), not processing date.

        Args:
            model: Filter by model number (None for all)
            days_back: Number of days to include (based on trim date)
            limit: Maximum number of records

        Returns:
            List of trend data dictionaries sorted by trim date
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            query = (
                session.query(
                    DBAnalysisResult.file_date,  # Trim date from file
                    DBAnalysisResult.model,
                    DBTrackResult.sigma_gradient,
                    DBTrackResult.sigma_threshold,
                    DBTrackResult.sigma_pass,
                    DBTrackResult.status,
                    DBTrackResult.unit_length,
                    DBTrackResult.linearity_spec,
                )
                .join(DBTrackResult)
                .filter(DBAnalysisResult.file_date >= cutoff_date)
            )

            if model:
                query = query.filter(DBAnalysisResult.model == model)

            results = (
                query
                .order_by(DBAnalysisResult.file_date.asc())  # Order by trim date
                .limit(limit)
                .all()
            )

            return [
                {
                    "date": r.file_date.strftime("%Y-%m-%d") if r.file_date else "",
                    "model": r.model,
                    "sigma_gradient": r.sigma_gradient,
                    "sigma_threshold": r.sigma_threshold,
                    "sigma_pass": r.sigma_pass,
                    "status": r.status.value if r.status else "UNKNOWN",
                    "unit_length": r.unit_length,
                    "linearity_spec": r.linearity_spec,
                }
                for r in results
            ]

    def search_for_export(
        self,
        model: Optional[str] = None,
        serial: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        limit: int = 500
    ) -> List[AnalysisResult]:
        """
        Search for units matching criteria for export selection.

        Supports partial matching on serial number (case-insensitive).
        Filters by trim date (file_date).

        Args:
            model: Filter by model number (exact match, None for all)
            serial: Filter by serial number (partial match, case-insensitive)
            date_from: Start of date range (inclusive)
            date_to: End of date range (inclusive)
            limit: Maximum number of results

        Returns:
            List of AnalysisResult objects sorted by trim date (newest first)
        """
        with self.session() as session:
            query = session.query(DBAnalysisResult)

            # Filter by model (exact match)
            if model and model != "All Models":
                query = query.filter(DBAnalysisResult.model == model)

            # Filter by serial (partial match, case-insensitive)
            if serial and serial.strip():
                serial_pattern = f"%{serial.strip()}%"
                query = query.filter(
                    func.lower(DBAnalysisResult.serial).like(func.lower(serial_pattern))
                )

            # Filter by date range (trim date)
            if date_from:
                query = query.filter(DBAnalysisResult.file_date >= date_from)
            if date_to:
                # Include the entire end date (until midnight)
                end_of_day = date_to.replace(hour=23, minute=59, second=59)
                query = query.filter(DBAnalysisResult.file_date <= end_of_day)

            # Order by trim date (newest first) and limit
            query = query.order_by(desc(DBAnalysisResult.file_date)).limit(limit)

            results = []
            for db_analysis in query.all():
                mapped = self._map_db_to_analysis(db_analysis)
                if mapped is not None:
                    results.append(mapped)

            return results

    # =========================================================================
    # Private Mapping Methods
    # =========================================================================

    def _map_analysis_to_db(self, analysis: AnalysisResult) -> DBAnalysisResult:
        """Map Pydantic AnalysisResult to SQLAlchemy model."""
        # Map system type. SystemType.UNKNOWN only reaches this trim-record
        # path on parser-error rows (FT files take save_final_test). Fall back
        # to B with a warning so misclassified records are visible in logs
        # rather than silently labeled.
        if analysis.metadata.system == SystemType.A:
            system_type = DBSystemType.A
        elif analysis.metadata.system == SystemType.B:
            system_type = DBSystemType.B
        elif analysis.metadata.system == SystemType.C:
            system_type = DBSystemType.C
        else:
            logger.warning(
                f"SystemType.UNKNOWN on trim record {analysis.metadata.filename} "
                f"(likely a parse-error row); storing as DBSystemType.B"
            )
            system_type = DBSystemType.B

        # Map overall status
        status_map = {
            AnalysisStatus.PASS: DBStatusType.PASS,
            AnalysisStatus.FAIL: DBStatusType.FAIL,
            AnalysisStatus.WARNING: DBStatusType.WARNING,
            AnalysisStatus.ERROR: DBStatusType.ERROR,
            AnalysisStatus.UNTRIMMED: DBStatusType.UNTRIMMED,
        }
        overall_status = status_map.get(analysis.overall_status, DBStatusType.ERROR)

        db_analysis = DBAnalysisResult(
            filename=analysis.metadata.filename,
            file_path=str(analysis.metadata.file_path),
            file_date=analysis.metadata.file_date,
            model=analysis.metadata.model,
            serial=analysis.metadata.serial,
            system=system_type,
            has_multi_tracks=analysis.metadata.has_multi_tracks,
            overall_status=overall_status,
            processing_time=analysis.processing_time,
            timestamp=utc_now(),
            unit_id=compute_unit_id(
                analysis.metadata.model,
                analysis.metadata.serial,
                analysis.metadata.file_date,
            ),
            data_quality=getattr(analysis, 'data_quality', 'good'),
            data_quality_issues=json.dumps(getattr(analysis, 'data_quality_issues', [])) if getattr(analysis, 'data_quality_issues', []) else None,
        )

        # Add track results
        for track in analysis.tracks:
            db_track = self._map_track_to_db(track)
            db_analysis.tracks.append(db_track)

        return db_analysis

    def _map_track_to_db(self, track: TrackData) -> DBTrackResult:
        """Map Pydantic TrackData to SQLAlchemy model."""
        status_map = {
            AnalysisStatus.PASS: DBStatusType.PASS,
            AnalysisStatus.FAIL: DBStatusType.FAIL,
            AnalysisStatus.WARNING: DBStatusType.WARNING,
            AnalysisStatus.ERROR: DBStatusType.ERROR,
            AnalysisStatus.UNTRIMMED: DBStatusType.UNTRIMMED,
        }
        status = status_map.get(track.status, DBStatusType.ERROR)

        risk_map = {
            RiskCategory.HIGH: DBRiskCategory.HIGH,
            RiskCategory.MEDIUM: DBRiskCategory.MEDIUM,
            RiskCategory.LOW: DBRiskCategory.LOW,
            RiskCategory.UNKNOWN: DBRiskCategory.UNKNOWN,
        }
        risk_category = risk_map.get(track.risk_category, DBRiskCategory.UNKNOWN)

        return DBTrackResult(
            track_id=track.track_id,
            status=status,
            travel_length=track.travel_length,
            linearity_spec=track.linearity_spec,
            sigma_gradient=track.sigma_gradient,
            untrimmed_sigma_gradient=track.untrimmed_sigma_gradient,
            sigma_threshold=track.sigma_threshold,
            sigma_pass=track.sigma_pass,
            unit_length=track.unit_length,
            untrimmed_resistance=track.untrimmed_resistance,
            trimmed_resistance=track.trimmed_resistance,
            measured_electrical_angle=track.measured_electrical_angle,
            optimal_offset=track.optimal_offset,
            final_linearity_error_shifted=track.linearity_error,
            linearity_pass=track.linearity_pass,
            linearity_fail_points=track.linearity_fail_points,
            failure_probability=track.failure_probability,
            risk_category=risk_category,
            is_anomaly=track.is_anomaly,  # Anomaly detection flag
            anomaly_reason=track.anomaly_reason,  # Reason for anomaly flag
            position_data=track.position_data,
            error_data=track.error_data,
            theory_data=getattr(track, 'theory_volts', None),
            test_volts=getattr(track, 'test_volts', None),
            upper_limits=track.upper_limits,  # Store position-dependent spec limits
            lower_limits=track.lower_limits,  # Store position-dependent spec limits
            untrimmed_positions=track.untrimmed_positions,  # Store untrimmed data for charts
            untrimmed_errors=track.untrimmed_errors,  # Store untrimmed data for charts
            # Trim effectiveness metrics
            resistance_change=track.resistance_change,
            resistance_change_percent=track.resistance_change_percent,
            trim_improvement_percent=track.trim_improvement_percent,
            untrimmed_rms_error=track.untrimmed_rms_error,
            untrimmed_error_max=getattr(track, 'untrimmed_error_max', None),
            trimmed_rms_error=track.trimmed_rms_error,
            max_error_reduction_percent=track.max_error_reduction_percent,
            # Max deviation metrics
            max_deviation=getattr(track, 'max_deviation', None),
            max_deviation_position=getattr(track, 'max_deviation_position', None),
            deviation_uniformity=getattr(track, 'deviation_uniformity', None),
            # Failure margin metrics
            max_violation=getattr(track, 'max_violation', None),
            avg_violation=getattr(track, 'avg_violation', None),
            margin_to_spec=getattr(track, 'margin_to_spec', None),
            # Spec-aware optimization fields (Phase 2)
            optimal_slope=getattr(track, 'optimal_slope', 0.0),
            station_compensation=getattr(track, 'station_compensation', None),
            linearity_type=getattr(track, 'linearity_type', None),
            raw_linearity_error=getattr(track, 'raw_linearity_error', None),
            optimized_linearity_error=getattr(track, 'optimized_linearity_error', None),
            raw_fail_points=getattr(track, 'raw_fail_points', None),
            # Trim difficulty (number of laser-trim passes recorded in file)
            trim_pass_count=getattr(track, 'trim_pass_count', None),
            # Composite trim-risk score (live-scored during processing for deployed models)
            composite_trim_risk_score=getattr(track, 'composite_trim_risk_score', None),
            # Computed metrics
            gradient_margin=track.gradient_margin,
            plot_path=str(track.plot_path) if track.plot_path else None,
        )

    @staticmethod
    def _map_status_enum(db_status, default=None):
        """Map DB status to AnalysisStatus, handling both enum and string values.

        SQLAlchemy Enum columns may return the actual enum, the enum name ('PASS'),
        or the enum value ('Pass') depending on the database and driver.
        """
        if default is None:
            default = AnalysisStatus.ERROR

        status_map = {
            DBStatusType.PASS: AnalysisStatus.PASS,
            DBStatusType.FAIL: AnalysisStatus.FAIL,
            DBStatusType.WARNING: AnalysisStatus.WARNING,
            DBStatusType.ERROR: AnalysisStatus.ERROR,
            DBStatusType.UNTRIMMED: AnalysisStatus.UNTRIMMED,
        }

        if isinstance(db_status, DBStatusType):
            return status_map.get(db_status, default)
        elif isinstance(db_status, str):
            # Check enum names (PASS, FAIL) and values (Pass, Fail)
            for db_enum, analysis_enum in status_map.items():
                if db_status == db_enum.name or db_status == db_enum.value:
                    return analysis_enum
        return default

    def _map_db_to_analysis(self, db_analysis: DBAnalysisResult) -> Optional[AnalysisResult]:
        """Map SQLAlchemy model back to Pydantic AnalysisResult.

        Returns None if the analysis has no valid tracks (corrupted data).
        """
        from laser_trim_analyzer.core.models import FileMetadata

        # Map system type. The DB enum stores A, B, or C (UNKNOWN parse-error
        # rows are written as B with a warning at save time — see
        # _map_analysis_to_db). The reverse mapping is therefore lossy by design.
        system_type = {
            DBSystemType.A: SystemType.A,
            DBSystemType.B: SystemType.B,
            DBSystemType.C: SystemType.C,
        }.get(db_analysis.system, SystemType.B)

        # Map status - handle both enum and string values from DB
        overall_status = self._map_status_enum(db_analysis.overall_status)

        # Create metadata
        metadata = FileMetadata(
            filename=db_analysis.filename,
            file_path=Path(db_analysis.file_path) if db_analysis.file_path else Path("."),
            file_date=db_analysis.file_date or datetime.now(),
            model=db_analysis.model,
            serial=db_analysis.serial,
            system=system_type,
            has_multi_tracks=db_analysis.has_multi_tracks,
        )

        # Map tracks - filter out None results from failed mappings
        tracks = [t for t in (self._map_db_to_track(t) for t in db_analysis.tracks) if t is not None]

        # If no valid tracks could be mapped, return None or create with ERROR status
        if not tracks and overall_status != AnalysisStatus.ERROR:
            logger.warning(f"Analysis {db_analysis.filename} has no valid tracks, skipping")
            return None

        # Restore data_quality state — the column is a JSON-encoded list and
        # was previously dropped on read, leaving every loaded record looking
        # "good" regardless of the suspect flag set at analysis time.
        dq_issues = []
        raw_issues = getattr(db_analysis, 'data_quality_issues', None)
        if raw_issues:
            try:
                decoded = json.loads(raw_issues)
                if isinstance(decoded, list):
                    dq_issues = decoded
            except (ValueError, TypeError) as e:
                logger.debug(f"Could not decode data_quality_issues for {db_analysis.filename}: {e}")
        dq_status = getattr(db_analysis, 'data_quality', None) or 'good'

        try:
            result = AnalysisResult(
                metadata=metadata,
                overall_status=overall_status,
                processing_time=db_analysis.processing_time or 0.0,
                tracks=tracks,
            )
            # AnalysisResult is Pydantic but data_quality fields are added
            # dynamically by the processor; preserve them across DB round-trip.
            result.data_quality = dq_status
            result.data_quality_issues = dq_issues
            return result
        except Exception as e:
            logger.error(f"Failed to create AnalysisResult for {db_analysis.filename}: {e}")
            return None

    def _map_db_to_track(self, db_track: DBTrackResult) -> Optional[TrackData]:
        """Map SQLAlchemy TrackResult back to Pydantic TrackData.

        Returns None if required fields are missing (corrupted/incomplete data).
        """
        # Use robust status mapping that handles both enum and string values
        status = self._map_status_enum(db_track.status)

        risk_map = {
            DBRiskCategory.HIGH: RiskCategory.HIGH,
            DBRiskCategory.MEDIUM: RiskCategory.MEDIUM,
            DBRiskCategory.LOW: RiskCategory.LOW,
            DBRiskCategory.UNKNOWN: RiskCategory.UNKNOWN,
        }
        risk_category = risk_map.get(db_track.risk_category, RiskCategory.UNKNOWN)

        # Sigma/linearity fields are optional. For UNTRIMMED tracks (no trim
        # run) they are legitimately None and we want to preserve that so the
        # GUI/exports can tell "no measurement" apart from "measured zero".
        # For pre-UNTRIMMED legacy rows with missing sigma, fall back to
        # defaults to keep loading non-fatal.
        is_untrimmed = status == AnalysisStatus.UNTRIMMED
        sigma_gradient = db_track.sigma_gradient
        sigma_threshold = db_track.sigma_threshold
        sigma_pass = db_track.sigma_pass

        if not is_untrimmed:
            if sigma_gradient is None:
                logger.warning(f"Track {db_track.track_id} has None sigma_gradient, using 0.0")
                sigma_gradient = 0.0
            if sigma_threshold is None:
                logger.warning(f"Track {db_track.track_id} has None sigma_threshold, using 0.01")
                sigma_threshold = 0.01
            if sigma_pass is None:
                sigma_pass = sigma_gradient <= sigma_threshold

        if is_untrimmed:
            linearity_error_val = None
            linearity_pass_val = None
            optimal_offset_val = None  # No trim ran → no offset was applied.
        else:
            linearity_error_val = abs(db_track.final_linearity_error_shifted or 0.0)
            linearity_pass_val = db_track.linearity_pass if db_track.linearity_pass is not None else False
            optimal_offset_val = db_track.optimal_offset or 0.0

        try:
            return TrackData(
                track_id=db_track.track_id or "default",
                status=status,
                travel_length=db_track.travel_length or 1.0,  # Default to 1.0 to avoid 0
                linearity_spec=db_track.linearity_spec or 0.01,
                sigma_gradient=sigma_gradient,
                sigma_threshold=sigma_threshold,
                sigma_pass=sigma_pass,
                optimal_offset=optimal_offset_val,
                linearity_error=linearity_error_val,
                linearity_pass=linearity_pass_val,
                linearity_fail_points=db_track.linearity_fail_points or 0,
                unit_length=db_track.unit_length,
                untrimmed_resistance=db_track.untrimmed_resistance,
                trimmed_resistance=db_track.trimmed_resistance,
                measured_electrical_angle=getattr(db_track, 'measured_electrical_angle', None),
                failure_probability=db_track.failure_probability,
                risk_category=risk_category,
                is_anomaly=db_track.is_anomaly or False,  # Retrieve anomaly flag
                anomaly_reason=db_track.anomaly_reason,  # Retrieve anomaly reason
                position_data=db_track.position_data,
                error_data=db_track.error_data,
                theory_volts=getattr(db_track, 'theory_data', None),
                test_volts=getattr(db_track, 'test_volts', None),
                upper_limits=db_track.upper_limits,  # Retrieve position-dependent spec limits
                lower_limits=db_track.lower_limits,  # Retrieve position-dependent spec limits
                untrimmed_positions=db_track.untrimmed_positions,  # Retrieve untrimmed data for charts
                untrimmed_errors=db_track.untrimmed_errors,  # Retrieve untrimmed data for charts
                # Trim effectiveness metrics
                resistance_change=db_track.resistance_change,
                trim_improvement_percent=db_track.trim_improvement_percent,
                untrimmed_rms_error=db_track.untrimmed_rms_error,
                trimmed_rms_error=db_track.trimmed_rms_error,
                max_error_reduction_percent=db_track.max_error_reduction_percent,
                # Phase 2 spec-aware fields
                optimal_slope=getattr(db_track, 'optimal_slope', 0.0),
                station_compensation=getattr(db_track, 'station_compensation', None),
                linearity_type=getattr(db_track, 'linearity_type', None),
                raw_linearity_error=getattr(db_track, 'raw_linearity_error', None),
                optimized_linearity_error=getattr(db_track, 'optimized_linearity_error', None),
                raw_fail_points=getattr(db_track, 'raw_fail_points', None),
                # Trim difficulty
                trim_pass_count=getattr(db_track, 'trim_pass_count', None),
                # Max deviation fields
                max_deviation=getattr(db_track, 'max_deviation', None),
                max_deviation_position=getattr(db_track, 'max_deviation_position', None),
                deviation_uniformity=getattr(db_track, 'deviation_uniformity', None),
                # Failure margin metrics
                max_violation=getattr(db_track, 'max_violation', None),
                avg_violation=getattr(db_track, 'avg_violation', None),
                margin_to_spec=getattr(db_track, 'margin_to_spec', None),
            )
        except Exception as e:
            logger.error(f"Failed to map track {db_track.track_id}: {e}")
            return None

    def _update_existing_analysis(
        self,
        session: Session,
        analysis: AnalysisResult
    ) -> int:
        """Update an existing analysis record."""
        # Find existing record by the DB UNIQUE-constraint key
        # (filename, file_date, model, serial).  Must match save_analysis /
        # save_batch -- if we re-query by (filename, file_path) here and the
        # path string differs from what's stored, we miss the row, fall
        # through to INSERT, and trip the UNIQUE constraint.
        existing = (
            session.query(DBAnalysisResult)
            .filter(
                DBAnalysisResult.filename == analysis.metadata.filename,
                DBAnalysisResult.file_date == analysis.metadata.file_date,
                DBAnalysisResult.model == analysis.metadata.model,
                DBAnalysisResult.serial == analysis.metadata.serial,
            )
            .first()
        )

        if existing:
            # Update ALL fields including model/serial (parsing may have changed)
            existing.file_path = str(analysis.metadata.file_path)
            existing.model = analysis.metadata.model
            existing.serial = analysis.metadata.serial
            existing.file_date = analysis.metadata.file_date
            existing.unit_id = compute_unit_id(
                analysis.metadata.model,
                analysis.metadata.serial,
                analysis.metadata.file_date,
            )
            # See _map_analysis_to_db for the UNKNOWN→B fallback rationale.
            if analysis.metadata.system == SystemType.A:
                existing.system = DBSystemType.A
            elif analysis.metadata.system == SystemType.B:
                existing.system = DBSystemType.B
            elif analysis.metadata.system == SystemType.C:
                existing.system = DBSystemType.C
            else:
                logger.warning(
                    f"SystemType.UNKNOWN on update of {analysis.metadata.filename}; "
                    f"keeping as DBSystemType.B"
                )
                existing.system = DBSystemType.B
            existing.has_multi_tracks = analysis.metadata.has_multi_tracks
            existing.processing_time = analysis.processing_time
            existing.timestamp = utc_now()

            # Map overall status
            status_map = {
                AnalysisStatus.PASS: DBStatusType.PASS,
                AnalysisStatus.FAIL: DBStatusType.FAIL,
                AnalysisStatus.WARNING: DBStatusType.WARNING,
                AnalysisStatus.ERROR: DBStatusType.ERROR,
                # Without UNTRIMMED here, re-saving an untrimmed file falls
                # through to the .get() default of ERROR, silently flipping
                # its overall_status. Must mirror the status_map in
                # _map_analysis_to_db.
                AnalysisStatus.UNTRIMMED: DBStatusType.UNTRIMMED,
            }
            existing.overall_status = status_map.get(analysis.overall_status, DBStatusType.ERROR)
            existing.data_quality = getattr(analysis, 'data_quality', None)
            issues = getattr(analysis, 'data_quality_issues', None) or []
            existing.data_quality_issues = json.dumps(issues) if issues else None

            # Delete old tracks explicitly and flush before adding new ones
            # This avoids unique constraint violations
            session.query(DBTrackResult).filter(
                DBTrackResult.analysis_id == existing.id
            ).delete()
            session.flush()

            # Add new tracks
            for track in analysis.tracks:
                db_track = self._map_track_to_db(track)
                db_track.analysis_id = existing.id
                session.add(db_track)

            # Flush to ensure changes are written
            session.flush()
            logger.debug(f"Updated analysis ID {existing.id}: status={analysis.overall_status.value}")
            return existing.id

        # If no existing record found, create new
        db_analysis = self._map_analysis_to_db(analysis)
        session.add(db_analysis)
        session.flush()
        return db_analysis.id

    # =========================================================================
    # Delete Operations
    # =========================================================================

    def delete_analysis(self, analysis_id: int) -> bool:
        """
        Delete an analysis record and all associated data.

        This removes:
        - The analysis record
        - All associated track results
        - The processed file record (to allow re-processing)
        - Any related alerts

        Args:
            analysis_id: Database ID of the analysis to delete

        Returns:
            True if deletion was successful, False if record not found
        """
        with self.session() as session:
            # Find the analysis record
            analysis = session.get(DBAnalysisResult, analysis_id)

            if not analysis:
                logger.warning(f"Analysis ID {analysis_id} not found for deletion")
                return False

            filename = analysis.filename

            # Delete associated tracks (cascade should handle this, but be explicit)
            session.query(DBTrackResult).filter(
                DBTrackResult.analysis_id == analysis_id
            ).delete()

            # Delete processed file record (allows re-processing of same file)
            session.query(DBProcessedFile).filter(
                DBProcessedFile.analysis_id == analysis_id
            ).delete()

            # Delete any related alerts
            session.query(DBQAAlert).filter(
                DBQAAlert.analysis_id == analysis_id
            ).delete()

            # Delete the analysis record itself
            session.delete(analysis)

            logger.info(f"Deleted analysis ID {analysis_id}: {filename}")
            return True

    def delete_analysis_by_filename(self, filename: str) -> bool:
        """
        Delete an analysis record by filename.

        Args:
            filename: The filename of the analysis to delete

        Returns:
            True if deletion was successful, False if record not found
        """
        with self.session() as session:
            analysis = session.query(DBAnalysisResult).filter(
                DBAnalysisResult.filename == filename
            ).first()

            if not analysis:
                logger.warning(f"Analysis with filename '{filename}' not found for deletion")
                return False

            analysis_id = analysis.id

            # Delete associated tracks
            session.query(DBTrackResult).filter(
                DBTrackResult.analysis_id == analysis_id
            ).delete()

            # Delete processed file record
            session.query(DBProcessedFile).filter(
                DBProcessedFile.analysis_id == analysis_id
            ).delete()

            # Delete any related alerts
            session.query(DBQAAlert).filter(
                DBQAAlert.analysis_id == analysis_id
            ).delete()

            # Delete the analysis record itself
            session.delete(analysis)

            logger.info(f"Deleted analysis by filename: {filename}")
            return True

    # =========================================================================
    # Utility Methods
    # =========================================================================

    def close(self) -> None:
        """Close the database connection."""
        if self._engine:
            self._engine.dispose()
            logger.info("Database connection closed")

    def get_database_path(self) -> Path:
        """Get the path to the database file."""
        return self.database_path

    def get_record_count(self) -> Dict[str, int]:
        """Get count of records in main tables."""
        with self.session() as session:
            return {
                "analyses": session.query(func.count(DBAnalysisResult.id)).scalar() or 0,
                "tracks": session.query(func.count(DBTrackResult.id)).scalar() or 0,
                "processed_files": session.query(func.count(DBProcessedFile.id)).scalar() or 0,
                "alerts": session.query(func.count(DBQAAlert.id)).scalar() or 0,
            }

    def get_models_list(self) -> List[str]:
        """
        Get a list of all unique model numbers in the database.

        Returns:
            List of model strings, sorted numerically
        """
        with self.session() as session:
            models = (
                session.query(DBAnalysisResult.model)
                .filter(DBAnalysisResult.model.isnot(None))
                .distinct()
                .all()
            )
            model_list = [m[0] for m in models if m[0] and m[0] != "Unknown"]
            return sorted(model_list, key=_model_sort_key)

    def get_models_with_sigma_data(self, days_back: int = 30) -> List[str]:
        """Models with at least one non-null sigma_gradient inside the
        day-range window. Used to populate the Drift tab's model dropdown
        — never-processed and dead models don't clutter the list.
        """
        cutoff = datetime.now() - timedelta(days=days_back)
        with self.session() as session:
            rows = (
                session.query(DBAnalysisResult.model)
                .join(DBTrackResult, DBTrackResult.analysis_id == DBAnalysisResult.id)
                .filter(
                    DBAnalysisResult.model.isnot(None),
                    DBAnalysisResult.model != "Unknown",
                    DBAnalysisResult.file_date >= cutoff,
                    DBTrackResult.sigma_gradient.isnot(None),
                )
                .distinct()
                .all()
            )
            return sorted([r[0] for r in rows if r[0]], key=_model_sort_key)

    def get_models_list_prioritized(
        self,
        mps_models: List[str] = None,
        recent_days: int = 90
    ) -> List[Dict[str, Any]]:
        """
        Get models sorted by priority: MPS first, then Recently Active, then Inactive.

        Each group is sorted numerically using _model_sort_key.

        Args:
            mps_models: List of models on MPS schedule (user-managed)
            recent_days: Days threshold for "recently active" (default 90)

        Returns:
            List of dicts: [{'model': str, 'status': 'mps'|'active'|'inactive', 'count': int, 'last_date': datetime}]
        """
        mps_models = mps_models or []
        mps_set = set(mps_models)
        cutoff_date = datetime.now() - timedelta(days=recent_days)

        with self.session() as session:
            # Get all models with their latest activity date and count
            results = (
                session.query(
                    DBAnalysisResult.model,
                    func.max(DBAnalysisResult.file_date).label('last_date'),
                    func.count(DBAnalysisResult.id).label('count')
                )
                .filter(DBAnalysisResult.model.isnot(None))
                .group_by(DBAnalysisResult.model)
                .all()
            )

            # Categorize models
            mps_list = []
            active_list = []
            inactive_list = []

            for model, last_date, count in results:
                if not model or model == "Unknown":
                    continue

                entry = {
                    'model': model,
                    'count': count,
                    'last_date': last_date
                }

                if model in mps_set:
                    entry['status'] = 'mps'
                    mps_list.append(entry)
                elif last_date and last_date >= cutoff_date:
                    entry['status'] = 'active'
                    active_list.append(entry)
                else:
                    entry['status'] = 'inactive'
                    inactive_list.append(entry)

            # Sort each group numerically
            mps_list.sort(key=lambda x: _model_sort_key(x['model']))
            active_list.sort(key=lambda x: _model_sort_key(x['model']))
            inactive_list.sort(key=lambda x: _model_sort_key(x['model']))

            return mps_list + active_list + inactive_list

    # =========================================================================
    # Trends Page Methods (Active Models Summary)
    # =========================================================================

    def get_unit_yield_trend(
        self,
        model: str,
        days_back: int = 90,
    ) -> List[Dict[str, Any]]:
        """Daily unit-level yield for one model.

        Pass rule: a unit passes only if every section's most recent
        attempt is PASS. Retrim tiebreak uses row `timestamp` (write
        time). NULL unit_id rows (junk serials, missing date) are
        excluded entirely.

        Returns rows of {"date", "total_units", "passed_units", "yield_pct"},
        sorted by date ascending. Days with no qualifying units omitted.
        """
        from laser_trim_analyzer.database.models import (
            AnalysisResult as DBAR,
            StatusType as DBStatusType,
        )
        from collections import defaultdict

        cutoff = datetime.now() - timedelta(days=days_back)
        with self.session() as s:
            # Pull (unit_id, serial, file_date, timestamp, overall_status)
            # for each qualifying row. We do the section-grouping and
            # latest-attempt-by-timestamp logic in Python because the
            # tie-breaking is awkward in SQL and the row count per
            # model+window is small (hundreds, not millions).
            rows = (
                s.query(
                    DBAR.unit_id,
                    DBAR.serial,
                    DBAR.file_date,
                    DBAR.timestamp,
                    DBAR.overall_status,
                )
                .filter(
                    DBAR.model == model,
                    DBAR.unit_id.isnot(None),
                    # Belt-and-suspenders: the backfill uses '' as a
                    # progress sentinel and converts it back to NULL at
                    # the end. If a process crash leaves '' in place,
                    # IS NOT NULL alone would let those rows through
                    # and collapse them into a phantom '' unit.
                    DBAR.unit_id != "",
                    DBAR.file_date >= cutoff,
                )
                .all()
            )

        # Group rows by unit_id, then within each unit by section (trailing
        # non-digit part of the serial after the shop number). For each
        # (unit, section) pick the row with the largest timestamp -- that's
        # the section's authoritative status.
        section_status_by_unit: Dict[str, Dict[str, Any]] = defaultdict(dict)
        unit_date: Dict[str, datetime] = {}

        def section_of(serial: Optional[str]) -> str:
            """Return the section suffix (everything after the leading digits).
            Default to '' if there's no suffix, so single-section units (just
            digits like '196') still group cleanly."""
            if not serial:
                return ""
            m = re.match(r"^\d+(.*)$", serial.strip(), re.ASCII)
            return (m.group(1) if m else "").upper()

        for unit_id, serial, file_date, ts, status in rows:
            section = section_of(serial)
            existing = section_status_by_unit[unit_id].get(section)
            if existing is None or (ts or datetime.min) > existing["ts"]:
                section_status_by_unit[unit_id][section] = {
                    "ts": ts or datetime.min,
                    "status": status,
                }
            # Remember the unit's date. The unit_id format embeds the
            # canonical date, so in practice all rows for one unit share
            # the same file_date. Use first-seen-wins to keep behavior
            # deterministic regardless of query result order, in case
            # of any future data drift.
            if file_date is not None and unit_id not in unit_date:
                unit_date[unit_id] = file_date

        # Apply the pass rule per unit and bucket by date.
        per_date: Dict[str, Dict[str, int]] = defaultdict(
            lambda: {"total": 0, "passed": 0}
        )
        for unit_id, sections in section_status_by_unit.items():
            fd = unit_date.get(unit_id)
            if fd is None:
                continue
            day = fd.strftime("%Y-%m-%d")
            per_date[day]["total"] += 1
            # All sections must be PASS for the unit to pass
            all_pass = all(
                _status_matches(sec["status"], DBStatusType.PASS)
                for sec in sections.values()
            )
            if all_pass:
                per_date[day]["passed"] += 1

        return [
            {
                "date": day,
                "total_units": counts["total"],
                "passed_units": counts["passed"],
                "yield_pct": (
                    counts["passed"] / counts["total"] * 100.0
                ) if counts["total"] > 0 else 0.0,
            }
            for day, counts in sorted(per_date.items())
        ]

    def get_active_models_summary(
        self,
        days_back: int = 90,
        min_samples: int = 5
    ) -> List[Dict[str, Any]]:
        """
        Get summary statistics for all models with recent activity.

        Filters by trim date (file_date), not processing date.

        Args:
            days_back: Only include models with files trimmed in this period
            min_samples: Minimum samples required for inclusion

        Returns:
            List of model summaries sorted by sample count descending
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Single query with join to get all data at once - no N+1 problem
            # Filter by trim date (file_date). UNTRIMMED tracks filtered at the
            # JOIN level so they don't pollute avg_sigma/threshold or deflate
            # pass rates (their sigma columns are NULL anyway).
            # `total` counts every analysis row (raw count). `trimmed_total`
            # excludes UNTRIMMED so pass_rate isn't deflated by test-sweep-only
            # files. Both are needed because the Failed column wants to know
            # the trimmed pool's failed count, not the total pool.
            model_data = (
                session.query(
                    DBAnalysisResult.model,
                    func.count(func.distinct(DBAnalysisResult.id)).label('total'),
                    func.count(func.distinct(case(
                        (DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name,
                         DBAnalysisResult.id),
                        else_=None
                    ))).label('trimmed_total'),
                    func.count(func.distinct(case(
                        (DBAnalysisResult.overall_status == DBStatusType.PASS, DBAnalysisResult.id),
                        else_=None
                    ))).label('passed'),
                    func.min(DBAnalysisResult.file_date).label('first_date'),
                    func.max(DBAnalysisResult.file_date).label('last_date'),
                    func.avg(DBTrackResult.sigma_gradient).label('avg_sigma'),
                    func.avg(DBTrackResult.sigma_threshold).label('avg_threshold'),
                    func.count(DBTrackResult.id).label('total_tracks'),
                    func.sum(case((DBTrackResult.sigma_pass == True, 1), else_=0)).label('sigma_passed'),
                    func.sum(case((DBTrackResult.linearity_pass == True, 1), else_=0)).label('linearity_passed'),
                )
                .outerjoin(
                    DBTrackResult,
                    and_(
                        DBAnalysisResult.id == DBTrackResult.analysis_id,
                        DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                    ),
                )
                .filter(
                    DBAnalysisResult.model.isnot(None),
                    DBAnalysisResult.file_date >= cutoff_date
                )
                .group_by(DBAnalysisResult.model)
                .having(func.count(func.distinct(DBAnalysisResult.id)) >= min_samples)
                .all()
            )

            results = []
            for row in model_data:
                model = row.model
                total = row.total
                trimmed_total = row.trimmed_total or 0
                passed = row.passed or 0
                total_tracks = row.total_tracks or 0
                sigma_passed = row.sigma_passed or 0
                linearity_passed = row.linearity_passed or 0

                if not model or total == 0 or model == "Unknown":
                    continue

                # pass_rate over trimmed pool only (UNTRIMMED files have no
                # pass/fail meaning). "failed" likewise counted within the
                # trimmed pool so failed + passed == trimmed_total.
                pass_rate = (passed / trimmed_total * 100) if trimmed_total > 0 else 0.0
                sigma_pass_rate = (sigma_passed / total_tracks * 100) if total_tracks > 0 else 0.0
                linearity_pass_rate = (linearity_passed / total_tracks * 100) if total_tracks > 0 else 0.0

                results.append({
                    "model": model,
                    "total": total,
                    "passed": passed,
                    "failed": trimmed_total - passed,
                    "pass_rate": pass_rate,
                    "sigma_pass_rate": sigma_pass_rate,
                    "linearity_pass_rate": linearity_pass_rate,
                    "avg_sigma": row.avg_sigma or 0,
                    "avg_threshold": row.avg_threshold or 0,
                    "first_date": row.first_date,
                    "last_date": row.last_date,
                })

            # Sort by sample count descending (for chart display)
            results.sort(key=lambda x: x["total"], reverse=True)
            return results

    def get_models_requiring_attention(
        self,
        days_back: int = 90,
        min_samples: int = 5,
        pass_rate_threshold: float = 80.0,
        trend_threshold: float = 10.0,
        rolling_days: int = 30,
        metric: str = "linearity",
    ) -> List[Dict[str, Any]]:
        """
        Get models that require attention based on alert criteria.

        Alert criteria:
        - Pass rate below threshold (default 80%)
        - Trending worse by threshold% over rolling period
        - High variance in recent samples

        Args:
            days_back: Period to analyze
            min_samples: Minimum samples for inclusion
            pass_rate_threshold: Alert if pass rate below this
            trend_threshold: Alert if trend worse by this %
            rolling_days: Rolling window for trend calculation
            metric: Which pass rate to use - "linearity", "sigma", or "overall"

        Returns:
            List of models requiring attention with alert details
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)
            rolling_cutoff = datetime.now() - timedelta(days=rolling_days)

            # Get active models
            active_models = self.get_active_models_summary(days_back, min_samples)

            # Select which rate to evaluate based on metric
            rate_key = {
                "linearity": "linearity_pass_rate",
                "sigma": "sigma_pass_rate",
                "overall": "pass_rate",
            }.get(metric, "linearity_pass_rate")
            rate_label = {
                "linearity": "Linearity",
                "sigma": "Sigma",
                "overall": "Overall",
            }.get(metric, "Linearity")

            alerts = []
            for model_data in active_models:
                model = model_data["model"]
                alert_reasons = []

                # Check 1: Low pass rate based on selected metric
                check_rate = model_data.get(rate_key, model_data["pass_rate"])
                if check_rate < pass_rate_threshold:
                    alert_reasons.append({
                        "type": "LOW_PASS_RATE",
                        "message": f"{rate_label} pass rate {check_rate:.1f}% is below {pass_rate_threshold}%",
                        "severity": "High" if check_rate < 70 else "Medium"
                    })

                # Check 2: Trending worse
                # Compare older period vs rolling period
                older_cutoff = cutoff_date
                older_end = rolling_cutoff

                # Use track-level linearity query when metric is "linearity"
                if metric == "linearity":
                    older_trend = (
                        session.query(
                            func.count(DBTrackResult.id),
                            func.sum(case((DBTrackResult.linearity_pass == True, 1), else_=0))
                        )
                        .join(DBAnalysisResult)
                        .filter(
                            DBAnalysisResult.model == model,
                            DBAnalysisResult.file_date >= older_cutoff,
                            DBAnalysisResult.file_date < older_end,
                            DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                        )
                        .first()
                    )
                    recent_trend = (
                        session.query(
                            func.count(DBTrackResult.id),
                            func.sum(case((DBTrackResult.linearity_pass == True, 1), else_=0))
                        )
                        .join(DBAnalysisResult)
                        .filter(
                            DBAnalysisResult.model == model,
                            DBAnalysisResult.file_date >= rolling_cutoff,
                            DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                        )
                        .first()
                    )
                else:
                    older_trend = (
                        session.query(
                            func.count(DBAnalysisResult.id),
                            func.sum(
                                case(
                                    (DBAnalysisResult.overall_status == DBStatusType.PASS, 1),
                                    else_=0
                                )
                            )
                        )
                        .filter(
                            DBAnalysisResult.model == model,
                            DBAnalysisResult.file_date >= older_cutoff,
                            DBAnalysisResult.file_date < older_end
                        )
                        .first()
                    )
                    recent_trend = (
                        session.query(
                            func.count(DBAnalysisResult.id),
                            func.sum(
                                case(
                                    (DBAnalysisResult.overall_status == DBStatusType.PASS, 1),
                                    else_=0
                                )
                            )
                        )
                        .filter(
                            DBAnalysisResult.model == model,
                            DBAnalysisResult.file_date >= rolling_cutoff
                        )
                        .first()
                    )

                older_count, older_passed = older_trend
                recent_count, recent_passed = recent_trend

                if older_count and older_count >= min_samples and recent_count and recent_count >= min_samples:
                    older_pct = (older_passed or 0) / older_count * 100
                    recent_pct = (recent_passed or 0) / recent_count * 100

                    if recent_pct < older_pct - trend_threshold:
                        alert_reasons.append({
                            "type": "TRENDING_WORSE",
                            "message": f"{rate_label} pass rate dropped from {older_pct:.1f}% to {recent_pct:.1f}% ({older_pct - recent_pct:.1f}% decline)",
                            "severity": "High" if (older_pct - recent_pct) > 20 else "Medium"
                        })

                # Check 3: High variance (use coefficient of variation)
                sigma_values = (
                    session.query(DBTrackResult.sigma_gradient)
                    .join(DBAnalysisResult)
                    .filter(
                        DBAnalysisResult.model == model,
                        DBAnalysisResult.file_date >= rolling_cutoff,
                        DBTrackResult.sigma_gradient.isnot(None),
                    )
                    .all()
                )

                if len(sigma_values) >= min_samples:
                    values = [v[0] for v in sigma_values if v[0] is not None]
                    if values:
                        import numpy as np
                        mean_val = np.mean(values)
                        std_val = np.std(values, ddof=1)
                        cv = (std_val / mean_val * 100) if mean_val > 0 else 0

                        # CV > 50% is high variance for sigma gradient
                        if cv > 50:
                            alert_reasons.append({
                                "type": "HIGH_VARIANCE",
                                "message": f"High variance in sigma gradient (CV={cv:.1f}%)",
                                "severity": "Medium"
                            })

                if alert_reasons:
                    alerts.append({
                        "model": model,
                        "pass_rate": model_data["pass_rate"],
                        "total_samples": model_data["total"],
                        "alerts": alert_reasons,
                        "severity": max(a["severity"] for a in alert_reasons)
                    })

            # Sort by severity (High first) then by pass rate (lowest first)
            severity_order = {"High": 0, "Medium": 1, "Low": 2}
            alerts.sort(key=lambda x: (severity_order.get(x["severity"], 3), x["pass_rate"]))

            return alerts

    def get_linearity_prioritization(
        self,
        days_back: int = 90,
        min_samples: int = 10,
    ) -> List[Dict[str, Any]]:
        """
        Get models ranked by improvement impact for linearity.

        Combines failure volume, near-miss count (easy wins), and trend
        to help users prioritize where to spend time.

        Args:
            days_back: Period to analyze
            min_samples: Minimum track count for inclusion

        Returns:
            List of models sorted by impact_score descending
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            model_data = (
                session.query(
                    DBAnalysisResult.model,
                    func.count(DBTrackResult.id).label('total_tracks'),
                    func.sum(case((DBTrackResult.linearity_pass == True, 1), else_=0)).label('lin_passed'),
                    func.sum(case((DBTrackResult.linearity_pass == False, 1), else_=0)).label('lin_failed'),
                    func.sum(case((DBTrackResult.sigma_pass == True, 1), else_=0)).label('sigma_passed'),
                    # Near-miss: failed linearity with only 1-2 fail points (easy wins)
                    func.sum(
                        case(
                            (and_(
                                DBTrackResult.linearity_pass == False,
                                DBTrackResult.linearity_fail_points <= 2,
                                DBTrackResult.linearity_fail_points > 0,
                            ), 1),
                            else_=0
                        )
                    ).label('near_miss_count'),
                    # Average fail points on failing tracks
                    func.avg(
                        case(
                            (DBTrackResult.linearity_pass == False, DBTrackResult.linearity_fail_points),
                            else_=None
                        )
                    ).label('avg_fail_points'),
                    # Trim effectiveness (may be NULL if not yet calculated)
                    func.avg(DBTrackResult.trim_improvement_percent).label('avg_trim_improvement'),
                    func.avg(DBTrackResult.resistance_change_percent).label('avg_resistance_change'),
                )
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(
                    DBAnalysisResult.model.isnot(None),
                    DBAnalysisResult.model != "Unknown",
                    DBAnalysisResult.file_date >= cutoff_date,
                    # UNTRIMMED tracks have no linearity outcome and would
                    # inflate the denominator for "fail rate" / "near-miss".
                    DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                )
                .group_by(DBAnalysisResult.model)
                .having(func.count(DBTrackResult.id) >= min_samples)
                .all()
            )

            results = []
            for row in model_data:
                total = row.total_tracks or 0
                lin_passed = row.lin_passed or 0
                lin_failed = row.lin_failed or 0
                sigma_passed = row.sigma_passed or 0
                near_miss = row.near_miss_count or 0

                if total == 0:
                    continue

                lin_rate = lin_passed / total * 100
                sigma_rate = sigma_passed / total * 100
                avg_fps = row.avg_fail_points or 0

                # Impact score: volume of failures + near-miss opportunity
                impact_score = (
                    lin_failed * 0.5
                    + near_miss * 0.3
                    + (1 - lin_rate / 100) * total * 0.2
                )

                # Generate recommendation
                recommendation = self._generate_recommendation(
                    near_miss, lin_failed, avg_fps,
                    row.avg_trim_improvement, row.avg_resistance_change,
                    lin_rate, sigma_rate
                )

                results.append({
                    "model": row.model,
                    "total_tracks": total,
                    "linearity_pass_rate": round(lin_rate, 1),
                    "sigma_pass_rate": round(sigma_rate, 1),
                    "failed_units": lin_failed,
                    "near_miss_count": near_miss,
                    "avg_fail_points": round(avg_fps, 1),
                    "avg_trim_improvement": round(row.avg_trim_improvement, 1) if row.avg_trim_improvement else None,
                    "avg_resistance_change": round(row.avg_resistance_change, 2) if row.avg_resistance_change else None,
                    "impact_score": round(impact_score, 1),
                    "recommendation": recommendation,
                })

            # Compute percentile ranks
            if results:
                sorted_by_rate = sorted(results, key=lambda x: x["linearity_pass_rate"])
                for i, r in enumerate(sorted_by_rate):
                    r["percentile_rank"] = round(i / len(sorted_by_rate) * 100, 0)

            # Sort by impact score descending
            results.sort(key=lambda x: x["impact_score"], reverse=True)
            return results

    def _generate_recommendation(
        self,
        near_miss: int,
        lin_failed: int,
        avg_fail_points: float,
        avg_trim_improvement: Optional[float],
        avg_resistance_change: Optional[float],
        lin_rate: float,
        sigma_rate: float,
    ) -> str:
        """Generate actionable recommendation based on model data."""
        if lin_failed == 0:
            return "Passing — monitor sigma trends"

        # Check near-miss ratio (easy wins)
        near_miss_ratio = near_miss / lin_failed if lin_failed > 0 else 0
        if near_miss_ratio > 0.3:
            return f"Easy-win potential — {near_miss} units within 2 fail points of passing"

        # Check if trim is not effective
        if avg_trim_improvement is not None and avg_trim_improvement < 30:
            return "Trim not effective — investigate incoming material quality"

        # Check excessive trimming
        if avg_resistance_change is not None and abs(avg_resistance_change) > 15:
            return "Excessive trimming — process may be over-cutting"

        # Check severity
        if avg_fail_points > 10:
            return "Severe failures — units far from spec, may need process change"

        # Sigma OK but linearity bad
        if sigma_rate > 80 and lin_rate < 70:
            return "Sigma healthy but linearity failing — check spec limits"

        return f"Review needed — {lin_failed} linearity failures"

    def get_linearity_margin_analysis(
        self,
        model: str,
        days_back: int = 90,
    ) -> Dict[str, Any]:
        """
        Get fail-point distribution for a specific model.

        Shows how close failing tracks are to passing — highlights easy wins.

        Args:
            model: Model number
            days_back: Period to analyze

        Returns:
            Dict with fail point distribution and easy-win analysis
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Get all failing tracks for this model
            failing_tracks = (
                session.query(DBTrackResult.linearity_fail_points)
                .join(DBAnalysisResult)
                .filter(
                    DBAnalysisResult.model == model,
                    DBAnalysisResult.file_date >= cutoff_date,
                    DBTrackResult.linearity_pass == False,
                    DBTrackResult.linearity_fail_points > 0,
                )
                .all()
            )

            total_tracks = (
                session.query(func.count(DBTrackResult.id))
                .join(DBAnalysisResult)
                .filter(
                    DBAnalysisResult.model == model,
                    DBAnalysisResult.file_date >= cutoff_date,
                )
                .scalar()
            ) or 0

            passing_tracks = (
                session.query(func.count(DBTrackResult.id))
                .join(DBAnalysisResult)
                .filter(
                    DBAnalysisResult.model == model,
                    DBAnalysisResult.file_date >= cutoff_date,
                    DBTrackResult.linearity_pass == True,
                )
                .scalar()
            ) or 0

            # Build distribution
            fail_points = [row[0] for row in failing_tracks]
            distribution = {
                "1_point": sum(1 for fp in fail_points if fp == 1),
                "2_points": sum(1 for fp in fail_points if fp == 2),
                "3_to_5": sum(1 for fp in fail_points if 3 <= fp <= 5),
                "6_to_10": sum(1 for fp in fail_points if 6 <= fp <= 10),
                "over_10": sum(1 for fp in fail_points if fp > 10),
            }

            easy_wins = distribution["1_point"] + distribution["2_points"]
            failing_count = len(fail_points)

            return {
                "model": model,
                "total_tracks": total_tracks,
                "passing_tracks": passing_tracks,
                "failing_tracks": failing_count,
                "fail_point_distribution": distribution,
                "easy_win_count": easy_wins,
                "easy_win_percent": round(easy_wins / failing_count * 100, 1) if failing_count > 0 else 0,
            }

    def get_near_miss_summary(self, days_back: int = 90) -> Dict[str, Any]:
        """
        Get overall near-miss analysis across all models.

        Returns:
            Dict with fail-point distribution, near-miss percentage,
            and top models by near-miss count.
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Get all failing tracks with fail point counts
            failing = (
                session.query(
                    DBTrackResult.linearity_fail_points,
                    DBAnalysisResult.model,
                )
                .join(DBAnalysisResult)
                .filter(
                    DBAnalysisResult.file_date >= cutoff_date,
                    DBTrackResult.linearity_pass == False,
                    DBTrackResult.linearity_fail_points > 0,
                )
                .all()
            )

            total_failing = len(failing)
            if total_failing == 0:
                return {
                    "total_failing": 0,
                    "distribution": {},
                    "near_miss_count": 0,
                    "near_miss_percent": 0,
                    "hard_fail_count": 0,
                    "hard_fail_percent": 0,
                    "top_near_miss_models": [],
                }

            # Distribution buckets
            buckets = {"1-3 points": 0, "4-10 points": 0, "11-50 points": 0, "50+ points": 0}
            near_miss_by_model = {}

            for fp, model in failing:
                if fp <= 3:
                    buckets["1-3 points"] += 1
                    near_miss_by_model[model] = near_miss_by_model.get(model, 0) + 1
                elif fp <= 10:
                    buckets["4-10 points"] += 1
                elif fp <= 50:
                    buckets["11-50 points"] += 1
                else:
                    buckets["50+ points"] += 1

            near_miss = buckets["1-3 points"]
            hard_fail = buckets["11-50 points"] + buckets["50+ points"]

            # Top models by near-miss count
            top_models = sorted(
                near_miss_by_model.items(), key=lambda x: x[1], reverse=True
            )[:10]

            return {
                "total_failing": total_failing,
                "distribution": buckets,
                "near_miss_count": near_miss,
                "near_miss_percent": round(near_miss / total_failing * 100, 1),
                "hard_fail_count": hard_fail,
                "hard_fail_percent": round(hard_fail / total_failing * 100, 1),
                "top_near_miss_models": [
                    {"model": m, "near_miss_count": c} for m, c in top_models
                ],
            }

    def get_trending_worse_models(
        self,
        days_back: int = 90,
        min_samples: int = 20,
        trend_threshold: float = 10.0,
        rolling_days: int = 30
    ) -> List[Dict[str, Any]]:
        """
        Get models whose pass rate is declining.

        Compares older period vs recent rolling period pass rates.

        Args:
            days_back: Total period to analyze
            min_samples: Minimum samples required for both periods
            trend_threshold: Minimum decline % to be considered "trending worse"
            rolling_days: Recent period for comparison

        Returns:
            List of models with decline info, sorted by decline magnitude (biggest first)
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)
            rolling_cutoff = datetime.now() - timedelta(days=rolling_days)

            # Get active models with sufficient data
            active_models = self.get_active_models_summary(days_back, min_samples)

            trending_worse = []
            for model_data in active_models:
                model = model_data["model"]

                # Skip if not enough total samples
                if model_data["total"] < min_samples:
                    continue

                # Get older period pass rate (from cutoff to rolling_cutoff)
                older_result = (
                    session.query(
                        func.count(DBAnalysisResult.id),
                        func.sum(
                            case(
                                (DBAnalysisResult.overall_status == DBStatusType.PASS, 1),
                                else_=0
                            )
                        )
                    )
                    .filter(
                        DBAnalysisResult.model == model,
                        DBAnalysisResult.file_date >= cutoff_date,
                        DBAnalysisResult.file_date < rolling_cutoff,
                        DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name,
                    )
                    .first()
                )

                # Get recent period pass rate (from rolling_cutoff to now)
                recent_result = (
                    session.query(
                        func.count(DBAnalysisResult.id),
                        func.sum(
                            case(
                                (DBAnalysisResult.overall_status == DBStatusType.PASS, 1),
                                else_=0
                            )
                        )
                    )
                    .filter(
                        DBAnalysisResult.model == model,
                        DBAnalysisResult.file_date >= rolling_cutoff,
                        DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name,
                    )
                    .first()
                )

                older_count, older_passed = older_result
                recent_count, recent_passed = recent_result

                # Need sufficient samples in both periods
                if not older_count or older_count < 5:
                    continue
                if not recent_count or recent_count < 5:
                    continue

                older_pct = (older_passed or 0) / older_count * 100
                recent_pct = (recent_passed or 0) / recent_count * 100
                decline = older_pct - recent_pct

                # Only include if declining by threshold
                if decline >= trend_threshold:
                    trending_worse.append({
                        "model": model,
                        "pass_rate": model_data["pass_rate"],
                        "older_pass_rate": older_pct,
                        "recent_pass_rate": recent_pct,
                        "decline": decline,
                        "total_samples": model_data["total"],
                        "older_samples": older_count,
                        "recent_samples": recent_count,
                    })

            # Sort by decline magnitude (biggest drop first)
            trending_worse.sort(key=lambda x: x["decline"], reverse=True)

            return trending_worse

    def get_model_trend_data(
        self,
        model: str,
        days_back: int = 90,
        rolling_window: int = 30
    ) -> Dict[str, Any]:
        """
        Get detailed trend data for a specific model.

        Filters by trim date (file_date), not processing date.

        Args:
            model: Model number
            days_back: Total days to include (based on trim date)
            rolling_window: Days for rolling average

        Returns:
            Dict with trend data for charts
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Get all track results for this model (including anomaly flag)
            # Filter by trim date (file_date)
            results = (
                session.query(
                    DBAnalysisResult.file_date,
                    DBAnalysisResult.overall_status,
                    DBTrackResult.sigma_gradient,
                    DBTrackResult.sigma_threshold,
                    DBTrackResult.sigma_pass,
                    DBTrackResult.is_anomaly,
                    DBTrackResult.final_linearity_error_shifted,
                    DBTrackResult.linearity_spec,
                    DBTrackResult.linearity_pass,
                    DBTrackResult.linearity_fail_points,
                    DBTrackResult.untrimmed_sigma_gradient,
                )
                .join(DBTrackResult)
                .filter(
                    DBAnalysisResult.model == model,
                    DBAnalysisResult.file_date >= cutoff_date,
                )
                .order_by(DBAnalysisResult.file_date.asc())
                .all()
            )

            if not results:
                return {
                    "model": model,
                    "data_points": [],
                    "rolling_averages": [],
                    "pass_rates_by_day": [],
                    "threshold": None,
                }

            # Extract data points
            data_points = []
            for file_date, status, sigma_gradient, sigma_threshold, sigma_pass, is_anomaly, \
                linearity_error, linearity_spec, linearity_pass, fail_points, \
                untrimmed_sigma_gradient in results:
                if file_date and (sigma_gradient is not None
                                  or untrimmed_sigma_gradient is not None):
                    data_points.append({
                        "date": file_date,
                        "sigma_gradient": sigma_gradient,
                        "untrimmed_sigma_gradient": untrimmed_sigma_gradient,
                        "sigma_threshold": sigma_threshold,
                        "sigma_pass": sigma_pass,
                        "status": status.value if hasattr(status, 'value') else str(status) if status else "UNKNOWN",
                        "is_anomaly": is_anomaly or False,
                        "linearity_error": linearity_error,
                        "linearity_spec": linearity_spec,
                        "linearity_pass": linearity_pass,
                        "fail_points": fail_points or 0,
                    })

            # Calculate threshold (use mode of thresholds)
            thresholds = [d["sigma_threshold"] for d in data_points if d["sigma_threshold"]]
            threshold = max(set(thresholds), key=thresholds.count) if thresholds else None

            # Calculate linearity spec (use mode of specs)
            linearity_specs = [d["linearity_spec"] for d in data_points if d["linearity_spec"]]
            linearity_spec = max(set(linearity_specs), key=linearity_specs.count) if linearity_specs else None

            # Calculate daily pass rates for rolling average
            from collections import defaultdict
            daily_data = defaultdict(lambda: {"passed": 0, "total": 0})

            for dp in data_points:
                day_key = dp["date"].strftime("%Y-%m-%d")
                daily_data[day_key]["total"] += 1
                if dp["sigma_pass"]:
                    daily_data[day_key]["passed"] += 1

            # Sort by date and calculate pass rates
            sorted_days = sorted(daily_data.keys())
            pass_rates_by_day = []
            for day in sorted_days:
                d = daily_data[day]
                pass_rates_by_day.append({
                    "date": day,
                    "pass_rate": (d["passed"] / d["total"] * 100) if d["total"] > 0 else 0,
                    "total": d["total"],
                })

            # Calculate daily LINEARITY pass rates
            linearity_daily_data = defaultdict(lambda: {"passed": 0, "total": 0})

            for dp in data_points:
                if dp["linearity_pass"] is not None:  # Only count units with linearity data
                    day_key = dp["date"].strftime("%Y-%m-%d")
                    linearity_daily_data[day_key]["total"] += 1
                    if dp["linearity_pass"]:
                        linearity_daily_data[day_key]["passed"] += 1

            # Sort by date and calculate linearity pass rates
            sorted_lin_days = sorted(linearity_daily_data.keys())
            linearity_pass_rates_by_day = []
            for day in sorted_lin_days:
                d = linearity_daily_data[day]
                linearity_pass_rates_by_day.append({
                    "date": day,
                    "pass_rate": (d["passed"] / d["total"] * 100) if d["total"] > 0 else 0,
                    "total": d["total"],
                })

            # Calculate rolling averages
            rolling_averages = []
            if len(pass_rates_by_day) >= 2:
                # Use the specified rolling window
                window_size = min(rolling_window, len(pass_rates_by_day))
                for i in range(len(pass_rates_by_day)):
                    start_idx = max(0, i - window_size + 1)
                    window = pass_rates_by_day[start_idx:i + 1]

                    total_passed = sum(d["pass_rate"] * d["total"] for d in window)
                    total_count = sum(d["total"] for d in window)
                    rolling_avg = total_passed / total_count if total_count > 0 else 0

                    rolling_averages.append({
                        "date": pass_rates_by_day[i]["date"],
                        "rolling_avg": rolling_avg,
                        "window_size": len(window),
                    })

            return {
                "model": model,
                "data_points": data_points,
                "rolling_averages": rolling_averages,
                "pass_rates_by_day": pass_rates_by_day,
                "linearity_pass_rates_by_day": linearity_pass_rates_by_day,
                "threshold": threshold,
                "linearity_spec": linearity_spec,
                "total_samples": len(data_points),
            }

    # =========================================================================
    # Final Test Methods - For post-assembly test data and comparison
    # =========================================================================

    @staticmethod
    def _coerce_optional_bool(value: Any) -> Optional[bool]:
        """Coerce common stored bool representations without making None pass."""
        if value is None:
            return None
        if isinstance(value, str):
            normalized = value.strip().lower()
            if normalized in {"1", "true", "pass", "passed", "yes"}:
                return True
            if normalized in {"0", "false", "fail", "failed", "no"}:
                return False
            return None
        return bool(value)

    @classmethod
    def _resolve_final_test_linearity_pass(
        cls,
        test_results: Dict[str, Any],
        tracks: List[Dict[str, Any]],
    ) -> Optional[bool]:
        """Resolve file-level FT linearity from corrected track-level results.

        Parser headers can say PASS before the analyzer applies slope/offset
        correction. Zero-tolerance linearity means one failed corrected track
        must make the whole Final Test fail.
        """
        track_values = [
            cls._coerce_optional_bool(track.get("linearity_pass"))
            for track in tracks
        ]
        known_track_values = [value for value in track_values if value is not None]

        if any(value is False for value in known_track_values):
            return False
        if known_track_values:
            return all(known_track_values)

        return cls._coerce_optional_bool(test_results.get("linearity_pass"))

    def save_final_test(
        self,
        metadata: Dict[str, Any],
        tracks: List[Dict[str, Any]],
        test_results: Dict[str, Any],
        file_hash: str,
        file_size: Optional[int] = None,
        file_modified_date: Optional[datetime] = None,
    ) -> int:
        """
        Save a Final Test result to the database.

        Args:
            metadata: Dict with filename, model, serial, test_date, etc.
            tracks: List of track data dicts with positions, errors, etc.
            test_results: Dict with pass/fail for each test type
            file_hash: SHA256 hash of the file
            file_size: On-disk size, recorded for the incremental scan's stat
                fast-path (without it every later scan re-hashes this file)
            file_modified_date: On-disk mtime, same purpose

        Returns:
            ID of saved FinalTestResult
        """
        from sqlalchemy.exc import IntegrityError
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            FinalTestTrack as DBFinalTestTrack,
        )

        # Use lock to prevent race conditions with SQLite
        with self._write_lock:
            try:
                with self.session() as session:
                    # Check for duplicate by file_hash
                    existing = (
                        session.query(DBFinalTestResult)
                        .filter(DBFinalTestResult.file_hash == file_hash)
                        .first()
                    )
                    if existing:
                        # Stamp the stat onto a legacy row while we're here —
                        # this file was fully read to get here, so record what
                        # it costs nothing to record.
                        if file_size is not None and existing.file_size is None:
                            existing.file_size = file_size
                            existing.file_modified_date = file_modified_date
                            session.commit()
                        logger.debug(f"Final test already exists: {metadata.get('filename')}")
                        return existing.id

                    # Determine overall status from corrected track-level
                    # linearity first. The raw FT header can be stale after
                    # analyzer correction; any corrected track failure wins.
                    linearity_pass = self._resolve_final_test_linearity_pass(test_results, tracks)
                    overall_status = (
                        DBStatusType.FAIL if linearity_pass is False else DBStatusType.PASS
                    )

                    # Find matching trim result
                    linked_trim_id, match_confidence, days_since_trim, match_method = self._find_matching_trim(
                        session,
                        metadata.get("model"),
                        metadata.get("serial"),
                        metadata.get("file_date") or metadata.get("test_date")
                    )

                    # Create FinalTestResult
                    db_result = DBFinalTestResult(
                        filename=metadata.get("filename", "unknown"),
                        file_path=str(metadata.get("file_path", "")),
                        file_hash=file_hash,
                        file_date=metadata.get("file_date"),
                        file_size=file_size,
                        file_modified_date=file_modified_date,
                        model=metadata.get("model", "unknown"),
                        serial=metadata.get("serial", "unknown"),
                        test_date=metadata.get("test_date"),
                        overall_status=overall_status,
                        linearity_pass=linearity_pass,
                        linearity_error=tracks[0].get("linearity_error") if tracks else None,
                        resistance_pass=test_results.get("resistance_pass"),
                        resistance_value=test_results.get("resistance_value"),
                        resistance_tolerance=test_results.get("resistance_tolerance"),
                        electrical_angle_pass=test_results.get("electrical_angle_pass"),
                        hysteresis_pass=test_results.get("hysteresis_pass"),
                        phasing_pass=test_results.get("phasing_pass"),
                        linked_trim_id=linked_trim_id,
                        match_confidence=match_confidence,
                        days_since_trim=days_since_trim,
                        match_method=match_method,
                    )

                    session.add(db_result)
                    session.flush()
                    result_id = db_result.id

                    # Add tracks
                    for track_data in tracks:
                        # Use electrical_angles as position_data (X-axis for charts)
                        # electrical_angles contains: inches for linear pots, degrees for rotary
                        position_values = track_data.get("electrical_angles") or track_data.get("positions")

                        db_track = DBFinalTestTrack(
                            final_test_id=result_id,
                            track_id=track_data.get("track_id", "default"),
                            status=DBStatusType.PASS if track_data.get("linearity_pass", True) else DBStatusType.FAIL,
                            linearity_spec=track_data.get("linearity_spec"),
                            linearity_error=track_data.get("linearity_error"),
                            linearity_pass=track_data.get("linearity_pass"),
                            linearity_fail_points=track_data.get("linearity_fail_points", 0),
                            position_data=position_values,
                            error_data=track_data.get("errors"),
                            theory_data=track_data.get("theory_values"),
                            electrical_angle_data=track_data.get("electrical_angles"),
                            upper_limits=track_data.get("upper_limits"),
                            lower_limits=track_data.get("lower_limits"),
                            max_deviation=track_data.get("max_deviation"),
                            max_deviation_position=track_data.get("max_deviation_angle"),
                            optimal_offset=track_data.get("optimal_offset"),
                            optimal_slope=track_data.get("optimal_slope"),
                            linearity_type=track_data.get("linearity_type"),
                        )
                        session.add(db_track)

                    session.commit()
                    logger.debug(f"Saved Final Test: {metadata.get('filename')} (ID: {result_id}, linked_trim: {linked_trim_id})")
                    return result_id

            except IntegrityError as e:
                # The unique constraint that can fire is on
                # (filename, file_date, model, serial), not on file_hash.
                # When the same file is reprocessed with edited content the
                # hash differs but the tuple still matches, so the previous
                # hash-only fallback couldn't find the existing row and the
                # error propagated to the user as "Error processing Final
                # Test ... UNIQUE constraint failed". Query by both keys.
                logger.warning(f"Final test duplicate detected (race condition): {metadata.get('filename')}")
                try:
                    with self.session() as session:
                        existing = (
                            session.query(DBFinalTestResult)
                            .filter(DBFinalTestResult.file_hash == file_hash)
                            .first()
                        )
                        if existing is None:
                            existing = (
                                session.query(DBFinalTestResult)
                                .filter(
                                    DBFinalTestResult.filename == metadata.get("filename"),
                                    DBFinalTestResult.file_date == metadata.get("file_date"),
                                    DBFinalTestResult.model == metadata.get("model"),
                                    DBFinalTestResult.serial == metadata.get("serial"),
                                )
                                .first()
                            )
                        if existing:
                            return existing.id
                except Exception:
                    logger.debug("FT duplicate-recovery query failed", exc_info=True)
                raise

    def get_ml_staleness(self) -> List[Dict[str, Any]]:
        """
        Get ML training staleness info for each trained model.

        Compares training_samples at training time vs current record count.
        Models with 50+ new records since training are flagged for retrain.

        Returns:
            List of dicts: model, training_date, training_samples,
            current_samples, new_since_training, needs_retrain, days_since_training
        """
        from laser_trim_analyzer.database.models import ModelMLState as DBModelMLState

        results = []
        with self.session() as session:
            # Get all trained models
            ml_states = session.query(DBModelMLState).filter(
                DBModelMLState.is_trained == True
            ).all()

            for state in ml_states:
                current_count = session.query(
                    func.count(DBAnalysisResult.id)
                ).filter(
                    DBAnalysisResult.model == state.model
                ).scalar() or 0

                training_samples = state.training_samples or 0
                new_records = max(0, current_count - training_samples)
                # tz-robust: training_date may be stored tz-aware (utc_now); strip
                # tzinfo so subtracting from naive datetime.now() never raises.
                _td = state.training_date
                if _td is not None and _td.tzinfo is not None:
                    _td = _td.replace(tzinfo=None)
                days_since = (datetime.now() - _td).days if _td else 999

                results.append({
                    "model": state.model,
                    "training_date": state.training_date,
                    "training_samples": training_samples,
                    "current_samples": current_count,
                    "new_since_training": new_records,
                    "needs_retrain": new_records >= 50,
                    "days_since_training": days_since,
                })

        # Sort by most stale first
        results.sort(key=lambda x: x["new_since_training"], reverse=True)
        return results

    def get_screening_recommendations(self, days_back: int = 90, min_samples: int = 20) -> List[Dict[str, Any]]:
        """
        Generate element screening recommendations based on failure patterns.

        Flags:
        - Near-miss rate >40%: candidate for in-process testing
        - Failure rate >50%: candidate for design review
        - High volume + high failure: candidate for incoming inspection

        Returns:
            List of recommendation dicts sorted by priority
        """
        recommendations = []

        try:
            priority_models = self.get_linearity_prioritization(
                days_back=days_back, min_samples=min_samples
            )
        except Exception:
            return recommendations

        for m in priority_models:
            model = m.get("model", "Unknown")
            lin_rate = m.get("linearity_pass_rate", 100)
            fail_rate = 100 - lin_rate
            failed = m.get("failed_units", 0)
            total = m.get("total_tracks", 0)
            near_miss = m.get("near_miss_count", 0)
            near_miss_ratio = near_miss / failed if failed > 0 else 0

            rec = {
                "model": model,
                "fail_rate": fail_rate,
                "total_units": total,
                "failed_units": failed,
                "near_miss_count": near_miss,
                "recommendations": [],
            }

            if fail_rate > 50:
                rec["recommendations"].append({
                    "type": "design_review",
                    "text": f"Design review — {fail_rate:.0f}% failure rate",
                    "priority": "high",
                })

            if near_miss_ratio > 0.4 and near_miss >= 3:
                rec["recommendations"].append({
                    "type": "in_process_testing",
                    "text": f"In-process testing candidate — {near_miss} near-miss ({near_miss_ratio:.0%} of failures)",
                    "priority": "medium",
                })

            if failed >= 10 and fail_rate > 30:
                rec["recommendations"].append({
                    "type": "incoming_inspection",
                    "text": f"Incoming element inspection — {failed} failures, high volume",
                    "priority": "medium",
                })

            if rec["recommendations"]:
                recommendations.append(rec)

        return recommendations

    def get_model_cpk(self, model: str, days_back: int = 90) -> Dict[str, Any]:
        """
        Calculate process capability (Cpk) for a model.

        Cpk = min(USL - mean, mean - LSL) / (3 * sigma)

        For sigma gradient: LSL=0, USL=threshold (one-sided: Cpk = (USL - mean) / (3*std))
        For linearity: based on pass rate as a capability proxy

        Args:
            model: Model number
            days_back: Period to analyze

        Returns:
            Dict with sigma_cpk, sigma_cpk_color, sample_count
        """
        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Get sigma gradient data for this model
            data = (
                session.query(
                    DBTrackResult.sigma_gradient,
                    DBTrackResult.sigma_threshold,
                )
                .join(DBAnalysisResult)
                .filter(
                    DBAnalysisResult.model == model,
                    DBAnalysisResult.file_date >= cutoff_date,
                    DBTrackResult.sigma_gradient.isnot(None),
                    DBTrackResult.sigma_threshold.isnot(None),
                )
                .all()
            )

            result = {
                "model": model,
                "sample_count": len(data),
                "sigma_cpk": None,
                "sigma_cpk_color": "gray",
            }

            if len(data) < 10:
                return result

            gradients = [float(d[0]) for d in data]
            threshold = float(data[0][1])  # Threshold is same for all tracks of a model

            import numpy as np
            mean = np.mean(gradients)
            std = np.std(gradients, ddof=1)  # Sample std dev

            if std > 0 and threshold > 0:
                # One-sided Cpk: process must stay BELOW threshold
                # Cpk = (USL - mean) / (3 * sigma)
                cpk = (threshold - mean) / (3 * std)
                result["sigma_cpk"] = round(cpk, 2)

                if cpk < 1.0:
                    result["sigma_cpk_color"] = "#e74c3c"  # Red — incapable
                elif cpk < 1.33:
                    result["sigma_cpk_color"] = "#f39c12"  # Yellow — marginal
                else:
                    result["sigma_cpk_color"] = "#27ae60"  # Green — capable

            return result

    @staticmethod
    def _normalize_serial(serial: str) -> str:
        """
        Normalize a serial number for fuzzy matching (selective).

        Handles common formatting differences between trim and FT files:
        - Strip leading zeros (007 -> 7)
        - Lowercase
        - Strip whitespace
        - Remove common prefixes (sn, s/n, #)
        - Strip known track-position suffixes only (A/B for dual-track,
          P/R for primary/redundant, T for test)
        - Do NOT strip other letters (25D, 31L stay as-is since they may
          be meaningful serial identifiers)
        """
        import re
        s = serial.lower().strip()
        s = re.sub(r'^(sn|s/n|s\.n\.|#)\s*', '', s)
        # Strip only known track-indicator suffixes
        s = re.sub(r'^(\d+)[abprt]$', r'\1', s)
        s = s.lstrip('0') or '0'
        return s

    @staticmethod
    def _normalize_serial_aggressive(serial: str) -> str:
        """
        Aggressively normalize a serial number — strips ALL trailing letters.

        Used as a fallback when selective normalization fails to find a match.
        May produce false matches (e.g. 25D matches 25E) but increases recall.
        """
        import re
        s = serial.lower().strip()
        s = re.sub(r'^(sn|s/n|s\.n\.|#)\s*', '', s)
        s = re.sub(r'^(\d+)[a-z]$', r'\1', s)
        s = s.lstrip('0') or '0'
        return s

    @staticmethod
    def _normalize_model(model: str) -> str:
        """
        Normalize a model number to its base form for variant matching.

        Strips trailing letter suffixes that indicate product variants:
        - 8275A, 8275B, 8275C → 8275
        - 8508-A, 8508-B → 8508
        - 7280-1-CT, 7280-1-AB → 7280-1

        Strips leading zeros in hyphenated suffixes:
        - 2475-08 → 2475-8
        - 8867-01 → 8867-1

        Does NOT strip numeric suffixes (8340-1 stays 8340-1) since
        those are distinct model configurations.
        """
        import re
        if not model:
            return model
        # Strip leading zeros in hyphenated numeric suffixes: "2475-08" → "2475-8"
        s = re.sub(r'-0+(\d)', r'-\1', model)
        # Strip trailing letter-only variant: "8275A" → "8275"
        s = re.sub(r'^(\d+)[A-Za-z]$', r'\1', s)
        # Strip trailing hyphen + letter(s) variant: "8508-A" → "8508", "7280-1-CT" → "7280-1"
        s = re.sub(r'^(\d+(?:-\d+)*)-[A-Za-z]+$', r'\1', s)
        # Strip trailing letters glued to a hyphenated numeric suffix:
        # "7953-1A" → "7953-1" (2026-07-13: FT files say "7953-1", trim files
        # say "7953-1A"/"7953-1B" — 197 recent FT records unlinkable without this)
        s = re.sub(r'^(\d+(?:-\d+)+)[A-Za-z]+$', r'\1', s)
        return s

    def _find_matching_trim(
        self,
        session: Session,
        model: Optional[str],
        serial: Optional[str],
        test_date: Optional[datetime]
    ) -> Tuple[Optional[int], Optional[float], Optional[int], Optional[str]]:
        """
        Find the matching trim result for a final test.

        Logic:
        1. Exact model + exact serial match (case-insensitive) — highest confidence
        2. Exact model + fuzzy serial match (strip zeros, prefixes, track suffixes)
        3. Normalized model + fuzzy serial match (8275A trim matches 8275 FT)

        Returns:
            Tuple of (trim_id, confidence, days_since_trim, match_method)
        """
        from laser_trim_analyzer.utils.constants import FINAL_TEST_MAX_DAYS_FROM_TRIM

        if not model or not serial or not test_date:
            return None, None, None, None

        serial_clean = serial.lower().strip()
        cutoff_date = test_date - timedelta(days=FINAL_TEST_MAX_DAYS_FROM_TRIM)

        # Attempt 1: Exact model + exact serial match (case-insensitive)
        candidates = (
            session.query(DBAnalysisResult)
            .filter(
                DBAnalysisResult.model == model,
                func.lower(DBAnalysisResult.serial) == serial_clean,
                DBAnalysisResult.file_date.isnot(None),
                DBAnalysisResult.file_date <= test_date,
                DBAnalysisResult.file_date >= cutoff_date,
            )
            .order_by(desc(DBAnalysisResult.file_date))
            .limit(5)
            .all()
        )

        if candidates:
            match = candidates[0]
            days_diff = (test_date - match.file_date).days
            confidence = self._calculate_match_confidence(days_diff, exact_serial=True)
            return match.id, confidence, days_diff, "exact"

        # Attempt 2: Exact model + fuzzy serial match
        ft_serial_norm = self._normalize_serial(serial)

        model_trims = (
            session.query(DBAnalysisResult.id, DBAnalysisResult.serial, DBAnalysisResult.file_date)
            .filter(
                DBAnalysisResult.model == model,
                DBAnalysisResult.file_date.isnot(None),
                DBAnalysisResult.file_date <= test_date,
                DBAnalysisResult.file_date >= cutoff_date,
            )
            .order_by(desc(DBAnalysisResult.file_date))
            .all()
        )

        for trim_id, trim_serial, trim_date in model_trims:
            if trim_serial and self._normalize_serial(trim_serial) == ft_serial_norm:
                days_diff = (test_date - trim_date).days
                confidence = self._calculate_match_confidence(days_diff, exact_serial=False)
                logger.debug(
                    f"Fuzzy match: FT serial '{serial}' → trim serial '{trim_serial}' "
                    f"(normalized: '{ft_serial_norm}'), {days_diff} days"
                )
                return trim_id, confidence, days_diff, "fuzzy_serial"

        # Attempt 2b: Exact model + aggressively normalized serial (strips all trailing letters)
        ft_serial_aggressive = self._normalize_serial_aggressive(serial)
        if ft_serial_aggressive != ft_serial_norm:
            for trim_id, trim_serial, trim_date in model_trims:
                if trim_serial and self._normalize_serial_aggressive(trim_serial) == ft_serial_aggressive:
                    days_diff = (test_date - trim_date).days
                    confidence = self._calculate_match_confidence(days_diff, exact_serial=False) * 0.90
                    logger.debug(
                        f"Aggressive fuzzy match: FT serial '{serial}' -> trim serial '{trim_serial}' "
                        f"(aggressive norm: '{ft_serial_aggressive}'), {days_diff} days"
                    )
                    return trim_id, confidence, days_diff, "fuzzy_serial_aggressive"

        # Attempt 3: Model variant matching — normalize model on both sides
        # This handles cases like FT model "8275" matching trim model "8275A"
        # or FT model "8508" matching trim model "8508-A"
        ft_model_norm = self._normalize_model(model)

        if ft_model_norm != model:
            # FT model itself has a suffix — try base model in trim
            variant_trims = (
                session.query(DBAnalysisResult.id, DBAnalysisResult.serial,
                              DBAnalysisResult.file_date, DBAnalysisResult.model)
                .filter(
                    DBAnalysisResult.model == ft_model_norm,
                    DBAnalysisResult.file_date.isnot(None),
                    DBAnalysisResult.file_date <= test_date,
                    DBAnalysisResult.file_date >= cutoff_date,
                )
                .order_by(desc(DBAnalysisResult.file_date))
                .all()
            )
            for trim_id, trim_serial, trim_date, trim_model in variant_trims:
                if trim_serial and self._normalize_serial(trim_serial) == ft_serial_norm:
                    days_diff = (test_date - trim_date).days
                    confidence = self._calculate_match_confidence(days_diff, exact_serial=False, model_variant=True)
                    logger.debug(
                        f"Model variant match: FT {model}/{serial} → trim {trim_model}/{trim_serial} "
                        f"(normalized model: '{ft_model_norm}'), {days_diff} days"
                    )
                    return trim_id, confidence, days_diff, "model_variant"

        # Try reverse: trim has variant suffixes, FT has base model
        # Find all trim models that normalize to our FT model
        # Use LIKE to find variants efficiently (e.g. "8275%" for FT model "8275")
        variant_trims = (
            session.query(DBAnalysisResult.id, DBAnalysisResult.serial,
                          DBAnalysisResult.file_date, DBAnalysisResult.model)
            .filter(
                DBAnalysisResult.model.like(f"{model}%"),
                DBAnalysisResult.model != model,  # Skip exact (already tried)
                DBAnalysisResult.file_date.isnot(None),
                DBAnalysisResult.file_date <= test_date,
                DBAnalysisResult.file_date >= cutoff_date,
            )
            .order_by(desc(DBAnalysisResult.file_date))
            .all()
        )

        for trim_id, trim_serial, trim_date, trim_model in variant_trims:
            # Verify this is actually a variant (normalizes to same base)
            if self._normalize_model(trim_model) != ft_model_norm:
                continue
            if trim_serial and self._normalize_serial(trim_serial) == ft_serial_norm:
                days_diff = (test_date - trim_date).days
                confidence = self._calculate_match_confidence(days_diff, exact_serial=False, model_variant=True)
                logger.debug(
                    f"Model variant match: FT {model}/{serial} → trim {trim_model}/{trim_serial} "
                    f"(base model: '{ft_model_norm}'), {days_diff} days"
                )
                return trim_id, confidence, days_diff, "model_variant"

        return None, None, None, None

    @staticmethod
    def _calculate_match_confidence(days_diff: int, exact_serial: bool = True,
                                     model_variant: bool = False) -> float:
        """Calculate match confidence based on time proximity and match quality.

        Confidence bands:
        - Exact model + exact serial, same week: 0.93-1.00
        - Exact model + fuzzy serial, same week: 0.84-0.90
        - Model variant + fuzzy serial, same week: 0.71-0.77
        - Any match beyond 30 days: drops significantly
        """
        # Time-based confidence. Decay beyond 30 days is 0.002/day so the
        # scale spans the full 180-day match window (0.40 ≈ 180d); the old
        # 0.007/day rate hit the 0.40 floor by day ~73 and couldn't tell a
        # 75-day link from a 175-day one.
        if days_diff <= 7:
            time_conf = 1.0 - (days_diff * 0.01)
        elif days_diff <= 30:
            time_conf = 0.9 - ((days_diff - 7) * 0.01)
        else:
            time_conf = 0.7 - ((days_diff - 30) * 0.002)

        # Match quality penalties (applied multiplicatively)
        if not exact_serial:
            time_conf *= 0.90  # was 0.95 — fuzzy serial is less certain

        if model_variant:
            time_conf *= 0.85  # model variant adds uncertainty

        return max(0.40, time_conf)

    def get_unmatched_ft_diagnostics(self, limit: int = 100) -> List[Dict[str, Any]]:
        """
        Diagnose why Final Test records are unmatched.

        For each unmatched FT record, checks:
        - Does the model exist in trim data?
        - Does the serial exist for that model?
        - Are there trims but outside the date window?

        Returns:
            List of dicts with FT info and reason for no match
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
        )
        from laser_trim_analyzer.utils.constants import FINAL_TEST_MAX_DAYS_FROM_TRIM

        results = []

        with self.session() as session:
            unmatched = (
                session.query(DBFinalTestResult)
                .filter(DBFinalTestResult.linked_trim_id.is_(None))
                .order_by(desc(DBFinalTestResult.file_date))
                .limit(limit)
                .all()
            )

            for ft in unmatched:
                diag = {
                    "ft_id": ft.id,
                    "filename": ft.filename,
                    "model": ft.model,
                    "serial": ft.serial,
                    "test_date": ft.file_date or ft.test_date,
                    "reason": "unknown",
                }

                # Check 1: Does model exist in trim data?
                model_exists = session.query(
                    func.count(DBAnalysisResult.id)
                ).filter(
                    DBAnalysisResult.model == ft.model
                ).scalar() or 0

                if model_exists == 0:
                    diag["reason"] = "no_model_in_trims"
                    diag["detail"] = f"Model '{ft.model}' has no trim records"
                    results.append(diag)
                    continue

                # Check 2: Does serial exist for this model?
                serial_clean = ft.serial.lower().strip() if ft.serial else ""
                serial_exists = session.query(
                    func.count(DBAnalysisResult.id)
                ).filter(
                    DBAnalysisResult.model == ft.model,
                    func.lower(DBAnalysisResult.serial) == serial_clean,
                ).scalar() or 0

                if serial_exists == 0:
                    # Check fuzzy match
                    ft_norm = self._normalize_serial(ft.serial) if ft.serial else ""
                    all_serials = session.query(
                        DBAnalysisResult.serial
                    ).filter(
                        DBAnalysisResult.model == ft.model
                    ).distinct().limit(5).all()
                    sample = [s[0] for s in all_serials]
                    diag["reason"] = "no_serial_match"
                    diag["detail"] = (
                        f"Serial '{ft.serial}' (norm: '{ft_norm}') not found. "
                        f"Sample trim serials: {sample}"
                    )
                    results.append(diag)
                    continue

                # Check 3: Serial exists but outside date window
                test_date = ft.file_date or ft.test_date
                if test_date:
                    cutoff = test_date - timedelta(days=FINAL_TEST_MAX_DAYS_FROM_TRIM)
                    in_window = session.query(
                        func.count(DBAnalysisResult.id)
                    ).filter(
                        DBAnalysisResult.model == ft.model,
                        func.lower(DBAnalysisResult.serial) == serial_clean,
                        DBAnalysisResult.file_date.isnot(None),
                        DBAnalysisResult.file_date <= test_date,
                        DBAnalysisResult.file_date >= cutoff,
                    ).scalar() or 0

                    if in_window == 0:
                        # Find nearest trim date
                        nearest = session.query(
                            DBAnalysisResult.file_date
                        ).filter(
                            DBAnalysisResult.model == ft.model,
                            func.lower(DBAnalysisResult.serial) == serial_clean,
                            DBAnalysisResult.file_date.isnot(None),
                        ).order_by(
                            func.abs(func.julianday(DBAnalysisResult.file_date) - func.julianday(test_date))
                        ).first()

                        if nearest and nearest[0]:
                            gap = abs((test_date - nearest[0]).days)
                            diag["reason"] = "outside_date_window"
                            diag["detail"] = (
                                f"Nearest trim is {gap} days away "
                                f"(max allowed: {FINAL_TEST_MAX_DAYS_FROM_TRIM})"
                            )
                        else:
                            diag["reason"] = "no_dated_trims"
                            diag["detail"] = "Matching trims have no file_date"
                    else:
                        diag["reason"] = "trim_after_test"
                        diag["detail"] = "Trims exist but all are after the FT date"
                else:
                    diag["reason"] = "no_test_date"
                    diag["detail"] = "FT record has no date"

                results.append(diag)

        return results

    def rematch_final_tests(self) -> Dict[str, int]:
        """
        Re-run matching for all Final Test records against current trim data.

        This is useful when trim files are imported after Final Test files,
        or when trim data has been updated.

        Returns:
            Dict with counts: new_matches, updated_matches, unchanged, total
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
        )

        stats = {"new_matches": 0, "updated_matches": 0, "unchanged": 0, "total": 0}

        with self._write_lock:
            with self.session() as session:
                # Get all Final Test records
                final_tests = session.query(DBFinalTestResult).all()
                stats["total"] = len(final_tests)

                for ft in final_tests:
                    # Get test date (prefer file_date, fall back to test_date)
                    test_date = ft.file_date or ft.test_date

                    # Find matching trim
                    new_trim_id, new_confidence, new_days, new_method = self._find_matching_trim(
                        session, ft.model, ft.serial, test_date
                    )

                    # Check if match changed
                    if new_trim_id != ft.linked_trim_id:
                        if ft.linked_trim_id is None and new_trim_id is not None:
                            stats["new_matches"] += 1
                        elif ft.linked_trim_id is not None and new_trim_id is not None:
                            stats["updated_matches"] += 1

                        # Update the record
                        ft.linked_trim_id = new_trim_id
                        ft.match_confidence = new_confidence
                        ft.days_since_trim = new_days
                        ft.match_method = new_method
                    else:
                        stats["unchanged"] += 1

                session.commit()
                logger.info(
                    f"Rematch complete: {stats['new_matches']} new, "
                    f"{stats['updated_matches']} updated, {stats['unchanged']} unchanged"
                )

        return stats

    def rematch_unlinked_final_tests(self, models=None) -> Dict[str, int]:
        """Link-only rematch pass over FT records that have NO trim link yet.

        Why this exists (2026-07-13): matching runs at FT save time, so an FT
        file processed in the same batch as — or any batch before — its trim
        file finds nothing and stays NULL forever. Nothing ever retried. This
        runs automatically after every processing batch.

        `models`: restrict the pass to FT records for these models (the models
        whose trims were just saved). A late trim can only create links for
        the models in that batch, so re-attempting the other ~100k
        permanently-unmatchable FT records every batch is pure waste — the
        work log read "Unlinked-FT rematch: 0 of 101,605 linked" after every
        single batch. Scoping is by NORMALIZED model family, not by exact
        name, so the model-variant stage ("8275" FT ↔ "8275A" trim) still
        works. None = every unlinked record (full pass).

        Bulk strategy: instead of the per-record query cascade (minutes for
        30k+ records), load every candidate trim ONCE, build serial-form
        indexes in memory, and answer each FT record with bisect lookups.
        Match semantics mirror _find_matching_trim exactly: newest trim at or
        before the FT date within FINAL_TEST_MAX_DAYS_FROM_TRIM, staged
        exact serial → fuzzy → aggressive → model-variant.

        Existing links are never touched — rematch_final_tests() re-evaluates
        everything if that is ever needed.
        """
        import bisect
        import time as _time
        from collections import defaultdict
        from sqlalchemy import update as sa_update
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
        )
        from laser_trim_analyzer.utils.constants import FINAL_TEST_MAX_DAYS_FROM_TRIM

        t0 = _time.time()
        stats = {"unlinked": 0, "new_matches": 0, "still_unmatched": 0,
                 "seconds": 0.0, "models": []}
        window = timedelta(days=FINAL_TEST_MAX_DAYS_FROM_TRIM)
        families = None
        if models is not None:
            families = {self._normalize_model(m) for m in models if m}
            if not families:
                return stats

        with self._write_lock:
            with self.session() as session:
                # Narrow column query, not full ORM entities: the loop reads
                # five fields, and materializing 100k+ FinalTestResult objects
                # (with their identity map) was minutes of the post-batch cost.
                pending = (
                    session.query(
                        DBFinalTestResult.id, DBFinalTestResult.model,
                        DBFinalTestResult.serial, DBFinalTestResult.file_date,
                        DBFinalTestResult.test_date,
                    )
                    .filter(DBFinalTestResult.linked_trim_id.is_(None))
                    .all()
                )
                if families is not None:
                    # Normalization is Python-side, so the family filter can't
                    # be pushed into SQL — but the rows are narrow tuples.
                    pending = [ft for ft in pending if ft.model
                               and self._normalize_model(ft.model) in families]
                stats["unlinked"] = len(pending)
                if not pending:
                    return stats

                # One pass over all trims → per-serial-form indexes of
                # (file_date, trim_id) sorted ascending, plus a per-family
                # index for variant matches ("8275" FT ↔ "8275A" trim).
                exact_ix: dict = defaultdict(list)    # (model, serial_lower) → [(date, id)]
                fuzzy_ix: dict = defaultdict(list)    # (model, norm_serial)  → [(date, id)]
                aggr_ix: dict = defaultdict(list)     # (model, aggr_serial)  → [(date, id)]
                variant_ix: dict = defaultdict(list)  # (norm_model, norm_serial) → [(date, id, model)]
                rows = session.query(
                    DBAnalysisResult.id, DBAnalysisResult.model,
                    DBAnalysisResult.serial, DBAnalysisResult.file_date,
                ).filter(
                    DBAnalysisResult.file_date.isnot(None),
                    DBAnalysisResult.serial.isnot(None),
                ).order_by(DBAnalysisResult.file_date).all()
                for tid, tmodel, tserial, tdate in rows:
                    if families is not None and self._normalize_model(tmodel) not in families:
                        continue    # can't match any in-scope FT record
                    s_low = tserial.lower().strip()
                    s_norm = self._normalize_serial(tserial)
                    exact_ix[(tmodel, s_low)].append((tdate, tid))
                    fuzzy_ix[(tmodel, s_norm)].append((tdate, tid))
                    aggr_ix[(tmodel, self._normalize_serial_aggressive(tserial))].append((tdate, tid))
                    variant_ix[(self._normalize_model(tmodel), s_norm)].append((tdate, tid, tmodel))

                def newest_in_window(entries, test_date, cutoff, model_ok=None):
                    """Rightmost entry with cutoff <= date <= test_date whose
                    model passes model_ok (None = any)."""
                    i = bisect.bisect_right(entries, (test_date, float("inf"))) - 1
                    while i >= 0:
                        e = entries[i]
                        if e[0] < cutoff:
                            return None
                        if model_ok is None or model_ok(e[2]):
                            return e
                        i -= 1
                    return None

                affected_models = set()
                link_updates: List[Dict[str, Any]] = []
                for ft in pending:
                    test_date = ft.file_date or ft.test_date
                    if not ft.model or not ft.serial or not test_date:
                        stats["still_unmatched"] += 1
                        continue
                    cutoff = test_date - window
                    s_low = ft.serial.lower().strip()
                    s_norm = self._normalize_serial(ft.serial)
                    hit = method = None
                    exact_serial = True
                    variant = False
                    penalty = 1.0
                    e = newest_in_window(exact_ix.get((ft.model, s_low), ()), test_date, cutoff)
                    if e:
                        hit, method = e, "exact"
                    if hit is None:
                        e = newest_in_window(fuzzy_ix.get((ft.model, s_norm), ()), test_date, cutoff)
                        if e:
                            hit, method, exact_serial = e, "fuzzy_serial", False
                    if hit is None:
                        # Parity with _find_matching_trim attempt 2b: the
                        # aggressive stage only runs when the FT serial itself
                        # changes under aggressive normalization (code-review
                        # finding #2, 2026-07-13 — the bulk path previously
                        # probed unconditionally and could link serial "123"
                        # to trim "123X", which save-time matching never does).
                        s_aggr = self._normalize_serial_aggressive(ft.serial)
                        if s_aggr != s_norm:
                            e = newest_in_window(
                                aggr_ix.get((ft.model, s_aggr), ()), test_date, cutoff)
                            if e:
                                hit, method, exact_serial, penalty = (
                                    e, "fuzzy_serial_aggressive", False, 0.90)
                    if hit is None:
                        # Variant stage — one side MUST be the base form
                        # (code-review finding #1, 2026-07-13 BLOCKER: keying
                        # only on the normalized family let FT "7953-1A" link
                        # to a "7953-1B" trim — SIBLING variants, a match class
                        # _find_matching_trim never makes. Attempt 3a matches
                        # trims whose model IS the base; attempt 3b matches
                        # variant trims only when the FT model is the base).
                        ft_norm = self._normalize_model(ft.model)
                        ft_is_base = (ft.model == ft_norm)
                        e = newest_in_window(
                            variant_ix.get((ft_norm, s_norm), ()),
                            test_date, cutoff,
                            model_ok=lambda m, _fn=ft_norm, _fm=ft.model, _fb=ft_is_base:
                                m != _fm and (_fb or m == _fn))
                        if e:
                            hit, method, exact_serial, variant = e, "model_variant", False, True
                    if hit is None:
                        stats["still_unmatched"] += 1
                        continue
                    days_diff = (test_date - hit[0]).days
                    link_updates.append({
                        "id": ft.id,
                        "linked_trim_id": hit[1],
                        "match_confidence": self._calculate_match_confidence(
                            days_diff, exact_serial=exact_serial,
                            model_variant=variant) * penalty,
                        "days_since_trim": days_diff,
                        "match_method": method,
                    })
                    affected_models.add(ft.model)
                    stats["new_matches"] += 1
                stats["models"] = sorted(affected_models)

                if link_updates:
                    # Bulk UPDATE ... WHERE id = :id (one statement, no ORM
                    # objects) — same four columns the loop used to assign.
                    session.execute(sa_update(DBFinalTestResult), link_updates)
                session.commit()

        stats["seconds"] = round(_time.time() - t0, 1)
        logger.info(
            "Unlinked-FT rematch (%s): %d of %d linked in %.1fs (%d still unmatched)",
            "all models" if families is None
            else f"{len(families)} model(s) from this batch",
            stats["new_matches"], stats["unlinked"], stats["seconds"],
            stats["still_unmatched"],
        )
        return stats

    def get_final_test(self, final_test_id: int) -> Optional[Dict[str, Any]]:
        """
        Get a Final Test result by ID.

        Returns:
            Dict with final test data, tracks, and linked trim info
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            FinalTestTrack as DBFinalTestTrack,
        )

        with self.session() as session:
            result = session.query(DBFinalTestResult).filter(
                DBFinalTestResult.id == final_test_id
            ).first()

            if not result:
                return None

            # Get tracks
            tracks = session.query(DBFinalTestTrack).filter(
                DBFinalTestTrack.final_test_id == final_test_id
            ).all()

            # Get linked trim if exists
            linked_trim = None
            if result.linked_trim_id:
                linked_trim = self._get_analysis_summary(session, result.linked_trim_id)

            return {
                "id": result.id,
                "filename": result.filename,
                "model": result.model,
                "serial": result.serial,
                "test_date": result.test_date,
                "file_date": result.file_date,
                "overall_status": result.overall_status.value if result.overall_status else "UNKNOWN",
                "linearity_pass": result.linearity_pass,
                "linearity_error": result.linearity_error,
                "resistance_pass": result.resistance_pass,
                "resistance_value": result.resistance_value,
                "linked_trim_id": result.linked_trim_id,
                "match_confidence": result.match_confidence,
                "days_since_trim": result.days_since_trim,
                "match_method": getattr(result, 'match_method', None),
                "linked_trim": linked_trim,
                "tracks": [
                    {
                        "track_id": t.track_id,
                        "status": t.status.value if t.status else "UNKNOWN",
                        "linearity_pass": t.linearity_pass,
                        "linearity_error": t.linearity_error,
                        "linearity_fail_points": t.linearity_fail_points,
                        # Use position_data if available, fall back to electrical_angle_data
                        "positions": t.position_data or t.electrical_angle_data or [],
                        "errors": t.error_data or [],
                        "theory_data": t.theory_data,
                        "electrical_angles": t.electrical_angle_data or [],
                        "upper_limits": t.upper_limits or [],
                        "lower_limits": t.lower_limits or [],
                        # Slope/offset/linearity_type are persisted on the track
                        # row but were previously omitted here, leaving callers
                        # (export, diagnostics) without the analyzer correction
                        # state. get_comparison_data already returns these.
                        "optimal_offset": getattr(t, "optimal_offset", None),
                        "optimal_slope": getattr(t, "optimal_slope", None),
                        "linearity_type": getattr(t, "linearity_type", None),
                    }
                    for t in tracks
                ],
            }

    def _get_analysis_summary(self, session: Session, analysis_id: int) -> Optional[Dict[str, Any]]:
        """Get a summary of an analysis result for linking."""
        result = session.query(DBAnalysisResult).filter(
            DBAnalysisResult.id == analysis_id
        ).first()

        if not result:
            return None

        return {
            "id": result.id,
            "filename": result.filename,
            "model": result.model,
            "serial": result.serial,
            "file_date": result.file_date,
            "overall_status": result.overall_status.value if result.overall_status else "UNKNOWN",
        }

    def search_final_tests(
        self,
        model: Optional[str] = None,
        serial: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        status: Optional[str] = None,
        linked_only: bool = False,
        limit: int = 500
    ) -> List[Dict[str, Any]]:
        """
        Search Final Test records with filters.

        Supports partial matching on serial number (case-insensitive).
        Filters by test date (file_date).

        Uses LEFT JOIN to fetch linked trim data in single query (avoids N+1).

        Args:
            model: Filter by model number (exact match, None for all)
            serial: Filter by serial number (partial match, case-insensitive)
            date_from: Start of date range (inclusive)
            date_to: End of date range (inclusive)
            status: Filter by status (Pass/Fail, None for all)
            linked_only: If True, only return records with linked trim
            limit: Maximum number of results

        Returns:
            List of Final Test result dicts sorted by test date (newest first)
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
        )

        with self.session() as session:
            # Use joinedload to fetch linked_trim in single query (avoids N+1)
            query = session.query(DBFinalTestResult).options(
                joinedload(DBFinalTestResult.linked_trim)
            )

            # Filter by model (exact match)
            if model and model != "All Models":
                query = query.filter(DBFinalTestResult.model == model)

            # Filter by serial (partial match, case-insensitive)
            if serial and serial.strip():
                serial_pattern = f"%{serial.strip()}%"
                query = query.filter(
                    func.lower(DBFinalTestResult.serial).like(func.lower(serial_pattern))
                )

            # Filter by date range (test date)
            if date_from:
                query = query.filter(DBFinalTestResult.file_date >= date_from)
            if date_to:
                # Include the entire end date (until midnight)
                end_of_day = date_to.replace(hour=23, minute=59, second=59)
                query = query.filter(DBFinalTestResult.file_date <= end_of_day)

            # Filter by status
            if status and status != "All":
                if status == "Pass":
                    query = query.filter(DBFinalTestResult.overall_status == DBStatusType.PASS)
                elif status == "Fail":
                    query = query.filter(DBFinalTestResult.overall_status == DBStatusType.FAIL)

            # Filter by linked status
            if linked_only:
                query = query.filter(DBFinalTestResult.linked_trim_id.isnot(None))

            # Order by test date (newest first) and limit
            results = query.order_by(desc(DBFinalTestResult.file_date)).limit(limit).all()

            pairs = []
            for ft in results:
                # Get linked trim info from pre-loaded relationship (no extra query)
                linked_trim = None
                if ft.linked_trim:
                    linked_trim = {
                        "id": ft.linked_trim.id,
                        "filename": ft.linked_trim.filename,
                        "model": ft.linked_trim.model,
                        "serial": ft.linked_trim.serial,
                        "file_date": ft.linked_trim.file_date,
                        "overall_status": ft.linked_trim.overall_status.value if ft.linked_trim.overall_status else "UNKNOWN",
                    }

                pairs.append({
                    "final_test_id": ft.id,
                    "final_test_filename": ft.filename,
                    "model": ft.model,
                    "serial": ft.serial,
                    "final_test_date": ft.file_date or ft.test_date,
                    "final_test_status": ft.overall_status.value if ft.overall_status else "UNKNOWN",
                    "linearity_pass": ft.linearity_pass,
                    "linked_trim_id": ft.linked_trim_id,
                    "linked_trim": linked_trim,
                    "match_confidence": ft.match_confidence,
                    "days_since_trim": ft.days_since_trim,
                    "match_method": getattr(ft, 'match_method', None),
                    "is_linked": ft.linked_trim_id is not None,
                })

            return pairs

    def get_comparison_pairs(
        self,
        model: Optional[str] = None,
        days_back: int = 90,
        linked_only: bool = False
    ) -> List[Dict[str, Any]]:
        """
        Get Final Test + Trim comparison pairs.

        Args:
            model: Filter by model (None = all models)
            days_back: How far back to look
            linked_only: If True, only return pairs with linked trim

        Returns:
            List of comparison pair dicts
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
        )

        with self.session() as session:
            cutoff_date = datetime.now() - timedelta(days=days_back)

            # Filter by test date (file_date) not processing date (timestamp)
            query = session.query(DBFinalTestResult).filter(
                DBFinalTestResult.file_date >= cutoff_date
            )

            if model:
                query = query.filter(DBFinalTestResult.model == model)

            if linked_only:
                query = query.filter(DBFinalTestResult.linked_trim_id.isnot(None))

            # Sort by test date (file_date)
            results = query.order_by(desc(DBFinalTestResult.file_date)).limit(500).all()

            # Pre-fetch all linked trims in a single query (avoids N+1)
            linked_trim_ids = [ft.linked_trim_id for ft in results if ft.linked_trim_id]
            linked_trims = {}
            if linked_trim_ids:
                trim_results = session.query(DBAnalysisResult).filter(
                    DBAnalysisResult.id.in_(linked_trim_ids)
                ).all()
                for trim in trim_results:
                    linked_trims[trim.id] = {
                        "id": trim.id,
                        "filename": trim.filename,
                        "model": trim.model,
                        "serial": trim.serial,
                        "file_date": trim.file_date,
                        "overall_status": trim.overall_status.value if trim.overall_status else "UNKNOWN",
                    }

            pairs = []
            for ft in results:
                # Get linked trim info from pre-fetched dict
                linked_trim = linked_trims.get(ft.linked_trim_id) if ft.linked_trim_id else None

                pairs.append({
                    "final_test_id": ft.id,
                    "final_test_filename": ft.filename,
                    "model": ft.model,
                    "serial": ft.serial,
                    "final_test_date": ft.file_date or ft.test_date,
                    "final_test_status": ft.overall_status.value if ft.overall_status else "UNKNOWN",
                    "linearity_pass": ft.linearity_pass,
                    "linked_trim_id": ft.linked_trim_id,
                    "linked_trim": linked_trim,
                    "match_confidence": ft.match_confidence,
                    "days_since_trim": ft.days_since_trim,
                    "match_method": getattr(ft, 'match_method', None),
                    "is_linked": ft.linked_trim_id is not None,
                })

            return pairs

    def get_comparison_data(
        self,
        final_test_id: int
    ) -> Optional[Dict[str, Any]]:
        """
        Get full comparison data for overlay chart.

        Returns both Final Test and linked Trim data with track details.
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            FinalTestTrack as DBFinalTestTrack,
        )

        with self.session() as session:
            # Get final test with tracks in single query (avoids N+1)
            ft = session.query(DBFinalTestResult).options(
                joinedload(DBFinalTestResult.tracks)
            ).filter(
                DBFinalTestResult.id == final_test_id
            ).first()

            if not ft:
                return None

            final_test_data = {
                "id": ft.id,
                "filename": ft.filename,
                "model": ft.model,
                "serial": ft.serial,
                "test_date": ft.file_date or ft.test_date,
                "status": ft.overall_status.value if ft.overall_status else "UNKNOWN",
                "linearity_pass": ft.linearity_pass,
                "linearity_error": ft.linearity_error,
                "tracks": [
                    {
                        "track_id": t.track_id,
                        # Use position_data if available, fall back to electrical_angle_data
                        "positions": t.position_data or t.electrical_angle_data or [],
                        "errors": t.error_data or [],
                        "electrical_angles": t.electrical_angle_data or [],
                        "upper_limits": t.upper_limits or [],
                        "lower_limits": t.lower_limits or [],
                        "linearity_error": t.linearity_error,
                        # Spec-aware correction fields (persisted by analyzer)
                        # so the compare chart can draw a corrected overlay.
                        "optimal_offset": t.optimal_offset if t.optimal_offset is not None else 0.0,
                        "optimal_slope": t.optimal_slope if t.optimal_slope is not None else 0.0,
                        "theory_data": t.theory_data,
                        "linearity_type": t.linearity_type,
                    }
                    for t in ft.tracks  # Use pre-loaded tracks
                ]
            }

            # Get linked trim data if exists (with tracks in single query)
            trim_data = None
            if ft.linked_trim_id:
                trim = session.query(DBAnalysisResult).options(
                    joinedload(DBAnalysisResult.tracks)
                ).filter(
                    DBAnalysisResult.id == ft.linked_trim_id
                ).first()

                if trim:
                    trim_data = {
                        "id": trim.id,
                        "filename": trim.filename,
                        "model": trim.model,
                        "serial": trim.serial,
                        "file_date": trim.file_date,
                        "status": trim.overall_status.value if trim.overall_status else "UNKNOWN",
                        "tracks": [
                            {
                                "track_id": t.track_id,
                                "positions": t.position_data or [],
                                "errors": t.error_data or [],
                                "upper_limits": t.upper_limits or [],
                                "lower_limits": t.lower_limits or [],
                                "optimal_offset": t.optimal_offset or 0,
                                "optimal_slope": t.optimal_slope if t.optimal_slope is not None else 0.0,
                                "theory_data": t.theory_data,
                                "linearity_error": t.final_linearity_error_shifted,
                                "linearity_pass": t.linearity_pass,
                                "sigma_gradient": t.sigma_gradient,
                                "sigma_pass": t.sigma_pass,
                            }
                            for t in trim.tracks  # Use pre-loaded tracks
                        ]
                    }

            return {
                "final_test": final_test_data,
                "trim": trim_data,
                "match_confidence": ft.match_confidence,
                "days_since_trim": ft.days_since_trim,
                "match_method": getattr(ft, 'match_method', None),
            }

    def get_final_test_models_list(self) -> List[str]:
        """
        Get list of unique models from Final Test results, sorted numerically.
        """
        from laser_trim_analyzer.database.models import FinalTestResult as DBFinalTestResult

        with self.session() as session:
            models = (
                session.query(DBFinalTestResult.model)
                .filter(DBFinalTestResult.model.isnot(None))
                .distinct()
                .all()
            )
            model_list = [m[0] for m in models if m[0]]
            return sorted(model_list, key=_model_sort_key)

    def get_final_tests_missing_tracks(self) -> List[Dict[str, Any]]:
        """
        Get Final Test records that have 0 tracks stored.

        Returns:
            List of dicts with id, filename, file_path, model
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            FinalTestTrack as DBFinalTestTrack,
        )

        with self.session() as session:
            # Subquery to count tracks per final test
            track_count_subq = (
                session.query(
                    DBFinalTestTrack.final_test_id,
                    func.count(DBFinalTestTrack.id).label('track_count')
                )
                .group_by(DBFinalTestTrack.final_test_id)
                .subquery()
            )

            # Get Final Tests with no tracks (LEFT JOIN where track_count is NULL)
            results = (
                session.query(DBFinalTestResult)
                .outerjoin(track_count_subq, DBFinalTestResult.id == track_count_subq.c.final_test_id)
                .filter(track_count_subq.c.track_count == None)
                .all()
            )
            return [
                {
                    "id": r.id,
                    "filename": r.filename,
                    "file_path": r.file_path,
                    "model": r.model,
                    "serial": r.serial,
                }
                for r in results
            ]

    def update_final_test_tracks(
        self,
        final_test_id: int,
        tracks: List[Dict[str, Any]]
    ) -> bool:
        """
        Update track data for an existing Final Test record.

        Used to fix records that were created before parser improvements.

        Args:
            final_test_id: ID of the Final Test record
            tracks: List of track data dicts

        Returns:
            True if successful
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            FinalTestTrack as DBFinalTestTrack,
        )

        with self._write_lock:
            try:
                with self.session() as session:
                    # Get existing record
                    result = session.get(DBFinalTestResult, final_test_id)
                    if not result:
                        logger.warning(f"Final Test ID {final_test_id} not found")
                        return False

                    # Delete existing tracks (if any)
                    session.query(DBFinalTestTrack).filter(
                        DBFinalTestTrack.final_test_id == final_test_id
                    ).delete()

                    # Add new tracks
                    for track_data in tracks:
                        position_values = track_data.get("electrical_angles") or track_data.get("positions")

                        db_track = DBFinalTestTrack(
                            final_test_id=final_test_id,
                            track_id=track_data.get("track_id", "default"),
                            status=DBStatusType.PASS if track_data.get("linearity_pass", True) else DBStatusType.FAIL,
                            linearity_spec=track_data.get("linearity_spec"),
                            linearity_error=track_data.get("linearity_error"),
                            linearity_pass=track_data.get("linearity_pass"),
                            linearity_fail_points=track_data.get("linearity_fail_points", 0),
                            position_data=position_values,
                            error_data=track_data.get("errors"),
                            theory_data=track_data.get("theory_values"),
                            electrical_angle_data=track_data.get("electrical_angles"),
                            upper_limits=track_data.get("upper_limits"),
                            lower_limits=track_data.get("lower_limits"),
                            max_deviation=track_data.get("max_deviation"),
                            max_deviation_position=track_data.get("max_deviation_angle"),
                            optimal_offset=track_data.get("optimal_offset"),
                            optimal_slope=track_data.get("optimal_slope"),
                            linearity_type=track_data.get("linearity_type"),
                        )
                        session.add(db_track)

                    # Update linearity_error on main record if tracks have it
                    if tracks and tracks[0].get("linearity_error") is not None:
                        result.linearity_error = tracks[0].get("linearity_error")

                    session.commit()
                    logger.info(f"Updated Final Test {final_test_id} with {len(tracks)} tracks")
                    return True

            except Exception as e:
                logger.error(f"Error updating Final Test tracks: {e}")
                return False

    def get_trim_records_missing_tracks(self, linked_only: bool = True) -> List[Dict[str, Any]]:
        """
        Get Trim (AnalysisResult) records that have no track data stored.

        Args:
            linked_only: If True, only return records that are linked to Final Tests

        Returns:
            List of record info dicts with id, filename, file_path, model, serial
        """
        with self.session() as session:
            # Subquery to count tracks per analysis
            track_count_subq = (
                session.query(
                    DBTrackResult.analysis_id,
                    func.count(DBTrackResult.id).label('track_count')
                )
                .group_by(DBTrackResult.analysis_id)
                .subquery()
            )

            # Base query for analyses with no tracks
            query = (
                session.query(DBAnalysisResult)
                .outerjoin(track_count_subq, DBAnalysisResult.id == track_count_subq.c.analysis_id)
                .filter(
                    (track_count_subq.c.track_count == None) |
                    (track_count_subq.c.track_count == 0)
                )
            )

            if linked_only:
                # Get IDs of analyses that are linked to Final Tests
                from laser_trim_analyzer.database.models import FinalTestResult as DBFinalTestResult
                linked_ids = (
                    session.query(DBFinalTestResult.linked_trim_id)
                    .filter(DBFinalTestResult.linked_trim_id != None)
                    .distinct()
                    .all()
                )
                linked_id_list = [lid[0] for lid in linked_ids]
                query = query.filter(DBAnalysisResult.id.in_(linked_id_list))

            results = query.all()

            return [
                {
                    "id": r.id,
                    "filename": r.filename,
                    "file_path": r.file_path,
                    "model": r.model,
                    "serial": r.serial,
                }
                for r in results
            ]

    def update_trim_tracks(
        self,
        analysis_id: int,
        tracks: List["TrackResult"]
    ) -> bool:
        """
        Update track data for an existing Trim (AnalysisResult) record.

        Used to fix records that were created before track data storage was added.

        Args:
            analysis_id: ID of the AnalysisResult record
            tracks: List of TrackResult objects from re-parsing

        Returns:
            True if successful
        """
        try:
            with self.session() as session:
                # Get existing record
                result = session.get(DBAnalysisResult, analysis_id)
                if not result:
                    logger.warning(f"Analysis ID {analysis_id} not found")
                    return False

                # Delete existing tracks (if any)
                session.query(DBTrackResult).filter(
                    DBTrackResult.analysis_id == analysis_id
                ).delete()

                # Add new tracks
                for track in tracks:
                    db_track = self._map_track_to_db(track)
                    db_track.analysis_id = analysis_id
                    session.add(db_track)

                session.commit()
                logger.info(f"Updated Analysis {analysis_id} with {len(tracks)} tracks")
                return True

        except Exception as e:
            logger.error(f"Error updating Trim tracks: {e}")
            return False

    def update_trim_tracks_from_final_test(
        self,
        analysis_id: int,
        ft_tracks: List[Dict[str, Any]]
    ) -> bool:
        """
        Update Trim (AnalysisResult) track data from Final Test format data.

        Used when "Trim" records actually point to Final Test files.
        Converts FT track format to TrackResult format.

        Args:
            analysis_id: ID of the AnalysisResult record
            ft_tracks: List of track dicts from Final Test parser

        Returns:
            True if successful
        """
        try:
            with self.session() as session:
                # Get existing record
                result = session.get(DBAnalysisResult, analysis_id)
                if not result:
                    logger.warning(f"Analysis ID {analysis_id} not found")
                    return False

                # Delete existing tracks (if any)
                session.query(DBTrackResult).filter(
                    DBTrackResult.analysis_id == analysis_id
                ).delete()

                # Add new tracks converted from FT format
                for ft_track in ft_tracks:
                    # Get position data (FT format uses electrical_angles)
                    positions = ft_track.get("electrical_angles") or ft_track.get("positions", [])
                    errors = ft_track.get("errors", [])
                    upper_limits = ft_track.get("upper_limits", [])
                    lower_limits = ft_track.get("lower_limits", [])

                    # Calculate linearity metrics
                    linearity_error = ft_track.get("linearity_error", 0.0)
                    linearity_spec = ft_track.get("linearity_spec", 0.02)
                    linearity_pass = ft_track.get("linearity_pass", True)

                    # Create TrackResult-compatible DB record
                    db_track = DBTrackResult(
                        analysis_id=analysis_id,
                        track_id=ft_track.get("track_id", "default"),
                        status=DBStatusType.PASS if linearity_pass else DBStatusType.FAIL,
                        # Sigma values - use defaults for FT data
                        sigma_gradient=0.0,
                        sigma_threshold=1.0,
                        sigma_pass=True,
                        # Linearity values
                        linearity_spec=linearity_spec,
                        final_linearity_error_shifted=linearity_error,
                        linearity_pass=linearity_pass,
                        linearity_fail_points=ft_track.get("linearity_fail_points", 0),
                        # Track data for charts
                        position_data=positions,
                        error_data=errors,
                        upper_limits=upper_limits,
                        lower_limits=lower_limits,
                        # Travel length from position range
                        travel_length=max(positions) - min(positions) if positions and len(positions) > 1 else 1.0,
                    )
                    session.add(db_track)

                session.commit()
                logger.info(f"Updated Analysis {analysis_id} with {len(ft_tracks)} tracks from FT format")
                return True

        except Exception as e:
            logger.error(f"Error updating Trim tracks from FT: {e}")
            return False

    # =========================================================================
    # Database Health & Cleanup
    # =========================================================================

    def scan_database_health(self) -> Dict[str, Any]:
        """
        Scan the entire database and return a health report.

        Identifies dirty/suspect records across multiple categories without
        modifying anything. Returns counts and record IDs for each issue.
        """
        health = {
            "total_analyses": 0,
            "total_tracks": 0,
            "issues": {},
            "total_dirty_records": 0,
        }

        with self.session() as session:
            health["total_analyses"] = (
                session.query(func.count(DBAnalysisResult.id)).scalar() or 0
            )
            health["total_tracks"] = (
                session.query(func.count(DBTrackResult.id)).scalar() or 0
            )

            dirty_ids = set()

            # 1. Unknown model
            unknown_model = session.query(DBAnalysisResult.id).filter(
                DBAnalysisResult.model == "Unknown"
            ).all()
            if unknown_model:
                ids = {r[0] for r in unknown_model}
                dirty_ids |= ids
                health["issues"]["unknown_model"] = {
                    "count": len(ids),
                    "label": "Unknown model (parser couldn't extract)",
                }

            # 2. Unknown serial
            unknown_serial = session.query(DBAnalysisResult.id).filter(
                DBAnalysisResult.serial == "Unknown"
            ).all()
            if unknown_serial:
                ids = {r[0] for r in unknown_serial}
                dirty_ids |= ids
                health["issues"]["unknown_serial"] = {
                    "count": len(ids),
                    "label": "Unknown serial number",
                }

            # 3. Missing file date
            null_date = session.query(DBAnalysisResult.id).filter(
                DBAnalysisResult.file_date.is_(None)
            ).all()
            if null_date:
                ids = {r[0] for r in null_date}
                dirty_ids |= ids
                health["issues"]["missing_file_date"] = {
                    "count": len(ids),
                    "label": "Missing file date",
                }

            # 4. ERROR status records
            error_records = session.query(DBAnalysisResult.id).filter(
                DBAnalysisResult.overall_status == DBStatusType.ERROR
            ).all()
            if error_records:
                ids = {r[0] for r in error_records}
                dirty_ids |= ids
                health["issues"]["error_status"] = {
                    "count": len(ids),
                    "label": "ERROR status (processing failed)",
                }

            # 5. Analyses with no tracks (orphaned)
            analyses_no_tracks = session.query(DBAnalysisResult.id).filter(
                ~exists().where(DBTrackResult.analysis_id == DBAnalysisResult.id)
            ).all()
            if analyses_no_tracks:
                ids = {r[0] for r in analyses_no_tracks}
                dirty_ids |= ids
                health["issues"]["no_tracks"] = {
                    "count": len(ids),
                    "label": "No track data (empty analyses)",
                }

            # 6. Track-level quality issues (negative sigma, all-zero data, etc.)
            bad_sigma = session.query(
                DBTrackResult.analysis_id
            ).filter(
                DBTrackResult.sigma_gradient < 0
            ).distinct().all()
            if bad_sigma:
                ids = {r[0] for r in bad_sigma}
                dirty_ids |= ids
                health["issues"]["negative_sigma"] = {
                    "count": len(ids),
                    "label": "Negative sigma gradient (impossible value)",
                }

            # 7. Tracks with no spec limits (can't determine pass/fail)
            no_limits = session.query(
                DBTrackResult.analysis_id
            ).filter(
                DBTrackResult.upper_limits.is_(None),
                DBTrackResult.lower_limits.is_(None),
                DBTrackResult.linearity_spec.is_(None),
            ).distinct().all()
            if no_limits:
                ids = {r[0] for r in no_limits}
                dirty_ids |= ids
                health["issues"]["no_spec_limits"] = {
                    "count": len(ids),
                    "label": "No spec limits (can't verify pass/fail)",
                }

            # 8. Already-flagged suspect quality
            suspect = session.query(DBAnalysisResult.id).filter(
                DBAnalysisResult.data_quality == "suspect"
            ).all()
            if suspect:
                ids = {r[0] for r in suspect}
                dirty_ids |= ids
                health["issues"]["suspect_quality"] = {
                    "count": len(ids),
                    "label": "Previously flagged as suspect",
                }

            health["total_dirty_records"] = len(dirty_ids)

        return health

    def retroactive_validate(self) -> Dict[str, Any]:
        """
        Retroactively validate ALL records in the database and update
        data_quality flags.

        Checks analysis-level and track-level quality issues, then
        updates the data_quality and data_quality_issues columns.

        Uses raw SQL updates to avoid SQLAlchemy dirty-tracking issues
        with JSON (list) columns that are unhashable.

        Returns summary of what was found and updated.
        """
        from sqlalchemy import update as sa_update

        summary = {"scanned": 0, "flagged": 0, "already_suspect": 0, "issues_by_type": {}}
        batch_size = 1000

        with self._write_lock:
            with self.session() as session:
                total = session.query(func.count(DBAnalysisResult.id)).scalar() or 0
                summary["scanned"] = total

                # Process in batches to avoid memory issues with large databases
                for offset in range(0, total, batch_size):
                    # Use read-only loading — we'll update via raw SQL to avoid
                    # SQLAlchemy dirty-tracking on JSON (list) columns
                    analyses = session.query(
                        DBAnalysisResult.id,
                        DBAnalysisResult.model,
                        DBAnalysisResult.serial,
                        DBAnalysisResult.file_date,
                        DBAnalysisResult.data_quality,
                    ).order_by(DBAnalysisResult.id).offset(offset).limit(batch_size).all()

                    for a_id, a_model, a_serial, a_file_date, a_dq in analyses:
                        issues = []

                        # Analysis-level checks
                        if a_model == "Unknown":
                            issues.append("Unknown model")
                        if a_serial == "Unknown":
                            issues.append("Unknown serial")
                        if a_file_date is None:
                            issues.append("Missing file date")

                        # Track-level checks — query track columns directly
                        tracks = session.query(
                            DBTrackResult.track_id,
                            DBTrackResult.sigma_gradient,
                            DBTrackResult.linearity_spec,
                            DBTrackResult.upper_limits,
                            DBTrackResult.lower_limits,
                            DBTrackResult.position_data,
                            DBTrackResult.error_data,
                        ).filter(
                            DBTrackResult.analysis_id == a_id
                        ).all()

                        if not tracks:
                            issues.append("No track data")

                        for t_id, t_sigma, t_lin_spec, t_upper, t_lower, t_pos, t_err in tracks:
                            tid = t_id or "?"

                            if t_sigma is not None and t_sigma < 0:
                                issues.append(f"{tid}: negative sigma_gradient ({t_sigma:.4f})")

                            if not t_upper and not t_lower and t_lin_spec is None:
                                issues.append(f"{tid}: no spec limits")

                            if t_err:
                                try:
                                    if all(v == 0 or v is None for v in t_err):
                                        issues.append(f"{tid}: all-zero error data")
                                except (TypeError, ValueError):
                                    issues.append(f"{tid}: corrupt error data")

                            if t_pos:
                                try:
                                    if len(t_pos) < 10:
                                        issues.append(f"{tid}: too few data points ({len(t_pos)})")
                                except TypeError:
                                    issues.append(f"{tid}: corrupt position data")

                            if t_pos and t_err:
                                try:
                                    if len(t_pos) != len(t_err):
                                        issues.append(
                                            f"{tid}: array mismatch (pos={len(t_pos)}, err={len(t_err)})"
                                        )
                                except TypeError:
                                    pass

                        # Update via raw SQL to avoid unhashable-list errors from JSON columns
                        if issues:
                            was_suspect = a_dq == "suspect"
                            session.execute(
                                sa_update(DBAnalysisResult)
                                .where(DBAnalysisResult.id == a_id)
                                .values(
                                    data_quality="suspect",
                                    data_quality_issues=", ".join(issues),
                                )
                            )
                            if was_suspect:
                                summary["already_suspect"] += 1
                            else:
                                summary["flagged"] += 1

                            for issue in issues:
                                category = issue.split(":")[0].strip() if ":" in issue else issue
                                summary["issues_by_type"][category] = summary["issues_by_type"].get(category, 0) + 1
                        else:
                            if a_dq == "suspect":
                                session.execute(
                                    sa_update(DBAnalysisResult)
                                    .where(DBAnalysisResult.id == a_id)
                                    .values(
                                        data_quality="good",
                                        data_quality_issues=None,
                                    )
                                )

                    session.flush()

                logger.info(
                    f"Retroactive validation: scanned {summary['scanned']}, "
                    f"flagged {summary['flagged']} new, "
                    f"{summary['already_suspect']} already suspect"
                )

        return summary

    def _collect_cleanup_ids(
        self,
        session,
        delete_non_mps: bool = False,
        mps_models: Optional[List[str]] = None,
        delete_before_date: Optional[datetime] = None,
        delete_suspect_quality: bool = False,
        delete_unknown: bool = False,
        delete_error_status: bool = False,
        delete_no_tracks: bool = False,
        delete_misclassified_ft: bool = False,
    ) -> tuple:
        """
        Collect record IDs matching cleanup criteria. Shared by preview and execute.

        Returns:
            (ids_to_delete set, by_reason dict)
        """
        ids_to_delete = set()
        by_reason = {}

        if delete_non_mps and mps_models:
            mps_set = set(m.strip() for m in mps_models if m.strip())
            non_mps = session.query(
                DBAnalysisResult.id, DBAnalysisResult.model
            ).filter(
                DBAnalysisResult.model.notin_(mps_set)
            ).all()
            non_mps_ids = {r[0] for r in non_mps}
            non_mps_models = sorted(set(r[1] for r in non_mps))
            ids_to_delete |= non_mps_ids
            by_reason["non_mps_models"] = {
                "count": len(non_mps_ids),
                "models": non_mps_models,
            }

        if delete_before_date:
            old_records = session.query(
                DBAnalysisResult.id
            ).filter(
                DBAnalysisResult.file_date < delete_before_date
            ).all()
            old_ids = {r[0] for r in old_records}
            ids_to_delete |= old_ids
            by_reason["before_date"] = {
                "count": len(old_ids),
                "date": delete_before_date.strftime("%Y-%m-%d"),
            }

        if delete_suspect_quality:
            suspect = session.query(
                DBAnalysisResult.id
            ).filter(
                DBAnalysisResult.data_quality == "suspect"
            ).all()
            suspect_ids = {r[0] for r in suspect}
            ids_to_delete |= suspect_ids
            by_reason["suspect_quality"] = {
                "count": len(suspect_ids),
            }

        if delete_unknown:
            unknown = session.query(
                DBAnalysisResult.id
            ).filter(
                or_(
                    DBAnalysisResult.model == "Unknown",
                    DBAnalysisResult.serial == "Unknown",
                )
            ).all()
            unknown_ids = {r[0] for r in unknown}
            ids_to_delete |= unknown_ids
            by_reason["unknown_model_serial"] = {
                "count": len(unknown_ids),
            }

        if delete_error_status:
            errors = session.query(
                DBAnalysisResult.id
            ).filter(
                DBAnalysisResult.overall_status == DBStatusType.ERROR
            ).all()
            error_ids = {r[0] for r in errors}
            ids_to_delete |= error_ids
            by_reason["error_status"] = {
                "count": len(error_ids),
            }

        if delete_no_tracks:
            no_tracks = session.query(
                DBAnalysisResult.id
            ).filter(
                ~exists().where(DBTrackResult.analysis_id == DBAnalysisResult.id)
            ).all()
            no_track_ids = {r[0] for r in no_tracks}
            ids_to_delete |= no_track_ids
            by_reason["no_tracks"] = {
                "count": len(no_track_ids),
            }

        if delete_misclassified_ft:
            # Find trim records that are actually Final Test files:
            # 1. Files from "Test Station" paths
            # 2. Files with _Redundant_ or _Primary_ in filename
            # 3. Files with "final" followed by a number in filename
            ft_patterns = [
                DBAnalysisResult.filename.like("%Test Station%"),
                DBAnalysisResult.filename.like("%test station%"),
                DBAnalysisResult.filename.like("%_Redundant_%"),
                DBAnalysisResult.filename.like("%_redundant_%"),
                DBAnalysisResult.filename.like("%_Primary_%"),
                DBAnalysisResult.filename.like("%_primary_%"),
            ]
            misclassified = session.query(
                DBAnalysisResult.id
            ).filter(
                or_(*ft_patterns)
            ).all()
            misc_ids = {r[0] for r in misclassified}

            # Also find "model final NNN" pattern files in trim table
            final_pattern = session.query(
                DBAnalysisResult.id
            ).filter(
                DBAnalysisResult.filename.like("% final %")
            ).all()
            final_ids = {r[0] for r in final_pattern}
            misc_ids |= final_ids

            ids_to_delete |= misc_ids
            by_reason["misclassified_ft"] = {
                "count": len(misc_ids),
            }

        return ids_to_delete, by_reason

    def preview_cleanup(
        self,
        delete_non_mps: bool = False,
        mps_models: Optional[List[str]] = None,
        delete_before_date: Optional[datetime] = None,
        delete_suspect_quality: bool = False,
        delete_unknown: bool = False,
        delete_error_status: bool = False,
        delete_no_tracks: bool = False,
        delete_misclassified_ft: bool = False,
    ) -> Dict[str, Any]:
        """
        Preview what a cleanup operation would delete WITHOUT actually deleting.

        Returns:
            Dict with counts and model lists for what would be deleted
        """
        preview = {
            "total_records": 0,
            "records_to_delete": 0,
            "models_to_delete": [],
            "by_reason": {},
        }

        with self.session() as session:
            preview["total_records"] = (
                session.query(func.count(DBAnalysisResult.id)).scalar() or 0
            )

            ids_to_delete, by_reason = self._collect_cleanup_ids(
                session,
                delete_non_mps=delete_non_mps,
                mps_models=mps_models,
                delete_before_date=delete_before_date,
                delete_suspect_quality=delete_suspect_quality,
                delete_unknown=delete_unknown,
                delete_error_status=delete_error_status,
                delete_no_tracks=delete_no_tracks,
                delete_misclassified_ft=delete_misclassified_ft,
            )

            preview["by_reason"] = by_reason
            preview["records_to_delete"] = len(ids_to_delete)

            if ids_to_delete:
                models = session.query(
                    DBAnalysisResult.model
                ).filter(
                    DBAnalysisResult.id.in_(ids_to_delete)
                ).distinct().all()
                preview["models_to_delete"] = sorted(m[0] for m in models)

        return preview

    def execute_cleanup(
        self,
        delete_non_mps: bool = False,
        mps_models: Optional[List[str]] = None,
        delete_before_date: Optional[datetime] = None,
        delete_suspect_quality: bool = False,
        delete_unknown: bool = False,
        delete_error_status: bool = False,
        delete_no_tracks: bool = False,
        delete_misclassified_ft: bool = False,
    ) -> Dict[str, int]:
        """
        Execute database cleanup — permanently delete matching records.

        Uses the same filters as preview_cleanup(). Deletes analysis records
        and associated tracks and alerts. Keeps processed_files records so
        the same bad files won't be reprocessed next time (the FK has
        ondelete=SET NULL so the link is safely cleared).

        Returns:
            Dict with deletion counts
        """
        deleted = {"analyses": 0, "tracks": 0, "alerts": 0}

        with self._write_lock:
            with self.session() as session:
                ids_to_delete, _ = self._collect_cleanup_ids(
                    session,
                    delete_non_mps=delete_non_mps,
                    mps_models=mps_models,
                    delete_before_date=delete_before_date,
                    delete_suspect_quality=delete_suspect_quality,
                    delete_unknown=delete_unknown,
                    delete_error_status=delete_error_status,
                    delete_no_tracks=delete_no_tracks,
                    delete_misclassified_ft=delete_misclassified_ft,
                )

                if not ids_to_delete:
                    return deleted

                # Delete in batches to avoid SQLite variable limits
                id_list = list(ids_to_delete)
                batch_size = 500

                for i in range(0, len(id_list), batch_size):
                    batch = id_list[i:i + batch_size]

                    deleted["tracks"] += session.query(DBTrackResult).filter(
                        DBTrackResult.analysis_id.in_(batch)
                    ).delete(synchronize_session=False)

                    deleted["alerts"] += session.query(DBQAAlert).filter(
                        DBQAAlert.analysis_id.in_(batch)
                    ).delete(synchronize_session=False)

                    # Keep processed_files records — prevents reprocessing
                    # the same bad files. FK ondelete=SET NULL clears the link.

                    deleted["analyses"] += session.query(DBAnalysisResult).filter(
                        DBAnalysisResult.id.in_(batch)
                    ).delete(synchronize_session=False)

                logger.info(
                    f"Database cleanup: deleted {deleted['analyses']} analyses, "
                    f"{deleted['tracks']} tracks, {deleted['alerts']} alerts "
                    f"(processed_files kept to prevent reprocessing)"
                )

        return deleted

    def count_skipped_files(self) -> int:
        """Count non-trim/non-FT files that were skipped and recorded."""
        with self.session() as session:
            return session.query(func.count(DBProcessedFile.id)).filter(
                DBProcessedFile.analysis_id.is_(None),
                DBProcessedFile.success == True,
            ).scalar() or 0

    def reset_skipped_files(self) -> int:
        """
        Remove processed_files entries for skipped non-trim files so they
        get re-evaluated on the next processing run.

        Only clears entries with analysis_id=NULL (no analysis was created),
        which are files that were detected as non-trim and skipped.

        Returns:
            Number of entries cleared
        """
        with self._write_lock:
            with self.session() as session:
                count = session.query(DBProcessedFile).filter(
                    DBProcessedFile.analysis_id.is_(None),
                    DBProcessedFile.success == True,
                ).delete(synchronize_session=False)

                logger.info(f"Reset {count} skipped file entries for reprocessing")

        return count

    def mark_file_skipped(self, filename: str, file_path: str,
                          file_hash: str, file_size: int,
                          file_modified_date) -> None:
        """Record a non-trim file so it's skipped on future processing runs."""
        with self._write_lock:
            with self.session() as session:
                existing = session.query(DBProcessedFile).filter(
                    DBProcessedFile.file_hash == file_hash
                ).first()
                if existing:
                    return

                session.add(DBProcessedFile(
                    filename=filename,
                    file_path=file_path,
                    file_hash=file_hash,
                    file_size=file_size,
                    file_modified_date=file_modified_date,
                    analysis_id=None,
                    success=True,
                ))

    def update_processed_file_stats(self, entries) -> Dict[str, int]:
        """Repair size/mtime on processed rows after a hash-confirm.

        entries: iterable of (file_hash, file_size, file_modified_date).
        Lets the incremental scan's stat fast-path work on the next run for
        rows whose recorded stat was missing or stale. Content identity is
        unchanged — only rows matched by their content hash are updated.

        Covers final_test_results and smoothness_results as well as
        processed_files (2026-08-29): FT rows never carried a stat, so the
        heal never reached them and every scan re-hashed the whole FT share.
        The first scan after this ships hash-confirms once, stamps the rows,
        and every scan after that is pure in-memory.

        Returns rows updated PER TABLE plus "total". Per-table counts, not one
        number, because the old single count hid the bug for six weeks: the
        scan logged "Repaired stat records for 150,938 processed files" while
        the UPDATE matched ~0 rows (FT files have no processed_files row).
        The caller logs queued-vs-updated so a silent no-op can't hide again.
        """
        from laser_trim_analyzer.database.models import (
            FinalTestResult as DBFinalTestResult,
            SmoothnessResult as DBSmoothnessResult,
        )

        tables = (("processed_files", DBProcessedFile),
                  ("final_test_results", DBFinalTestResult),
                  ("smoothness_results", DBSmoothnessResult))
        counts: Dict[str, int] = {name: 0 for name, _ in tables}
        with self._write_lock:
            with self.session() as session:
                for file_hash, file_size, file_modified_date in entries:
                    for name, model in tables:
                        counts[name] += session.query(model).filter(
                            model.file_hash == file_hash
                        ).update({
                            model.file_size: file_size,
                            model.file_modified_date: file_modified_date,
                        })
        counts["total"] = sum(counts[name] for name, _ in tables)
        return counts

    def recompute_overall_statuses(self, dry_run: bool = True,
                                   batch_size: int = 1000) -> Dict[str, Any]:
        """Re-grade every analysis' overall_status from its tracks' STORED
        pass flags, using the current (correct) rule. (2026-07-07, M4.)

        Why: ~42%% of historical rows are WARNING with three different
        historical meanings (old rule labeled linearity-FAILs as Warning;
        later gate changes were never backfilled). Linearity is a zero-
        tolerance customer requirement — a linearity-FAIL presenting as
        "Warning" is a misclassification, not a cosmetic quirk.

        Rule (mirrors analyzer._determine + processor rollup):
          track: FAIL if linearity_pass is False; else PASS if sigma_pass is
                 True; else WARNING. UNTRIMMED tracks excluded from judging.
          analysis: all-PASS -> PASS; any FAIL -> FAIL; else WARNING.

        Safety: analyses where any judged track has linearity_pass = NULL are
        SKIPPED (never regraded) — those are the un-evaluated/empty-array rows
        that Fix Missing Tracks must repair first. ERROR and UNTRIMMED
        analyses are untouched. dry_run=True only counts.

        Returns {"examined", "changed", "skipped_null_flags", "transitions":
        {"OLD->NEW": n}, "sample_changed_ids": [...]}.
        """
        from collections import defaultdict

        out: Dict[str, Any] = {"examined": 0, "changed": 0,
                               "skipped_null_flags": 0,
                               "transitions": defaultdict(int),
                               "sample_changed_ids": []}
        updates: List[tuple] = []  # (analysis_id, new_status)

        with self.session() as session:
            rows = (session.query(
                        DBAnalysisResult.id, DBAnalysisResult.overall_status,
                        DBTrackResult.status, DBTrackResult.linearity_pass,
                        DBTrackResult.sigma_pass)
                    .join(DBTrackResult,
                          DBTrackResult.analysis_id == DBAnalysisResult.id)
                    .filter(DBAnalysisResult.overall_status.notin_(
                        [DBStatusType.UNTRIMMED, DBStatusType.ERROR]))
                    .order_by(DBAnalysisResult.id)
                    .yield_per(5000))

            current: Dict[int, Any] = {}
            tracks_by_analysis: Dict[int, list] = {}
            for aid, overall, tstatus, lin, sig in rows:
                current[aid] = overall
                tracks_by_analysis.setdefault(aid, []).append((tstatus, lin, sig))

            for aid, tracks in tracks_by_analysis.items():
                out["examined"] += 1
                judged = [(lin, sig) for (tstatus, lin, sig) in tracks
                          if getattr(tstatus, "name", str(tstatus)) != "UNTRIMMED"]
                if not judged:
                    continue
                if any(lin is None for (lin, _sig) in judged):
                    out["skipped_null_flags"] += 1
                    continue
                track_statuses = [
                    DBStatusType.FAIL if lin is False
                    else (DBStatusType.PASS if sig is True else DBStatusType.WARNING)
                    for (lin, sig) in judged
                ]
                if all(s == DBStatusType.PASS for s in track_statuses):
                    new = DBStatusType.PASS
                elif any(s == DBStatusType.FAIL for s in track_statuses):
                    new = DBStatusType.FAIL
                else:
                    new = DBStatusType.WARNING

                old = current[aid]
                old_name = getattr(old, "name", str(old))
                if old_name != new.name:
                    out["changed"] += 1
                    out["transitions"][f"{old_name}->{new.name}"] += 1
                    if len(out["sample_changed_ids"]) < 10:
                        out["sample_changed_ids"].append(aid)
                    updates.append((aid, new))

        out["transitions"] = dict(out["transitions"])
        if dry_run or not updates:
            return out

        # Execute in batches; partial progress is preserved on error (same
        # philosophy as backfill_max_deviation).
        with self._write_lock:
            with self.session() as session:
                for i in range(0, len(updates), batch_size):
                    for aid, new in updates[i:i + batch_size]:
                        session.query(DBAnalysisResult).filter(
                            DBAnalysisResult.id == aid
                        ).update({DBAnalysisResult.overall_status: new},
                                 synchronize_session=False)
                    session.commit()
        logger.info(f"Status recompute: {out['changed']} of {out['examined']} "
                    f"regraded; {out['skipped_null_flags']} skipped (NULL flags); "
                    f"transitions={out['transitions']}")
        return out

    def backfill_max_deviation(self, batch_size: int = 1000) -> int:
        """
        Backfill max_deviation, max_deviation_position, and deviation_uniformity
        for existing tracks that have error_data but no max_deviation.

        Commits in batches so that partial progress is preserved if an error
        occurs mid-way.  This is intentional — a backfill that saves 900 of
        1000 rows is better than one that saves 0.

        Returns:
            Number of tracks updated (may be partial on error)
        """
        import json
        import statistics as stats_module

        updated = 0
        with self._write_lock:
            session = self._SessionFactory()
            try:
                # Get total count first
                total = session.execute(text(
                    "SELECT COUNT(*) FROM track_results "
                    "WHERE max_deviation IS NULL AND error_data IS NOT NULL"
                )).scalar()

                if total == 0:
                    logger.info("No tracks need max_deviation backfill")
                    return 0

                logger.info(f"Backfilling max_deviation for {total} tracks...")

                last_id = 0
                while True:
                    rows = session.execute(text(
                        "SELECT id, error_data, position_data, optimal_offset "
                        "FROM track_results "
                        "WHERE max_deviation IS NULL AND error_data IS NOT NULL "
                        "AND id > :last_id "
                        "ORDER BY id LIMIT :limit"
                    ), {"limit": batch_size, "last_id": last_id}).fetchall()

                    if not rows:
                        break
                    last_id = rows[-1].id

                    for row in rows:
                        try:
                            errors = json.loads(row.error_data) if isinstance(row.error_data, str) else row.error_data
                            positions = json.loads(row.position_data) if isinstance(row.position_data, str) else row.position_data
                            opt_offset = row.optimal_offset or 0.0

                            if not errors or not positions:
                                continue

                            shifted = [e + opt_offset for e in errors]
                            abs_errs = [abs(e) for e in shifted]
                            max_dev = max(abs_errs)
                            max_idx = abs_errs.index(max_dev)
                            max_dev_pos = positions[max_idx] if max_idx < len(positions) else None

                            dev_unif = None
                            if len(abs_errs) > 1:
                                mean_abs = stats_module.mean(abs_errs)
                                if mean_abs > 0:
                                    dev_unif = stats_module.stdev(abs_errs) / mean_abs

                            session.execute(text(
                                "UPDATE track_results SET "
                                "max_deviation = :max_dev, "
                                "max_deviation_position = :max_dev_pos, "
                                "deviation_uniformity = :dev_unif "
                                "WHERE id = :id"
                            ), {
                                "max_dev": max_dev,
                                "max_dev_pos": max_dev_pos,
                                "dev_unif": dev_unif,
                                "id": row.id
                            })
                            updated += 1
                        except Exception as e:
                            logger.warning(f"Failed to backfill track {row.id}: {e}")

                    session.commit()
                    logger.info(f"Backfilled {updated}/{total} tracks...")

            except Exception as e:
                session.rollback()
                logger.error(f"Backfill error after {updated} updates: {e}")
            finally:
                session.close()

        logger.info(f"Backfill complete: {updated} tracks updated")
        return updated

    # =========================================================================
    # Model Specifications
    # =========================================================================

    @staticmethod
    def _spec_to_dict(s: "ModelSpec") -> Dict[str, Any]:
        return {
            "id": s.id,
            "model": s.model,
            "element_type": s.element_type,
            "product_class": s.product_class,
            "linearity_type": s.linearity_type,
            "linearity_spec_text": s.linearity_spec_text,
            "linearity_spec_pct": s.linearity_spec_pct,
            "total_resistance_min": s.total_resistance_min,
            "total_resistance_max": s.total_resistance_max,
            "electrical_angle": s.electrical_angle,
            "electrical_angle_tol": s.electrical_angle_tol,
            "electrical_angle_tol_type": getattr(s, "electrical_angle_tol_type", None),
            "electrical_angle_unit": s.electrical_angle_unit,
            "output_smoothness": s.output_smoothness,
            "circuit_type": s.circuit_type,
            "open_closed": getattr(s, "open_closed", None) or s.circuit_type,
            "aliases": getattr(s, "aliases", None),
            "exclude_points": getattr(s, "exclude_points", None),
            "exclude_points_ft": getattr(s, "exclude_points_ft", None),
            "notes": s.notes,
        }

    @staticmethod
    def _parse_aliases(aliases_str: Optional[str]) -> List[str]:
        """Parse pipe-separated aliases into a trimmed list of non-empty tokens."""
        if not aliases_str:
            return []
        return [a.strip() for a in aliases_str.split("|") if a.strip()]

    def get_all_model_specs(self) -> List[Dict[str, Any]]:
        """Get all model specs as dicts."""
        with self.session() as session:
            specs = session.query(ModelSpec).order_by(ModelSpec.model).all()
            return [self._spec_to_dict(s) for s in specs]

    def get_model_spec(self, model: str) -> Optional[Dict[str, Any]]:
        """
        Get spec for a specific model. Checks both the primary `model` column
        and the pipe-separated `aliases` column, so `1621501` and `2001621501`
        can share a single spec row.
        """
        if not model:
            return None
        model = model.strip()
        with self.session() as session:
            # Primary match first
            spec = session.query(ModelSpec).filter(
                ModelSpec.model == model
            ).first()
            if spec:
                return self._spec_to_dict(spec)

            # Fallback: search aliases. SQLite's LIKE is case-insensitive by
            # default for ASCII; we wrap with the delimiter to avoid matching
            # prefixes/suffixes ('21501' should not match '1621501').
            like_pattern = f"%|{model}|%"
            # Also match at start/end without a leading/trailing pipe
            candidates = session.query(ModelSpec).filter(
                ModelSpec.aliases.isnot(None),
                ModelSpec.aliases != "",
            ).all()
            for c in candidates:
                if model in self._parse_aliases(c.aliases):
                    return self._spec_to_dict(c)
            return None

    def resolve_spec_for_ft(self, model: Optional[str], serial: Optional[str]) -> Optional[Dict[str, Any]]:
        """
        Resolve a model spec for a Final Test record.

        Multi-section parts (e.g. 8508) store their spec as per-section rows:
        8508-A, 8508-B, 8508-C, 8508-D. But FT files for the same product are
        labeled as model='8508' with the section baked into the serial by the
        operator (e.g. serial='31B' for section B, SN31). This helper tries
        the section-specific spec first, then falls back to the plain model.

        Resolution order:
          1. If serial ends in a letter AND get_model_spec(model-letter) exists,
             return that row.
          2. Otherwise return get_model_spec(model).
        """
        if not model:
            return None

        if serial:
            # Trailing letter on the serial — e.g., '31B', '1004a'.
            # Uppercase it so '31b' and '31B' both resolve to '-B'.
            import re as _re
            m = _re.match(r'^.*?([A-Za-z])\s*$', str(serial))
            if m:
                section_letter = m.group(1).upper()
                section_model = f"{model}-{section_letter}"
                section_spec = self.get_model_spec(section_model)
                if section_spec:
                    return section_spec

        # Fallback: plain model lookup (covers single-section parts).
        return self.get_model_spec(model)

    def save_model_spec(self, data: Dict[str, Any]) -> Tuple[int, bool]:
        """Create or update a model spec. Returns (spec_id, was_update)."""
        with self._write_lock:
            with self.session() as session:
                existing = session.query(ModelSpec).filter(
                    ModelSpec.model == data["model"]
                ).first()

                if existing:
                    for key, value in data.items():
                        if key not in ("id", "model", "created_at", "updated_at"):
                            setattr(existing, key, value)
                    # updated_at handled automatically by onupdate=utc_now
                    session.flush()
                    return existing.id, True
                else:
                    spec = ModelSpec(**{k: v for k, v in data.items() if k != "id"})
                    session.add(spec)
                    session.flush()
                    return spec.id, False

    def delete_model_spec(self, model: str) -> bool:
        """Delete a model spec. Returns True if found and deleted."""
        with self._write_lock:
            with self.session() as session:
                spec = session.query(ModelSpec).filter(
                    ModelSpec.model == model
                ).first()
                if spec:
                    session.delete(spec)
                    return True
                return False

    def get_distinct_element_types(self) -> List[str]:
        """Get all distinct element types from model_specs."""
        with self.session() as session:
            results = session.query(ModelSpec.element_type).filter(
                ModelSpec.element_type.isnot(None)
            ).distinct().order_by(ModelSpec.element_type).all()
            return [r[0] for r in results]

    def get_distinct_product_classes(self) -> List[str]:
        """Get all distinct product classes from model_specs."""
        with self.session() as session:
            results = session.query(ModelSpec.product_class).filter(
                ModelSpec.product_class.isnot(None)
            ).distinct().order_by(ModelSpec.product_class).all()
            return [r[0] for r in results]

    @staticmethod
    def _parse_angle_string(angle_text: Optional[str]) -> Tuple[Optional[float], Optional[float], Optional[str], Optional[str]]:
        """
        Parse a single angle-spec string into (value, tol, unit, tol_type).

        Handles many formats:
          '1.31" ± .005"'        symmetric tolerance
          '.665" +/-.005"'       symmetric tolerance
          '150° ± 1°'            symmetric tolerance
          '350° Min'             one-sided (floor; slope may go up)
          '340° Max'             one-sided (ceiling; slope may go down)
          '89° - 91°'            range (midpoint ± half-range)
          '2.812" - 2.832"'      range
          '±45°', '+/- 27.5°'    bilateral (±N from center)
          '120°', '1.25"'        nominal only, no tolerance
          'See ATP-10312-DS'     reference doc — returns all Nones
          'SEE CHARTS'           reference doc — returns all Nones

        Returns (angle_val, angle_tol, angle_unit, angle_tol_type).
        All None if the text is empty or a reference-doc string.
        """
        import re as _re

        angle_val = None
        angle_tol = None
        angle_unit = None
        angle_tol_type = None

        if not angle_text:
            return angle_val, angle_tol, angle_unit, angle_tol_type

        txt = angle_text.strip()
        if not txt:
            return angle_val, angle_tol, angle_unit, angle_tol_type

        txt_lower = txt.lower()

        # Reference-doc strings: store nothing (don't pull a part
        # number out of the string and call it an angle).
        if (txt_lower.startswith("see ") or
            "see chart" in txt_lower or
            "see table" in txt_lower or
            "see atp" in txt_lower):
            return None, None, None, None

        has_deg = '°' in txt or 'deg' in txt_lower
        has_inch = '"' in txt
        unit_guess = "deg" if has_deg else ("in" if has_inch else None)

        # Bilateral: starts with ± or +/- (e.g. '±45°', '+/- 27.5°')
        bi_match = _re.match(r'^\s*(?:[±]|\+/?-)\s*([\d.]+)', txt)

        # "Min" or "Max" qualifier anywhere in the text.
        has_min = bool(_re.search(r'\bmin\b', txt_lower))
        has_max = bool(_re.search(r'\bmax\b', txt_lower))

        # Range form: "89° - 91°" or "2.812" - 2.832""
        range_match = _re.search(r'([\d.]+)[°"]?\s*[-–]\s*([\d.]+)', txt)

        # Symmetric form: "N ± M" or "N +/- M"
        sym_match = _re.search(r'([\d.]+)[°"]?\s*(?:[±]|\+/?-)\s*([\d.]+)', txt)

        # Priority: symmetric > range > bilateral > min/max > plain
        if sym_match and not (bi_match and bi_match.start() == 0 and '±' not in txt[:3]):
            try:
                angle_val = float(sym_match.group(1))
                angle_tol = float(sym_match.group(2))
                angle_tol_type = "symmetric"
                angle_unit = unit_guess or "in"
            except ValueError:
                pass

        if angle_val is None and range_match:
            try:
                lo = float(range_match.group(1))
                hi = float(range_match.group(2))
                if hi > lo:
                    angle_val = (lo + hi) / 2.0
                    angle_tol = (hi - lo) / 2.0
                    angle_tol_type = "range"
                    angle_unit = unit_guess or "in"
            except ValueError:
                pass

        if angle_val is None and bi_match:
            try:
                angle_val = float(bi_match.group(1))
                angle_tol = None
                angle_tol_type = "bilateral"
                angle_unit = unit_guess or "deg"
            except ValueError:
                pass

        if angle_val is None and has_min:
            num_match = _re.search(r'([\d.]+)', txt)
            if num_match:
                try:
                    angle_val = float(num_match.group(1))
                    angle_tol = None
                    angle_tol_type = "min"
                    angle_unit = unit_guess or "in"
                except ValueError:
                    pass

        if angle_val is None and has_max:
            num_match = _re.search(r'([\d.]+)', txt)
            if num_match:
                try:
                    angle_val = float(num_match.group(1))
                    angle_tol = None
                    angle_tol_type = "max"
                    angle_unit = unit_guess or "in"
                except ValueError:
                    pass

        if angle_val is None:
            num_match = _re.search(r'([\d.]+)', txt)
            if num_match:
                try:
                    angle_val = float(num_match.group(1))
                    angle_tol = None
                    angle_tol_type = None
                    angle_unit = unit_guess or "in"
                except ValueError:
                    pass

        return angle_val, angle_tol, angle_unit, angle_tol_type

    @staticmethod
    def _split_multi_section_angle(angle_text: Optional[str]) -> List[Tuple[List[str], str]]:
        """
        If the angle text describes multiple sections with different specs,
        split it into [(sections, per_section_angle_text), ...].

        Example inputs that trigger splitting:
          'Section A, B & C = 60° +/-.3°\\nSection D = 66.66° +/-.3°'
            -> [(['A','B','C'], '60° +/-.3°'), (['D'], '66.66° +/-.3°')]
          'Sections A, B = 60° ± .3°; Section C = 66° ± .3°'
            -> [(['A','B'], '60° ± .3°'), (['C'], '66° ± .3°')]

        Returns empty list when the text is NOT a multi-section spec — caller
        should then treat the whole string as a single spec.
        """
        import re as _re

        if not angle_text:
            return []

        txt = angle_text.strip()

        # Must contain at least two occurrences of "Section" (case-insensitive)
        # to qualify as multi-section. One "Section X = Y" row is technically
        # possible but pointless to split.
        if len(_re.findall(r'\bsections?\b', txt, _re.IGNORECASE)) < 2:
            return []

        # Split on newlines OR on semicolons — the real-world Excel has
        # '\n' but users may type ';' too.
        raw_parts = [p.strip() for p in _re.split(r'[\n\r;]+', txt) if p.strip()]

        out: List[Tuple[List[str], str]] = []
        for part in raw_parts:
            # Match: "Section(s) A, B & C = <spec text>"
            m = _re.match(
                r'^\s*Sections?\s+([A-Za-z0-9 ,&/]+?)\s*=\s*(.+)$',
                part,
                _re.IGNORECASE,
            )
            if not m:
                continue
            sections_str = m.group(1).strip()
            spec_text = m.group(2).strip()
            # Break 'A, B & C' into ['A','B','C']. Accept ',', '&', ' and '.
            tokens = _re.split(r'[,&/]|\band\b', sections_str, flags=_re.IGNORECASE)
            sections = [t.strip().upper() for t in tokens if t.strip()]
            # Only keep single-letter section labels (A-Z). Drop anything weird
            # to stay conservative.
            sections = [s for s in sections if _re.match(r'^[A-Z]$', s)]
            if sections and spec_text:
                out.append((sections, spec_text))

        return out

    def import_model_specs_from_excel(self, file_path: str) -> Dict[str, int]:
        """
        Import model specs from the reference Excel file.
        Merges: updates existing, adds new, never deletes.

        Returns: {"updated": N, "added": N, "skipped": N}
        """
        import re
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True)
        result = {"updated": 0, "added": 0, "skipped": 0}

        # Collect data from all three sheets
        model_data = {}  # model -> dict of fields

        # Sheet 1: Model Reference (primary, most complete)
        if "Model Reference" in wb.sheetnames:
            ws = wb["Model Reference"]

            # Detect column positions from header row instead of hardcoding.
            # This handles spreadsheets with or without an extra leading column.
            col_map = {}
            header_aliases = {
                "model": "model",
                "element type": "element_type",
                "linearity": "linearity",
                "total resistance": "resistance",
                "electrical angle": "angle",
                "output smoothness": "smoothness",
                "open/closed": "open_closed",
                "product class": "product_class",
                "aliases": "aliases",
            }
            for header_row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                if not header_row:
                    break
                for idx, cell in enumerate(header_row):
                    if cell is None:
                        continue
                    key = str(cell).strip().lower()
                    if key in header_aliases:
                        col_map[header_aliases[key]] = idx
            logger.debug(f"Model Reference column map: {col_map}")

            if "model" not in col_map:
                logger.warning("Model Reference sheet has no 'Model' column header — skipping")
            else:
                def _cell(row, field):
                    """Get a cell value by field name, or None if column missing."""
                    idx = col_map.get(field)
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return None
                    return str(row[idx]).strip() or None

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row:
                        continue
                    model = _cell(row, "model")
                    if not model:
                        continue

                    element_type = _cell(row, "element_type")
                    linearity_text = _cell(row, "linearity")
                    resistance_text = _cell(row, "resistance")
                    angle_text = _cell(row, "angle")
                    smoothness = _cell(row, "smoothness")
                    open_closed = _cell(row, "open_closed")
                    product_class = _cell(row, "product_class")
                    aliases_raw = _cell(row, "aliases")

                    # Parse linearity type from text
                    linearity_type = None
                    linearity_pct = None
                    if linearity_text:
                        lt_lower = linearity_text.lower()
                        # Extract type: look for (Absolute), (Independent), etc.
                        type_match = re.search(
                            r'\(?(Absolute|Independent|Term Base|Zero-Based|VR Max)\)?',
                            linearity_text, re.IGNORECASE
                        )
                        if type_match:
                            linearity_type = type_match.group(1)
                            # Normalize case
                            type_map = {"absolute": "Absolute", "independent": "Independent",
                                        "term base": "Term Base", "zero-based": "Zero-Based",
                                        "vr max": "VR Max"}
                            linearity_type = type_map.get(linearity_type.lower(), linearity_type)
                        elif any(kw in lt_lower for kw in
                                 ['see chart', 'see table', 'function', 'trim according',
                                  'logarithmic', 'logaithmic', 'bowtie', 'no linearity']):
                            linearity_type = "Custom"

                        # Extract percentage: handle ± N.N%, +/-N.N%, +/-.N%
                        # Try ± first, then +/- variants
                        pct_match = re.search(r'[±]\s*(\d*\.?\d+)\s*%', linearity_text)
                        if not pct_match:
                            pct_match = re.search(r'\+/?-?\s*(\d*\.?\d+)\s*%', linearity_text)
                        if pct_match:
                            try:
                                linearity_pct = float(pct_match.group(1))
                            except ValueError:
                                pass

                    # Parse resistance: "950 - 1,050 Ω" → min=950, max=1050
                    r_min = None
                    r_max = None
                    if resistance_text:
                        r_match = re.search(
                            r'([\d,]+\.?\d*)\s*[-–]\s*([\d,]+\.?\d*)',
                            resistance_text
                        )
                        if r_match:
                            try:
                                r_min = float(r_match.group(1).replace(',', ''))
                                r_max = float(r_match.group(2).replace(',', ''))
                            except ValueError:
                                pass

                    # Parse angle — either a single spec or a multi-section spec.
                    # Multi-section example from Excel (model 8508):
                    #   'Section A, B & C = 60° +/-.3°\nSection D = 66.66° +/-.3°'
                    # In that case we emit one spec row per section letter so the
                    # trim files (which come in as 8508-A, 8508-B, ...) each find
                    # the matching spec via plain model-name lookup.
                    sections = self._split_multi_section_angle(angle_text)

                    # Normalize aliases: accept '|' or ',' as separator, dedupe
                    # and drop empties. Stored as pipe-separated in DB.
                    aliases_norm = None
                    if aliases_raw and aliases_raw not in ("None", "nan"):
                        tokens = re.split(r'[|,]', aliases_raw)
                        clean = []
                        seen = set()
                        for t in tokens:
                            t = t.strip()
                            if t and t not in seen and t != model:
                                seen.add(t)
                                clean.append(t)
                        if clean:
                            aliases_norm = " | ".join(clean)

                    # Shared fields common to every section row for this source row.
                    shared = {
                        "element_type": element_type if element_type and element_type != 'None' else None,
                        "product_class": product_class if product_class and product_class != 'None' else None,
                        "linearity_type": linearity_type,
                        "linearity_spec_text": linearity_text if linearity_text and linearity_text != 'None' else None,
                        "linearity_spec_pct": linearity_pct,
                        "total_resistance_min": r_min,
                        "total_resistance_max": r_max,
                        "output_smoothness": smoothness if smoothness and smoothness != 'None' else None,
                        # Write open_closed to both new and legacy fields so GUIs
                        # reading either column keep working.
                        "open_closed": open_closed if open_closed and open_closed != 'None' else None,
                        "circuit_type": open_closed if open_closed and open_closed != 'None' else None,
                        "aliases": aliases_norm,
                    }

                    if sections:
                        # Multi-section model: emit one row per section letter.
                        for section_letters, per_section_text in sections:
                            angle_val, angle_tol, angle_unit, angle_tol_type = \
                                self._parse_angle_string(per_section_text)
                            for letter in section_letters:
                                section_model = f"{model}-{letter}"
                                model_data[section_model] = {
                                    "model": section_model,
                                    **shared,
                                    "electrical_angle": angle_val,
                                    "electrical_angle_tol": angle_tol,
                                    "electrical_angle_tol_type": angle_tol_type,
                                    "electrical_angle_unit": angle_unit,
                                }
                        logger.info(
                            f"Model specs: expanded {model!r} into "
                            f"{sum(len(s) for s, _ in sections)} section rows"
                        )
                    else:
                        # Normal single-spec row.
                        angle_val, angle_tol, angle_unit, angle_tol_type = \
                            self._parse_angle_string(angle_text)
                        model_data[model] = {
                            "model": model,
                            **shared,
                            "electrical_angle": angle_val,
                            "electrical_angle_tol": angle_tol,
                            "electrical_angle_tol_type": angle_tol_type,
                            "electrical_angle_unit": angle_unit,
                        }

        # Sheet 2: Element Type (supplement — broader coverage)
        if "Element Type" in wb.sheetnames:
            ws = wb["Element Type"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                model = str(row[0]).strip() if row[0] else None
                etype = str(row[1]).strip() if row[1] else None
                if model and etype and etype != 'None':
                    if model not in model_data:
                        model_data[model] = {"model": model, "element_type": etype}
                    elif not model_data[model].get("element_type"):
                        model_data[model]["element_type"] = etype

        # Sheet 3: Product Class (supplement — broadest coverage)
        if "Product Class" in wb.sheetnames:
            ws = wb["Product Class"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                model = str(row[0]).strip() if row[0] else None
                pclass = str(row[1]).strip() if row[1] else None
                if model and pclass and pclass != 'None':
                    if model not in model_data:
                        model_data[model] = {"model": model, "product_class": pclass}
                    elif not model_data[model].get("product_class"):
                        model_data[model]["product_class"] = pclass

        wb.close()

        # Save to database (merge logic — save_model_spec handles upsert atomically)
        for model_name, data in model_data.items():
            try:
                _, was_update = self.save_model_spec(data)
                if was_update:
                    result["updated"] += 1
                else:
                    result["added"] += 1
            except Exception as e:
                logger.warning(f"Skipping model spec {model_name}: {e}")
                result["skipped"] += 1

        logger.info(
            f"Model specs import: {result['added']} added, "
            f"{result['updated']} updated, {result['skipped']} skipped"
        )
        return result

    # =========================================================================
    # Output Smoothness Methods
    # =========================================================================

    def save_smoothness_result(
        self, metadata: Dict[str, Any], tracks: List[Dict[str, Any]], file_hash: str,
        file_size: Optional[int] = None,
        file_modified_date: Optional[datetime] = None,
    ) -> int:
        """Save an Output Smoothness result. Returns ID.

        file_size / file_modified_date feed the incremental scan's stat
        fast-path (see save_final_test).
        """
        from laser_trim_analyzer.database.models import (
            SmoothnessResult as DBSmoothnessResult,
            SmoothnessTrack as DBSmoothnessTrack,
        )

        # Precompute aggregates (used for both insert and upsert paths)
        overall_status = DBStatusType.PASS
        for track in tracks:
            if track.get("smoothness_pass") is False:
                overall_status = DBStatusType.FAIL
                break

        max_smooth = max((t.get("max_smoothness", 0) or 0 for t in tracks), default=0)
        avg_smooth = sum(t.get("avg_smoothness", 0) or 0 for t in tracks) / len(tracks) if tracks else 0
        spec = metadata.get("smoothness_spec") or (tracks[0].get("smoothness_spec") if tracks else None)
        passes = all(t.get("smoothness_pass", True) for t in tracks) if tracks else None

        with self._write_lock:
            try:
                with self.session() as session:
                    existing = session.query(DBSmoothnessResult).filter(
                        DBSmoothnessResult.file_hash == file_hash
                    ).first()
                    if existing:
                        # UPSERT: the old code silently returned here without
                        # updating anything. That meant records imported before
                        # the parser fix kept their zeroed values forever, even
                        # when reprocessed. Now we overwrite the parent row's
                        # aggregate fields and replace the child tracks so a
                        # reprocess actually refreshes the stored data.
                        existing.overall_status = overall_status
                        existing.smoothness_spec = spec
                        existing.max_smoothness_value = max_smooth
                        existing.avg_smoothness_value = avg_smooth
                        existing.smoothness_pass = passes
                        if file_size is not None:
                            existing.file_size = file_size
                            existing.file_modified_date = file_modified_date
                        if metadata.get("file_date"):
                            existing.file_date = metadata.get("file_date")
                        if metadata.get("test_date"):
                            existing.test_date = metadata.get("test_date")
                        if metadata.get("element_label"):
                            existing.element_label = metadata.get("element_label")

                        # Replace the per-track rows
                        session.query(DBSmoothnessTrack).filter(
                            DBSmoothnessTrack.smoothness_id == existing.id
                        ).delete(synchronize_session=False)

                        for track_data in tracks:
                            db_track = DBSmoothnessTrack(
                                smoothness_id=existing.id,
                                track_id=track_data.get("track_id", "default"),
                                status=DBStatusType.PASS if track_data.get("smoothness_pass", True) else DBStatusType.FAIL,
                                smoothness_spec=track_data.get("smoothness_spec"),
                                max_smoothness=track_data.get("max_smoothness"),
                                avg_smoothness=track_data.get("avg_smoothness"),
                                smoothness_pass=track_data.get("smoothness_pass"),
                                position_data=track_data.get("positions"),
                                smoothness_data=track_data.get("smoothness_values"),
                            )
                            session.add(db_track)

                        logger.debug(
                            f"Updated Smoothness: {metadata.get('filename')} "
                            f"(ID: {existing.id}, max={max_smooth:.4f}, spec={spec}, "
                            f"tracks={len(tracks)})"
                        )
                        return existing.id

                    linked_trim_id, match_confidence, days_since_trim, match_method = self._find_matching_trim(
                        session, metadata.get("model"), metadata.get("serial"),
                        metadata.get("file_date") or metadata.get("test_date")
                    )

                    db_result = DBSmoothnessResult(
                        filename=metadata.get("filename", "unknown"),
                        file_path=str(metadata.get("file_path", "")),
                        file_hash=file_hash,
                        file_date=metadata.get("file_date"),
                        file_size=file_size,
                        file_modified_date=file_modified_date,
                        model=metadata.get("model", "unknown"),
                        serial=metadata.get("serial", "unknown"),
                        element_label=metadata.get("element_label"),
                        test_date=metadata.get("test_date"),
                        overall_status=overall_status,
                        smoothness_spec=spec,
                        max_smoothness_value=max_smooth,
                        avg_smoothness_value=avg_smooth,
                        smoothness_pass=passes,
                        linked_trim_id=linked_trim_id,
                        match_confidence=match_confidence,
                        match_method=match_method,
                        days_since_trim=days_since_trim,
                    )
                    session.add(db_result)
                    session.flush()
                    result_id = db_result.id

                    for track_data in tracks:
                        db_track = DBSmoothnessTrack(
                            smoothness_id=result_id,
                            track_id=track_data.get("track_id", "default"),
                            status=DBStatusType.PASS if track_data.get("smoothness_pass", True) else DBStatusType.FAIL,
                            smoothness_spec=track_data.get("smoothness_spec"),
                            max_smoothness=track_data.get("max_smoothness"),
                            avg_smoothness=track_data.get("avg_smoothness"),
                            smoothness_pass=track_data.get("smoothness_pass"),
                            position_data=track_data.get("positions"),
                            smoothness_data=track_data.get("smoothness_values"),
                        )
                        session.add(db_track)

                    logger.debug(f"Saved Smoothness: {metadata.get('filename')} (ID: {result_id})")
                    return result_id

            except IntegrityError:
                logger.warning(f"Smoothness duplicate: {metadata.get('filename')}")
                with self.session() as session:
                    existing = session.query(DBSmoothnessResult).filter(
                        DBSmoothnessResult.file_hash == file_hash
                    ).first()
                    return existing.id if existing else -1

    def get_smoothness_files_missing_tracks(self) -> List[Dict[str, Any]]:
        """
        Get Output Smoothness records that have 0 tracks stored.

        Used to repair records imported by the older code that wrote the
        result row but did not persist the per-position arrays needed to
        render the chart.

        Returns:
            List of dicts with id, filename, file_path, model, serial.
        """
        from laser_trim_analyzer.database.models import (
            SmoothnessResult as DBSmoothnessResult,
            SmoothnessTrack as DBSmoothnessTrack,
        )

        with self.session() as session:
            track_count_subq = (
                session.query(
                    DBSmoothnessTrack.smoothness_id,
                    func.count(DBSmoothnessTrack.id).label('track_count')
                )
                .group_by(DBSmoothnessTrack.smoothness_id)
                .subquery()
            )

            results = (
                session.query(DBSmoothnessResult)
                .outerjoin(
                    track_count_subq,
                    DBSmoothnessResult.id == track_count_subq.c.smoothness_id,
                )
                .filter(track_count_subq.c.track_count == None)
                .all()
            )
            return [
                {
                    "id": r.id,
                    "filename": r.filename,
                    "file_path": r.file_path,
                    "model": r.model,
                    "serial": r.serial,
                }
                for r in results
            ]

    def update_smoothness_tracks(
        self,
        smoothness_id: int,
        tracks: List[Dict[str, Any]],
    ) -> bool:
        """
        Replace the per-track data for an existing Smoothness record.

        Used to fix records that were imported before the smoothness_tracks
        write was added to save_smoothness_result.
        """
        from laser_trim_analyzer.database.models import (
            SmoothnessTrack as DBSmoothnessTrack,
        )

        if not tracks:
            return False

        with self._write_lock:
            try:
                with self.session() as session:
                    # Delete any existing (likely zero) tracks first
                    session.query(DBSmoothnessTrack).filter(
                        DBSmoothnessTrack.smoothness_id == smoothness_id
                    ).delete(synchronize_session=False)

                    for track_data in tracks:
                        db_track = DBSmoothnessTrack(
                            smoothness_id=smoothness_id,
                            track_id=track_data.get("track_id", "default"),
                            status=DBStatusType.PASS if track_data.get("smoothness_pass", True) else DBStatusType.FAIL,
                            smoothness_spec=track_data.get("smoothness_spec"),
                            max_smoothness=track_data.get("max_smoothness"),
                            avg_smoothness=track_data.get("avg_smoothness"),
                            smoothness_pass=track_data.get("smoothness_pass"),
                            position_data=track_data.get("positions"),
                            smoothness_data=track_data.get("smoothness_values"),
                        )
                        session.add(db_track)
                    return True
            except Exception as e:
                logger.error(f"update_smoothness_tracks({smoothness_id}) failed: {e}")
                return False

    def search_smoothness_results(
        self, model: Optional[str] = None, limit: int = 500
    ) -> List[Dict[str, Any]]:
        """Search Output Smoothness results."""
        from laser_trim_analyzer.database.models import SmoothnessResult as DBSmoothnessResult

        with self.session() as session:
            query = session.query(DBSmoothnessResult)
            if model and model != "All Models":
                query = query.filter(DBSmoothnessResult.model == model)
            results = query.order_by(desc(DBSmoothnessResult.file_date)).limit(limit).all()
            return [
                {
                    "id": r.id, "filename": r.filename, "model": r.model,
                    "serial": r.serial, "element_label": r.element_label,
                    "file_date": r.file_date, "test_date": r.test_date,
                    "overall_status": r.overall_status.value if r.overall_status else "UNKNOWN",
                    "smoothness_spec": r.smoothness_spec,
                    "max_smoothness_value": r.max_smoothness_value,
                    "avg_smoothness_value": r.avg_smoothness_value,
                    "smoothness_pass": r.smoothness_pass,
                    "linked_trim_id": r.linked_trim_id,
                    "match_confidence": r.match_confidence,
                    "match_method": r.match_method,
                }
                for r in results
            ]

    def get_smoothness_result(self, result_id: int) -> Optional[Dict[str, Any]]:
        """Get a single Output Smoothness result by ID with tracks."""
        from laser_trim_analyzer.database.models import (
            SmoothnessResult as DBSmoothnessResult,
            SmoothnessTrack as DBSmoothnessTrack,
        )
        with self.session() as session:
            result = session.query(DBSmoothnessResult).filter(
                DBSmoothnessResult.id == result_id
            ).first()
            if not result:
                return None
            tracks = session.query(DBSmoothnessTrack).filter(
                DBSmoothnessTrack.smoothness_id == result_id
            ).all()
            return {
                "id": result.id, "filename": result.filename,
                "model": result.model, "serial": result.serial,
                "element_label": result.element_label,
                "file_date": result.file_date, "test_date": result.test_date,
                "overall_status": result.overall_status.value if result.overall_status else "UNKNOWN",
                "smoothness_spec": result.smoothness_spec,
                "max_smoothness_value": result.max_smoothness_value,
                "smoothness_pass": result.smoothness_pass,
                "linked_trim_id": result.linked_trim_id,
                "match_method": result.match_method,
                "match_confidence": result.match_confidence,
                "tracks": [
                    {
                        "track_id": t.track_id,
                        "smoothness_spec": t.smoothness_spec,
                        "max_smoothness": t.max_smoothness,
                        "smoothness_pass": t.smoothness_pass,
                        "positions": t.position_data or [],
                        "smoothness_values": t.smoothness_data or [],
                    }
                    for t in tracks
                ],
            }

    def get_smoothness_stats(self, days_back: int = 90) -> Dict[str, Any]:
        """Get Output Smoothness dashboard statistics."""
        from laser_trim_analyzer.database.models import SmoothnessResult as DBSmoothnessResult
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            total = session.query(func.count(DBSmoothnessResult.id)).filter(
                DBSmoothnessResult.file_date >= cutoff
            ).scalar() or 0
            if total == 0:
                return {"total": 0, "pass_rate": 0, "linked_count": 0, "link_rate": 0}
            passed = session.query(func.count(DBSmoothnessResult.id)).filter(
                DBSmoothnessResult.file_date >= cutoff,
                DBSmoothnessResult.smoothness_pass == True,
            ).scalar() or 0
            linked = session.query(func.count(DBSmoothnessResult.id)).filter(
                DBSmoothnessResult.file_date >= cutoff,
                DBSmoothnessResult.linked_trim_id.isnot(None),
            ).scalar() or 0
            return {
                "total": total,
                "pass_rate": round(passed / total * 100, 1),
                "linked_count": linked,
                "link_rate": round(linked / total * 100, 1),
            }

    def get_smoothness_stats_by_model(
        self, model: Optional[str] = None, days_back: int = 90
    ) -> List[Dict[str, Any]]:
        """Get Output Smoothness statistics grouped by model.

        Args:
            model: Optional model filter. If given, return stats for that model only.
            days_back: Number of days to look back (default 90).

        Returns:
            List of dicts sorted by pass_rate ascending (worst first), then margin.
        """
        from laser_trim_analyzer.database.models import SmoothnessResult as DBSmoothnessResult

        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)

            query = session.query(
                DBSmoothnessResult.model,
                func.count(DBSmoothnessResult.id).label("count"),
                func.sum(
                    case(
                        (DBSmoothnessResult.smoothness_pass == True, 1),
                        else_=0,
                    )
                ).label("passed"),
                func.avg(DBSmoothnessResult.max_smoothness_value).label("avg_max_smoothness"),
                func.max(DBSmoothnessResult.max_smoothness_value).label("worst_case"),
                func.avg(DBSmoothnessResult.smoothness_spec).label("spec_limit"),
            ).filter(
                DBSmoothnessResult.file_date >= cutoff,
            ).group_by(DBSmoothnessResult.model)

            if model is not None:
                query = query.filter(DBSmoothnessResult.model == model)

            rows = query.all()

            results: List[Dict[str, Any]] = []
            for row in rows:
                count = row.count
                passed = row.passed or 0
                pass_rate = round(passed / count * 100, 1) if count else 0.0
                avg_max = round(row.avg_max_smoothness, 4) if row.avg_max_smoothness is not None else 0.0
                worst = round(row.worst_case, 4) if row.worst_case is not None else 0.0
                spec = round(row.spec_limit, 4) if row.spec_limit is not None else 0.0
                margin = round(spec - avg_max, 4)

                results.append({
                    "model": row.model,
                    "count": count,
                    "passed": passed,
                    "pass_rate": pass_rate,
                    "avg_max_smoothness": avg_max,
                    "worst_case": worst,
                    "spec_limit": spec,
                    "margin": margin,
                })

            results.sort(key=lambda r: (r["pass_rate"], r["margin"]))
            return results

    # =========================================================================
    # Cpk / Analytics Queries (Phase 4)
    # =========================================================================

    def get_linearity_deviations_for_cpk(
        self, model: str, days_back: int = 90
    ) -> List[float]:
        """Get raw linearity deviation values for Cpk calculation."""
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            results = session.query(
                DBTrackResult.final_linearity_error_shifted
            ).join(DBAnalysisResult).filter(
                DBAnalysisResult.model == model,
                DBAnalysisResult.file_date >= cutoff,
                DBTrackResult.final_linearity_error_shifted.isnot(None),
            ).all()
            return [r[0] for r in results]

    def get_cpk_by_model(self, days_back: int = 90) -> List[Dict[str, Any]]:
        """Calculate Cpk for each model that has a linearity spec defined."""
        from laser_trim_analyzer.core.cpk import calculate_cpk

        with self.session() as session:
            specs = session.query(
                ModelSpec.model, ModelSpec.linearity_spec_pct
            ).filter(
                ModelSpec.linearity_spec_pct.isnot(None)
            ).all()

        results = []
        for model_name, spec_pct in specs:
            devs = self.get_linearity_deviations_for_cpk(model_name, days_back)
            if len(devs) < 10:
                continue
            cpk_result = calculate_cpk(devs, spec_pct)
            results.append({
                "model": model_name,
                "cpk": cpk_result.cpk,
                "ppk": cpk_result.ppk,
                "rating": cpk_result.rating,
                "n_samples": cpk_result.n_samples,
                "spec_pct": spec_pct,
                "mean": cpk_result.mean,
            })
        results.sort(key=lambda x: x["cpk"] if x["cpk"] is not None else 999)
        return results

    def get_cpk_trend_for_model(
        self, model: str, spec_limit_pct: float,
        days_back: int = 180, period: str = "month"
    ) -> List[Dict[str, Any]]:
        """Get Cpk trend over time for a specific model."""
        from laser_trim_analyzer.core.cpk import calculate_cpk_trend
        from collections import defaultdict

        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            if period == "week":
                period_expr = func.strftime('%Y-W%W', DBAnalysisResult.file_date)
            else:
                period_expr = func.strftime('%Y-%m', DBAnalysisResult.file_date)

            results = session.query(
                period_expr.label("period"),
                DBTrackResult.final_linearity_error_shifted,
            ).join(DBAnalysisResult).filter(
                DBAnalysisResult.model == model,
                DBAnalysisResult.file_date >= cutoff,
                DBTrackResult.final_linearity_error_shifted.isnot(None),
            ).order_by(period_expr).all()

        period_data = defaultdict(list)
        for r in results:
            period_data[r.period].append(r.final_linearity_error_shifted)

        deviations_by_period = sorted(period_data.items())
        return calculate_cpk_trend(deviations_by_period, spec_limit_pct)

    def get_model_scorecard_data(
        self, model: str, days_back: int = 90
    ) -> Dict[str, Any]:
        """Get comprehensive scorecard data for a single model."""
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)

            total = session.query(func.count(DBAnalysisResult.id)).filter(
                DBAnalysisResult.model == model,
                DBAnalysisResult.file_date >= cutoff,
            ).scalar() or 0

            passed = session.query(func.count(DBAnalysisResult.id)).filter(
                DBAnalysisResult.model == model,
                DBAnalysisResult.file_date >= cutoff,
                DBAnalysisResult.overall_status == DBStatusType.PASS,
            ).scalar() or 0

            pass_rate = (passed / total * 100) if total > 0 else 0

            avg_dev = session.query(
                func.avg(DBTrackResult.final_linearity_error_shifted)
            ).join(DBAnalysisResult).filter(
                DBAnalysisResult.model == model,
                DBAnalysisResult.file_date >= cutoff,
                DBTrackResult.final_linearity_error_shifted.isnot(None),
            ).scalar()

            spec_obj = session.query(ModelSpec).filter(
                ModelSpec.model == model
            ).first()

            # Extract spec data while still in session to avoid detached instance errors
            spec_data = None
            spec_linearity_pct = None
            if spec_obj:
                spec_data = {
                    "element_type": spec_obj.element_type,
                    "product_class": spec_obj.product_class,
                    "linearity_type": spec_obj.linearity_type,
                    "linearity_spec_pct": spec_obj.linearity_spec_pct,
                }
                spec_linearity_pct = spec_obj.linearity_spec_pct

        cpk_data = None
        if spec_linearity_pct:
            from laser_trim_analyzer.core.cpk import calculate_cpk
            devs = self.get_linearity_deviations_for_cpk(model, days_back)
            if len(devs) >= 10:
                cpk_result = calculate_cpk(devs, spec_linearity_pct)
                cpk_data = cpk_result.to_dict()

        # Drift status from ML state
        drift_status = None
        try:
            from laser_trim_analyzer.database.models import ModelMLState
            with self.session() as session:
                ml_state = session.query(ModelMLState).filter(
                    ModelMLState.model == model
                ).first()
                if ml_state:
                    drift_status = "drifting" if ml_state.is_drifting else "stable"
        except Exception:
            pass

        return {
            "model": model,
            "total": total,
            "passed": passed,
            "pass_rate": pass_rate,
            "avg_deviation": avg_dev,
            "cpk": cpk_data,
            "drift_status": drift_status,
            "spec": spec_data,
        }

    def get_company_yield_trend(
        self, days_back: int = 365, period: str = "week"
    ) -> Dict[str, Any]:
        """Company-wide yield trend, overall and split by trim system (A/B/C).

        The v6 Dashboard's company-trend section (2026-07-06). Differs from
        get_yield_trend in two ways: UNTRIMMED test-sweeps are excluded from
        BOTH numerator and denominator (they have no trim verdict — dashboard
        convention), and results are additionally grouped by system so the
        LTS3 (C) ramp can be compared against A/B.

        Returns {"periods": [...ordered period keys...],
                 "company":  [{"period","total","passed","pass_rate"}...],
                 "by_system": {"A": [rows...], "B": [...], "C": [...]}}.
        Systems with no data in the window are omitted from by_system.
        """
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            if period == "week":
                period_expr = func.strftime('%Y-W%W', DBAnalysisResult.file_date)
            else:
                period_expr = func.strftime('%Y-%m', DBAnalysisResult.file_date)

            # Numerator = LINEARITY-ACCEPTED (PASS + WARNING): linearity is the
            # zero-tolerance customer requirement; WARNING units passed it
            # (sigma is an internal drift-watch flag, not a disposition). The
            # company trend answers "is our yield moving" on the customer basis.
            rows = session.query(
                period_expr.label("period"),
                DBAnalysisResult.system,
                func.count(DBAnalysisResult.id).label("total"),
                func.sum(
                    case((DBAnalysisResult.overall_status.in_(
                        [DBStatusType.PASS, DBStatusType.WARNING]), 1), else_=0)
                ).label("accepted"),
            ).filter(
                DBAnalysisResult.file_date >= cutoff,
                DBAnalysisResult.overall_status != DBStatusType.UNTRIMMED.name,
            ).group_by(period_expr, DBAnalysisResult.system) \
             .order_by(period_expr).all()

            periods: List[str] = []
            company: Dict[str, Dict[str, int]] = {}
            by_system: Dict[str, Dict[str, Dict[str, int]]] = {}
            for r in rows:
                if r.period not in company:
                    periods.append(r.period)
                    company[r.period] = {"total": 0, "accepted": 0}
                company[r.period]["total"] += r.total or 0
                company[r.period]["accepted"] += r.accepted or 0
                sys_val = r.system.value if isinstance(r.system, DBSystemType) else str(r.system)
                by_system.setdefault(sys_val, {})[r.period] = {
                    "total": r.total or 0, "accepted": r.accepted or 0}

            def _series(bucket: Dict[str, Dict[str, int]]) -> List[Dict[str, Any]]:
                out = []
                for p in periods:
                    d = bucket.get(p)
                    total = d["total"] if d else 0
                    accepted = d["accepted"] if d else 0
                    out.append({"period": p, "total": total, "accepted": accepted,
                                "linearity_yield": (accepted / total * 100) if total else None})
                return out

            # Honesty metadata for the chart: the newest period is usually
            # PARTIAL (the week/month is still filling), and the whole dataset
            # has a vintage (batch-loaded data can lag production by weeks).
            data_through = session.query(func.max(DBAnalysisResult.file_date)).scalar()
            partial_last = False
            if periods and data_through is not None:
                fmt = '%Y-W%W' if period == "week" else '%Y-%m'
                if periods[-1] == data_through.strftime(fmt):
                    # Last bucket = the bucket the data ends in. It's PARTIAL
                    # unless the data runs to the bucket's calendar end.
                    if period == "week":
                        partial_last = data_through.weekday() != 6  # %W: Mon–Sun
                    else:
                        import calendar
                        last_day = calendar.monthrange(data_through.year,
                                                       data_through.month)[1]
                        partial_last = data_through.day != last_day

            return {
                "periods": periods,
                "company": _series(company),
                "by_system": {s: _series(b) for s, b in sorted(by_system.items())},
                "partial_last": partial_last,
                "data_through": data_through,
            }

    def get_yield_trend(
        self, days_back: int = 180, period: str = "week"
    ) -> List[Dict[str, Any]]:
        """Get overall yield (pass rate) trend across all models."""
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            if period == "week":
                period_expr = func.strftime('%Y-W%W', DBAnalysisResult.file_date)
            else:
                period_expr = func.strftime('%Y-%m', DBAnalysisResult.file_date)

            results = session.query(
                period_expr.label("period"),
                func.count(DBAnalysisResult.id).label("total"),
                func.sum(
                    case((DBAnalysisResult.overall_status == DBStatusType.PASS, 1), else_=0)
                ).label("passed"),
            ).filter(
                DBAnalysisResult.file_date >= cutoff,
            ).group_by(period_expr).order_by(period_expr).all()

            return [
                {
                    "period": r.period,
                    "total": r.total,
                    "passed": r.passed or 0,
                    "pass_rate": ((r.passed or 0) / r.total * 100) if r.total > 0 else 0,
                }
                for r in results
            ]

    def get_comparative_model_trends(
        self, models: List[str], days_back: int = 90, period: str = "week"
    ) -> Dict[str, List[Dict[str, Any]]]:
        """Get pass rate trends for multiple models for overlay comparison."""
        result = {}
        for model in models:
            with self.session() as session:
                cutoff = datetime.now() - timedelta(days=days_back)
                if period == "week":
                    period_expr = func.strftime('%Y-W%W', DBAnalysisResult.file_date)
                else:
                    period_expr = func.strftime('%Y-%m', DBAnalysisResult.file_date)

                rows = session.query(
                    period_expr.label("period"),
                    func.count(DBAnalysisResult.id).label("total"),
                    func.sum(
                        case((DBAnalysisResult.overall_status == DBStatusType.PASS, 1), else_=0)
                    ).label("passed"),
                ).filter(
                    DBAnalysisResult.model == model,
                    DBAnalysisResult.file_date >= cutoff,
                ).group_by(period_expr).order_by(period_expr).all()

                result[model] = [
                    {
                        "period": r.period,
                        "total": r.total,
                        "pass_rate": ((r.passed or 0) / r.total * 100) if r.total > 0 else 0,
                    }
                    for r in rows
                ]
        return result

    def get_drift_events_timeline(self, days_back: int = 180) -> List[Dict[str, Any]]:
        """Get drift detection events for timeline visualization.

        Returns the date drift was first detected (drift_start_date), not when the
        ML state row was last updated. Falls back to updated_date for legacy rows
        that were written before drift_start_date was populated.
        """
        from laser_trim_analyzer.database.models import ModelMLState
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            # Use drift_start_date where available; updated_date as fallback for old rows.
            effective_date = func.coalesce(
                ModelMLState.drift_start_date, ModelMLState.updated_date
            )
            results = session.query(ModelMLState).filter(
                effective_date >= cutoff,
                ModelMLState.is_drifting == True,
            ).order_by(effective_date).all()
            return [
                {
                    "model": r.model,
                    # Prefer the actual detection date; fall back to updated_date only if missing.
                    "date": (r.drift_start_date or r.updated_date).isoformat()
                    if (r.drift_start_date or r.updated_date) else None,
                    "direction": r.drift_direction,
                }
                for r in results
            ]

    def get_trim_difficulty_by_model(
        self,
        days_back: int = 90,
        min_units: int = 5,
        limit: int = 30,
    ) -> List[Dict[str, Any]]:
        """Aggregate trim_pass_count per model.

        Higher avg passes = unit was harder to trim to spec.
        retrim_rate = % of units that needed >1 trim cycle.

        FT-derived rows have NULL trim_pass_count and are excluded.
        UNTRIMMED tracks (test-sweep-only files) carry trim_pass_count=0
        but never went through a laser-trim cycle, so they would skew the
        average toward "easy" and inflate the retrim_rate denominator;
        excluded by status.
        Models with fewer than `min_units` qualifying records are excluded
        so a single outlier doesn't dominate the chart.

        Returns rows sorted by avg_passes descending (worst first).
        """
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            results = session.query(
                DBAnalysisResult.model,
                func.count(DBTrackResult.id).label("count"),
                func.avg(DBTrackResult.trim_pass_count).label("avg_passes"),
                func.max(DBTrackResult.trim_pass_count).label("max_passes"),
                func.sum(
                    case((DBTrackResult.trim_pass_count > 1, 1), else_=0)
                ).label("retrims"),
                # Avg max_error_reduction_percent — distinguishes models
                # where retrimming actually helps (high reduction) from
                # those where extra passes don't improve outcomes (low
                # reduction, process root-cause issue).
                func.avg(
                    DBTrackResult.max_error_reduction_percent
                ).label("avg_error_reduction"),
            ).join(DBTrackResult).filter(
                DBAnalysisResult.file_date >= cutoff,
                DBTrackResult.trim_pass_count.isnot(None),
                DBTrackResult.status != DBStatusType.UNTRIMMED.name,
            ).group_by(DBAnalysisResult.model).having(
                func.count(DBTrackResult.id) >= min_units
            ).order_by(desc("avg_passes")).limit(limit).all()

            return [
                {
                    "model": r.model,
                    "count": int(r.count or 0),
                    "avg_passes": float(r.avg_passes or 0.0),
                    "max_passes": int(r.max_passes or 0),
                    "retrim_rate": (float(r.retrims or 0) / float(r.count)) * 100.0
                    if r.count else 0.0,
                    "avg_error_reduction": (
                        float(r.avg_error_reduction)
                        if r.avg_error_reduction is not None
                        else None
                    ),
                }
                for r in results
            ]

    def get_anomaly_rate_by_model(
        self,
        days_back: int = 90,
        min_samples: int = 10,
    ) -> List[Dict[str, Any]]:
        """Aggregate anomaly flag rate per model over a window.

        is_anomaly is set per-track when the trim has the linear-slope
        signature of a true trim failure (vs random noise). Rolling it
        up per model surfaces persistent setup issues — e.g. a model
        with 12 anomalies in 30 days likely has a fixture or operator
        problem rather than random material variation.

        Args:
            days_back: window length in days, anchored to file_date.
            min_samples: minimum total tracks per model to include in
                the result so a single anomaly on a low-volume model
                doesn't dominate the ranking.

        Returns:
            List of dicts sorted by anomaly_rate descending. Each dict:
                model, total_tracks, anomaly_count, anomaly_rate
                (percent), last_anomaly_date.
        """
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            rows = (
                session.query(
                    DBAnalysisResult.model,
                    func.count(DBTrackResult.id).label("total_tracks"),
                    func.sum(
                        case((DBTrackResult.is_anomaly == True, 1), else_=0)
                    ).label("anomaly_count"),
                    func.max(
                        case(
                            (DBTrackResult.is_anomaly == True,
                             DBAnalysisResult.file_date),
                            else_=None,
                        )
                    ).label("last_anomaly_date"),
                )
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(
                    DBAnalysisResult.model.isnot(None),
                    DBAnalysisResult.model != "Unknown",
                    DBAnalysisResult.file_date >= cutoff,
                    # Exclude UNTRIMMED tracks (is_anomaly defaults False) from the rate denominator.
                    DBTrackResult.status != DBStatusType.UNTRIMMED.name,
                )
                .group_by(DBAnalysisResult.model)
                .having(func.count(DBTrackResult.id) >= min_samples)
                .all()
            )

            results = []
            for r in rows:
                total = int(r.total_tracks or 0)
                anom = int(r.anomaly_count or 0)
                rate = (anom / total * 100.0) if total else 0.0
                results.append({
                    "model": r.model,
                    "total_tracks": total,
                    "anomaly_count": anom,
                    "anomaly_rate": rate,
                    "last_anomaly_date": r.last_anomaly_date,
                })
            results.sort(key=lambda r: -r["anomaly_rate"])
            return results

    # Columns the Process Drift view supports. Mapping from the
    # user-facing label to the SQLAlchemy column attribute on
    # DBTrackResult, plus a unit string for the chart axis.
    _PROCESS_DRIFT_METRICS: Dict[str, Dict[str, Any]] = {
        "untrimmed_resistance": {
            "label": "Untrimmed Resistance",
            "unit": "Ω",
            "fmt": "{:.0f}",
        },
        "trimmed_resistance": {
            "label": "Trimmed Resistance",
            "unit": "Ω",
            "fmt": "{:.0f}",
        },
        "measured_electrical_angle": {
            "label": "Measured Electrical Angle",
            "unit": "°",
            "fmt": "{:.2f}",
        },
        "trim_pass_count": {
            "label": "Trim Passes",
            "unit": "passes",
            "fmt": "{:.2f}",
        },
    }

    def get_process_drift_table(
        self,
        metric: str,
        baseline_days: int = 90,
        recent_days: int = 14,
        min_baseline_samples: int = 20,
        min_recent_samples: int = 5,
        z_threshold: float = 2.0,
    ) -> List[Dict[str, Any]]:
        """Per-model drift snapshot for one physical metric.

        Compares each model's recent mean of `metric` against its older
        baseline mean+stdev; rows with |z| >= z_threshold are flagged as
        drifting. Returns delta_pct (signed % change) and a downsampled
        time series (≤80 (date_iso, value) tuples) per model so the GUI
        can render a sparkline alongside the numeric stats.

        Useful for spotting:
            * starting resistance shifting (carbon batch / tooling)
            * measured electrical angle shifting (setup / fixture)
            * trim passes creeping up (process degradation)

        Args:
            metric: One of self._PROCESS_DRIFT_METRICS keys.
            baseline_days: How far back the 'normal' window goes.
            recent_days: How recent the 'is it different now?' window is.
                Must be < baseline_days.
            min_baseline_samples / min_recent_samples: filters out models
                without enough data to make a meaningful comparison.
            z_threshold: |z| above this counts as drifting.

        Returns:
            List of dicts sorted by abs(z_score) descending. Each dict:
                model, baseline_mean, baseline_std, baseline_n,
                recent_mean, recent_n, delta, delta_pct, z_score,
                direction, is_drifting, series.
        """
        if metric not in self._PROCESS_DRIFT_METRICS:
            raise ValueError(
                f"Unknown drift metric {metric!r}; "
                f"choose from {list(self._PROCESS_DRIFT_METRICS)}"
            )
        if recent_days >= baseline_days:
            raise ValueError("recent_days must be less than baseline_days")

        column = getattr(DBTrackResult, metric)
        now = datetime.now()
        recent_cutoff = now - timedelta(days=recent_days)
        baseline_start = now - timedelta(days=baseline_days)

        with self.session() as session:
            rows = (
                session.query(
                    DBAnalysisResult.model,
                    column,
                    DBAnalysisResult.file_date,
                )
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(
                    DBAnalysisResult.model.isnot(None),
                    DBAnalysisResult.model != "Unknown",
                    DBAnalysisResult.file_date >= baseline_start,
                    column.isnot(None),
                )
                .order_by(DBAnalysisResult.file_date.asc())
                .all()
            )

        # Bucket into baseline/recent and collect a series for the sparkline.
        per_model: Dict[str, Dict[str, Any]] = {}
        for model, value, file_date in rows:
            if value is None or file_date is None:
                continue
            try:
                value = float(value)
            except (TypeError, ValueError):
                continue
            entry = per_model.setdefault(
                model, {"baseline": [], "recent": [], "series": []}
            )
            entry["series"].append((file_date.isoformat(), value))
            bucket = "recent" if file_date >= recent_cutoff else "baseline"
            entry[bucket].append(value)

        results = []
        for model, buckets in per_model.items():
            base = buckets["baseline"]
            recent = buckets["recent"]
            if len(base) < min_baseline_samples:
                continue
            if len(recent) < min_recent_samples:
                continue
            base_mean = sum(base) / len(base)
            if len(base) > 1:
                var = sum((x - base_mean) ** 2 for x in base) / (len(base) - 1)
                base_std = var ** 0.5
            else:
                base_std = 0.0
            recent_mean = sum(recent) / len(recent)
            delta = recent_mean - base_mean
            z = (delta / base_std) if base_std > 0 else 0.0
            delta_pct = (delta / base_mean * 100.0) if base_mean != 0 else 0.0
            direction = "up" if delta > 0 else ("down" if delta < 0 else "stable")
            is_drifting = abs(z) >= z_threshold and base_std > 0

            # Downsample series to ≤80 points to keep sparkline rendering cheap
            series = buckets["series"]
            if len(series) > 80:
                step = len(series) // 80
                series = series[::step]

            results.append({
                "model": model,
                "baseline_mean": base_mean,
                "baseline_std": base_std,
                "baseline_n": len(base),
                "recent_mean": recent_mean,
                "recent_n": len(recent),
                "delta": delta,
                "delta_pct": delta_pct,
                "z_score": z,
                "direction": direction,
                "is_drifting": is_drifting,
                "series": series,
            })

        results.sort(key=lambda r: abs(r["z_score"]), reverse=True)
        return results

    def get_model_drift_dashboard(
        self,
        model: str,
        days_back: int = 90,
        recent_days: int = 14,
    ) -> Dict[str, Any]:
        """Per-model drift dashboard data: sigma series + 3 process metric
        series with baseline/recent stats. Single round-trip per panel.
        """
        if recent_days >= days_back:
            raise ValueError("recent_days must be less than days_back")

        cutoff = datetime.now() - timedelta(days=days_back)
        recent_cutoff = datetime.now() - timedelta(days=recent_days)

        with self.session() as session:
            rows = (
                session.query(
                    DBAnalysisResult.file_date,
                    DBTrackResult.sigma_gradient,
                    DBTrackResult.untrimmed_resistance,
                    DBTrackResult.measured_electrical_angle,
                    DBTrackResult.trim_pass_count,
                )
                .join(DBTrackResult, DBTrackResult.analysis_id == DBAnalysisResult.id)
                .filter(
                    DBAnalysisResult.model == model,
                    DBAnalysisResult.file_date >= cutoff,
                )
                .order_by(DBAnalysisResult.file_date.asc())
                .all()
            )

        sigma_series: List[Tuple[str, float]] = []
        process_series: Dict[str, List[Tuple[str, float]]] = {
            "untrimmed_resistance": [],
            "measured_electrical_angle": [],
            "trim_pass_count": [],
        }
        process_baseline: Dict[str, List[float]] = {k: [] for k in process_series}
        process_recent: Dict[str, List[float]] = {k: [] for k in process_series}
        retrim_rate_series: List[Tuple[str, int]] = []

        for file_date, sigma, ur, mea, tpc in rows:
            if file_date is None:
                continue
            iso = file_date.isoformat()
            if sigma is not None:
                sigma_series.append((iso, float(sigma)))
            for metric, value in (
                ("untrimmed_resistance", ur),
                ("measured_electrical_angle", mea),
                ("trim_pass_count", tpc),
            ):
                if value is None:
                    continue
                try:
                    value = float(value)
                except (TypeError, ValueError):
                    continue
                process_series[metric].append((iso, value))
                if file_date >= recent_cutoff:
                    process_recent[metric].append(value)
                else:
                    process_baseline[metric].append(value)

            # Retrim rate: one (iso_date, 0_or_1) per row that was actually trimmed.
            # NULL trim_pass_count = pre-feature data (not parsed yet).
            # trim_pass_count == 0 = test-sweep-only file (no laser trim ran).
            # Both are excluded so the rate denominator only counts real trim runs.
            if tpc is not None:
                try:
                    tpc_int = int(tpc)
                except (TypeError, ValueError):
                    tpc_int = None
                if tpc_int is not None and tpc_int > 0:
                    retrim_rate_series.append((iso, 1 if tpc_int > 1 else 0))

        process: Dict[str, Dict[str, Any]] = {}
        for metric in process_series:
            base = process_baseline[metric]
            recent = process_recent[metric]
            base_mean = (sum(base) / len(base)) if base else None
            recent_mean = (sum(recent) / len(recent)) if recent else None
            if base and len(base) > 1:
                var = sum((x - base_mean) ** 2 for x in base) / (len(base) - 1)
                base_std = var ** 0.5
            else:
                base_std = 0.0
            delta = (
                recent_mean - base_mean
                if base_mean is not None and recent_mean is not None
                else None
            )
            z = (delta / base_std) if (delta is not None and base_std > 0) else None
            delta_pct = (
                (delta / base_mean * 100.0)
                if delta is not None and base_mean and base_mean != 0
                else None
            )
            process[metric] = {
                "series": process_series[metric],
                "baseline_mean": base_mean,
                "baseline_n": len(base),
                "recent_mean": recent_mean,
                "recent_n": len(recent),
                "delta": delta,
                "delta_pct": delta_pct,
                "z_score": z,
                "is_drifting": z is not None and abs(z) >= 2.0,
            }

        return {
            "model": model,
            "unit_count": len(rows),
            "baseline_cutoff_date": recent_cutoff.isoformat(),
            "sigma_series": sigma_series,
            "process": {
                **process,
                "retrim_rate_series": retrim_rate_series,
            },
        }

    def get_drift_state_for_models(
        self,
        days_back: int = 30,
        max_series_points: int = 60,
    ) -> Dict[str, Dict[str, Any]]:
        """Per-model drift state from the DB only.

        For each model with sigma data in the window, returns:
            is_drifting, direction, drift_start_date,
            sigma_series (list of (iso_date, sigma) tuples).

        The CUSUM score and threshold live on the in-memory DriftDetector;
        the caller is expected to join those in at render time.
        """
        from laser_trim_analyzer.database.models import ModelMLState

        cutoff = datetime.now() - timedelta(days=days_back)
        with self.session() as session:
            # Pull sigma series for every model in the window
            rows = (
                session.query(
                    DBAnalysisResult.model,
                    DBAnalysisResult.file_date,
                    DBTrackResult.sigma_gradient,
                )
                .join(DBTrackResult, DBTrackResult.analysis_id == DBAnalysisResult.id)
                .filter(
                    DBAnalysisResult.model.isnot(None),
                    DBAnalysisResult.model != "Unknown",
                    DBAnalysisResult.file_date >= cutoff,
                    DBTrackResult.sigma_gradient.isnot(None),
                )
                .order_by(DBAnalysisResult.file_date.asc())
                .all()
            )
            # Pull drift state rows and extract attributes while in session
            ml_state_data = {}
            for s in session.query(ModelMLState).all():
                ml_state_data[s.model] = {
                    "is_drifting": bool(s.is_drifting),
                    "direction": s.drift_direction,
                    "drift_start_date": s.drift_start_date,
                }

        per_model: Dict[str, Dict[str, Any]] = {}
        for model, file_date, sigma in rows:
            if not model or file_date is None or sigma is None:
                continue
            entry = per_model.setdefault(
                model, {"sigma_series": []}
            )
            entry["sigma_series"].append((file_date.isoformat(), float(sigma)))

        result: Dict[str, Dict[str, Any]] = {}
        for model, entry in per_model.items():
            series = entry["sigma_series"]
            # Downsample to keep sparkline rendering cheap
            if len(series) > max_series_points:
                step = len(series) // max_series_points
                series = series[::step]
            state = ml_state_data.get(model)
            result[model] = {
                "model": model,
                "is_drifting": state["is_drifting"] if state else False,
                "direction": state["direction"] if state else None,
                "drift_start_date": state["drift_start_date"] if state else None,
                "sigma_series": series,
            }
        return result

    def get_failure_mode_summary(self, days_back: int = 90) -> List[Dict[str, Any]]:
        """Categorize failures by mode: linearity only, sigma only, or both."""
        with self.session() as session:
            cutoff = datetime.now() - timedelta(days=days_back)
            # Get track-level fail data
            results = session.query(
                DBTrackResult.linearity_pass,
                DBTrackResult.sigma_pass,
                func.count(DBTrackResult.id).label("count"),
            ).join(DBAnalysisResult).filter(
                DBAnalysisResult.file_date >= cutoff,
                or_(
                    DBTrackResult.linearity_pass == False,
                    DBTrackResult.sigma_pass == False,
                ),
            ).group_by(
                DBTrackResult.linearity_pass,
                DBTrackResult.sigma_pass,
            ).all()

            modes = {}
            for r in results:
                lin_fail = r.linearity_pass is False
                sig_fail = r.sigma_pass is False
                if lin_fail and sig_fail:
                    modes["Both Fail"] = modes.get("Both Fail", 0) + r.count
                elif lin_fail:
                    modes["Linearity Fail"] = modes.get("Linearity Fail", 0) + r.count
                elif sig_fail:
                    modes["Sigma Fail"] = modes.get("Sigma Fail", 0) + r.count

            return [{"mode": m, "count": c} for m, c in modes.items() if c > 0]

    def get_spec_discrepancies(self, tolerance_pct: float = 5.0) -> List[Dict[str, Any]]:
        """
        Compare file-parsed linearity specs against model_specs reference.

        Flags models where the spec parsed from trim files differs from the
        engineering reference by more than tolerance_pct percent.

        Returns:
            List of dicts with model, file_spec, reference_spec, difference_pct
        """
        results = []

        with self.session() as session:
            file_specs = (
                session.query(
                    DBAnalysisResult.model,
                    func.avg(DBTrackResult.linearity_spec).label('avg_file_spec'),
                    func.min(DBTrackResult.linearity_spec).label('min_file_spec'),
                    func.max(DBTrackResult.linearity_spec).label('max_file_spec'),
                    func.count(DBTrackResult.id).label('sample_count'),
                )
                .join(DBTrackResult, DBAnalysisResult.id == DBTrackResult.analysis_id)
                .filter(
                    DBTrackResult.linearity_spec.isnot(None),
                    DBTrackResult.linearity_spec > 0,
                )
                .group_by(DBAnalysisResult.model)
                .all()
            )

            for row in file_specs:
                ref = session.query(ModelSpec).filter(
                    ModelSpec.model == row.model
                ).first()

                if not ref or not ref.linearity_spec_pct:
                    continue

                ref_spec = ref.linearity_spec_pct / 100.0  # Convert % to decimal
                file_spec = row.avg_file_spec

                if ref_spec > 0:
                    diff_pct = abs(file_spec - ref_spec) / ref_spec * 100
                else:
                    diff_pct = 0

                if diff_pct > tolerance_pct:
                    results.append({
                        "model": row.model,
                        "file_spec_avg": round(file_spec, 6),
                        "file_spec_min": round(row.min_file_spec, 6),
                        "file_spec_max": round(row.max_file_spec, 6),
                        "reference_spec_pct": ref.linearity_spec_pct,
                        "reference_spec_decimal": round(ref_spec, 6),
                        "difference_pct": round(diff_pct, 1),
                        "sample_count": row.sample_count,
                        "linearity_type": ref.linearity_type,
                    })

        results.sort(key=lambda x: x["difference_pct"], reverse=True)
        return results


# Global instance for convenience
_db_manager: Optional[DatabaseManager] = None


def get_database() -> DatabaseManager:
    """Get the global database manager instance."""
    global _db_manager
    if _db_manager is None:
        _db_manager = DatabaseManager()
    return _db_manager


def reset_database() -> None:
    """Reset the global database manager (for testing)."""
    global _db_manager
    if _db_manager:
        _db_manager.close()
    _db_manager = None
