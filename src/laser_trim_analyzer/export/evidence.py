"""Spec 3c — Evidence export. The daily 'what I hand engineers' payoff (foundations Q8)."""
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from laser_trim_analyzer.ml.drift_types import WATCHED_METRICS, metric_label

# "Recent" window length (days of DATA, anchored to the model's latest file_date).
RECENT_DAYS = 30


# Shared with the drift engine (single source): values beyond this many
# baseline-σ are suspect data, not process signal. Evidence displays EXCLUDE
# and disclose them; the detector winsorizes them (drift_training).
from laser_trim_analyzer.ml.drift_types import SUSPECT_SIGMA_GATE  # noqa: E402,F401


def compute_recent_means(db, model: str, recent_days: int = RECENT_DAYS,
                         with_meta: bool = False):
    """LAST CLOSED LOT median per watched metric (lot mode, 2026-07-13).

    'Recent' used to be an outlier-gated mean over the newest days of units;
    with the detector observing LOTS, every surface (pills, drift table,
    triage shift, verdict, Excel pack) must compare baseline vs the SAME
    observation the detector saw — the last closed lot's median. Returns
    metric -> float|None; with_meta=True returns (means, meta) where
    meta[metric] = {"n": lot size, "excluded": None, "lot_end": date} —
    'excluded' is retained for column compatibility (medians need no gate).
    `recent_days` is accepted for API compatibility and unused.
    """
    from laser_trim_analyzer.ml.lots import get_model_lots
    from laser_trim_analyzer.ml.drift_types import WATCHED_METRICS

    out: dict = {}
    meta: dict = {}
    floor = None
    try:
        req = db.get_baseline_requalification(model)
        if req:
            floor = datetime.fromisoformat(str(req[0])[:19])
    except Exception:
        pass
    for metric in WATCHED_METRICS:
        try:
            lots = [l for l in get_model_lots(db, model, metric, after=floor)
                    if not l.is_open()]
        except Exception:
            lots = []
        if lots:
            last = lots[-1]
            out[metric] = last.median
            meta[metric] = {"n": last.n, "excluded": None,
                            "lot_end": last.end}
        else:
            out[metric] = None
            meta[metric] = {"n": 0, "excluded": None, "lot_end": None}
    return (out, meta) if with_meta else out

def _sigma_shift(ms, recent_val) -> Optional[float]:
    """Honest (recent − baseline) / baseline_std — the SAME number the Triage
    cards and Model-page pills show. MetricStatus.magnitude is alert-type
    scaled (step: √N·shift/σ; slow drift: cusum/h) and is NOT comparable
    across metrics; labeling it 'Δσ' made exports contradict the screen."""
    if recent_val is None or ms.baseline_std is None or ms.baseline_std <= 0:
        return None
    return (recent_val - ms.baseline_mean) / ms.baseline_std


def build_summary_text(model: str, status, recent_means: Optional[dict] = None,
                       recent_meta: Optional[dict] = None) -> str:
    """Paste-ready text: model + per-metric baseline-vs-recent + alert. Q8 traceable.

    `recent_means` (metric -> float|None) supplies the data-derived recent values; the
    hydrated detector's own recent_mean is always None, so without this the summary
    reads 'recent n/a' for every metric.
    """
    recent_means = recent_means or {}
    lines = [f"Drift summary — model {model}",
             f"Overall: {status.overall_tier.name.replace('_', ' ').title()}"
             + (f" (worst: {metric_label(status.worst_metric)})" if status.worst_metric else ""), ""]
    for m in WATCHED_METRICS:
        ms = status.per_metric.get(m)
        if ms is None:
            continue
        recent_val = recent_means.get(m) if recent_means.get(m) is not None else ms.recent_mean
        recent = f"{recent_val:.4g}" if recent_val is not None else "n/a"
        tier = ms.tier.name.replace("_", " ").title()
        shift = _sigma_shift(ms, recent_val)
        shift_txt = f"shift {shift:+.2f}σ" if shift is not None else "shift n/a"
        lot_n = (recent_meta or {}).get(m, {}).get("n") if recent_meta else None
        lot_txt = f" (lot of {lot_n})" if lot_n else ""
        lines.append(f"- {metric_label(m)}: baseline lots {ms.baseline_mean:.4g} ± {ms.baseline_std:.4g}, "
                     f"last lot {recent}{lot_txt}, {shift_txt} [{tier}]")
    return "\n".join(lines)


def export_evidence_pack(db, model: str, out_path, window_days: Optional[int] = None) -> Path:
    """Write a traceable evidence workbook for `model`.

    Sheets:
      * Drift evidence — per-metric baseline vs recent, honest σ-shift, tier
      * Unit history  — one row per unit/track over the FULL record (default;
        pass window_days to restrict), every watched metric + pass flags.
        This is the 'export the model to Excel and read its whole history'
        workflow — judging whether the process is moving for better or worse.
      * Monthly summary — units, pass rate, and metric means per month
    """
    import pandas as pd
    from sqlalchemy import func, case
    from laser_trim_analyzer.database.models import (
        AnalysisResult as DBAR, TrackResult as DBTR, StatusType)
    from laser_trim_analyzer.ml.manager import get_model_drift_status

    status = get_model_drift_status(db, model)
    recent_means, recent_meta = compute_recent_means(db, model, with_meta=True)
    try:
        _req = db.get_baseline_requalification(model)
    except Exception:
        _req = None
    baseline_since = str(_req[0])[:10] if _req else "full history"
    metric_rows = []
    for m in WATCHED_METRICS:
        ms = status.per_metric.get(m)
        if ms is None:
            # An absent row was ambiguous between "fine" and "never checked"
            # (work convo 2026-07-10) — say NOT TRAINED and still show the
            # recent mean, which needs no baseline.
            metric_rows.append({"Metric": metric_label(m), "Tier": "NOT TRAINED",
                                "Baseline since": baseline_since,
                                "Alert": "", "Baseline mean": None,
                                "Baseline std": None,
                                "Last lot median": recent_means.get(m),
                                "Lot size": recent_meta.get(m, {}).get("n"),
                                "Suspect excluded": None, "Shift (σ)": None,
                                "Alert magnitude (scaled)": None})
            continue
        recent_val = recent_means.get(m) if recent_means.get(m) is not None else ms.recent_mean
        mm = recent_meta.get(m, {})
        metric_rows.append({"Metric": metric_label(m), "Tier": ms.tier.name,
                            "Baseline since": baseline_since,
                            "Alert": ms.alert_type.value if ms.alert_type else "",
                            "Baseline mean": ms.baseline_mean, "Baseline std": ms.baseline_std,
                            "Last lot median": recent_val,
                            "Lot size": mm.get("n"),
                            # Traceability: how many recent values were gated
                            # out as suspect (scale anomalies) before the mean.
                            "Suspect excluded": mm.get("excluded"),
                            # Matches the UI (Triage cards / pills). The old
                            # 'Delta_sigma' column exported alert-scaled
                            # magnitude, which contradicted the screen.
                            "Shift (σ)": _sigma_shift(ms, recent_val),
                            "Alert magnitude (scaled)": ms.magnitude})

    cutoff = datetime.now() - timedelta(days=window_days) if window_days else None
    with db.session() as s:
        # Final-test outcome per trim unit (the other half of the trim-vs-FT
        # workflow — "did this unit pass at final test"; user request 2026-07-08).
        from laser_trim_analyzer.database.models import FinalTestResult as DBFT
        ft_by_trim: dict = {}
        for ft in (s.query(DBFT.linked_trim_id, DBFT.overall_status,
                           DBFT.file_date, DBFT.match_confidence)
                   .filter(DBFT.model == model, DBFT.linked_trim_id.isnot(None)).all()):
            # Keep the newest FT per trim record.
            prev = ft_by_trim.get(ft[0])
            # NULL date sorts OLDEST: a dated FT record always displaces a
            # dateless one (work convo 2026-07-10 — the old comparison kept
            # the stale record whenever prev had no date).
            if (prev is None or prev[1] is None
                    or (ft[2] is not None and ft[2] > prev[1])):
                ft_by_trim[ft[0]] = (getattr(ft[1], "value", str(ft[1])), ft[2], ft[3])

        # OUTER join (work convo 2026-07-10): an inner join silently dropped
        # every analysis with no track rows (ERROR files) from the history —
        # Monthly counted them, Unit history didn't, and the mismatch read as
        # missing data. ERROR rows now appear with blank metrics.
        q = (s.query(DBAR.id, DBAR.serial, DBAR.file_date, DBAR.overall_status, DBAR.system,
                     DBTR.track_id, DBTR.sigma_gradient, DBTR.untrimmed_sigma_gradient,
                     DBTR.final_linearity_error_shifted, DBTR.untrimmed_error_max,
                     DBTR.untrimmed_resistance, DBTR.resistance_change_percent,
                     DBTR.measured_electrical_angle, DBTR.trim_pass_count,
                     DBTR.composite_trim_risk_score, DBTR.sigma_pass, DBTR.linearity_pass,
                     DBTR.trimmed_resistance, DBTR.sigma_threshold, DBTR.linearity_spec,
                     DBTR.linearity_fail_points, DBTR.optimal_offset,
                     DBTR.trim_improvement_percent, DBAR.filename)
             .outerjoin(DBTR, DBTR.analysis_id == DBAR.id).filter(DBAR.model == model))
        if cutoff is not None:
            q = q.filter(DBAR.file_date >= cutoff)
        def _d(v):
            # Day-granularity data: '2026-05-27 00:00:00' in a cell is noise
            # (work finding #6, 2026-07-10).
            return v.strftime("%Y-%m-%d") if v is not None else None

        unit_rows = []
        for r in q.order_by(DBAR.file_date.desc()).all():
            ft = ft_by_trim.get(r[0])
            unit_rows.append({
                "Serial": r[1], "Date": _d(r[2]),
                "Status": getattr(r[3], "value", str(r[3])),
                "System": getattr(r[4], "value", str(r[4])),
                "Track": r[5],
                "Sigma gradient": r[6], "Untrimmed sigma": r[7],
                "Linearity error": r[8], "Untrimmed error max": r[9],
                "Untrimmed resistance": r[10],
                # Side-by-side untrimmed vs trimmed (user question 2026-07-10:
                # "no way to look at the untrimmed vs trimmed resistance").
                "Trimmed resistance": r[17],
                "Resistance change %": r[11],
                "Electrical angle": r[12], "Trim passes": r[13],
                "Composite risk": r[14],
                "Sigma pass": r[15], "Linearity pass": r[16],
                "Sigma threshold": r[18], "Linearity spec": r[19],
                "Fail points": r[20], "Optimal offset": r[21],
                "Trim improvement %": r[22], "Filename": r[23],
                "FT result": ft[0] if ft else None,
                "FT date": _d(ft[1]) if ft else None,
                "FT match %": (round(ft[2] * 100) if ft and ft[2] is not None
                               else None)})

        # Final-test units sheet (work findings #3/#8, 2026-07-10): every FT
        # record for the model — the "did it pass at the last station" list.
        ft_rows_sheet = []
        ftq = s.query(DBFT.serial, DBFT.file_date, DBFT.overall_status,
                      DBFT.linked_trim_id, DBFT.match_confidence)\
               .filter(DBFT.model == model)
        if cutoff is not None:
            ftq = ftq.filter(DBFT.file_date >= cutoff)
        for r in ftq.order_by(DBFT.file_date.desc()).all():
            ft_rows_sheet.append({
                "Serial": r[0], "Test date": _d(r[1]),
                "Result": getattr(r[2], "value", str(r[2])),
                "Linked to trim record": "yes" if r[3] is not None else "no",
                "Match %": (round(r[4] * 100) if r[4] is not None else None)})

        # Output smoothness sheet (work finding #16): the model's smoothness
        # records with spec and verdict.
        smooth_rows = []
        try:
            from laser_trim_analyzer.database.models import SmoothnessResult as DBSR
            sq = s.query(DBSR.serial, DBSR.file_date, DBSR.max_smoothness_value,
                         DBSR.smoothness_spec, DBSR.smoothness_pass)\
                  .filter(DBSR.model == model)
            if cutoff is not None:
                sq = sq.filter(DBSR.file_date >= cutoff)
            for r in sq.order_by(DBSR.file_date.desc()).all():
                smooth_rows.append({
                    "Serial": r[0], "Date": _d(r[1]),
                    "Max smoothness": r[2], "Spec": r[3],
                    "Pass": r[4]})
        except Exception:
            pass

        # Monthly rollup: UNIT-level counts/pass rate (overall_status is the
        # per-unit verdict; UNTRIMMED test-sweeps excluded from the rate), plus
        # track-metric means for the better-or-worse read.
        month = func.strftime("%Y-%m", DBAR.file_date)
        unit_q = (s.query(
                    month.label("month"),
                    func.count(DBAR.id).label("units"),
                    func.sum(case((DBAR.overall_status == StatusType.PASS, 1), else_=0)).label("passed"),
                    func.sum(case((DBAR.overall_status == StatusType.FAIL, 1), else_=0)).label("failed"),
                    func.sum(case((DBAR.overall_status == StatusType.WARNING, 1), else_=0)).label("warned"),
                  )
                  .filter(DBAR.model == model,
                          DBAR.overall_status != StatusType.UNTRIMMED)
                  .group_by(month).order_by(month).all())
        mean_q = (s.query(
                    month.label("month"),
                    func.avg(DBTR.sigma_gradient), func.avg(DBTR.untrimmed_sigma_gradient),
                    func.avg(DBTR.final_linearity_error_shifted), func.avg(DBTR.untrimmed_error_max),
                    func.avg(DBTR.resistance_change_percent), func.avg(DBTR.measured_electrical_angle),
                    func.avg(DBTR.untrimmed_resistance), func.avg(DBTR.trimmed_resistance),
                  )
                  .join(DBTR, DBTR.analysis_id == DBAR.id)
                  .filter(DBAR.model == model)
                  .group_by(month).all())
        means_by_month = {r[0]: r[1:] for r in mean_q}

        # Smoothness rollup per month (user question 2026-07-10: "no way to
        # tell the average output smoothness / if we struggle with it").
        smooth_by_month: dict = {}
        try:
            smonth = func.strftime("%Y-%m", DBSR.file_date)
            sm_q = (s.query(smonth.label("month"),
                            func.count(DBSR.id),
                            func.avg(DBSR.max_smoothness_value),
                            func.avg(case((DBSR.smoothness_pass == True, 1.0), else_=0.0)))
                    .filter(DBSR.model == model).group_by(smonth).all())
            smooth_by_month = {r[0]: (r[1], r[2],
                                      (r[3] * 100.0 if r[3] is not None else None))
                               for r in sm_q if r[0]}
        except Exception:
            pass

        # UNIT-cohort stats per month (QA audit 2026-07-13: first-pass vs
        # final vs rework — the attempt-basis sheet alone hid rework cost).
        try:
            from laser_trim_analyzer.core.yield_stats import compute_unit_yield_monthly
            unit_cohorts = compute_unit_yield_monthly(db, model)
        except Exception:
            unit_cohorts = {}

        # Iterate the UNION of months — a month of pure test sweeps or
        # smoothness-only tests used to vanish from the sheet entirely.
        units_by_month = {r.month: r for r in unit_q if r.month}
        all_months = sorted(set(units_by_month) | set(means_by_month)
                            | set(smooth_by_month) | set(unit_cohorts))
        monthly_rows = []
        for mo in all_months:
            r = units_by_month.get(mo)
            mm = means_by_month.get(mo, (None,) * 8)
            sm = smooth_by_month.get(mo, (None, None, None))
            uc = unit_cohorts.get(mo, {})
            units = (r.units or 0) if r else 0
            passed = (r.passed or 0) if r else 0
            warned = (r.warned or 0) if r else 0
            failed = (r.failed or 0) if r else 0
            accepted = passed + warned
            monthly_rows.append({
                "Month": mo, "Units": units,
                "Passed": passed, "Watch (sigma)": warned,
                "Failed": failed,
                # Customer basis: linearity is the zero-tolerance requirement;
                # Watch units passed it (sigma = internal drift flag only).
                "Linearity yield %": (accepted / units * 100) if units else None,
                "Clean pass %": (passed / units * 100) if units else None,
                "Mean sigma gradient": mm[0], "Mean untrimmed sigma": mm[1],
                "Mean linearity error": mm[2], "Mean untrimmed error max": mm[3],
                "Mean resistance change %": mm[4], "Mean electrical angle": mm[5],
                "Mean untrimmed resistance": mm[6], "Mean trimmed resistance": mm[7],
                "Smoothness tests": sm[0], "Mean max smoothness": sm[1],
                "Smoothness pass %": sm[2],
                # Unit cohort (units whose FIRST attempt fell in this month;
                # unit = shop # + day, sections rolled up zero-tolerance).
                "Units started (distinct)": uc.get("units"),
                "First-pass yield %": uc.get("first_pass_yield"),
                "Final yield %": uc.get("final_yield"),
                "Trims per section": uc.get("attempts_per_section"),
                "Units reworked": uc.get("rework_units"),
            })

    out_path = Path(out_path)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xl:
        pd.DataFrame(metric_rows).to_excel(xl, sheet_name="Drift evidence", index=False)
        udf = pd.DataFrame(unit_rows)
        udf.to_excel(xl, sheet_name="Unit history", index=False)
        _ft_cols = ["Serial", "Test date", "Result", "Linked to trim record", "Match %"]
        pd.DataFrame(ft_rows_sheet, columns=(None if ft_rows_sheet else _ft_cols))\
            .to_excel(xl, sheet_name="Final test units", index=False)
        _sm_cols = ["Serial", "Date", "Max smoothness", "Spec", "Pass"]
        pd.DataFrame(smooth_rows, columns=(None if smooth_rows else _sm_cols))\
            .to_excel(xl, sheet_name="Smoothness", index=False)
        pd.DataFrame(monthly_rows).to_excel(xl, sheet_name="Monthly summary", index=False)

        # Color the disposition columns (work finding #7): linearity is the
        # zero-tolerance customer rule — red/green beats True/False on a page
        # of 2,000 rows. Status + FT result get the same treatment.
        try:
            from openpyxl.styles import PatternFill
            red = PatternFill("solid", start_color="FFC7CE")
            green = PatternFill("solid", start_color="C6EFCE")
            amber = PatternFill("solid", start_color="FFEB9C")
            ws = xl.book["Unit history"]
            cols = {c: i + 1 for i, c in enumerate(udf.columns)}
            for row in range(2, ws.max_row + 1):
                lp = ws.cell(row=row, column=cols["Linearity pass"])
                if lp.value is True or lp.value == 1:
                    lp.fill = green
                elif lp.value is False or lp.value == 0:
                    lp.fill = red
                st = ws.cell(row=row, column=cols["Status"])
                if st.value == "FAIL":
                    st.fill = red
                elif st.value == "WARNING":
                    st.fill = amber
                elif st.value == "PASS":
                    st.fill = green
                if "FT result" in cols:
                    ftc = ws.cell(row=row, column=cols["FT result"])
                    if ftc.value == "FAIL":
                        ftc.fill = red
                    elif ftc.value == "PASS":
                        ftc.fill = green
        except Exception:
            pass   # styling is best-effort; the data itself is already written
    return out_path
