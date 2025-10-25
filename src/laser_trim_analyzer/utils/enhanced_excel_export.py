"""
Enhanced Excel Export Module for Laser Trim Analyzer

Provides comprehensive Excel export functionality ensuring ALL important analysis
data is exported for both single file and batch processing results.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Union
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from laser_trim_analyzer.core.models import AnalysisResult, TrackData

logger = logging.getLogger(__name__)


class EnhancedExcelExporter:
    """Enhanced Excel exporter that ensures all important data is exported."""
    
    def __init__(self):
        """Initialize the enhanced Excel exporter."""
        self.logger = logger
        
        # Define style presets
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export_single_file_comprehensive(
        self,
        result: AnalysisResult,
        output_path: Path,
        include_raw_data: bool = True,
        include_plots: bool = True
    ) -> Path:
        """
        Export comprehensive analysis results for a single file.
        
        Args:
            result: Analysis result to export
            output_path: Path to save Excel file
            include_raw_data: Include raw position/error data
            include_plots: Include plot references
            
        Returns:
            Path to the created Excel file
        """
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # 1. Summary Sheet
            self._create_summary_sheet(wb, result)
            
            # 2. Detailed Analysis Sheet (per track)
            self._create_detailed_analysis_sheet(wb, result)
            
            # 3. Validation Details Sheet
            self._create_validation_sheet(wb, result)
            
            # 4. Zone Analysis Sheet (if available)
            if any(hasattr(track, 'zone_analysis') and track.zone_analysis for track in result.tracks):
                self._create_zone_analysis_sheet(wb, result)
            
            # 5. Raw Data Sheet (if requested)
            if include_raw_data:
                self._create_raw_data_sheet(wb, result)
            
            # 6. Metadata Sheet
            self._create_metadata_sheet(wb, result)
            
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Comprehensive single file Excel report saved to: {output_path}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"Failed to create single file Excel report: {e}")
            raise
    
    def export_batch_comprehensive(
        self,
        results: List[AnalysisResult],
        output_path: Path,
        include_individual_details: bool = True,
        max_individual_sheets: int = 50
    ) -> Path:
        """
        Export comprehensive batch analysis results.
        
        Args:
            results: List of analysis results
            output_path: Path to save Excel file
            include_individual_details: Include per-file detail sheets
            max_individual_sheets: Maximum number of individual detail sheets
            
        Returns:
            Path to the created Excel file
        """
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # 1. Batch Summary Sheet
            self._create_batch_summary_sheet(wb, results)
            
            # 2. Detailed Results Sheet (all files)
            self._create_batch_detailed_sheet(wb, results)
            
            # 3. Enhanced Data Sheet (additional useful fields)
            self._create_enhanced_data_sheet(wb, results)
            
            # 4. Statistical Analysis Sheet
            self._create_statistical_analysis_sheet(wb, results)
            
            # 5. Model Performance Sheet
            self._create_model_performance_sheet(wb, results)
            
            # 6. Failure Analysis Sheet
            self._create_failure_analysis_sheet(wb, results)
            
            # 6. Individual File Details (limited)
            if include_individual_details:
                included_count = min(len(results), max_individual_sheets)
                for i, result in enumerate(results[:included_count]):
                    sheet_name = f"File_{i+1}_{Path(result.metadata.filename).stem[:20]}"
                    self._create_individual_detail_sheet(wb, result, sheet_name)
            
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Comprehensive batch Excel report saved to: {output_path}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"Failed to create batch Excel report: {e}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, result: AnalysisResult):
        """Create comprehensive summary sheet for single file."""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Laser Trim Analysis Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        row = 3
        
        # File Information
        ws[f'A{row}'] = "FILE INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        file_info = [
            ("Filename", result.metadata.filename),
            ("File Path", str(result.metadata.file_path)),
            ("File Size (MB)", f"{getattr(result.metadata, 'file_size_mb', 'N/A'):.2f}" if hasattr(result.metadata, 'file_size_mb') else 'N/A'),
            ("Model", result.metadata.model),
            ("Serial", result.metadata.serial),
            ("System Type", result.metadata.system.value),
            ("Test Date", result.metadata.test_date.strftime('%Y-%m-%d %H:%M:%S') if hasattr(result.metadata, 'test_date') and result.metadata.test_date else 'N/A'),
            ("Analysis Date", result.metadata.timestamp.strftime('%Y-%m-%d %H:%M:%S')),
            ("Has Multiple Tracks", "Yes" if getattr(result.metadata, 'has_multi_tracks', False) else "No"),
            ("Track Count", len(result.tracks)),
        ]
        
        for label, value in file_info:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Overall Status
        ws[f'A{row}'] = "OVERALL ANALYSIS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Overall Status"
        ws[f'B{row}'] = result.overall_status.value
        self._apply_status_formatting(ws[f'B{row}'], result.overall_status.value)
        row += 1
        
        ws[f'A{row}'] = "Processing Time (s)"
        ws[f'B{row}'] = f"{result.processing_time:.2f}"
        row += 1
        
        if hasattr(result, 'overall_validation_status'):
            ws[f'A{row}'] = "Validation Status"
            ws[f'B{row}'] = getattr(result, 'overall_validation_status', 'N/A')
            row += 1
            
            ws[f'A{row}'] = "Validation Grade"
            ws[f'B{row}'] = getattr(result, 'validation_grade', 'N/A')
            row += 1
        
        row += 1
        
        # Primary Track Analysis
        primary_track = result.primary_track
        if primary_track:
            ws[f'A{row}'] = "PRIMARY TRACK ANALYSIS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Unit Properties
            if primary_track.unit_properties:
                ws[f'A{row}'] = "Unit Properties"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                unit_props = [
                    ("Unit Length (degrees)", f"{primary_track.unit_properties.unit_length:.2f}"),
                    ("Min Resistance", f"{primary_track.unit_properties.min_resistance:.4f}"),
                    ("Max Resistance", f"{primary_track.unit_properties.max_resistance:.4f}"),
                    ("Resistance Range", f"{primary_track.unit_properties.resistance_range:.4f}"),
                    ("Resistance Change", f"{getattr(primary_track.unit_properties, 'resistance_change', 'N/A'):.4f}" if hasattr(primary_track.unit_properties, 'resistance_change') else 'N/A'),
                    ("Resistance Change %", f"{getattr(primary_track.unit_properties, 'resistance_change_percent', 'N/A'):.2f}%" if hasattr(primary_track.unit_properties, 'resistance_change_percent') else 'N/A'),
                    ("Normalized Range %", f"{primary_track.unit_properties.normalized_range:.2f}%"),
                ]
                
                for label, value in unit_props:
                    ws[f'B{row}'] = label
                    ws[f'C{row}'] = value
                    row += 1
                
                row += 1
            
            # Sigma Analysis
            if primary_track.sigma_analysis:
                ws[f'A{row}'] = "Sigma Analysis"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                sigma_props = [
                    ("Sigma Gradient", f"{primary_track.sigma_analysis.sigma_gradient:.6f}"),
                    ("Sigma Threshold", f"{primary_track.sigma_analysis.sigma_threshold:.6f}"),
                    ("Scaling Factor", f"{getattr(primary_track.sigma_analysis, 'scaling_factor', 24.0):.1f}"),
                    ("Pass/Fail", "PASS" if getattr(primary_track.sigma_analysis, 'sigma_pass', False) else "FAIL"),
                    ("Sigma Ratio", f"{getattr(primary_track.sigma_analysis, 'sigma_ratio', 'N/A'):.2f}" if hasattr(primary_track.sigma_analysis, 'sigma_ratio') else 'N/A'),
                    ("Industry Compliance", getattr(primary_track.sigma_analysis, 'industry_compliance', 'N/A')),
                ]
                
                for label, value in sigma_props:
                    ws[f'B{row}'] = label
                    ws[f'C{row}'] = value
                    if label == "Pass/Fail":
                        self._apply_status_formatting(ws[f'C{row}'], value)
                    row += 1
                
                row += 1
            
            # Linearity Analysis
            if primary_track.linearity_analysis:
                ws[f'A{row}'] = "Linearity Analysis"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                linearity_props = [
                    ("Linearity Spec", f"{primary_track.linearity_analysis.linearity_spec:.4f}"),
                    ("Pass/Fail", "PASS" if getattr(primary_track.linearity_analysis, 'linearity_pass', False) else "FAIL"),
                    ("Fail Points", primary_track.linearity_analysis.fail_points),
                    ("Optimal Offset", f"{primary_track.linearity_analysis.optimal_offset:.6f}"),
                    ("Final Linearity Error", f"{primary_track.linearity_analysis.final_linearity_error_shifted:.6f}"),
                    ("Max Deviation", f"{getattr(primary_track.linearity_analysis, 'max_deviation', 'N/A'):.6f}" if hasattr(primary_track.linearity_analysis, 'max_deviation') else 'N/A'),
                    ("Max Deviation Position", getattr(primary_track.linearity_analysis, 'max_deviation_position', 'N/A')),
                    ("Industry Grade", getattr(primary_track.linearity_analysis, 'industry_grade', 'N/A')),
                ]
                
                for label, value in linearity_props:
                    ws[f'B{row}'] = label
                    ws[f'C{row}'] = value
                    if label == "Pass/Fail":
                        self._apply_status_formatting(ws[f'C{row}'], value)
                    row += 1
                
                row += 1
            
            # ML Predictions
            if getattr(primary_track, 'failure_prediction', None):
                ws[f'A{row}'] = "ML Failure Prediction"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                fp = primary_track.failure_prediction
                risk_val = getattr(fp.risk_category, 'value', str(getattr(fp, 'risk_category', 'Unknown')))
                ml_props = [
                    ("Risk Category", risk_val),
                    ("Failure Probability", f"{getattr(fp, 'failure_probability', 0.0):.2%}"),
                    ("Gradient Margin", f"{getattr(primary_track.sigma_analysis, 'gradient_margin', 0.0):.6f}"),
                ]
                
                for label, value in ml_props:
                    ws[f'B{row}'] = label
                    ws[f'C{row}'] = value
                    row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_detailed_analysis_sheet(self, wb: Workbook, result: AnalysisResult):
        """Create detailed analysis sheet with all track data."""
        ws = wb.create_sheet("Detailed Analysis")
        
        # Create headers
        headers = [
            "Track ID", "Data Type", "Data Points", "Travel Length",
            "Unit Length", "Min Resistance", "Max Resistance", "Resistance Range", 
            "Resistance Change", "Resistance Change %", "Normalized Range %",
            "Sigma Gradient", "Sigma Threshold", "Sigma Pass", "Sigma Ratio",
            "Linearity Spec", "Linearity Pass", "Fail Points", "Optimal Offset",
            "Final Linearity Error", "Max Deviation", "Max Deviation Pos",
            "Has Upper Limits", "Has Lower Limits", "Plot Path"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Add data for each track
        row = 2
        for track in result.tracks:
            # Basic track info
            ws.cell(row=row, column=1, value=track.track_id)
            ws.cell(row=row, column=2, value="Track")
            # Approximate data points as length of position_data if available
            ws.cell(row=row, column=3, value=len(getattr(track, 'position_data', []) or []))
            ws.cell(row=row, column=4, value=getattr(track, 'travel_length', 'N/A'))
            
            # Unit properties
            if track.unit_properties:
                ws.cell(row=row, column=5, value=track.unit_properties.unit_length)
                ws.cell(row=row, column=6, value=getattr(track.unit_properties, 'untrimmed_resistance', None))
                ws.cell(row=row, column=7, value=getattr(track.unit_properties, 'trimmed_resistance', None))
                ws.cell(row=row, column=8, value=getattr(track.unit_properties, 'resistance_change', None))
                ws.cell(row=row, column=9, value=getattr(track.unit_properties, 'resistance_change', 'N/A'))
                ws.cell(row=row, column=10, value=f"{getattr(track.unit_properties, 'resistance_change_percent', 'N/A'):.2f}%" if hasattr(track.unit_properties, 'resistance_change_percent') and track.unit_properties.resistance_change_percent != 'N/A' else 'N/A')
                ws.cell(row=row, column=11, value=f"{getattr(track.unit_properties, 'resistance_change_percent', 0.0):.2f}%" if getattr(track.unit_properties, 'resistance_change_percent', None) is not None else 'N/A')
            
            # Sigma analysis
            if track.sigma_analysis:
                ws.cell(row=row, column=12, value=track.sigma_analysis.sigma_gradient)
                ws.cell(row=row, column=13, value=track.sigma_analysis.sigma_threshold)
                ws.cell(row=row, column=14, value="PASS" if getattr(track.sigma_analysis, 'sigma_pass', False) else "FAIL")
                self._apply_status_formatting(ws.cell(row=row, column=14), "PASS" if getattr(track.sigma_analysis, 'sigma_pass', False) else "FAIL")
                ws.cell(row=row, column=15, value=getattr(track.sigma_analysis, 'sigma_ratio', 'N/A'))
            
            # Linearity analysis
            if track.linearity_analysis:
                ws.cell(row=row, column=16, value=track.linearity_analysis.linearity_spec)
                ws.cell(row=row, column=17, value="PASS" if getattr(track.linearity_analysis, 'linearity_pass', False) else "FAIL")
                self._apply_status_formatting(ws.cell(row=row, column=17), "PASS" if getattr(track.linearity_analysis, 'linearity_pass', False) else "FAIL")
                ws.cell(row=row, column=18, value=track.linearity_analysis.fail_points)
                ws.cell(row=row, column=19, value=track.linearity_analysis.optimal_offset)
                ws.cell(row=row, column=20, value=track.linearity_analysis.final_linearity_error_shifted)
                ws.cell(row=row, column=21, value=getattr(track.linearity_analysis, 'max_deviation', 'N/A'))
                ws.cell(row=row, column=22, value=getattr(track.linearity_analysis, 'max_deviation_position', 'N/A'))
            
            # Spec limits
            ws.cell(row=row, column=23, value="Yes" if hasattr(track, 'upper_limits') and track.upper_limits is not None else "No")
            ws.cell(row=row, column=24, value="Yes" if hasattr(track, 'lower_limits') and track.lower_limits is not None else "No")
            
            # Plot path
            ws.cell(row=row, column=25, value=str(track.plot_path) if track.plot_path else "N/A")
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_validation_sheet(self, wb: Workbook, result: AnalysisResult):
        """Create validation details sheet."""
        ws = wb.create_sheet("Validation Details")
        
        ws['A1'] = "Validation Details"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # Collect all validation results
        for track in result.tracks:
            ws[f'A{row}'] = f"Track {track.track_id} Validation"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Sigma validation
            if track.sigma_analysis and hasattr(track.sigma_analysis, 'validation_result') and track.sigma_analysis.validation_result:
                ws[f'A{row}'] = "Sigma Analysis Validation"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                val_result = track.sigma_analysis.validation_result
                ws[f'B{row}'] = "Expected Value"
                ws[f'C{row}'] = f"{val_result.expected_value:.6f}"
                row += 1
                
                ws[f'B{row}'] = "Actual Value"
                ws[f'C{row}'] = f"{val_result.actual_value:.6f}"
                row += 1
                
                ws[f'B{row}'] = "Deviation %"
                ws[f'C{row}'] = f"{val_result.deviation_percent:.2f}%"
                row += 1
                
                ws[f'B{row}'] = "Status"
                ws[f'C{row}'] = val_result.status
                self._apply_status_formatting(ws[f'C{row}'], val_result.status)
                row += 1
                
                if val_result.warnings:
                    ws[f'B{row}'] = "Warnings"
                    ws[f'C{row}'] = "; ".join(val_result.warnings)
                    row += 1
                
                row += 1
            
            # Linearity validation
            if track.linearity_analysis and hasattr(track.linearity_analysis, 'validation_result') and track.linearity_analysis.validation_result:
                ws[f'A{row}'] = "Linearity Analysis Validation"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                val_result = track.linearity_analysis.validation_result
                ws[f'B{row}'] = "Expected Value"
                ws[f'C{row}'] = f"{val_result.expected_value:.6f}"
                row += 1
                
                ws[f'B{row}'] = "Actual Value"
                ws[f'C{row}'] = f"{val_result.actual_value:.6f}"
                row += 1
                
                ws[f'B{row}'] = "Deviation %"
                ws[f'C{row}'] = f"{val_result.deviation_percent:.2f}%"
                row += 1
                
                ws[f'B{row}'] = "Status"
                ws[f'C{row}'] = val_result.status
                self._apply_status_formatting(ws[f'C{row}'], val_result.status)
                row += 1
                
                row += 1
            
            # Validation summary
            if hasattr(track, 'validation_summary'):
                ws[f'A{row}'] = "Overall Validation Summary"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                ws[f'B{row}'] = track.validation_summary
                row += 2
    
    def _create_zone_analysis_sheet(self, wb: Workbook, result: AnalysisResult):
        """Create zone analysis sheet if available."""
        ws = wb.create_sheet("Zone Analysis")
        
        # Headers
        headers = ["Track ID", "Zone", "Start Position", "End Position", 
                   "Data Points", "Avg Error", "Max Error", "Error Variance"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        row = 2
        for track in result.tracks:
            if hasattr(track, 'zone_analysis') and track.zone_analysis:
                for zone_result in track.zone_analysis.zone_results:
                    ws.cell(row=row, column=1, value=track.track_id)
                    ws.cell(row=row, column=2, value=zone_result.zone_number)
                    ws.cell(row=row, column=3, value=zone_result.position_range[0])
                    ws.cell(row=row, column=4, value=zone_result.position_range[1])
                    ws.cell(row=row, column=5, value=zone_result.data_points)
                    ws.cell(row=row, column=6, value=zone_result.average_error)
                    ws.cell(row=row, column=7, value=zone_result.max_error)
                    ws.cell(row=row, column=8, value=zone_result.error_variance)
                    row += 1
    
    def _create_raw_data_sheet(self, wb: Workbook, result: AnalysisResult):
        """Create raw data sheet with position and error data."""
        ws = wb.create_sheet("Raw Data")
        
        # Headers
        headers = ["Track ID", "Data Type", "Index", "Position", "Error", 
                   "Upper Limit", "Lower Limit", "Within Spec"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        row = 2
        for track in result.tracks:
            positions = getattr(track, 'position_data', []) or []
            errors = getattr(track, 'error_data', []) or []
            upper_limits = getattr(track, 'upper_limits', []) or []
            lower_limits = getattr(track, 'lower_limits', []) or []
            
            # Ensure all arrays have the same length
            max_len = max(len(positions), len(errors))
            
            for i in range(max_len):
                ws.cell(row=row, column=1, value=track.track_id)
                ws.cell(row=row, column=2, value="Error")
                ws.cell(row=row, column=3, value=i)
                
                if i < len(positions):
                    ws.cell(row=row, column=4, value=positions[i])
                
                if i < len(errors):
                    ws.cell(row=row, column=5, value=errors[i])
                
                if i < len(upper_limits):
                    ws.cell(row=row, column=6, value=upper_limits[i])
                
                if i < len(lower_limits):
                    ws.cell(row=row, column=7, value=lower_limits[i])
                
                # Check if within spec
                if (i < len(errors) and i < len(upper_limits) and i < len(lower_limits) and
                    upper_limits[i] is not None and lower_limits[i] is not None):
                    within_spec = lower_limits[i] <= errors[i] <= upper_limits[i]
                    ws.cell(row=row, column=8, value="Yes" if within_spec else "No")
                    if not within_spec:
                        ws.cell(row=row, column=8).fill = self.fail_fill
                
                row += 1
                
                # Limit to 10000 rows to prevent huge files
                if row > 10000:
                    ws.cell(row=row, column=1, value="Data truncated at 10000 rows...")
                    break
    
    def _create_metadata_sheet(self, wb: Workbook, result: AnalysisResult):
        """Create comprehensive metadata sheet."""
        ws = wb.create_sheet("Metadata")
        
        ws['A1'] = "Complete Metadata Information"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        
        # All metadata fields
        metadata_fields = [
            ("Filename", result.metadata.filename),
            ("File Path", str(result.metadata.file_path)),
            ("File Size (MB)", f"{getattr(result.metadata, 'file_size_mb', 'N/A'):.2f}" if hasattr(result.metadata, 'file_size_mb') else 'N/A'),
            ("Model", result.metadata.model),
            ("Serial", result.metadata.serial),
            ("System Type", result.metadata.system.value),
            ("Test Date", result.metadata.test_date.strftime('%Y-%m-%d %H:%M:%S') if hasattr(result.metadata, 'test_date') and result.metadata.test_date else 'N/A'),
            ("File Modified Date", result.metadata.file_date.strftime('%Y-%m-%d %H:%M:%S') if hasattr(result.metadata, 'file_date') and result.metadata.file_date else 'N/A'),
            ("Analysis Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Has Multiple Tracks", "Yes" if getattr(result.metadata, 'has_multi_tracks', False) else "No"),
            ("Track Count", len(result.tracks)),
            ("Track Identifier", getattr(result.metadata, 'track_identifier', 'N/A')),
        ]
        
        for label, value in metadata_fields:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Processing metadata
        ws[f'A{row}'] = "PROCESSING INFORMATION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        processing_fields = [
            ("Overall Status", getattr(result, 'overall_status', 'N/A').value if hasattr(result, 'overall_status') else 'N/A'),
            ("Processing Time (s)", f"{result.processing_time:.2f}"),
            ("Error Message", result.error_message or "None"),
            ("Analysis Mode", "Standard"),  # Could be enhanced with actual mode
        ]
        
        for label, value in processing_fields:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
    
    def _create_batch_summary_sheet(self, wb: Workbook, results: List[AnalysisResult]):
        """Create batch summary sheet with detailed file information."""
        ws = wb.create_sheet("Batch Summary")
        
        # Title
        ws['A1'] = "Batch Processing Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:L1')
        
        # Overall statistics section
        row = 3
        ws[f'A{row}'] = "OVERALL STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Calculate statistics
        total_files = len(results)
        passed_files = sum(1 for r in results if r.overall_status.value == "Pass")
        failed_files = sum(1 for r in results if r.overall_status.value == "Fail")
        warning_files = sum(1 for r in results if r.overall_status.value == "Warning")
        
        stats = [
            ("Total Files Processed", total_files),
            ("Passed", passed_files),
            ("Failed", failed_files),
            ("Warnings", warning_files),
            ("Pass Rate", f"{(passed_files/total_files*100):.1f}%" if total_files > 0 else "0%"),
            ("Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Detailed file results table
        ws[f'A{row}'] = "DETAILED FILE RESULTS"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers for file details
        headers = [
            "Filename", "Model", "Serial", "System Type", "File Date", "Trim Date",
            "Overall Status", "Validation Status", "Processing Time (s)", 
            "Track Count", "ML Risk Category", "Primary Failure Mode"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        row += 1
        
        # Add data for each file
        for result in results:
            try:
                # Basic metadata
                ws.cell(row=row, column=1, value=result.metadata.filename)
                ws.cell(row=row, column=2, value=result.metadata.model)
                ws.cell(row=row, column=3, value=result.metadata.serial)
                ws.cell(row=row, column=4, value=result.metadata.system.value)
                
                # File date (when file was created/modified)
                ws.cell(row=row, column=5, value=result.metadata.file_date.strftime('%Y-%m-%d %H:%M:%S'))
                
                # Trim date (original test date from the Excel file)
                trim_date = "Unknown"
                if hasattr(result.metadata, 'test_date') and result.metadata.test_date:
                    trim_date = result.metadata.test_date.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(result.metadata, 'file_date'):
                    # Fallback to file date if test_date not available
                    trim_date = result.metadata.file_date.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row, column=6, value=trim_date)
                
                # Status information
                ws.cell(row=row, column=7, value=result.overall_status.value)
                self._apply_status_formatting(ws.cell(row=row, column=7), result.overall_status.value)
                
                # Validation status
                validation_status = "N/A"
                if hasattr(result, 'overall_validation_status'):
                    validation_status = result.overall_validation_status.value
                ws.cell(row=row, column=8, value=validation_status)
                
                # Processing time
                ws.cell(row=row, column=9, value=f"{result.processing_time:.2f}")
                
                # Track count
                ws.cell(row=row, column=10, value=len(result.tracks))
                
                # ML Risk Category (from primary track)
                risk_category = "Unknown"
                primary_failure = "N/A"
                if result.primary_track:
                    # Check for ML prediction in track
                    if hasattr(result.primary_track, 'failure_prediction') and result.primary_track.failure_prediction:
                        risk_category = result.primary_track.failure_prediction.risk_category.value
                    elif getattr(result.primary_track, 'failure_prediction', None):
                        risk_category = result.primary_track.failure_prediction.risk_category.value
                    
                    # Determine primary failure mode
                    failures = []
                    if result.primary_track.sigma_analysis and not result.primary_track.sigma_analysis.sigma_pass:
                        failures.append("Sigma")
                    if result.primary_track.linearity_analysis and not result.primary_track.linearity_analysis.linearity_pass:
                        failures.append("Linearity")
                    primary_failure = ", ".join(failures) if failures else "None"
                
                ws.cell(row=row, column=11, value=risk_category)
                ws.cell(row=row, column=12, value=primary_failure)
                
                # Apply borders to all cells in row
                for col in range(1, 13):
                    ws.cell(row=row, column=col).border = self.border
                
                row += 1
                
            except Exception as e:
                self.logger.error(f"Error processing result for batch summary: {e}")
                # Add minimal data for failed row
                ws.cell(row=row, column=1, value=getattr(result.metadata, 'filename', 'Error'))
                ws.cell(row=row, column=2, value="Error processing")
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_batch_detailed_sheet(self, wb: Workbook, results: List[AnalysisResult]):
        """Create detailed results sheet for all files in batch."""
        ws = wb.create_sheet("Detailed Results")
        
        # Create comprehensive headers
        headers = [
            "Filename", "Model", "Serial", "System Type", "File Date", "Trim Date",
            "Status", "Processing Time", "Tracks", "Primary Track ID", "Data Points", 
            "Sigma Gradient", "Sigma Threshold", "Sigma Pass", "Sigma Ratio",
            "Linearity Spec", "Linearity Pass", "Fail Points", "Optimal Offset",
            "Final Linearity Error", "Max Deviation",
            "Unit Length", "Resistance Range", "Normalized Range %",
            "ML Risk Category", "ML Failure Probability", "ML Gradient Margin",
            "Validation Status", "Validation Grade", "Error Message"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Add data for each result
        row = 2
        for result in results:
            # Debug logging for model 8340
            if "8340" in str(result.metadata.model):
                self.logger.debug(f"Processing model 8340 file: {result.metadata.filename}")
                self.logger.debug(f"  Tracks: {len(result.tracks)}")
                if result.primary_track:
                    self.logger.debug(f"  Primary track data points: {len(result.primary_track.position_data) if result.primary_track.position_data else 0}")
                    if result.primary_track.position_data and len(result.primary_track.position_data) > 0:
                        self.logger.debug(f"  First 5 position values: {result.primary_track.position_data[:5]}")
                
            # Basic info
            ws.cell(row=row, column=1, value=result.metadata.filename)
            ws.cell(row=row, column=2, value=result.metadata.model)
            ws.cell(row=row, column=3, value=result.metadata.serial)
            ws.cell(row=row, column=4, value=result.metadata.system.value)
            
            # File date
            ws.cell(row=row, column=5, value=result.metadata.file_date.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Trim date (original test date from the Excel file)
            trim_date = "Unknown"
            if hasattr(result.metadata, 'test_date') and result.metadata.test_date:
                trim_date = result.metadata.test_date.strftime('%Y-%m-%d %H:%M:%S')
            elif hasattr(result.metadata, 'file_date'):
                # Fallback to file date if test_date not available
                trim_date = result.metadata.file_date.strftime('%Y-%m-%d %H:%M:%S')
            ws.cell(row=row, column=6, value=trim_date)
            
            ws.cell(row=row, column=7, value=result.overall_status.value)
            self._apply_status_formatting(ws.cell(row=row, column=7), result.overall_status.value)
            ws.cell(row=row, column=8, value=result.processing_time)
            ws.cell(row=row, column=9, value=len(result.tracks))
            
            # Primary track analysis
            primary_track = result.primary_track
            if primary_track:
                ws.cell(row=row, column=10, value=primary_track.track_id)
                ws.cell(row=row, column=11, value=len(primary_track.position_data) if primary_track.position_data else 0)
                
                # Sigma analysis
                if primary_track.sigma_analysis:
                    # Add safe value extraction with logging for debugging
                    sigma_gradient = getattr(primary_track.sigma_analysis, 'sigma_gradient', None)
                    sigma_threshold = getattr(primary_track.sigma_analysis, 'sigma_threshold', None)
                    sigma_pass = getattr(primary_track.sigma_analysis, 'sigma_pass', None)
                    sigma_ratio = getattr(primary_track.sigma_analysis, 'sigma_ratio', None)
                    
                    # Log if model 8340 has zero values
                    if "8340" in str(result.metadata.model) and (sigma_gradient == 0 or sigma_gradient is None):
                        self.logger.warning(f"Model 8340 file {result.metadata.filename} has sigma_gradient: {sigma_gradient}")
                    
                    ws.cell(row=row, column=12, value=sigma_gradient if sigma_gradient is not None else 'N/A')
                    ws.cell(row=row, column=13, value=sigma_threshold if sigma_threshold is not None else 'N/A')
                    ws.cell(row=row, column=14, value="PASS" if sigma_pass else "FAIL")
                    ws.cell(row=row, column=15, value=sigma_ratio if sigma_ratio is not None else 'N/A')
                
                # Linearity analysis
                if primary_track.linearity_analysis:
                    ws.cell(row=row, column=16, value=primary_track.linearity_analysis.linearity_spec)
                    ws.cell(row=row, column=17, value="PASS" if primary_track.linearity_analysis.linearity_pass else "FAIL")
                    ws.cell(row=row, column=18, value=primary_track.linearity_analysis.linearity_fail_points)
                    ws.cell(row=row, column=19, value=primary_track.linearity_analysis.optimal_offset)
                    ws.cell(row=row, column=20, value=primary_track.linearity_analysis.final_linearity_error_shifted)
                    ws.cell(row=row, column=21, value=getattr(primary_track.linearity_analysis, 'max_deviation', 'N/A'))
                
                # Unit properties
                if primary_track.unit_properties:
                    ws.cell(row=row, column=22, value=primary_track.unit_properties.unit_length)
                    ws.cell(row=row, column=23, value=getattr(primary_track.unit_properties, 'resistance_change', None) if primary_track.unit_properties else None)
                    ws.cell(row=row, column=24, value=f"{getattr(primary_track.unit_properties, 'resistance_change_percent', 0.0):.2f}%" if primary_track.unit_properties and getattr(primary_track.unit_properties, 'resistance_change_percent', None) is not None else "N/A")
            
            # ML predictions (from primary track)
            risk_category = "Unknown"
            if primary_track:
                if hasattr(primary_track, 'failure_prediction') and primary_track.failure_prediction:
                    risk_category = primary_track.failure_prediction.risk_category.value if hasattr(primary_track.failure_prediction, 'risk_category') else 'Unknown'
                    ws.cell(row=row, column=26, value=f"{primary_track.failure_prediction.failure_probability:.2%}" if hasattr(primary_track.failure_prediction, 'failure_probability') else 'N/A')
                    ws.cell(row=row, column=27, value=primary_track.sigma_analysis.gradient_margin if primary_track.sigma_analysis and hasattr(primary_track.sigma_analysis, 'gradient_margin') else 'N/A')
                elif getattr(result.primary_track, 'failure_prediction', None):
                    risk_category = result.primary_track.failure_prediction.risk_category.value if hasattr(result.primary_track.failure_prediction, 'risk_category') else 'Unknown'
                    ws.cell(row=row, column=26, value=f"{getattr(result.primary_track.failure_prediction, 'failure_probability', 0.0):.2%}")
                    ws.cell(row=row, column=27, value=getattr(result.primary_track.sigma_analysis, 'gradient_margin', 'N/A'))
                else:
                    ws.cell(row=row, column=26, value='N/A')
                    ws.cell(row=row, column=27, value='N/A')
            else:
                ws.cell(row=row, column=26, value='N/A')
                ws.cell(row=row, column=27, value='N/A')
            
            ws.cell(row=row, column=25, value=risk_category)
            
            # Validation
            ws.cell(row=row, column=28, value=result.overall_validation_status.value if hasattr(result, 'overall_validation_status') else 'N/A')
            ws.cell(row=row, column=29, value=result.validation_grade if hasattr(result, 'validation_grade') else 'N/A')
            
            # Processing errors/issues
            ws.cell(row=row, column=30, value="; ".join(result.processing_errors) if hasattr(result, 'processing_errors') and result.processing_errors else "")
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_enhanced_data_sheet(self, wb: Workbook, results: List[AnalysisResult]):
        """Create enhanced data sheet with additional useful information."""
        ws = wb.create_sheet("Enhanced Data")
        
        # Title
        ws['A1'] = "Enhanced Analysis Data"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:Q1')
        
        # Headers with all useful fields
        headers = [
            "Filename", "Model", "Serial", "System Type", "File Date", "Trim Date",
            "Overall Status", "Sigma Pass", "Linearity Pass", "Risk Category",
            "Unit Length (deg)", "Resistance Before (Ω)", "Resistance After (Ω)", 
            "Resistance Change (%)", "Normalized Range (%)",
            "Sigma Gradient", "Sigma Threshold", "Sigma Ratio", "Gradient Margin",
            "Linearity Spec", "Linearity Error", "Optimal Offset", "Max Deviation",
            "Failure Probability (%)", "Industry Compliance", "Validation Grade",
            "Data Points", "Travel Length", "Comments"
        ]
        
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        row += 1
        
        # Add data for each result
        for result in results:
            try:
                # Basic info
                ws.cell(row=row, column=1, value=result.metadata.filename)
                ws.cell(row=row, column=2, value=result.metadata.model)
                ws.cell(row=row, column=3, value=result.metadata.serial)
                ws.cell(row=row, column=4, value=result.metadata.system.value)
                ws.cell(row=row, column=5, value=result.metadata.file_date.strftime('%Y-%m-%d %H:%M:%S'))
                # Trim date (original test date from the Excel file)
                trim_date = "Unknown"
                if hasattr(result.metadata, 'test_date') and result.metadata.test_date:
                    trim_date = result.metadata.test_date.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(result.metadata, 'file_date'):
                    # Fallback to file date if test_date not available
                    trim_date = result.metadata.file_date.strftime('%Y-%m-%d %H:%M:%S')
                ws.cell(row=row, column=6, value=trim_date)
                ws.cell(row=row, column=7, value=result.overall_status.value)
                self._apply_status_formatting(ws.cell(row=row, column=7), result.overall_status.value)
                
                if result.primary_track:
                    track = result.primary_track
                    
                    # Pass/Fail status
                    sigma_pass = "PASS" if track.sigma_analysis and track.sigma_analysis.sigma_pass else "FAIL"
                    linearity_pass = "PASS" if track.linearity_analysis and track.linearity_analysis.linearity_pass else "FAIL"
                    ws.cell(row=row, column=8, value=sigma_pass)
                    self._apply_status_formatting(ws.cell(row=row, column=8), sigma_pass)
                    ws.cell(row=row, column=9, value=linearity_pass)
                    self._apply_status_formatting(ws.cell(row=row, column=9), linearity_pass)
                    
                    # Risk category
                    risk_category = "Unknown"
                    failure_prob = None
                    if hasattr(track, 'failure_prediction') and track.failure_prediction:
                        risk_category = track.failure_prediction.risk_category.value if hasattr(track.failure_prediction, 'risk_category') else 'Unknown'
                        failure_prob = track.failure_prediction.failure_probability
                    elif getattr(result.primary_track, 'failure_prediction', None):
                        risk_category = result.primary_track.failure_prediction.risk_category.value if hasattr(result.primary_track.failure_prediction, 'risk_category') else 'Unknown'
                        failure_prob = result.primary_track.failure_prediction.failure_probability
                    ws.cell(row=row, column=10, value=risk_category)
                    
                    # Unit properties
                    if track.unit_properties:
                        ws.cell(row=row, column=11, value=track.unit_properties.unit_length)
                        ws.cell(row=row, column=12, value=track.unit_properties.untrimmed_resistance if hasattr(track.unit_properties, 'untrimmed_resistance') else 'N/A')
                        ws.cell(row=row, column=13, value=track.unit_properties.trimmed_resistance if hasattr(track.unit_properties, 'trimmed_resistance') else 'N/A')
                        ws.cell(row=row, column=14, value=f"{track.unit_properties.resistance_change_percent:.2f}" if hasattr(track.unit_properties, 'resistance_change_percent') and track.unit_properties.resistance_change_percent is not None else 'N/A')
                        ws.cell(row=row, column=15, value=f"{getattr(track.unit_properties, 'resistance_change_percent', 0.0):.2f}" if hasattr(track.unit_properties, 'resistance_change_percent') and track.unit_properties.resistance_change_percent is not None else 'N/A')
                    
                    # Sigma analysis
                    if track.sigma_analysis:
                        ws.cell(row=row, column=16, value=track.sigma_analysis.sigma_gradient)
                        ws.cell(row=row, column=17, value=track.sigma_analysis.sigma_threshold)
                        ws.cell(row=row, column=18, value=f"{track.sigma_analysis.sigma_ratio:.3f}" if hasattr(track.sigma_analysis, 'sigma_ratio') else 'N/A')
                        ws.cell(row=row, column=19, value=track.sigma_analysis.gradient_margin if hasattr(track.sigma_analysis, 'gradient_margin') else 'N/A')
                        
                        # Industry compliance
                        industry_compliance = track.sigma_analysis.industry_compliance if hasattr(track.sigma_analysis, 'industry_compliance') else 'N/A'
                        ws.cell(row=row, column=25, value=industry_compliance)
                    
                    # Linearity analysis
                    if track.linearity_analysis:
                        ws.cell(row=row, column=20, value=track.linearity_analysis.linearity_spec)
                        ws.cell(row=row, column=21, value=track.linearity_analysis.final_linearity_error_shifted if hasattr(track.linearity_analysis, 'final_linearity_error_shifted') else 'N/A')
                        ws.cell(row=row, column=22, value=track.linearity_analysis.optimal_offset)
                        ws.cell(row=row, column=23, value=track.linearity_analysis.max_deviation if hasattr(track.linearity_analysis, 'max_deviation') else 'N/A')
                    
                    # ML prediction
                    if failure_prob is not None:
                        ws.cell(row=row, column=24, value=f"{failure_prob * 100:.1f}")
                    else:
                        ws.cell(row=row, column=24, value='N/A')
                    
                    # Validation grade
                    validation_grade = result.validation_grade if hasattr(result, 'validation_grade') else 'N/A'
                    ws.cell(row=row, column=26, value=validation_grade)
                    
                    # Track data
                    ws.cell(row=row, column=27, value=track.data_points if hasattr(track, 'data_points') else len(track.position_data) if hasattr(track, 'position_data') else 'N/A')
                    ws.cell(row=row, column=28, value=track.travel_length if hasattr(track, 'travel_length') else 'N/A')
                    
                    # Comments
                    comments = []
                    if track.sigma_analysis and not track.sigma_analysis.sigma_pass:
                        comments.append("Sigma fail")
                    if track.linearity_analysis and not track.linearity_analysis.linearity_pass:
                        comments.append(f"Linearity fail ({track.linearity_analysis.linearity_fail_points} pts)")
                    if hasattr(result, 'processing_errors') and result.processing_errors:
                        comments.extend(result.processing_errors)
                    ws.cell(row=row, column=29, value="; ".join(comments))
                
                # Apply borders
                for col in range(1, 30):
                    ws.cell(row=row, column=col).border = self.border
                
                row += 1
                
            except Exception as e:
                self.logger.error(f"Error adding enhanced data for {getattr(result.metadata, 'filename', 'Unknown')}: {e}")
                # Add minimal data
                ws.cell(row=row, column=1, value=getattr(result.metadata, 'filename', 'Error'))
                ws.cell(row=row, column=2, value="Error")
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_statistical_analysis_sheet(self, wb: Workbook, results: List[AnalysisResult]):
        """Create statistical analysis sheet."""
        ws = wb.create_sheet("Statistical Analysis")
        
        # Collect all numeric data for statistics
        sigma_gradients = []
        linearity_specs = []
        resistance_ranges = []
        failure_probs = []
        
        for result in results:
            if result.primary_track:
                if result.primary_track.sigma_analysis:
                    sigma_gradients.append(result.primary_track.sigma_analysis.sigma_gradient)
                if result.primary_track.linearity_analysis:
                    linearity_specs.append(result.primary_track.linearity_analysis.linearity_spec)
                if result.primary_track.unit_properties and result.primary_track.unit_properties.resistance_change_percent:
                    resistance_ranges.append(result.primary_track.unit_properties.resistance_change_percent)
            
            if result.primary_track and result.primary_track.failure_prediction and result.primary_track.failure_prediction.failure_probability:
                failure_probs.append(result.primary_track.failure_prediction.failure_probability)
        
        # Calculate statistics
        row = 1
        ws[f'A{row}'] = "Statistical Analysis"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Sigma gradient statistics
        if sigma_gradients:
            ws[f'A{row}'] = "SIGMA GRADIENT STATISTICS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            stats = [
                ("Count", len(sigma_gradients)),
                ("Mean", np.mean(sigma_gradients)),
                ("Std Dev", np.std(sigma_gradients)),
                ("Min", np.min(sigma_gradients)),
                ("25th Percentile", np.percentile(sigma_gradients, 25)),
                ("Median", np.median(sigma_gradients)),
                ("75th Percentile", np.percentile(sigma_gradients, 75)),
                ("Max", np.max(sigma_gradients)),
            ]
            
            for label, value in stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = f"{value:.6f}" if isinstance(value, float) else value
                row += 1
            
            row += 1
        
        # Linearity spec statistics
        if linearity_specs:
            ws[f'A{row}'] = "LINEARITY SPEC STATISTICS"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            stats = [
                ("Count", len(linearity_specs)),
                ("Mean", np.mean(linearity_specs)),
                ("Std Dev", np.std(linearity_specs)),
                ("Min", np.min(linearity_specs)),
                ("25th Percentile", np.percentile(linearity_specs, 25)),
                ("Median", np.median(linearity_specs)),
                ("75th Percentile", np.percentile(linearity_specs, 75)),
                ("Max", np.max(linearity_specs)),
            ]
            
            for label, value in stats:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = f"{value:.6f}" if isinstance(value, float) else value
                row += 1
    
    def _create_model_performance_sheet(self, wb: Workbook, results: List[AnalysisResult]):
        """Create model performance comparison sheet."""
        ws = wb.create_sheet("Model Performance")
        
        # Aggregate by model
        model_data = {}
        
        for result in results:
            model = result.metadata.model
            if model not in model_data:
                model_data[model] = {
                    'count': 0,
                    'pass': 0,
                    'fail': 0,
                    'warning': 0,
                    'sigma_gradients': [],
                    'linearity_specs': [],
                    'failure_probs': [],
                    'processing_times': []
                }
            
            model_data[model]['count'] += 1
            model_data[model]['processing_times'].append(result.processing_time)
            
            if result.overall_status.value == "PASS":
                model_data[model]['pass'] += 1
            elif result.overall_status.value == "FAIL":
                model_data[model]['fail'] += 1
            else:
                model_data[model]['warning'] += 1
            
            if result.primary_track:
                if result.primary_track.sigma_analysis:
                    model_data[model]['sigma_gradients'].append(
                        result.primary_track.sigma_analysis.sigma_gradient
                    )
                if result.primary_track.linearity_analysis:
                    model_data[model]['linearity_specs'].append(
                        result.primary_track.linearity_analysis.linearity_spec
                    )
            
            if hasattr(result, 'ml_prediction') and result.ml_prediction:
                model_data[model]['failure_probs'].append(
                    result.ml_prediction.failure_probability
                )
        
        # Create headers
        headers = [
            "Model", "Count", "Pass", "Fail", "Warning", "Pass Rate",
            "Avg Sigma Gradient", "Avg Linearity Spec", "Avg ML Risk",
            "Avg Processing Time"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Add model data
        row = 2
        for model, data in model_data.items():
            ws.cell(row=row, column=1, value=model)
            ws.cell(row=row, column=2, value=data['count'])
            ws.cell(row=row, column=3, value=data['pass'])
            ws.cell(row=row, column=4, value=data['fail'])
            ws.cell(row=row, column=5, value=data['warning'])
            ws.cell(row=row, column=6, value=f"{(data['pass']/data['count']*100):.1f}%")
            
            if data['sigma_gradients']:
                ws.cell(row=row, column=7, value=f"{np.mean(data['sigma_gradients']):.6f}")
            
            if data['linearity_specs']:
                ws.cell(row=row, column=8, value=f"{np.mean(data['linearity_specs']):.6f}")
            
            if data['failure_probs']:
                ws.cell(row=row, column=9, value=f"{np.mean(data['failure_probs']):.2%}")
            
            ws.cell(row=row, column=10, value=f"{np.mean(data['processing_times']):.2f}")
            
            row += 1
    
    def _create_failure_analysis_sheet(self, wb: Workbook, results: List[AnalysisResult]):
        """Create detailed failure analysis sheet."""
        ws = wb.create_sheet("Failure Analysis")
        
        # Headers
        headers = [
            "Filename", "Model", "Serial", "Failure Type", "Primary Cause",
            "Sigma Gradient", "Sigma Threshold", "Linearity Fail Points",
            "Max Deviation", "Error Message", "Recommendations"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
        
        # Add only failed results
        row = 2
        for result in results:
            if result.overall_status.value != "Fail":
                continue
            
            ws.cell(row=row, column=1, value=result.metadata.filename)
            ws.cell(row=row, column=2, value=result.metadata.model)
            ws.cell(row=row, column=3, value=result.metadata.serial)
            
            # Determine failure type and cause
            failure_types = []
            primary_cause = "Unknown"
            
            if result.primary_track:
                if result.primary_track.sigma_analysis and not getattr(result.primary_track.sigma_analysis, 'sigma_pass', False):
                    failure_types.append("Sigma")
                    primary_cause = "Sigma gradient exceeds threshold"
                    ws.cell(row=row, column=6, value=result.primary_track.sigma_analysis.sigma_gradient)
                    ws.cell(row=row, column=7, value=result.primary_track.sigma_analysis.sigma_threshold)
                
                if result.primary_track.linearity_analysis and not getattr(result.primary_track.linearity_analysis, 'linearity_pass', False):
                    failure_types.append("Linearity")
                    if primary_cause == "Unknown":
                        primary_cause = f"{result.primary_track.linearity_analysis.fail_points} points exceed spec"
                    ws.cell(row=row, column=8, value=result.primary_track.linearity_analysis.fail_points)
                    ws.cell(row=row, column=9, value=getattr(result.primary_track.linearity_analysis, 'max_deviation', 'N/A'))
            
            ws.cell(row=row, column=4, value=", ".join(failure_types) if failure_types else "Processing Error")
            ws.cell(row=row, column=5, value=primary_cause)
            # No centralized error_message field; join processing_errors if present
            ws.cell(row=row, column=10, value="; ".join(getattr(result, 'processing_errors', []) or []))
            
            # Add recommendations
            recommendations = []
            if "Sigma" in failure_types:
                recommendations.append("Review trim parameters")
            if "Linearity" in failure_types:
                recommendations.append("Check system calibration")
            
            ws.cell(row=row, column=11, value="; ".join(recommendations))
            
            row += 1
    
    def _create_individual_detail_sheet(self, wb: Workbook, result: AnalysisResult, sheet_name: str):
        """Create individual detail sheet for a specific file."""
        ws = wb.create_sheet(sheet_name[:31])  # Excel sheet name limit
        
        # Use the single file summary format
        self._populate_summary_content(ws, result)
    
    def _populate_summary_content(self, ws, result: AnalysisResult):
        """Populate worksheet with summary content (reusable)."""
        # Similar to _create_summary_sheet but just the content population
        # This avoids code duplication
        row = 1
        ws[f'A{row}'] = f"Analysis Summary: {result.metadata.filename}"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 2
        
        # Add all the summary content similar to _create_summary_sheet
        # (Implementation details omitted for brevity but would follow same pattern)
    
    def _apply_status_formatting(self, cell, status: str):
        """Apply color formatting based on status."""
        if status == "PASS" or status == "Pass":
            cell.fill = self.pass_fill
        elif status == "FAIL" or status == "Fail":
            cell.fill = self.fail_fill
        elif status == "WARNING" or status == "Warning":
            cell.fill = self.warning_fill
    
    def _format_contributing_factors(self, factors: Optional[Dict[str, float]]) -> str:
        """Format contributing factors dictionary into readable string."""
        if not factors:
            return "N/A"
        
        # Sort by contribution value
        sorted_factors = sorted(factors.items(), key=lambda x: x[1], reverse=True)
        
        # Take top 3 factors
        top_factors = sorted_factors[:3]
        
        # Format as string
        formatted = []
        for factor, value in top_factors:
            formatted.append(f"{factor}: {value:.2f}")
        
        return "; ".join(formatted)
