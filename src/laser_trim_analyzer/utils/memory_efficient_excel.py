"""
Memory-efficient Excel reading utilities.

Provides functions for reading large Excel files with minimal memory usage.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, Union, Iterator, List, Dict, Any
import logging
import gc
import psutil
import openpyxl
from contextlib import contextmanager

logger = logging.getLogger(__name__)


def read_excel_memory_efficient(
    file_path: Union[str, Path],
    sheet_name: Optional[Union[str, int]] = None,
    chunk_size: int = 10000,
    usecols: Optional[List[Union[int, str]]] = None,
    max_rows: Optional[int] = None
) -> pd.DataFrame:
    """
    Read Excel file with memory-efficient chunking.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name or index (default: first sheet)
        chunk_size: Number of rows to read at once
        usecols: Columns to read (default: all)
        max_rows: Maximum rows to read (default: all)
        
    Returns:
        Complete DataFrame with all data
    """
    file_path = Path(file_path)
    file_size_mb = file_path.stat().st_size / (1024 * 1024)
    
    logger.info(f"Reading Excel file ({file_size_mb:.1f} MB) with chunking: {file_path.name}")
    
    # For small files, use standard pandas reader
    if file_size_mb < 10:
        return pd.read_excel(file_path, sheet_name=sheet_name, usecols=usecols, nrows=max_rows)
    
    # For large files, use openpyxl with chunking
    chunks = []
    rows_read = 0
    
    try:
        # Open workbook in read-only mode
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # Get the worksheet
        if sheet_name is None:
            ws = wb.active
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
        
        # Get column indices if specific columns requested
        if usecols:
            if isinstance(usecols[0], str):
                # Convert column letters to indices
                col_indices = [openpyxl.utils.column_index_from_string(col) - 1 for col in usecols]
            else:
                col_indices = usecols
        else:
            col_indices = None
        
        # Read in chunks
        current_chunk = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            # Check max rows limit
            if max_rows and rows_read >= max_rows:
                break
            
            # Filter columns if specified
            if col_indices:
                row_data = [row[i] if i < len(row) else None for i in col_indices]
            else:
                row_data = list(row)
            
            current_chunk.append(row_data)
            rows_read += 1
            
            # Process chunk when it reaches chunk_size
            if len(current_chunk) >= chunk_size:
                chunk_df = pd.DataFrame(current_chunk)
                chunks.append(chunk_df)
                current_chunk = []
                
                # Force garbage collection
                gc.collect()
                
                # Check memory usage
                memory_percent = psutil.virtual_memory().percent
                if memory_percent > 85:
                    logger.warning(f"High memory usage: {memory_percent:.1f}%")
                    # Reduce chunk size for remaining data
                    chunk_size = max(1000, chunk_size // 2)
        
        # Process final chunk
        if current_chunk:
            chunks.append(pd.DataFrame(current_chunk))
        
        # Close workbook
        wb.close()
        
        # Combine all chunks
        if chunks:
            logger.info(f"Combining {len(chunks)} chunks into final DataFrame")
            df = pd.concat(chunks, ignore_index=True)
            
            # Clean up chunks
            del chunks
            gc.collect()
            
            # Set column names if provided
            if usecols and isinstance(usecols[0], str):
                df.columns = usecols
            
            return df
        else:
            return pd.DataFrame()
            
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        raise


@contextmanager
def excel_reader_context(file_path: Union[str, Path], sheet_name: Optional[Union[str, int]] = None):
    """
    Context manager for reading Excel files with automatic resource cleanup.
    
    Usage:
        with excel_reader_context('large_file.xlsx') as reader:
            for chunk in reader.iter_chunks(chunk_size=5000):
                process_chunk(chunk)
    """
    file_path = Path(file_path)
    wb = None
    
    try:
        # Open workbook
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        # Get worksheet
        if sheet_name is None:
            ws = wb.active
        elif isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
        
        # Create reader object
        reader = ExcelChunkReader(ws)
        yield reader
        
    finally:
        # Ensure workbook is closed
        if wb:
            wb.close()
        gc.collect()


class ExcelChunkReader:
    """Helper class for reading Excel data in chunks."""
    
    def __init__(self, worksheet):
        self.worksheet = worksheet
        self._row_iterator = None
    
    def iter_chunks(self, chunk_size: int = 10000, usecols: Optional[List[Union[int, str]]] = None) -> Iterator[pd.DataFrame]:
        """
        Iterate over Excel data in chunks.
        
        Args:
            chunk_size: Number of rows per chunk
            usecols: Columns to include
            
        Yields:
            DataFrame chunks
        """
        # Get column indices
        if usecols and isinstance(usecols[0], str):
            col_indices = [openpyxl.utils.column_index_from_string(col) - 1 for col in usecols]
        else:
            col_indices = usecols
        
        current_chunk = []
        
        for row in self.worksheet.iter_rows(values_only=True):
            # Filter columns
            if col_indices:
                row_data = [row[i] if i < len(row) else None for i in col_indices]
            else:
                row_data = list(row)
            
            current_chunk.append(row_data)
            
            if len(current_chunk) >= chunk_size:
                # Yield chunk as DataFrame
                chunk_df = pd.DataFrame(current_chunk)
                if usecols and isinstance(usecols[0], str):
                    chunk_df.columns = usecols
                
                yield chunk_df
                
                # Clear chunk and collect garbage
                current_chunk = []
                gc.collect()
        
        # Yield final chunk
        if current_chunk:
            chunk_df = pd.DataFrame(current_chunk)
            if usecols and isinstance(usecols[0], str):
                chunk_df.columns = usecols
            yield chunk_df
    
    def read_sample(self, n_rows: int = 1000) -> pd.DataFrame:
        """Read a sample of rows from the beginning of the sheet."""
        rows = []
        for i, row in enumerate(self.worksheet.iter_rows(values_only=True)):
            if i >= n_rows:
                break
            rows.append(list(row))
        return pd.DataFrame(rows)


def estimate_memory_usage(file_path: Union[str, Path], sample_size: int = 1000) -> Dict[str, Any]:
    """
    Estimate memory usage for reading an Excel file.
    
    Args:
        file_path: Path to Excel file
        sample_size: Number of rows to sample
        
    Returns:
        Dictionary with memory estimates
    """
    file_path = Path(file_path)
    file_size_mb = file_path.stat().st_size / (1024 * 1024)
    
    # Read sample to estimate row size
    sample_df = read_excel_memory_efficient(file_path, max_rows=sample_size)
    
    if len(sample_df) == 0:
        return {
            'file_size_mb': file_size_mb,
            'estimated_memory_mb': 0,
            'recommended_chunk_size': 10000,
            'available_memory_mb': psutil.virtual_memory().available / (1024 * 1024)
        }
    
    # Estimate memory per row
    sample_memory = sample_df.memory_usage(deep=True).sum()
    memory_per_row = sample_memory / len(sample_df)
    
    # Estimate total rows (rough approximation)
    estimated_total_rows = int(file_size_mb * 1024 * 1024 / (memory_per_row * 10))  # Factor of 10 for Excel overhead
    
    # Calculate estimates
    estimated_memory_mb = (memory_per_row * estimated_total_rows) / (1024 * 1024)
    available_memory_mb = psutil.virtual_memory().available / (1024 * 1024)
    
    # Recommend chunk size based on available memory
    if available_memory_mb > estimated_memory_mb * 2:
        recommended_chunk_size = 50000
    elif available_memory_mb > estimated_memory_mb:
        recommended_chunk_size = 20000
    else:
        # Memory constrained - use smaller chunks
        recommended_chunk_size = min(10000, int(available_memory_mb * 1024 * 1024 / (memory_per_row * 10)))
    
    return {
        'file_size_mb': file_size_mb,
        'estimated_rows': estimated_total_rows,
        'memory_per_row_bytes': memory_per_row,
        'estimated_memory_mb': estimated_memory_mb,
        'available_memory_mb': available_memory_mb,
        'recommended_chunk_size': max(1000, recommended_chunk_size),
        'memory_constrained': available_memory_mb < estimated_memory_mb
    }