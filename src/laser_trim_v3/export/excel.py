"""
Excel Export for Laser Trim Analyzer v3.

Clean, focused Excel export that includes:
- Analysis summary sheet
- Track details sheet
- Optional: Raw data sheet

Simplified from v2's multi-format export system.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Optional, Dict, Any, Union
from dataclasses import dataclass

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from laser_trim_v3.core.models import AnalysisResult, AnalysisStatus, TrackData

logger = logging.getLogger(__name__)


@dataclass
class ExportConfig:
    """Configuration for Excel export."""
    include_raw_data: bool = False
    include_charts: bool = True
    include_summary: bool = True
    sheet_name_prefix: str = ""


class ExcelExportError(Exception):
    """Excel export error."""
    pass


# Style definitions
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def export_single_result(
    result: AnalysisResult,
    output_path: Union[str, Path],
    config: Optional[ExportConfig] = None,
) -> Path:
    """
    Export a single analysis result to Excel.

    Args:
        result: Analysis result to export
        output_path: Output file path
        config: Export configuration

    Returns:
        Path to created file

    Raises:
        ExcelExportError: If export fails
    """
    if not HAS_OPENPYXL:
        raise ExcelExportError("openpyxl not installed. Run: pip install openpyxl")

    config = config or ExportConfig()
    output_path = Path(output_path)

    try:
        wb = Workbook()

        # Summary sheet
        if config.include_summary:
            _create_summary_sheet(wb, result)

        # Track details sheet
        _create_tracks_sheet(wb, result)

        # Raw data sheet (optional)
        if config.include_raw_data:
            _create_raw_data_sheet(wb, result)

        # Remove default empty sheet if we created others
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb.save(output_path)
        logger.info(f"Exported to: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Excel export failed: {e}")
        raise ExcelExportError(f"Export failed: {e}") from e


def export_batch_results(
    results: List[AnalysisResult],
    output_path: Union[str, Path],
    config: Optional[ExportConfig] = None,
) -> Path:
    """
    Export multiple analysis results to a single Excel file.

    Args:
        results: List of analysis results
        output_path: Output file path
        config: Export configuration

    Returns:
        Path to created file
    """
    if not HAS_OPENPYXL:
        raise ExcelExportError("openpyxl not installed. Run: pip install openpyxl")

    if not results:
        raise ExcelExportError("No results to export")

    config = config or ExportConfig()
    output_path = Path(output_path)

    try:
        wb = Workbook()

        # Batch summary sheet
        _create_batch_summary_sheet(wb, results)

        # All results sheet
        _create_all_results_sheet(wb, results)

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info(f"Exported batch ({len(results)} files) to: {output_path}")
        return output_path

    except Exception as e:
        logger.error(f"Batch export failed: {e}")
        raise ExcelExportError(f"Export failed: {e}") from e


def _create_summary_sheet(wb: "Workbook", result: AnalysisResult) -> None:
    """Create summary sheet with file info and overall results."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws["A1"] = "Laser Trim Analysis Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # File info section
    ws["A3"] = "File Information"
    ws["A3"].font = Font(bold=True, size=12)

    # Use file_date which is the trim date (we set file_date = test_date in parser)
    trim_date = result.metadata.file_date or result.metadata.test_date

    info_rows = [
        ("Filename:", result.metadata.filename),
        ("Model:", result.metadata.model),
        ("Serial:", result.metadata.serial),
        ("System:", result.metadata.system.value),
        ("Trim Date:", trim_date.strftime("%Y-%m-%d %H:%M") if trim_date else "N/A"),
        ("Analysis Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Processing Time:", f"{result.processing_time:.2f} seconds"),
    ]

    row = 4
    for label, value in info_rows:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Overall status section
    row += 1
    ws[f"A{row}"] = "Overall Status"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    status_cell = ws[f"B{row}"]
    status_cell.value = result.overall_status.value
    status_cell.font = Font(bold=True, size=14)

    if result.overall_status == AnalysisStatus.PASS:
        status_cell.fill = PASS_FILL
    elif result.overall_status == AnalysisStatus.FAIL:
        status_cell.fill = FAIL_FILL
    else:
        status_cell.fill = WARNING_FILL

    ws[f"A{row}"] = "Status:"
    ws[f"A{row}"].font = Font(bold=True)

    # Track summary section
    row += 2
    ws[f"A{row}"] = "Track Summary"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    row += 1

    # Track summary headers
    headers = ["Track", "Status", "Sigma Gradient", "Threshold", "Sigma Pass",
               "Linearity Error", "Linearity Pass", "Risk", "Failure Prob"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    row += 1

    # Track data rows
    for track in result.tracks:
        cells = [
            track.track_id,
            track.status.value,
            f"{track.sigma_gradient:.6f}",
            f"{track.sigma_threshold:.6f}",
            "PASS" if track.sigma_pass else "FAIL",
            f"{track.linearity_error:.6f}",
            "PASS" if track.linearity_pass else "FAIL",
            track.risk_category.value,
            f"{track.failure_probability:.1%}" if track.failure_probability is not None else "N/A",
        ]

        for col, value in enumerate(cells, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

            # Color status cells
            if col == 2:  # Status column
                if value == "PASS":
                    cell.fill = PASS_FILL
                elif value == "FAIL":
                    cell.fill = FAIL_FILL

        row += 1

    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_tracks_sheet(wb: "Workbook", result: AnalysisResult) -> None:
    """Create detailed track analysis sheet."""
    ws = wb.create_sheet("Track Details")

    row = 1
    for track in result.tracks:
        # Track header
        ws[f"A{row}"] = f"Track {track.track_id}"
        ws[f"A{row}"].font = Font(bold=True, size=12)
        row += 1

        # Status banner
        status_cell = ws[f"A{row}"]
        status_cell.value = track.status.value
        status_cell.font = Font(bold=True, size=14)
        if track.status == AnalysisStatus.PASS:
            status_cell.fill = PASS_FILL
        elif track.status == AnalysisStatus.FAIL:
            status_cell.fill = FAIL_FILL
        else:
            status_cell.fill = WARNING_FILL
        row += 2

        # Sigma analysis section
        ws[f"A{row}"] = "Sigma Analysis"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        sigma_data = [
            ("Sigma Gradient:", f"{track.sigma_gradient:.6f}"),
            ("Sigma Threshold:", f"{track.sigma_threshold:.6f}"),
            ("Sigma Pass:", "YES" if track.sigma_pass else "NO"),
            ("Margin:", f"{((track.sigma_threshold - track.sigma_gradient) / track.sigma_threshold * 100):.1f}%" if track.sigma_threshold > 0 else "N/A"),
        ]

        for label, value in sigma_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1

        # Linearity analysis section
        ws[f"A{row}"] = "Linearity Analysis"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        linearity_data = [
            ("Optimal Offset:", f"{track.optimal_offset:.6f}"),
            ("Max Error:", f"{track.linearity_error:.6f}"),
            ("Linearity Spec:", f"{track.linearity_spec:.6f}"),
            ("Fail Points:", str(track.linearity_fail_points)),
            ("Linearity Pass:", "YES" if track.linearity_pass else "NO"),
        ]

        for label, value in linearity_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1

        # Risk assessment section
        ws[f"A{row}"] = "Risk Assessment"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

        risk_data = [
            ("Failure Probability:", f"{track.failure_probability:.1%}" if track.failure_probability is not None else "N/A"),
            ("Risk Category:", track.risk_category.value),
        ]

        for label, value in risk_data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            row += 1

        row += 1

        # Unit properties (if available)
        if track.travel_length or track.unit_length:
            ws[f"A{row}"] = "Unit Properties"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            if track.travel_length:
                ws[f"A{row}"] = "Travel Length:"
                ws[f"B{row}"] = f"{track.travel_length}"
                row += 1
            if track.unit_length:
                ws[f"A{row}"] = "Unit Length:"
                ws[f"B{row}"] = f"{track.unit_length}"
                row += 1
            if track.untrimmed_resistance:
                ws[f"A{row}"] = "Untrimmed R:"
                ws[f"B{row}"] = f"{track.untrimmed_resistance}"
                row += 1
            if track.trimmed_resistance:
                ws[f"A{row}"] = "Trimmed R:"
                ws[f"B{row}"] = f"{track.trimmed_resistance}"
                row += 1

        row += 2  # Space before next track

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20


def _create_raw_data_sheet(wb: "Workbook", result: AnalysisResult) -> None:
    """Create raw data sheet with position/error data."""
    ws = wb.create_sheet("Raw Data")

    # For each track, add position and error columns
    col = 1
    for track in result.tracks:
        if not track.position_data or not track.error_data:
            continue

        # Headers
        ws.cell(row=1, column=col).value = f"Track {track.track_id} Position"
        ws.cell(row=1, column=col).font = HEADER_FONT
        ws.cell(row=1, column=col).fill = HEADER_FILL

        ws.cell(row=1, column=col + 1).value = f"Track {track.track_id} Error"
        ws.cell(row=1, column=col + 1).font = HEADER_FONT
        ws.cell(row=1, column=col + 1).fill = HEADER_FILL

        # Data
        for row_idx, (pos, err) in enumerate(zip(track.position_data, track.error_data), 2):
            ws.cell(row=row_idx, column=col).value = pos
            ws.cell(row=row_idx, column=col + 1).value = err

        # Upper/lower limits if available
        if track.upper_limits:
            ws.cell(row=1, column=col + 2).value = f"Track {track.track_id} Upper"
            ws.cell(row=1, column=col + 2).font = HEADER_FONT
            ws.cell(row=1, column=col + 2).fill = HEADER_FILL
            for row_idx, val in enumerate(track.upper_limits, 2):
                ws.cell(row=row_idx, column=col + 2).value = val

        if track.lower_limits:
            ws.cell(row=1, column=col + 3).value = f"Track {track.track_id} Lower"
            ws.cell(row=1, column=col + 3).font = HEADER_FONT
            ws.cell(row=1, column=col + 3).fill = HEADER_FILL
            for row_idx, val in enumerate(track.lower_limits, 2):
                ws.cell(row=row_idx, column=col + 3).value = val

        col += 5  # Move to next track columns


def _create_batch_summary_sheet(wb: "Workbook", results: List[AnalysisResult]) -> None:
    """Create batch summary sheet."""
    ws = wb.create_sheet("Batch Summary", 0)

    # Title
    ws["A1"] = "Batch Analysis Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # Summary stats - properly count each status type
    total = len(results)
    passed = sum(1 for r in results if r.overall_status == AnalysisStatus.PASS)
    warnings = sum(1 for r in results if r.overall_status == AnalysisStatus.WARNING)
    failed = sum(1 for r in results if r.overall_status == AnalysisStatus.FAIL)
    errors = sum(1 for r in results if r.overall_status == AnalysisStatus.ERROR)
    pass_rate = (passed / total * 100) if total > 0 else 0

    stats = [
        ("Total Files:", total),
        ("Passed:", passed),
        ("Warnings:", f"{warnings} (partial pass - e.g., pass linearity but fail sigma)"),
        ("Failed:", failed),
        ("Errors:", errors),
        ("Pass Rate:", f"{pass_rate:.1f}%"),
        ("Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    row = 3
    for label, value in stats:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Adjust widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 50


def _create_all_results_sheet(wb: "Workbook", results: List[AnalysisResult]) -> None:
    """Create sheet with all results in tabular format."""
    ws = wb.create_sheet("All Results")

    # Headers - filename at the end for better readability
    # For multi-track: (max)=worst case, (min)=most restrictive, (sum)=total across tracks
    headers = [
        "Model", "Serial", "System", "Trim Date", "Status",
        "Tracks", "Sigma Gradient (max)", "Sigma Threshold (min)", "Sigma Margin %",
        "Sigma Pass", "Linearity Error (max)", "Linearity Spec", "Fail Points (sum)",
        "Linearity Pass", "Risk Category (max)", "Failure Prob (max)", "Travel Length",
        "Untrimmed R", "Trimmed R", "R Change %",
        "Filename"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, result in enumerate(results, 2):
        # Get track data - for multi-track files, aggregate appropriately
        sigma_gradient = 0
        sigma_threshold = 0
        sigma_margin_pct = 0
        linearity_error = 0
        linearity_spec = 0
        fail_points = 0
        travel_length = 0
        sigma_pass = True
        linearity_pass = True
        risk = "Unknown"
        fail_prob = None
        untrimmed_r = None
        trimmed_r = None
        r_change_pct = None

        if result.tracks:
            num_tracks = len(result.tracks)

            # For single track: use direct values
            # For multi-track: use max sigma (worst case), sum fail points
            if num_tracks == 1:
                track = result.tracks[0]
                sigma_gradient = track.sigma_gradient
                sigma_threshold = track.sigma_threshold
                linearity_error = track.linearity_error
                linearity_spec = track.linearity_spec
                fail_points = track.linearity_fail_points
                travel_length = track.travel_length
                sigma_pass = track.sigma_pass
                linearity_pass = track.linearity_pass
                risk = track.risk_category.value
                fail_prob = track.failure_probability
                untrimmed_r = track.untrimmed_resistance
                trimmed_r = track.trimmed_resistance
                r_change_pct = track.resistance_change_percent
            else:
                # Multi-track: use worst-case values for sigma, sum for fail points
                sigma_gradient = max(t.sigma_gradient for t in result.tracks)
                sigma_threshold = min(t.sigma_threshold for t in result.tracks)  # Most restrictive
                linearity_error = max(t.linearity_error for t in result.tracks)
                linearity_spec = result.tracks[0].linearity_spec  # Should be same for all
                fail_points = sum(t.linearity_fail_points for t in result.tracks)
                travel_length = result.tracks[0].travel_length  # Should be same
                sigma_pass = all(t.sigma_pass for t in result.tracks)
                linearity_pass = all(t.linearity_pass for t in result.tracks)
                # Use worst risk category
                risk_order = {"Low": 0, "Medium": 1, "High": 2, "Unknown": -1}
                risk = max([t.risk_category.value for t in result.tracks], key=lambda x: risk_order.get(x, -1))
                # Max failure probability (worst case)
                probs = [t.failure_probability for t in result.tracks if t.failure_probability is not None]
                fail_prob = max(probs) if probs else None
                # Use first track's resistance (typically same for multi-track units)
                untrimmed_r = result.tracks[0].untrimmed_resistance
                trimmed_r = result.tracks[0].trimmed_resistance
                r_change_pct = result.tracks[0].resistance_change_percent

            # Calculate sigma margin percentage
            if sigma_threshold > 0:
                sigma_margin_pct = ((sigma_threshold - sigma_gradient) / sigma_threshold) * 100

        # Use file_date which is the trim date (we set file_date = test_date in parser)
        trim_date = result.metadata.file_date or result.metadata.test_date

        values = [
            result.metadata.model,
            result.metadata.serial,
            result.metadata.system.value,
            trim_date.strftime("%Y-%m-%d") if trim_date else "",
            result.overall_status.value,
            len(result.tracks),
            f"{sigma_gradient:.6f}",
            f"{sigma_threshold:.6f}",
            f"{sigma_margin_pct:.1f}%",
            "PASS" if sigma_pass else "FAIL",
            f"{linearity_error:.6f}",
            f"{linearity_spec:.6f}",
            fail_points,
            "PASS" if linearity_pass else "FAIL",
            risk,
            f"{fail_prob:.1%}" if fail_prob is not None else "N/A",
            f"{travel_length:.1f}" if travel_length else "",
            f"{untrimmed_r:.2f}" if untrimmed_r is not None else "",
            f"{trimmed_r:.2f}" if trimmed_r is not None else "",
            f"{r_change_pct:.2f}%" if r_change_pct is not None else "",
            result.metadata.filename,  # Filename at end
        ]

        status_col = 5  # Status is now column 5

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = value
            cell.border = THIN_BORDER

            # Color status column
            if col == status_col:
                if value == "Pass":
                    cell.fill = PASS_FILL
                elif value == "Fail" or value == "Error":
                    cell.fill = FAIL_FILL
                elif value == "Warning":
                    cell.fill = WARNING_FILL

    # Auto-filter
    num_cols = len(headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{len(results) + 1}"

    # Adjust column widths (21 columns now)
    widths = [12, 12, 8, 12, 8, 6, 14, 14, 12, 10, 14, 12, 10, 12, 12, 10, 12, 12, 12, 10, 35]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def generate_export_filename(
    result: AnalysisResult,
    suffix: str = "",
    extension: str = ".xlsx"
) -> str:
    """Generate a descriptive export filename."""
    base = f"{result.metadata.model}_{result.metadata.serial}"
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    status = result.overall_status.value

    if suffix:
        return f"{base}_{status}_{suffix}_{date_str}{extension}"
    return f"{base}_{status}_{date_str}{extension}"


def generate_batch_export_filename(
    results: List[AnalysisResult],
    prefix: str = "batch_export"
) -> str:
    """Generate filename for batch export."""
    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    count = len(results)
    return f"{prefix}_{count}_files_{date_str}.xlsx"
