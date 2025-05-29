"""
Trend Reporter for Laser Trim AI System

This module generates comprehensive trend reports from historical data,
providing insights and visualizations for continuous improvement.

Author: Laser Trim AI System
Date: 2024
"""

import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from database.database_manager import DatabaseManager
from database.historical_analyzer import HistoricalAnalyzer
from core.config import Config


class TrendReporter:
    """Generates trend reports and visualizations from historical data."""

    def __init__(self, db_manager: DatabaseManager, analyzer: HistoricalAnalyzer, config: Config):
        """
        Initialize trend reporter.

        Args:
            db_manager: Database manager instance
            analyzer: Historical analyzer instance
            config: System configuration
        """
        self.db = db_manager
        self.analyzer = analyzer
        self.config = config
        self.logger = logging.getLogger(__name__)

        # Set visualization style
        plt.style.use('seaborn-v0_8-darkgrid')
        sns.set_palette("husl")

    def generate_comprehensive_report(self, output_dir: Path, days_back: int = 30) -> Path:
        """
        Generate comprehensive trend report with visualizations.

        Args:
            output_dir: Directory to save report
            days_back: Number of days to analyze

        Returns:
            Path to generated report
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        report_name = f'trend_report_{timestamp}.xlsx'
        report_path = output_dir / report_name

        # Create workbook
        wb = Workbook()

        # Generate report sections
        self._add_summary_sheet(wb, days_back)
        self._add_model_trends_sheet(wb, days_back)
        self._add_anomaly_analysis_sheet(wb, days_back)
        self._add_ml_performance_sheet(wb)
        self._add_recommendations_sheet(wb, days_back)

        # Save workbook
        wb.save(report_path)

        # Generate visualizations
        viz_dir = output_dir / f'trend_visualizations_{timestamp}'
        viz_dir.mkdir(exist_ok=True)
        self._generate_visualizations(viz_dir, days_back)

        self.logger.info(f"Generated trend report: {report_path}")
        return report_path

    def _add_summary_sheet(self, wb: Workbook, days_back: int):
        """Add executive summary sheet."""
        ws = wb.active
        ws.title = "Executive Summary"

        # Header
        ws['A1'] = "Laser Trim AI System - Trend Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Period: Last {days_back} days"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Get summary statistics
        stats = self._get_summary_statistics(days_back)

        # Key metrics
        row = 5
        ws[f'A{row}'] = "KEY METRICS"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 2
        metrics = [
            ("Total Units Analyzed", stats['total_units']),
            ("Overall Pass Rate", f"{stats['overall_pass_rate']:.1%}"),
            ("Average Sigma Gradient", f"{stats['avg_sigma_gradient']:.4f}"),
            ("High Risk Units", f"{stats['high_risk_count']} ({stats['high_risk_percentage']:.1%})"),
            ("Models Analyzed", stats['unique_models']),
            ("Total Anomalies Detected", stats['total_anomalies'])
        ]

        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'C{row}'] = str(value)
            row += 1

        # Trend summary
        row += 2
        ws[f'A{row}'] = "TREND SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 2
        ws[f'A{row}'] = "• " + stats['trend_summary']

        # Format cells
        self._format_summary_cells(ws)

    def _get_summary_statistics(self, days_back: int) -> Dict[str, Any]:
        """Get summary statistics for the report period."""
        df = self.db.get_historical_data(days_back=days_back)

        if df.empty:
            return {
                'total_units': 0,
                'overall_pass_rate': 0,
                'avg_sigma_gradient': 0,
                'high_risk_count': 0,
                'high_risk_percentage': 0,
                'unique_models': 0,
                'total_anomalies': 0,
                'trend_summary': 'No data available for analysis'
            }

        # Calculate statistics
        total_units = len(df)
        pass_rate = (df['sigma_pass'] & df['linearity_pass']).mean()
        avg_sigma = df['sigma_gradient'].mean()
        high_risk_count = len(df[df['risk_category'] == 'High'])
        high_risk_pct = high_risk_count / total_units if total_units > 0 else 0

        # Get anomaly count
        anomaly_summary = self.db.get_anomaly_summary(days_back)
        total_anomalies = sum(anomaly_summary['by_type'].values())

        # Determine trend
        if len(df) > 7:
            recent_pass_rate = df.tail(len(df) // 3)['sigma_pass'].mean()
            older_pass_rate = df.head(len(df) // 3)['sigma_pass'].mean()

            if recent_pass_rate > older_pass_rate + 0.05:
                trend = "Performance is improving"
            elif recent_pass_rate < older_pass_rate - 0.05:
                trend = "Performance is declining - attention needed"
            else:
                trend = "Performance is stable"
        else:
            trend = "Insufficient data for trend analysis"

        return {
            'total_units': total_units,
            'overall_pass_rate': pass_rate,
            'avg_sigma_gradient': avg_sigma,
            'high_risk_count': high_risk_count,
            'high_risk_percentage': high_risk_pct,
            'unique_models': df['model'].nunique(),
            'total_anomalies': total_anomalies,
            'trend_summary': trend
        }

    def _add_model_trends_sheet(self, wb: Workbook, days_back: int):
        """Add model-specific trends sheet."""
        ws = wb.create_sheet("Model Trends")

        # Get unique models
        df = self.db.get_historical_data(days_back=days_back)
        if df.empty:
            ws['A1'] = "No data available"
            return

        models = df['model'].value_counts().head(10).index.tolist()

        # Header
        ws['A1'] = "Model Performance Trends"
        ws['A1'].font = Font(bold=True, size=14)

        # Create comparison table
        row = 3
        headers = ['Model', 'Units', 'Pass Rate', 'Avg Sigma', 'Trend', 'Risk Assessment']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)

        row += 1
        for model in models:
            model_df = df[df['model'] == model]
            pass_rate = (model_df['sigma_pass'] & model_df['linearity_pass']).mean()

            # Get trend
            trend_analysis = self.analyzer.analyze_model_trends(model, days_back)
            if 'pass_rate_trend' in trend_analysis and trend_analysis['pass_rate_trend'].get('direction'):
                trend = trend_analysis['pass_rate_trend']['direction'].capitalize()
            else:
                trend = 'Stable'

            # Risk assessment
            high_risk_rate = (model_df['risk_category'] == 'High').mean()
            if high_risk_rate > 0.1:
                risk = 'High'
            elif high_risk_rate > 0.05:
                risk = 'Medium'
            else:
                risk = 'Low'

            ws.cell(row=row, column=1, value=model)
            ws.cell(row=row, column=2, value=len(model_df))
            ws.cell(row=row, column=3, value=f"{pass_rate:.1%}")
            ws.cell(row=row, column=4, value=f"{model_df['sigma_gradient'].mean():.4f}")
            ws.cell(row=row, column=5, value=trend)
            ws.cell(row=row, column=6, value=risk)

            row += 1

        # Add chart
        self._add_trend_chart(ws, df, models[:5], row + 2)

    def _add_trend_chart(self, ws, df: pd.DataFrame, models: List[str], start_row: int):
        """Add trend chart to worksheet."""
        # Prepare daily data
        chart_data = []
        for model in models:
            model_df = df[df['model'] == model].copy()
            model_df['date'] = pd.to_datetime(model_df['timestamp']).dt.date
            daily = model_df.groupby('date')['sigma_pass'].mean()

            for date, pass_rate in daily.items():
                chart_data.append({
                    'Date': date,
                    'Model': model,
                    'Pass Rate': pass_rate
                })

        chart_df = pd.DataFrame(chart_data)

        # Write data to sheet
        ws[f'A{start_row}'] = "Daily Pass Rates"
        ws[f'A{start_row}'].font = Font(bold=True)

        # Pivot for chart
        pivot = chart_df.pivot(index='Date', columns='Model', values='Pass Rate')

        # Write headers
        row = start_row + 2
        ws.cell(row=row, column=1, value='Date')
        for col, model in enumerate(pivot.columns, 2):
            ws.cell(row=row, column=col, value=model)

        # Write data
        for idx, date in enumerate(pivot.index):
            row = start_row + 3 + idx
            ws.cell(row=row, column=1, value=date)
            for col, model in enumerate(pivot.columns, 2):
                value = pivot.loc[date, model]
                if pd.notna(value):
                    ws.cell(row=row, column=col, value=value)

    def _add_anomaly_analysis_sheet(self, wb: Workbook, days_back: int):
        """Add anomaly analysis sheet."""
        ws = wb.create_sheet("Anomaly Analysis")

        # Get anomaly data
        anomaly_summary = self.db.get_anomaly_summary(days_back)
        clusters = self.analyzer.detect_anomaly_clusters(days_back)

        # Header
        ws['A1'] = "Anomaly Analysis"
        ws['A1'].font = Font(bold=True, size=14)

        # Summary
        row = 3
        ws[f'A{row}'] = "Anomaly Summary"
        ws[f'A{row}'].font = Font(bold=True)

        row += 2
        ws[f'A{row}'] = "By Type:"
        row += 1
        for atype, count in anomaly_summary['by_type'].items():
            ws[f'A{row}'] = f"  • {atype}: {count}"
            row += 1

        row += 1
        ws[f'A{row}'] = "By Severity:"
        row += 1
        for severity, count in anomaly_summary['by_severity'].items():
            ws[f'A{row}'] = f"  • {severity}: {count}"
            row += 1

        # Clusters
        row += 2
        ws[f'A{row}'] = "Anomaly Clusters"
        ws[f'A{row}'].font = Font(bold=True)

        row += 2
        ws[f'A{row}'] = clusters.get('interpretation', 'No cluster analysis available')

        if clusters.get('clusters'):
            row += 2
            for cluster in clusters['clusters']:
                ws[f'A{row}'] = f"Cluster {cluster['cluster_id']}: {cluster['size']} anomalies"
                ws[f'B{row}'] = f"Models: {', '.join(cluster['models_affected'])}"
                row += 1

    def _add_ml_performance_sheet(self, wb: Workbook):
        """Add ML model performance sheet."""
        ws = wb.create_sheet("ML Performance")

        # Get ML accuracy data
        accuracy_stats = self.db.get_ml_model_accuracy()

        # Header
        ws['A1'] = "Machine Learning Model Performance"
        ws['A1'].font = Font(bold=True, size=14)

        # Create table
        row = 3
        headers = ['Model Name', 'Accuracy', 'Predictions', 'Avg Confidence']
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)

        row += 1
        for model_name, stats in accuracy_stats.items():
            ws.cell(row=row, column=1, value=model_name)
            ws.cell(row=row, column=2, value=f"{stats['accuracy']:.1%}")
            ws.cell(row=row, column=3, value=stats['total_predictions'])
            ws.cell(row=row, column=4, value=f"{stats['avg_confidence']:.1%}")
            row += 1

        # Add interpretation
        row += 2
        ws[f'A{row}'] = "Performance Summary:"
        row += 1

        if accuracy_stats:
            avg_accuracy = np.mean([s['accuracy'] for s in accuracy_stats.values()])
            ws[f'A{row}'] = f"• Average model accuracy: {avg_accuracy:.1%}"
            row += 1

            if avg_accuracy > 0.9:
                ws[f'A{row}'] = "• ML models are performing well"
            elif avg_accuracy > 0.8:
                ws[f'A{row}'] = "• ML models show good performance with room for improvement"
            else:
                ws[f'A{row}'] = "• ML models need retraining with more data"
        else:
            ws[f'A{row}'] = "• No ML performance data available"

    def _add_recommendations_sheet(self, wb: Workbook, days_back: int):
        """Add recommendations sheet."""
        ws = wb.create_sheet("Recommendations")

        # Get top models by volume
        df = self.db.get_historical_data(days_back=days_back)
        if df.empty:
            ws['A1'] = "No data available for recommendations"
            return

        top_models = df['model'].value_counts().head(5).index.tolist()

        # Header
        ws['A1'] = "Improvement Recommendations"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        for model in top_models:
            recommendations = self.analyzer.generate_improvement_recommendations(model)

            if recommendations:
                ws[f'A{row}'] = f"Model {model}"
                ws[f'A{row}'].font = Font(bold=True, size=12)
                row += 1

                for rec in recommendations:
                    ws[f'A{row}'] = f"• {rec.get('recommendation', 'No recommendation')}"
                    ws[f'B{row}'] = f"Priority: {rec.get('priority', 'unknown')}"
                    ws[f'C{row}'] = f"Impact: {rec.get('expected_impact', 'unknown')}"
                    row += 1

                row += 1

    def _format_summary_cells(self, ws):
        """Format summary worksheet cells."""
        # Apply formatting
        for row in ws.iter_rows(min_row=5, max_row=20, min_col=1, max_col=3):
            for cell in row:
                if cell.column == 1 and cell.value and isinstance(cell.value, str) and not cell.value.startswith('•'):
                    cell.font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['C'].width = 20

    def _generate_visualizations(self, viz_dir: Path, days_back: int):
        """Generate visualization images."""
        # Overall trend plot
        self._plot_overall_trend(viz_dir / 'overall_trend.png', days_back)

        # Model comparison
        self._plot_model_comparison(viz_dir / 'model_comparison.png', days_back)

        # Anomaly heatmap
        self._plot_anomaly_heatmap(viz_dir / 'anomaly_heatmap.png', days_back)

        # Risk distribution
        self._plot_risk_distribution(viz_dir / 'risk_distribution.png', days_back)

    def _plot_overall_trend(self, output_path: Path, days_back: int):
        """Plot overall performance trend."""
        df = self.db.get_historical_data(days_back=days_back)
        if df.empty:
            return

        # Prepare daily data
        df['date'] = pd.to_datetime(df['timestamp']).dt.date
        daily = df.groupby('date').agg({
            'sigma_pass': 'mean',
            'sigma_gradient': 'mean',
            'risk_category': lambda x: (x == 'High').mean()
        })

        # Create figure
        fig, axes = plt.subplots(3, 1, figsize=(12, 10), sharex=True)

        # Pass rate
        axes[0].plot(daily.index, daily['sigma_pass'] * 100, 'b-', linewidth=2)
        axes[0].set_ylabel('Pass Rate (%)')
        axes[0].set_title('Daily Performance Trends')
        axes[0].grid(True, alpha=0.3)

        # Sigma gradient
        axes[1].plot(daily.index, daily['sigma_gradient'], 'g-', linewidth=2)
        axes[1].axhline(y=self.config.SIGMA_THRESHOLD, color='r', linestyle='--', label='Threshold')
        axes[1].set_ylabel('Avg Sigma Gradient')
        axes[1].legend()
        axes[1].grid(True, alpha=0.3)

        # High risk rate
        axes[2].plot(daily.index, daily['risk_category'] * 100, 'r-', linewidth=2)
        axes[2].set_ylabel('High Risk Rate (%)')
        axes[2].set_xlabel('Date')
        axes[2].grid(True, alpha=0.3)

        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

    def _plot_model_comparison(self, output_path: Path, days_back: int):
        """Plot model comparison."""
        df = self.db.get_historical_data(days_back=days_back)
        if df.empty:
            return

        # Get top models
        top_models = df['model'].value_counts().head(5).index.tolist()

        # Calculate metrics for each model
        model_metrics = []
        for model in top_models:
            model_df = df[df['model'] == model]
            model_metrics.append({
                'model': model,
                'pass_rate': (model_df['sigma_pass'] & model_df['linearity_pass']).mean() * 100,
                'avg_sigma': model_df['sigma_gradient'].mean(),
                'high_risk_rate': (model_df['risk_category'] == 'High').mean() * 100
            })

        metrics_df = pd.DataFrame(model_metrics)

        # Create figure
        fig, axes = plt.subplots(1, 3, figsize=(15, 5))

        # Pass rate comparison
        axes[0].bar(metrics_df['model'], metrics_df['pass_rate'])
        axes[0].set_ylabel('Pass Rate (%)')
        axes[0].set_title('Pass Rate by Model')
        axes[0].tick_params(axis='x', rotation=45)

        # Sigma gradient comparison
        axes[1].bar(metrics_df['model'], metrics_df['avg_sigma'])
        axes[1].axhline(y=self.config.SIGMA_THRESHOLD, color='r', linestyle='--', label='Threshold')
        axes[1].set_ylabel('Average Sigma Gradient')
        axes[1].set_title('Sigma Gradient by Model')
        axes[1].tick_params(axis='x', rotation=45)
        axes[1].legend()

        # High risk rate comparison
        axes[2].bar(metrics_df['model'], metrics_df['high_risk_rate'], color='red', alpha=0.7)
        axes[2].set_ylabel('High Risk Rate (%)')
        axes[2].set_title('High Risk Units by Model')
        axes[2].tick_params(axis='x', rotation=45)

        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

    def _plot_anomaly_heatmap(self, output_path: Path, days_back: int):
        """Plot anomaly heatmap."""
        # Get anomaly data with model and date
        query = '''
            SELECT 
                DATE(a.timestamp) as date,
                f.model,
                COUNT(*) as anomaly_count
            FROM anomalies a
            JOIN file_results f ON a.file_id = f.id
            WHERE a.timestamp > datetime('now', ?)
            GROUP BY DATE(a.timestamp), f.model
        '''

        with self.db.get_connection() as conn:
            df = pd.read_sql_query(query, conn, params=(f'-{days_back} days',))

        if df.empty:
            return

        # Pivot for heatmap
        pivot = df.pivot(index='model', columns='date', values='anomaly_count').fillna(0)

        # Create heatmap
        plt.figure(figsize=(12, 8))
        sns.heatmap(pivot, cmap='YlOrRd', annot=True, fmt='.0f', cbar_kws={'label': 'Anomaly Count'})
        plt.title('Anomaly Heatmap by Model and Date')
        plt.xlabel('Date')
        plt.ylabel('Model')
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

    def _plot_risk_distribution(self, output_path: Path, days_back: int):
        """Plot risk distribution."""
        df = self.db.get_historical_data(days_back=days_back)
        if df.empty:
            return

        # Create figure
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

        # Overall risk distribution
        risk_counts = df['risk_category'].value_counts()
        colors = {'Low': 'green', 'Medium': 'yellow', 'High': 'red'}
        ax1.pie(risk_counts.values, labels=risk_counts.index, autopct='%1.1f%%',
                colors=[colors.get(cat, 'gray') for cat in risk_counts.index])
        ax1.set_title('Overall Risk Distribution')

        # Risk distribution over time
        df['date'] = pd.to_datetime(df['timestamp']).dt.date
        risk_by_date = df.groupby(['date', 'risk_category']).size().unstack(fill_value=0)

        if not risk_by_date.empty:
            risk_by_date.plot(kind='area', stacked=True, ax=ax2,
                              color=['green', 'yellow', 'red'], alpha=0.7)
            ax2.set_ylabel('Number of Units')
            ax2.set_xlabel('Date')
            ax2.set_title('Risk Distribution Over Time')
            ax2.legend(title='Risk Category')

        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

    def generate_quick_report(self, model: str, output_path: Path) -> Dict[str, Any]:
        """
        Generate a quick report for a specific model.

        Args:
            model: Model number
            output_path: Path to save report

        Returns:
            Dictionary with report summary
        """
        # Analyze model
        analysis = self.analyzer.analyze_model_trends(model, days_back=30)
        recommendations = self.analyzer.generate_improvement_recommendations(model)
        cost_impact = self.analyzer.calculate_cost_impact(model, days_back=30)

        # Create simple text report
        report_lines = [
            f"Quick Report for Model {model}",
            "=" * 50,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "PERFORMANCE SUMMARY",
            "-" * 20,
            f"Data Points: {analysis.get('data_points', 0)}",
            f"Average Sigma: {analysis.get('sigma_analysis', {}).get('mean', 0):.4f}",
            f"Pass Rate Trend: {analysis.get('pass_rate_trend', {}).get('interpretation', 'Unknown')}",
            "",
            "RECOMMENDATIONS",
            "-" * 20
        ]

        for rec in recommendations[:3]:  # Top 3 recommendations
            report_lines.append(f"• {rec.get('recommendation', 'No recommendation')}")
            report_lines.append(f"  Priority: {rec.get('priority', 'unknown')}")
            report_lines.append(f"  Impact: {rec.get('expected_impact', 'unknown')}")
            report_lines.append("")

        report_lines.extend([
            "COST IMPACT",
            "-" * 20,
            f"Period: {cost_impact.get('period', 'Unknown')}",
            f"Total Cost Impact: ${cost_impact.get('costs', {}).get('total', 0):,.2f}",
            f"Cost per Unit: ${cost_impact.get('cost_per_unit', 0):.2f}",
            "",
            "POTENTIAL SAVINGS",
            "-" * 20,
            f"• Reduce failures by 10%: ${cost_impact.get('potential_savings', {}).get('reduce_failures_10pct', 0):,.2f}",
            f"• Eliminate high risk units: ${cost_impact.get('potential_savings', {}).get('eliminate_high_risk', 0):,.2f}"
        ])

        # Write report
        with open(output_path, 'w') as f:
            f.write('\n'.join(report_lines))

        return {
            'model': model,
            'report_path': str(output_path),
            'summary': {
                'trend': analysis.get('pass_rate_trend', {}).get('interpretation', 'Unknown'),
                'recommendations': len(recommendations),
                'cost_impact': cost_impact.get('costs', {}).get('total', 0)
            }
        }