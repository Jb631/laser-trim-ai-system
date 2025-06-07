"""
Optimized batch file processing system with parallel processing, memory management,
and resource pooling capabilities.
"""

import asyncio
import concurrent.futures
import multiprocessing as mp
import os
import psutil
import queue
import threading
import time
from collections import deque
from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Dict, Iterator, List, Optional, Set, Tuple, Union
import logging
import weakref

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..core.models import ProcessingResult, FileData, TrackData
from ..core.exceptions import ProcessingError, CancellationError
from ..utils.excel_utils import ExcelReader
from ..utils.performance_monitor import PerformanceMonitor

logger = logging.getLogger(__name__)


@dataclass
class BatchConfig:
    """Configuration for batch processing optimization."""
    max_workers: Optional[int] = None
    batch_size: Optional[int] = None
    memory_limit_mb: int = 2048
    chunk_size: int = 10000
    enable_streaming: bool = True
    enable_caching: bool = True
    cache_size_mb: int = 512
    timeout_seconds: int = 300
    retry_attempts: int = 3
    adaptive_parallelism: bool = True
    progress_callback: Optional[Callable[[float, str], None]] = None


@dataclass
class ProcessingStats:
    """Statistics for batch processing operations."""
    total_files: int = 0
    processed_files: int = 0
    failed_files: int = 0
    total_bytes: int = 0
    processed_bytes: int = 0
    start_time: float = field(default_factory=time.time)
    end_time: Optional[float] = None
    peak_memory_mb: float = 0
    average_file_time: float = 0
    
    @property
    def elapsed_time(self) -> float:
        """Get elapsed processing time."""
        end = self.end_time or time.time()
        return end - self.start_time
    
    @property
    def throughput(self) -> float:
        """Get processing throughput in files/second."""
        if self.elapsed_time > 0:
            return self.processed_files / self.elapsed_time
        return 0.0
    
    @property
    def success_rate(self) -> float:
        """Get success rate as percentage."""
        total = self.processed_files + self.failed_files
        if total > 0:
            return (self.processed_files / total) * 100
        return 0.0


class ResourcePool:
    """Thread-safe resource pool for reusable objects."""
    
    def __init__(self, factory: Callable[[], Any], max_size: int = 10):
        self._factory = factory
        self._max_size = max_size
        self._pool: deque = deque()
        self._lock = threading.Lock()
        self._created = 0
        
    @contextmanager
    def acquire(self):
        """Acquire a resource from the pool."""
        resource = None
        
        # Try to get from pool
        with self._lock:
            if self._pool:
                resource = self._pool.popleft()
        
        # Create new if needed
        if resource is None and self._created < self._max_size:
            with self._lock:
                if self._created < self._max_size:
                    resource = self._factory()
                    self._created += 1
        
        # Wait for available resource if at limit
        while resource is None:
            time.sleep(0.01)
            with self._lock:
                if self._pool:
                    resource = self._pool.popleft()
        
        try:
            yield resource
        finally:
            # Return to pool
            with self._lock:
                self._pool.append(resource)


class MemoryManager:
    """Manages memory usage during batch processing."""
    
    def __init__(self, limit_mb: int = 2048):
        self.limit_bytes = limit_mb * 1024 * 1024
        self._current_usage = 0
        self._lock = threading.Lock()
        self._condition = threading.Condition(self._lock)
        
    def request_memory(self, size_bytes: int, timeout: float = 30.0) -> bool:
        """Request memory allocation."""
        deadline = time.time() + timeout
        
        with self._condition:
            while self._current_usage + size_bytes > self.limit_bytes:
                remaining = deadline - time.time()
                if remaining <= 0:
                    return False
                self._condition.wait(timeout=remaining)
            
            self._current_usage += size_bytes
            return True
    
    def release_memory(self, size_bytes: int):
        """Release allocated memory."""
        with self._condition:
            self._current_usage = max(0, self._current_usage - size_bytes)
            self._condition.notify_all()
    
    @contextmanager
    def allocate(self, size_bytes: int):
        """Context manager for memory allocation."""
        acquired = self.request_memory(size_bytes)
        if not acquired:
            raise MemoryError(f"Could not allocate {size_bytes} bytes")
        
        try:
            yield
        finally:
            self.release_memory(size_bytes)


class DataStreamProcessor:
    """Processes data in streaming fashion to minimize memory usage."""
    
    def __init__(self, chunk_size: int = 10000):
        self.chunk_size = chunk_size
        
    def process_excel_stream(self, file_path: Path, 
                           processor: Callable[[pd.DataFrame], Any]) -> Iterator[Any]:
        """Process Excel file in chunks."""
        try:
            # Use openpyxl for memory-efficient reading
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            for sheet in wb.worksheets:
                if self._is_data_sheet(sheet.title):
                    yield from self._process_sheet_chunks(sheet, processor)
                    
            wb.close()
            
        except Exception as e:
            logger.error(f"Error streaming {file_path}: {e}")
            raise ProcessingError(f"Failed to stream {file_path}: {e}")
    
    def _is_data_sheet(self, sheet_name: str) -> bool:
        """Check if sheet contains data."""
        data_indicators = ['data', 'results', 'measurements', 'test']
        return any(indicator in sheet_name.lower() for indicator in data_indicators)
    
    def _process_sheet_chunks(self, sheet: Worksheet, 
                            processor: Callable[[pd.DataFrame], Any]) -> Iterator[Any]:
        """Process worksheet in chunks."""
        rows = []
        header = None
        
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                header = row
                continue
                
            rows.append(row)
            
            if len(rows) >= self.chunk_size:
                df_chunk = pd.DataFrame(rows, columns=header)
                yield processor(df_chunk)
                rows = []
        
        # Process remaining rows
        if rows and header:
            df_chunk = pd.DataFrame(rows, columns=header)
            yield processor(df_chunk)


class AdaptiveWorkerPool:
    """Dynamically adjusts worker count based on system resources."""
    
    def __init__(self, min_workers: int = 1, max_workers: Optional[int] = None):
        self.min_workers = min_workers
        self.max_workers = max_workers or mp.cpu_count()
        self._current_workers = min_workers
        self._executor: Optional[concurrent.futures.ProcessPoolExecutor] = None
        self._lock = threading.Lock()
        self._performance_history = deque(maxlen=10)
        
    def _calculate_optimal_workers(self) -> int:
        """Calculate optimal worker count based on system state."""
        cpu_percent = psutil.cpu_percent(interval=0.1)
        memory_percent = psutil.virtual_memory().percent
        
        # Reduce workers if system is under stress
        if cpu_percent > 80 or memory_percent > 85:
            return max(self.min_workers, self._current_workers - 1)
        
        # Increase workers if resources available
        if cpu_percent < 50 and memory_percent < 60:
            return min(self.max_workers, self._current_workers + 1)
        
        return self._current_workers
    
    def adjust_workers(self):
        """Adjust worker pool size based on performance."""
        optimal = self._calculate_optimal_workers()
        
        if optimal != self._current_workers:
            with self._lock:
                if self._executor:
                    self._executor.shutdown(wait=False)
                
                self._current_workers = optimal
                self._executor = concurrent.futures.ProcessPoolExecutor(
                    max_workers=self._current_workers
                )
                
                logger.info(f"Adjusted worker pool to {self._current_workers} workers")
    
    @property
    def executor(self) -> concurrent.futures.ProcessPoolExecutor:
        """Get the current executor."""
        if self._executor is None:
            with self._lock:
                if self._executor is None:
                    self._executor = concurrent.futures.ProcessPoolExecutor(
                        max_workers=self._current_workers
                    )
        return self._executor
    
    def shutdown(self):
        """Shutdown the worker pool."""
        with self._lock:
            if self._executor:
                self._executor.shutdown(wait=True)
                self._executor = None


class OptimizedBatchProcessor:
    """High-performance batch file processor with advanced optimization features."""
    
    def __init__(self, config: Optional[BatchConfig] = None):
        self.config = config or BatchConfig()
        self._setup_components()
        self._cancellation_event = threading.Event()
        self._progress_lock = threading.Lock()
        self._stats = ProcessingStats()
        
    def _setup_components(self):
        """Initialize processing components."""
        # Memory management
        self.memory_manager = MemoryManager(self.config.memory_limit_mb)
        
        # Worker pool
        self.worker_pool = AdaptiveWorkerPool(
            min_workers=1,
            max_workers=self.config.max_workers
        )
        
        # Resource pools
        self.excel_reader_pool = ResourcePool(
            factory=lambda: ExcelReader(),
            max_size=self.config.max_workers or mp.cpu_count()
        )
        
        # Data streaming
        self.stream_processor = DataStreamProcessor(self.config.chunk_size)
        
        # Performance monitoring
        self.performance_monitor = PerformanceMonitor()
        
        # Result cache
        self._result_cache: Dict[str, ProcessingResult] = {}
        self._cache_lock = threading.Lock()
        
    def process_batch(self, file_paths: List[Path], 
                     process_func: Callable[[Path], ProcessingResult],
                     parallel: bool = True) -> Dict[str, ProcessingResult]:
        """
        Process a batch of files with optimization.
        
        Args:
            file_paths: List of file paths to process
            process_func: Function to process each file
            parallel: Whether to use parallel processing
            
        Returns:
            Dictionary mapping file paths to results
        """
        self._reset_stats(file_paths)
        results = {}
        
        try:
            # Calculate optimal batch size
            batch_size = self._calculate_batch_size(file_paths)
            
            # Process in batches
            for i in range(0, len(file_paths), batch_size):
                if self._cancellation_event.is_set():
                    raise CancellationError("Batch processing cancelled")
                
                batch = file_paths[i:i + batch_size]
                
                if parallel and len(batch) > 1:
                    batch_results = self._process_parallel(batch, process_func)
                else:
                    batch_results = self._process_sequential(batch, process_func)
                
                results.update(batch_results)
                
                # Adjust workers if using adaptive parallelism
                if self.config.adaptive_parallelism and parallel:
                    self.worker_pool.adjust_workers()
        
        finally:
            self._stats.end_time = time.time()
            
        return results
    
    def _calculate_batch_size(self, file_paths: List[Path]) -> int:
        """Calculate optimal batch size based on system resources."""
        if self.config.batch_size:
            return self.config.batch_size
        
        # Estimate based on available memory and file sizes
        total_size = sum(p.stat().st_size for p in file_paths if p.exists())
        avg_size = total_size / len(file_paths) if file_paths else 0
        
        # Assume each file needs 10x its size in memory during processing
        memory_per_file = avg_size * 10
        
        if memory_per_file > 0:
            available_memory = self.memory_manager.limit_bytes
            batch_size = int(available_memory / memory_per_file)
            
            # Constrain to reasonable bounds
            batch_size = max(1, min(batch_size, 100))
        else:
            batch_size = 10
        
        logger.info(f"Calculated batch size: {batch_size}")
        return batch_size
    
    def _process_parallel(self, file_paths: List[Path], 
                         process_func: Callable[[Path], ProcessingResult]) -> Dict[str, ProcessingResult]:
        """Process files in parallel."""
        results = {}
        futures = {}
        
        with self.performance_monitor.measure("parallel_batch"):
            executor = self.worker_pool.executor
            
            # Submit tasks
            for file_path in file_paths:
                if self._cancellation_event.is_set():
                    break
                    
                # Check cache first
                if self.config.enable_caching:
                    cached = self._get_cached_result(str(file_path))
                    if cached:
                        results[str(file_path)] = cached
                        self._update_progress(file_path, success=True)
                        continue
                
                # Submit for processing
                future = executor.submit(self._process_file_wrapped, file_path, process_func)
                futures[future] = file_path
            
            # Collect results
            for future in concurrent.futures.as_completed(futures):
                if self._cancellation_event.is_set():
                    # Cancel remaining futures
                    for f in futures:
                        f.cancel()
                    raise CancellationError("Batch processing cancelled")
                
                file_path = futures[future]
                
                try:
                    result = future.result(timeout=self.config.timeout_seconds)
                    results[str(file_path)] = result
                    
                    # Cache result
                    if self.config.enable_caching:
                        self._cache_result(str(file_path), result)
                    
                    self._update_progress(file_path, success=True)
                    
                except Exception as e:
                    logger.error(f"Failed to process {file_path}: {e}")
                    self._update_progress(file_path, success=False)
                    
                    # Retry if configured
                    if self.config.retry_attempts > 0:
                        results[str(file_path)] = self._retry_processing(
                            file_path, process_func, self.config.retry_attempts
                        )
        
        return results
    
    def _process_sequential(self, file_paths: List[Path], 
                          process_func: Callable[[Path], ProcessingResult]) -> Dict[str, ProcessingResult]:
        """Process files sequentially."""
        results = {}
        
        with self.performance_monitor.measure("sequential_batch"):
            for file_path in file_paths:
                if self._cancellation_event.is_set():
                    raise CancellationError("Batch processing cancelled")
                
                # Check cache first
                if self.config.enable_caching:
                    cached = self._get_cached_result(str(file_path))
                    if cached:
                        results[str(file_path)] = cached
                        self._update_progress(file_path, success=True)
                        continue
                
                try:
                    result = self._process_file_wrapped(file_path, process_func)
                    results[str(file_path)] = result
                    
                    # Cache result
                    if self.config.enable_caching:
                        self._cache_result(str(file_path), result)
                    
                    self._update_progress(file_path, success=True)
                    
                except Exception as e:
                    logger.error(f"Failed to process {file_path}: {e}")
                    self._update_progress(file_path, success=False)
                    
                    # Retry if configured
                    if self.config.retry_attempts > 0:
                        results[str(file_path)] = self._retry_processing(
                            file_path, process_func, self.config.retry_attempts
                        )
        
        return results
    
    def _process_file_wrapped(self, file_path: Path, 
                            process_func: Callable[[Path], ProcessingResult]) -> ProcessingResult:
        """Wrapper for file processing with resource management."""
        file_size = file_path.stat().st_size if file_path.exists() else 0
        
        # Allocate memory for processing
        with self.memory_manager.allocate(file_size * 10):  # Assume 10x size needed
            # Use streaming if enabled and file is large
            if self.config.enable_streaming and file_size > 50 * 1024 * 1024:  # 50MB
                return self._process_file_streaming(file_path, process_func)
            else:
                return process_func(file_path)
    
    def _process_file_streaming(self, file_path: Path, 
                              process_func: Callable[[Path], ProcessingResult]) -> ProcessingResult:
        """Process file using streaming to minimize memory usage."""
        # This is a simplified example - actual implementation would depend on
        # the specific processing requirements
        logger.info(f"Processing {file_path} in streaming mode")
        
        # For now, fall back to regular processing
        # In a real implementation, this would process the file in chunks
        return process_func(file_path)
    
    def _retry_processing(self, file_path: Path, process_func: Callable[[Path], ProcessingResult],
                         attempts: int) -> Optional[ProcessingResult]:
        """Retry failed processing with exponential backoff."""
        for attempt in range(attempts):
            try:
                time.sleep(2 ** attempt)  # Exponential backoff
                return self._process_file_wrapped(file_path, process_func)
            except Exception as e:
                if attempt == attempts - 1:
                    logger.error(f"Failed to process {file_path} after {attempts} attempts: {e}")
                    return None
        
        return None
    
    def _reset_stats(self, file_paths: List[Path]):
        """Reset processing statistics."""
        self._stats = ProcessingStats(
            total_files=len(file_paths),
            total_bytes=sum(p.stat().st_size for p in file_paths if p.exists())
        )
    
    def _update_progress(self, file_path: Path, success: bool):
        """Update processing progress."""
        with self._progress_lock:
            if success:
                self._stats.processed_files += 1
                self._stats.processed_bytes += file_path.stat().st_size if file_path.exists() else 0
            else:
                self._stats.failed_files += 1
            
            # Update memory usage
            current_memory = psutil.Process().memory_info().rss / 1024 / 1024
            self._stats.peak_memory_mb = max(self._stats.peak_memory_mb, current_memory)
            
            # Calculate progress
            total_processed = self._stats.processed_files + self._stats.failed_files
            progress = total_processed / self._stats.total_files if self._stats.total_files > 0 else 0
            
            # Notify progress callback
            if self.config.progress_callback:
                self.config.progress_callback(
                    progress,
                    f"Processed {total_processed}/{self._stats.total_files} files"
                )
    
    def _get_cached_result(self, file_path: str) -> Optional[ProcessingResult]:
        """Get cached result if available."""
        with self._cache_lock:
            return self._result_cache.get(file_path)
    
    def _cache_result(self, file_path: str, result: ProcessingResult):
        """Cache processing result."""
        with self._cache_lock:
            # Simple cache size management
            cache_size = sum(
                len(str(r)) for r in self._result_cache.values()
            )
            
            # Clear cache if too large
            if cache_size > self.config.cache_size_mb * 1024 * 1024:
                self._result_cache.clear()
            
            self._result_cache[file_path] = result
    
    def cancel(self):
        """Cancel ongoing processing."""
        self._cancellation_event.set()
    
    def get_stats(self) -> ProcessingStats:
        """Get current processing statistics."""
        return self._stats
    
    def cleanup(self):
        """Cleanup resources."""
        self.worker_pool.shutdown()
        self._result_cache.clear()
        

class BatchProcessingPipeline:
    """
    High-level pipeline for batch processing with optimizations.
    """
    
    def __init__(self, config: Optional[BatchConfig] = None):
        self.config = config or BatchConfig()
        self.processor = OptimizedBatchProcessor(config)
        
    async def process_directory_async(self, directory: Path, 
                                    pattern: str = "*.xls*",
                                    process_func: Callable[[Path], ProcessingResult]) -> Dict[str, ProcessingResult]:
        """Process all matching files in a directory asynchronously."""
        file_paths = list(directory.glob(pattern))
        
        # Sort by size for better load balancing
        file_paths.sort(key=lambda p: p.stat().st_size if p.exists() else 0)
        
        # Process in background
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            None,
            self.processor.process_batch,
            file_paths,
            process_func,
            True
        )
    
    def create_processing_stream(self, file_paths: List[Path],
                               process_func: Callable[[Path], ProcessingResult]) -> Iterator[Tuple[Path, ProcessingResult]]:
        """Create a stream of processed results."""
        for file_path in file_paths:
            if self.processor._cancellation_event.is_set():
                break
                
            try:
                result = process_func(file_path)
                yield file_path, result
            except Exception as e:
                logger.error(f"Error processing {file_path}: {e}")
                yield file_path, None
    
    def optimize_for_memory(self):
        """Optimize settings for memory-constrained environments."""
        self.config.enable_streaming = True
        self.config.batch_size = 5
        self.config.chunk_size = 1000
        self.config.cache_size_mb = 128
        
    def optimize_for_speed(self):
        """Optimize settings for maximum speed."""
        self.config.enable_streaming = False
        self.config.batch_size = 50
        self.config.max_workers = mp.cpu_count()
        self.config.cache_size_mb = 1024
        
    def cleanup(self):
        """Cleanup pipeline resources."""
        self.processor.cleanup()


# Utility functions for common batch operations

def create_batch_processor(memory_limit_mb: int = 2048,
                         max_workers: Optional[int] = None,
                         enable_caching: bool = True) -> OptimizedBatchProcessor:
    """Create a configured batch processor."""
    config = BatchConfig(
        memory_limit_mb=memory_limit_mb,
        max_workers=max_workers,
        enable_caching=enable_caching,
        adaptive_parallelism=True
    )
    
    return OptimizedBatchProcessor(config)


def process_files_in_parallel(file_paths: List[Path],
                            process_func: Callable[[Path], ProcessingResult],
                            progress_callback: Optional[Callable[[float, str], None]] = None) -> Dict[str, ProcessingResult]:
    """Convenience function for parallel file processing."""
    config = BatchConfig(
        progress_callback=progress_callback,
        adaptive_parallelism=True
    )
    
    processor = OptimizedBatchProcessor(config)
    
    try:
        return processor.process_batch(file_paths, process_func, parallel=True)
    finally:
        processor.cleanup()


def estimate_processing_time(file_paths: List[Path],
                           avg_file_time: float = 1.0) -> float:
    """Estimate total processing time for a batch of files."""
    total_size = sum(p.stat().st_size for p in file_paths if p.exists())
    avg_size = total_size / len(file_paths) if file_paths else 0
    
    # Adjust time estimate based on file size
    size_factor = (avg_size / (1024 * 1024)) / 10  # Normalize to 10MB
    adjusted_time = avg_file_time * max(0.5, size_factor)
    
    # Account for parallelism
    num_workers = mp.cpu_count()
    parallel_time = (len(file_paths) * adjusted_time) / num_workers
    
    # Add overhead
    overhead = len(file_paths) * 0.1  # 0.1s per file overhead
    
    return parallel_time + overhead